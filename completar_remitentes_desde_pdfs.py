#!/usr/bin/env python3
"""Completa Solicitante Promovente / Representante Legal desde PDFs de descargas.

Esta etapa se ejecuta después del procesamiento SATyS y ANTES de la
reconciliación global. Sólo actúa sobre filas de ``TrámitesCRT.xlsx`` donde
alguno de estos campos esté vacío o sea ``SIN REMITENTE``:

- ``Solicitante Promovente``
- ``Representante Legal``

La columna ``1711`` identifica el expediente. Se concilia contra las carpetas
originales de ``descargas`` (incluidos Internos), se recorren recursivamente
TODOS sus PDF y se reutiliza la extracción tolerante de ``extraer_operador.py``.

Reglas de seguridad:
- nunca sobrescribe un valor válido ya existente en el Excel;
- si varios PDF entregan valores incompatibles para el mismo campo, no elige
  arbitrariamente: deja el campo pendiente y registra el conflicto;
- PDFs sin capa de texto quedan auditados, pero no se inventa OCR/valores;
- ``descargas`` es sólo lectura;
- el Excel se sustituye de forma atómica y con backup cuando hay cambios.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import shutil
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl

import extraer_operador as extractor
import resolver_sin_operador_rpc_publico as sor
from configuracion_local import ruta_configurada
from guardado_seguro import reemplazar_desde_temporal
from proceso_lock import ProcesoLock, LockOcupadoError

log = logging.getLogger("SATyS-Remitentes-PDF")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)

DEFAULT_EXCEL = ruta_configurada("excel", "TrámitesCRT.xlsx")
DEFAULT_DESCARGAS = ruta_configurada("descargas", "descargas")
DEFAULT_LOGS = ruta_configurada("logs", "logs")
SHEETS_DEFAULT = ("Turnados recibidos", "Internos")


def _texto(value: Any) -> str:
    return str(value or "").strip()


def _normalizar(value: Any) -> str:
    text = unicodedata.normalize("NFD", _texto(value))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"[^A-Z0-9]+", " ", text.upper())
    return re.sub(r"\s+", " ", text).strip()


def _es_faltante(value: Any) -> bool:
    return _normalizar(value) in {"", "SIN REMITENTE"}


def _filas_reales(ws, max_col: int) -> list[int]:
    cells = getattr(ws, "_cells", None)
    if isinstance(cells, dict):
        return sorted({
            row
            for (row, col), cell in cells.items()
            if row >= 2 and col <= max_col and cell.value not in (None, "")
        })
    return list(range(2, ws.max_row + 1))


@dataclass
class Pendiente:
    sheet: str
    row: int
    identificador: str
    col_solicitante: int
    col_representante: int
    solicitante_anterior: str
    representante_anterior: str


@dataclass
class ResultadoFila:
    sheet: str
    row: int
    identificador: str
    estado: str = "pendiente"
    motivo: str = ""
    fuentes: list[str] = field(default_factory=list)
    pdfs_revisados: int = 0
    pdfs_sin_texto: int = 0
    pdfs_error: int = 0
    solicitante_anterior: str = ""
    solicitante_nuevo: str = ""
    solicitante_fuentes: list[str] = field(default_factory=list)
    solicitante_conflictos: list[str] = field(default_factory=list)
    representante_anterior: str = ""
    representante_nuevo: str = ""
    representante_fuentes: list[str] = field(default_factory=list)
    representante_conflictos: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row": self.row,
            "identificador": self.identificador,
            "estado": self.estado,
            "motivo": self.motivo,
            "fuentes": self.fuentes,
            "pdfs_revisados": self.pdfs_revisados,
            "pdfs_sin_texto": self.pdfs_sin_texto,
            "pdfs_error": self.pdfs_error,
            "solicitante_anterior": self.solicitante_anterior,
            "solicitante_nuevo": self.solicitante_nuevo,
            "solicitante_fuentes": self.solicitante_fuentes,
            "solicitante_conflictos": self.solicitante_conflictos,
            "representante_anterior": self.representante_anterior,
            "representante_nuevo": self.representante_nuevo,
            "representante_fuentes": self.representante_fuentes,
            "representante_conflictos": self.representante_conflictos,
        }


def leer_pendientes(wb) -> list[Pendiente]:
    pendientes: list[Pendiente] = []
    for sheet_name in SHEETS_DEFAULT:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {
            _normalizar(cell.value): cell.column
            for cell in ws[1]
            if _texto(cell.value)
        }
        col_id = headers.get("1711")
        col_sol = headers.get("SOLICITANTE PROMOVENTE")
        col_rep = headers.get("REPRESENTANTE LEGAL")
        if not col_id or not col_sol or not col_rep:
            log.warning(
                "Hoja %s sin 1711/Solicitante Promovente/Representante Legal; se omite.",
                sheet_name,
            )
            continue
        max_col = max(col_id, col_sol, col_rep)
        for row in _filas_reales(ws, max_col):
            identificador = _texto(ws.cell(row=row, column=col_id).value)
            if not identificador:
                continue
            solicitante = _texto(ws.cell(row=row, column=col_sol).value)
            representante = _texto(ws.cell(row=row, column=col_rep).value)
            if not (_es_faltante(solicitante) or _es_faltante(representante)):
                continue
            pendientes.append(Pendiente(
                sheet=sheet_name,
                row=row,
                identificador=identificador,
                col_solicitante=col_sol,
                col_representante=col_rep,
                solicitante_anterior=solicitante,
                representante_anterior=representante,
            ))
    return pendientes


def _fuentes_para(
    pendiente: Pendiente,
    descargas_base: Path,
    indice_metadata: dict[str, list[sor.MetadataEntrada]],
) -> list[Path]:
    dummy = sor.PendienteExcel(
        sheet=pendiente.sheet,
        row=pendiente.row,
        identificador=pendiente.identificador,
        ruta_actual="",
        col_ruta=0,
    )
    entradas = sor.buscar_fuentes_descargas(dummy, [], indice_metadata)
    candidatos: dict[str, Path] = {}
    for entrada in entradas:
        if entrada.carpeta.is_dir():
            candidatos[str(entrada.carpeta.resolve())] = entrada.carpeta

    # Fallback explícito para expedientes cuyo metadata falte/esté dañado.
    directo = descargas_base / pendiente.identificador
    if directo.is_dir():
        candidatos[str(directo.resolve())] = directo
    if pendiente.sheet == "Internos":
        internos = descargas_base / "internos"
        if internos.is_dir():
            try:
                for bandeja in internos.iterdir():
                    candidato = bandeja / pendiente.identificador
                    if candidato.is_dir():
                        candidatos[str(candidato.resolve())] = candidato
            except OSError:
                pass

    def mtime(path: Path) -> float:
        try:
            return path.stat().st_mtime
        except OSError:
            return 0.0

    return sorted(candidatos.values(), key=lambda p: (mtime(p), str(p).casefold()))


def _pdfs_de_fuentes(fuentes: Iterable[Path]) -> list[Path]:
    encontrados: dict[str, Path] = {}
    for carpeta in fuentes:
        try:
            for path in carpeta.rglob("*"):
                if path.is_file() and path.suffix.strip().casefold() == ".pdf":
                    encontrados[str(path.resolve())] = path
        except OSError as exc:
            log.warning("No se pudo recorrer %s: %s", carpeta, exc)
    return sorted(encontrados.values(), key=lambda p: str(p).casefold())


def _resolver_valores(candidatos: list[tuple[Path, dict[str, Any]]]) -> tuple[str, list[str], list[str]]:
    """Devuelve valor único seguro, fuentes y conflictos entre PDFs."""
    grupos: dict[str, list[tuple[Path, str]]] = {}
    for pdf, campo in candidatos:
        if campo.get("estado") != "encontrado":
            continue
        valor = _texto(campo.get("valor"))
        if not valor or _es_faltante(valor):
            continue
        clave = extractor.clave_comparacion(valor)
        if not clave:
            continue
        grupos.setdefault(clave, []).append((pdf, valor))

    if not grupos:
        return "", [], []
    if len(grupos) > 1:
        conflictos = []
        for items in grupos.values():
            valor = max((v for _, v in items), key=len)
            archivos = ", ".join(sorted({p.name for p, _ in items}))
            conflictos.append(f"{valor} [{archivos}]")
        return "", [], sorted(conflictos)

    items = next(iter(grupos.values()))
    # La forma más completa (mayor longitud) preserva acentos/espaciado del PDF.
    valor = max((v for _, v in items), key=len)
    fuentes = sorted({str(p) for p, _ in items})
    return valor, fuentes, []


def _guardar_reportes(logs_dir: Path, payload: dict[str, Any]) -> tuple[Path, Path]:
    logs_dir.mkdir(parents=True, exist_ok=True)
    sello = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = logs_dir / f"completar_remitentes_pdf_{sello}.json"
    csv_path = logs_dir / f"completar_remitentes_pdf_{sello}.csv"
    latest = logs_dir / "completar_remitentes_pdf_ultimo.json"
    data = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    json_path.write_text(data, encoding="utf-8")
    latest.write_text(data, encoding="utf-8")

    headers = [
        "sheet", "row", "identificador", "estado", "motivo", "pdfs_revisados",
        "pdfs_sin_texto", "pdfs_error", "solicitante_anterior", "solicitante_nuevo",
        "representante_anterior", "representante_nuevo", "fuentes",
        "solicitante_fuentes", "solicitante_conflictos", "representante_fuentes",
        "representante_conflictos",
    ]
    with csv_path.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=headers)
        writer.writeheader()
        for item in payload.get("resultados", []):
            row = dict(item)
            for key in (
                "fuentes", "solicitante_fuentes", "solicitante_conflictos",
                "representante_fuentes", "representante_conflictos",
            ):
                row[key] = " | ".join(row.get(key) or [])
            writer.writerow({key: row.get(key, "") for key in headers})
    return json_path, csv_path


def completar(
    *,
    excel_path: Path,
    descargas_base: Path,
    logs_dir: Path,
    dry_run: bool = False,
    crear_backup: bool = True,
) -> dict[str, Any]:
    excel_path = Path(excel_path)
    descargas_base = Path(descargas_base)
    logs_dir = Path(logs_dir)
    if not excel_path.is_file():
        raise FileNotFoundError(f"No existe Excel: {excel_path}")
    if not descargas_base.is_dir():
        raise FileNotFoundError(f"No existe descargas: {descargas_base}")

    wb = openpyxl.load_workbook(excel_path)
    pendientes = leer_pendientes(wb)
    log.info("🧾 Remitentes PDF: %d fila(s) con campos vacíos/SIN REMITENTE.", len(pendientes))

    _entradas, indice_metadata = sor.construir_indice_metadata(descargas_base)
    resultados: list[ResultadoFila] = []
    cambios_solicitante = 0
    cambios_representante = 0

    for idx, pendiente in enumerate(pendientes, start=1):
        log.info(
            "[REMITENTES %d/%d] %s/%s",
            idx, len(pendientes), pendiente.sheet, pendiente.identificador,
        )
        item = ResultadoFila(
            sheet=pendiente.sheet,
            row=pendiente.row,
            identificador=pendiente.identificador,
            solicitante_anterior=pendiente.solicitante_anterior,
            representante_anterior=pendiente.representante_anterior,
        )
        resultados.append(item)

        fuentes = _fuentes_para(pendiente, descargas_base, indice_metadata)
        item.fuentes = [str(p) for p in fuentes]
        if not fuentes:
            item.estado = "sin_fuente_descargas"
            item.motivo = "no_se_encontro_carpeta_origen_para_1711"
            continue
        pdfs = _pdfs_de_fuentes(fuentes)
        if not pdfs:
            item.estado = "sin_pdfs"
            item.motivo = "carpeta_descargas_sin_pdf"
            continue

        candidatos_sol: list[tuple[Path, dict[str, Any]]] = []
        candidatos_rep: list[tuple[Path, dict[str, Any]]] = []
        for pdf in pdfs:
            item.pdfs_revisados += 1
            try:
                extraido = extractor.extraer_de_pdf(pdf)
            except Exception as exc:
                item.pdfs_error += 1
                log.warning("PDF %s falló: %s", pdf, exc)
                continue
            if extraido.get("estado_general") == "requiere_ocr":
                item.pdfs_sin_texto += 1
            candidatos_sol.append((pdf, extraido.get("solicitante") or {}))
            candidatos_rep.append((pdf, extraido.get("representante") or {}))

        sol, sol_fuentes, sol_conflictos = _resolver_valores(candidatos_sol)
        rep, rep_fuentes, rep_conflictos = _resolver_valores(candidatos_rep)
        item.solicitante_fuentes = sol_fuentes
        item.solicitante_conflictos = sol_conflictos
        item.representante_fuentes = rep_fuentes
        item.representante_conflictos = rep_conflictos

        ws = wb[pendiente.sheet]
        cambios_fila = 0
        if _es_faltante(pendiente.solicitante_anterior) and sol:
            item.solicitante_nuevo = sol
            if not dry_run:
                ws.cell(row=pendiente.row, column=pendiente.col_solicitante, value=sol)
            cambios_solicitante += 1
            cambios_fila += 1
        if _es_faltante(pendiente.representante_anterior) and rep:
            item.representante_nuevo = rep
            if not dry_run:
                ws.cell(row=pendiente.row, column=pendiente.col_representante, value=rep)
            cambios_representante += 1
            cambios_fila += 1

        if cambios_fila:
            item.estado = "dry_run_actualizable" if dry_run else "actualizado"
            item.motivo = f"campos_actualizados={cambios_fila}"
        elif sol_conflictos or rep_conflictos:
            item.estado = "conflicto_entre_pdfs"
            item.motivo = "se_conserva_excel_por_inconsistencia"
        else:
            item.estado = "sin_resultado_pdf"
            item.motivo = "no_se_encontraron_valores_seguros_en_los_pdf"

    cambios = cambios_solicitante + cambios_representante
    backup_path: Path | None = None
    if cambios and not dry_run:
        if crear_backup:
            sello = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = excel_path.with_name(
                f"{excel_path.stem}_backup_pre_remitentes_pdf_{sello}{excel_path.suffix}"
            )
            shutil.copy2(excel_path, backup_path)
        temporal = excel_path.with_name(f".{excel_path.name}.tmp_remitentes_pdf")
        wb.save(temporal)
        reemplazar_desde_temporal(temporal, excel_path)
        log.info(
            "💾 TrámitesCRT.xlsx actualizado desde PDFs: %d Solicitante(s), %d Representante(s).",
            cambios_solicitante, cambios_representante,
        )
    wb.close()

    payload = {
        "fecha": datetime.now().isoformat(),
        "excel": str(excel_path),
        "descargas": str(descargas_base),
        "total_filas_pendientes": len(pendientes),
        "cambios_solicitante": cambios_solicitante,
        "cambios_representante": cambios_representante,
        "cambios_totales": cambios,
        "filas_actualizadas": sum(1 for r in resultados if r.estado in {"actualizado", "dry_run_actualizable"}),
        "filas_conflicto": sum(1 for r in resultados if r.estado == "conflicto_entre_pdfs"),
        "filas_sin_fuente": sum(1 for r in resultados if r.estado == "sin_fuente_descargas"),
        "filas_sin_pdfs": sum(1 for r in resultados if r.estado == "sin_pdfs"),
        "filas_sin_resultado": sum(1 for r in resultados if r.estado == "sin_resultado_pdf"),
        "backup_excel": str(backup_path) if backup_path else "",
        "dry_run": dry_run,
        "resultados": [r.as_dict() for r in resultados],
    }
    json_path, csv_path = _guardar_reportes(logs_dir, payload)
    payload["reporte_json"] = str(json_path)
    payload["reporte_csv"] = str(csv_path)
    log.info(
        "✅ Remitentes PDF: %d campo(s) actualizable(s), %d fila(s) con conflicto. Reporte: %s",
        cambios, payload["filas_conflicto"], json_path,
    )
    return payload


def construir_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Completa Solicitante Promovente/Representante Legal vacíos o SIN REMITENTE "
            "buscando en todos los PDF de cada expediente en descargas."
        )
    )
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--descargas", type=Path, default=DEFAULT_DESCARGAS)
    parser.add_argument("--logs-dir", type=Path, default=DEFAULT_LOGS)
    parser.add_argument("--sin-backup", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--sin-lock",
        action="store_true",
        help="Omite lock global SATyS; sólo para pruebas controladas.",
    )
    return parser


def main() -> int:
    args = construir_parser().parse_args()
    lock = None
    try:
        if not args.sin_lock:
            lock = ProcesoLock(proceso="completar_remitentes_desde_pdfs.py")
            lock.adquirir()
            log.info("🔒 Lock SATyS adquirido/heredado para completar remitentes desde PDF.")
        completar(
            excel_path=args.excel,
            descargas_base=args.descargas,
            logs_dir=args.logs_dir,
            dry_run=args.dry_run,
            crear_backup=not args.sin_backup,
        )
        return 0
    except LockOcupadoError as exc:
        log.error("No se inicia completar remitentes PDF: %s", exc)
        return 3
    except Exception as exc:
        log.exception("Fallo fatal completando remitentes desde PDF: %s", exc)
        return 1
    finally:
        if lock is not None:
            lock.liberar()


if __name__ == "__main__":
    raise SystemExit(main())
