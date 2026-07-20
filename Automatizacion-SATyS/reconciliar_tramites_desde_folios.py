#!/usr/bin/env python3
"""Reconcilia TrámitesCRT.xlsx usando Folios_Datos_Completos.xlsx como fuente.

Objetivos:
- una fila única por número de Registro CRT;
- permitir varios registros con el mismo folio/Memo;
- completar las columnas equivalentes del Excel maestro;
- escribir Ruta también para registros en ``_sin_operador``;
- eliminar únicamente filas fantasma creadas por subcarpetas de ZIP;
- conservar columnas manuales y las tres columnas administradas por SharePoint.
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

REGISTRO_RE = re.compile(r"^CRT\d{2}-\d+$", re.IGNORECASE)
FORMATO_RE = re.compile(r"\bR(?:0(?:0[1-9]|1\d|2[0-7]))\b", re.IGNORECASE)
SHEET_TRAMITES = "Turnados recibidos"
SHEET_FOLIOS = "Datos_Completos"


def _texto(value: Any) -> str:
    return str(value or "").strip()


def _registro(value: Any) -> str:
    text = _texto(value).upper()
    return text if REGISTRO_RE.fullmatch(text) else ""


def _primero(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return ""


def _ruta_desde_output(value: Any) -> str:
    r"""Convierte ``\output\...`` a la ruta relativa guardada en el maestro."""
    text = _texto(value).replace("/", "\\")
    text = text.lstrip("\\")
    if text.lower().startswith("output\\"):
        text = text[len("output\\"):]
    return text


def _headers(ws) -> dict[str, int]:
    return {
        _texto(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if _texto(ws.cell(1, col).value)
    }


def _row_dict(ws, row_number: int, headers: dict[str, int]) -> dict[str, Any]:
    return {name: ws.cell(row_number, col).value for name, col in headers.items()}


def _es_fila_fantasma(values: list[Any], headers: dict[str, int]) -> bool:
    """Detecta filas creadas al procesar carpetas internas de ZIP como trámites.

    Solo se elimina una fila sin Registro cuando sus únicos valores útiles están
    en Memo/Volante y/o NOTAS_VICTOR. Las filas manuales con más información se
    conservan.
    """
    registro_col = headers.get("1711", 4) - 1
    if _registro(values[registro_col]):
        return False

    permitidas = {
        headers.get("Memo/Volante", 5) - 1,
        headers.get("NOTAS_VICTOR", 41) - 1,
    }
    utiles = {i for i, value in enumerate(values) if value not in (None, "")}
    return bool(utiles) and utiles.issubset(permitidas)


def reconciliar(
    tramites_path: str | Path,
    folios_path: str | Path,
    *,
    crear_backup: bool = True,
) -> dict[str, Any]:
    tramites_path = Path(tramites_path)
    folios_path = Path(folios_path)
    if not tramites_path.exists():
        raise FileNotFoundError(f"No existe el Excel maestro: {tramites_path}")
    if not folios_path.exists():
        raise FileNotFoundError(f"No existe el Excel consolidado: {folios_path}")

    wb_t = openpyxl.load_workbook(tramites_path)
    if SHEET_TRAMITES not in wb_t.sheetnames:
        wb_t.close()
        raise KeyError(f"No existe la hoja '{SHEET_TRAMITES}' en {tramites_path}")
    ws_t = wb_t[SHEET_TRAMITES]
    headers_t = _headers(ws_t)

    required = {
        "1711", "Memo/Volante", "Solicitante Promovente", "Representante Legal",
        "Asunto", "Tipo Trámite", "Fecha de creación", "FECHA LÍMITE", "Ruta",
    }
    missing_headers = sorted(required - set(headers_t))
    if missing_headers:
        wb_t.close()
        raise ValueError(f"Faltan columnas en TrámitesCRT.xlsx: {', '.join(missing_headers)}")

    wb_f = openpyxl.load_workbook(folios_path, read_only=True, data_only=True)
    if SHEET_FOLIOS not in wb_f.sheetnames:
        wb_f.close()
        wb_t.close()
        raise KeyError(f"No existe la hoja '{SHEET_FOLIOS}' en {folios_path}")
    ws_f = wb_f[SHEET_FOLIOS]
    headers_f = _headers(ws_f)

    source_rows: list[dict[str, Any]] = []
    source_seen: set[str] = set()
    source_duplicates: list[str] = []
    for row_num in range(2, ws_f.max_row + 1):
        item = _row_dict(ws_f, row_num, headers_f)
        registro = _registro(item.get("registro") or item.get("metadata_satys.registro"))
        if not registro:
            continue
        if registro in source_seen:
            source_duplicates.append(registro)
            continue
        source_seen.add(registro)
        item["registro"] = registro
        source_rows.append(item)
    wb_f.close()

    # Indexar filas existentes por Registro. Las duplicadas se consolidan usando
    # la primera; sus valores manuales se combinan cuando la primera está vacía.
    existing_by_record: dict[str, list[Any]] = {}
    target_duplicates: list[str] = []
    manual_rows: list[list[Any]] = []
    phantom_removed = 0
    max_col = ws_t.max_column

    for row_num in range(2, ws_t.max_row + 1):
        values = [ws_t.cell(row_num, col).value for col in range(1, max_col + 1)]
        registro = _registro(values[headers_t["1711"] - 1])
        if registro:
            if registro in existing_by_record:
                target_duplicates.append(registro)
                current = existing_by_record[registro]
                for i, value in enumerate(values):
                    if current[i] in (None, "") and value not in (None, ""):
                        current[i] = value
            else:
                existing_by_record[registro] = values
        elif _es_fila_fantasma(values, headers_t):
            phantom_removed += 1
        elif any(value not in (None, "") for value in values):
            manual_rows.append(values)

    output_rows: list[list[Any]] = []
    appended = 0
    updated = 0
    format_headers = [f"R{i:03d}" for i in range(1, 28)]

    for item in source_rows:
        registro = item["registro"]
        values = list(existing_by_record.pop(registro, [None] * max_col))
        if all(value in (None, "") for value in values):
            appended += 1
        else:
            updated += 1

        def setv(header: str, value: Any) -> None:
            col = headers_t.get(header)
            if col:
                values[col - 1] = value if value not in (None, "") else None

        asunto = _primero(
            item,
            "metadata_satys.asunto",
            "metadata_tramite_nuevo.asunto",
        )
        tipo = _primero(
            item,
            "metadata_satys.tipo_tramite",
            "metadata_tramite_nuevo.tipo_tramite",
        )
        fecha = _primero(
            item,
            "metadata_satys.fecha_registro",
            "metadata_satys.fecha_folio_opc",
            "metadata_tramite_nuevo.fecha_registro",
        )
        fecha_limite = _primero(
            item,
            "metadata_tramite_nuevo.plazo_atencion",
            "metadata_satys.plazo_atencion",
        )

        setv("1711", registro)
        setv("Memo/Volante", _primero(item, "folio", "metadata_satys.folio", "metadata_tramite_nuevo.folio"))
        setv(
            "Solicitante Promovente",
            _primero(
                item,
                "metadata_satys.solicitante",
                "metadata_tramite_nuevo.solicitante",
                "nombre_operador",
                "metadata_satys.nombre_operador",
                "metadata_tramite_nuevo.nombre_operador",
            ),
        )
        setv(
            "Representante Legal",
            _primero(
                item,
                "representante_legal",
                "metadata_satys.representante_legal",
                "metadata_tramite_nuevo.representante_legal",
            ),
        )
        setv("Asunto", asunto)
        setv("Tipo Trámite", tipo)
        setv("Fecha de creación", fecha)
        setv("FECHA LÍMITE", fecha_limite)
        setv("Ruta", _ruta_desde_output(item.get("output")))

        # Recalcular banderas de formato; evita marcas heredadas cuando antes se
        # sobrescribía una fila de otro registro con el mismo folio.
        for header in format_headers:
            setv(header, None)
        formatos = {match.group(0).upper() for match in FORMATO_RE.finditer(_texto(asunto))}
        for header in formatos:
            setv(header, 1)

        output_rows.append(values)

    # Conservar registros válidos que solo existan en el maestro y filas manuales
    # no clasificadas como fantasmas. Actualmente no debería haber extras, pero
    # esta medida evita pérdida de datos ante una fuente parcial.
    target_only = sorted(existing_by_record)
    output_rows.extend(existing_by_record[key] for key in target_only)
    output_rows.extend(manual_rows)

    old_max_row = ws_t.max_row
    for r_offset, values in enumerate(output_rows, start=2):
        for col, value in enumerate(values, start=1):
            ws_t.cell(r_offset, col).value = value

    new_last = len(output_rows) + 1
    if old_max_row > new_last:
        ws_t.delete_rows(new_last + 1, old_max_row - new_last)

    # Guardado atómico y respaldo antes de sustituir el maestro.
    backup_path = None
    if crear_backup:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = tramites_path.with_name(f"{tramites_path.stem}_backup_{stamp}{tramites_path.suffix}")
        shutil.copy2(tramites_path, backup_path)

    temp_path = tramites_path.with_name(f".{tramites_path.name}.tmp")
    wb_t.save(temp_path)
    wb_t.close()
    os.replace(temp_path, tramites_path)

    valid_final = len(source_rows) + len(target_only)
    routes_blank = sum(1 for item in source_rows if not _ruta_desde_output(item.get("output")))
    return {
        "source_records": len(source_rows),
        "updated": updated,
        "appended": appended,
        "phantom_removed": phantom_removed,
        "target_only_preserved": len(target_only),
        "manual_rows_preserved": len(manual_rows),
        "source_duplicates": sorted(set(source_duplicates)),
        "target_duplicates": sorted(set(target_duplicates)),
        "valid_final": valid_final,
        "routes_blank": routes_blank,
        "backup": str(backup_path) if backup_path else "",
        "output": str(tramites_path),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Reconcilia TrámitesCRT.xlsx desde Folios_Datos_Completos.xlsx")
    parser.add_argument("--tramites", default="TrámitesCRT.xlsx")
    parser.add_argument("--folios", default="output/Folios_Datos_Completos.xlsx")
    parser.add_argument("--sin-backup", action="store_true")
    args = parser.parse_args()

    result = reconciliar(args.tramites, args.folios, crear_backup=not args.sin_backup)
    print("Reconciliación terminada:")
    for key, value in result.items():
        print(f"  {key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
