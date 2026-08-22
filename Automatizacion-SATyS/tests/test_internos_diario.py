#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import sys
import tempfile
import threading
import time
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import automatizar_registros_diario as monitor
import extraer_registros_documentos as extractor
import main_procesar
import Parte1_descarga
import Parte4_excel
import estado_descargas


class InternosDiarioTest(unittest.TestCase):
    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def test_objetivos_conservan_cada_pareja_bandeja_folio_entre_etapas(self):
        resumen = {
            "por_bandeja": [
                {"bandeja": "En proceso", "folios": ["148326", "147390"]},
                {"bandeja": "Atendidos", "folios": ["148326", "190823"]},
            ]
        }
        objetivos = monitor.construir_objetivos_internos(
            resumen,
            ["148326", "190823"],
        )
        self.assertEqual(objetivos, [
            {"bandeja": "En proceso", "folio": "148326"},
            {"bandeja": "Atendidos", "folio": "148326"},
            {"bandeja": "Atendidos", "folio": "190823"},
        ])

        path = self.tmp / "objetivos.json"
        monitor.guardar_objetivos_internos(path, objetivos)
        self.assertEqual(main_procesar.cargar_objetivos_internos(path), objetivos)
        self.assertEqual(Parte1_descarga._cargar_objetivos_internos(path), objetivos)

    def test_folio_unico_conserva_todas_sus_bandejas_y_fuerza_sin_email(self):
        resumen = {
            "por_bandeja": [
                {"bandeja": "En proceso", "folios": ["148326"]},
                {"bandeja": "Atendidos", "folios": ["148326", "190823"]},
                {"bandeja": "Fuera de tiempo", "folios": ["148326"]},
            ]
        }
        objetivos = monitor.seleccionar_objetivos_folio_internos(resumen, " 148326 ")
        self.assertEqual(objetivos, [
            {"bandeja": "En proceso", "folio": "148326"},
            {"bandeja": "Atendidos", "folio": "148326"},
            {"bandeja": "Fuera de tiempo", "folio": "148326"},
        ])

        args = SimpleNamespace(
            folio_internos="148326",
            solo_internos=False,
            sin_email=False,
        )
        monitor.configurar_modo_folio_internos(args)
        self.assertTrue(args.solo_internos)
        self.assertTrue(args.sin_email)

        with self.assertRaisesRegex(LookupError, "no aparece"):
            monitor.seleccionar_objetivos_folio_internos(resumen, "999999")
        with self.assertRaisesRegex(ValueError, "1 y 15"):
            monitor.seleccionar_objetivos_folio_internos(resumen, "ABC")

    def test_valida_excel_y_output_del_folio_unico(self):
        output_a = self.tmp / "output" / "operador" / "148326"
        output_b = self.tmp / "output" / "_sin_operador" / "148326"
        output_a.mkdir(parents=True)
        output_b.mkdir(parents=True)
        log_path = self.tmp / "procesamiento_log_internos.json"
        log_path.write_text(json.dumps({
            "resultados": [
                {
                    "bandeja_internos": "En proceso",
                    "folio_tabla_internos": "148326",
                    "excel_ok": True,
                    "output_dir": str(output_a),
                },
                {
                    "bandeja_internos": "Atendidos",
                    "folio_tabla_internos": "148326",
                    "excel_ok": True,
                    "sin_operador_dir": str(output_b),
                },
            ]
        }), encoding="utf-8")

        objetivos = [
            {"bandeja": "En proceso", "folio": "148326"},
            {"bandeja": "Atendidos", "folio": "148326"},
        ]
        with patch.object(
            monitor,
            "cargar_folios_internos_procesados_excel",
            return_value=({"148326"}, {}),
        ):
            validacion = monitor.validar_salidas_folio_internos(
                folio="148326",
                objetivos=objetivos,
                procesamiento_log=log_path,
                excel_path=self.tmp / "TrámitesCRT.xlsx",
                sheet="Internos",
                header_folio="Folio Internos",
                project_dir=self.tmp,
            )

        self.assertTrue(validacion["ok"], validacion)
        self.assertCountEqual(
            validacion["output_dirs"],
            [str(output_a.resolve()), str(output_b.resolve())],
        )

    def test_folio_unico_falla_si_no_quedo_en_excel_o_output(self):
        log_path = self.tmp / "procesamiento_log_internos.json"
        log_path.write_text(json.dumps({
            "resultados": [{
                "bandeja_internos": "Atendidos",
                "folio_tabla_internos": "148326",
                "excel_ok": False,
                "output_dir": str(self.tmp / "no_existe"),
            }]
        }), encoding="utf-8")

        with patch.object(
            monitor,
            "cargar_folios_internos_procesados_excel",
            return_value=(set(), {}),
        ):
            validacion = monitor.validar_salidas_folio_internos(
                folio="148326",
                objetivos=[{"bandeja": "Atendidos", "folio": "148326"}],
                procesamiento_log=log_path,
                excel_path=self.tmp / "TrámitesCRT.xlsx",
                sheet="Internos",
                header_folio="Folio Internos",
                project_dir=self.tmp,
            )

        self.assertFalse(validacion["ok"])
        self.assertEqual(len(validacion["errores"]), 3)

    def test_auditoria_internos_exige_metadata_y_todos_los_archivos(self):
        carpeta = self.tmp / "descargas" / "internos" / "atendidos" / "190823"
        carpeta.mkdir(parents=True)
        (carpeta / "documento_1.pdf").write_bytes(b"pdf-1")
        (carpeta / "documento_2.csv").write_bytes(b"csv-2")
        metadata = {
            "folio": "CRT26-000001",
            "estado": "OK",
            "coincide": True,
            "total_archivos_encontrados": 2,
            "total_archivos_ok": 2,
            "total_archivos_error": 0,
            "metadatos_satys": {
                "bandeja_internos": "Atendidos",
                "folio_tabla_internos": "190823",
            },
            "archivos": [
                {"archivo": "documento_1.pdf", "ok": True},
                {"archivo": "documento_2.csv", "ok": True},
            ],
        }
        (carpeta / "metadata_completo.json").write_text(
            json.dumps(metadata),
            encoding="utf-8",
        )

        self.assertTrue(estado_descargas.objetivo_internos_descarga_completa(
            self.tmp / "descargas", "Atendidos", "190823"
        ))
        (carpeta / "documento_2.csv").unlink()
        self.assertFalse(estado_descargas.objetivo_internos_descarga_completa(
            self.tmp / "descargas", "Atendidos", "190823"
        ))

    def test_auditoria_internos_reintenta_zip_residual(self):
        carpeta = self.tmp / "descargas" / "internos" / "fuera_de_tiempo" / "121195"
        carpeta.mkdir(parents=True)
        (carpeta / "pendiente.zip").write_bytes(b"zip-incompleto")
        metadata = {
            "folio": "121195",
            "estado": "OK",
            "coincide": True,
            "total_archivos_encontrados": 1,
            "total_archivos_ok": 1,
            "total_archivos_error": 0,
            "metadatos_satys": {
                "bandeja_internos": "Fuera de tiempo",
                "folio_tabla_internos": "121195",
            },
            "archivos": [{"archivo": "pendiente.zip", "ok": True}],
        }
        (carpeta / "metadata_completo.json").write_text(
            json.dumps(metadata),
            encoding="utf-8",
        )

        self.assertFalse(estado_descargas.objetivo_internos_esta_completo(
            self.tmp / "descargas", "Fuera de tiempo", "121195"
        ))

    def test_metadata_historica_sin_marca_no_redescarga_pero_si_reprocesa_postproceso(self):
        carpeta = self.tmp / "descargas" / "internos" / "atendidos" / "190822"
        carpeta.mkdir(parents=True)
        (carpeta / "documento.pdf").write_bytes(b"pdf")
        metadata = {
            "folio": "190822",
            "estado": "OK",
            "coincide": True,
            "total_archivos_encontrados": 1,
            "total_archivos_ok": 1,
            "total_archivos_error": 0,
            "metadatos_satys": {
                "bandeja_internos": "Atendidos",
                "folio_tabla_internos": "190822",
            },
            "archivos": [{"archivo": "documento.pdf", "ok": True}],
        }
        (carpeta / "metadata_completo.json").write_text(
            json.dumps(metadata), encoding="utf-8"
        )

        self.assertTrue(estado_descargas.objetivo_internos_descarga_completa(
            self.tmp / "descargas", "Atendidos", "190822"
        ))
        self.assertFalse(estado_descargas.objetivo_internos_esta_completo(
            self.tmp / "descargas", "Atendidos", "190822"
        ))

    def test_postproceso_pendiente_no_obliga_a_redescargar_pero_si_a_reprocesar(self):
        carpeta = self.tmp / "descargas" / "internos" / "atendidos" / "190823"
        carpeta.mkdir(parents=True)
        (carpeta / "documento.pdf").write_bytes(b"pdf")
        metadata = {
            "folio": "190823",
            "estado": "OK",
            "coincide": True,
            "total_archivos_encontrados": 1,
            "total_archivos_ok": 1,
            "total_archivos_error": 0,
            "metadatos_satys": {
                "bandeja_internos": "Atendidos",
                "folio_tabla_internos": "190823",
            },
            "archivos": [{"archivo": "documento.pdf", "ok": True}],
            "postproceso_internos": {
                "ok": False,
                "estado": "PENDIENTE",
                "excel_ok": False,
                "organizado_ok": False,
            },
        }
        (carpeta / "metadata_completo.json").write_text(
            json.dumps(metadata), encoding="utf-8"
        )

        self.assertTrue(estado_descargas.objetivo_internos_descarga_completa(
            self.tmp / "descargas", "Atendidos", "190823"
        ))
        self.assertFalse(estado_descargas.objetivo_internos_esta_completo(
            self.tmp / "descargas", "Atendidos", "190823"
        ))

        with patch.object(main_procesar, "ORGANIZAR_DESCARGAS", True):
            ok = main_procesar.actualizar_postproceso_internos(
                carpeta,
                {"excel_ok": True, "organizado_ok": True, "output_dir": "/tmp/out"},
            )
        self.assertTrue(ok)
        self.assertTrue(estado_descargas.objetivo_internos_esta_completo(
            self.tmp / "descargas", "Atendidos", "190823"
        ))

    def test_metadata_nueva_internos_nace_con_postproceso_pendiente(self):
        carpeta = self.tmp / "metadata_pendiente"
        metadata = Parte1_descarga.guardar_metadata_completo(
            "190823", "190823", carpeta,
            {"bandeja_internos": "Atendidos", "folio_tabla_internos": "190823"},
            {},
            [{"archivo": "documento.pdf", "ok": True}],
            "INTERNOS_DOCUMENTOS_ANEXOS",
        )
        self.assertEqual(metadata["postproceso_internos"]["estado"], "PENDIENTE")
        self.assertFalse(metadata["postproceso_internos"]["ok"])

    def test_metadata_distingue_documentos_portal_de_archivos_extraidos_del_zip(self):
        carpeta = self.tmp / "zip_expandido"
        resultados = [
            {
                "archivo": f"miembro_{indice}.csv",
                "ok": True,
                "indice_documento_portal": 1,
                "total_documentos_portal": 1,
            }
            for indice in range(3)
        ]

        metadata = Parte1_descarga.guardar_metadata_completo(
            "190823", "190823", carpeta,
            {"bandeja_internos": "Atendidos", "folio_tabla_internos": "190823"},
            {}, resultados, "INTERNOS_DOCUMENTOS_ANEXOS",
        )

        self.assertEqual(metadata["estado"], "OK")
        self.assertEqual(metadata["total_archivos_ok"], 3)
        self.assertEqual(metadata["total_documentos_portal"], 1)
        self.assertEqual(metadata["total_documentos_portal_ok"], 1)
        self.assertTrue(metadata["documentos_portal_completos"])

    def test_metadata_marca_parcial_si_no_recorrio_todos_los_documentos_portal(self):
        carpeta = self.tmp / "conteo_incompleto"
        resultados = [{
            "archivo": "documento_1.pdf",
            "ok": True,
            "indice_documento_portal": 1,
            "total_documentos_portal": 2,
        }]

        metadata = Parte1_descarga.guardar_metadata_completo(
            "190823", "190823", carpeta,
            {"bandeja_internos": "Atendidos", "folio_tabla_internos": "190823"},
            {}, resultados, "INTERNOS_DOCUMENTOS_ANEXOS",
        )

        self.assertEqual(metadata["estado"], "PARCIAL")
        self.assertFalse(metadata["coincide"])
        self.assertFalse(metadata["documentos_portal_completos"])

    def test_clasificacion_internos_no_usa_excel_y_mantiene_bandejas(self):
        resumen = {
            "por_bandeja": [
                {"bandeja": "En proceso", "folios": ["148326"]},
                {"bandeja": "Atendidos", "folios": ["148326"]},
                {"bandeja": "Fuera de tiempo", "folios": ["148326"]},
            ]
        }
        with patch.object(
            monitor,
            "objetivo_internos_esta_completo",
            side_effect=lambda _base, bandeja, _folio: bandeja == "En proceso",
        ):
            inventario, completos, pendientes = monitor.clasificar_objetivos_internos(
                resumen,
                self.tmp / "descargas",
            )

        self.assertEqual(len(inventario), 3)
        self.assertEqual(completos, [{"bandeja": "En proceso", "folio": "148326"}])
        self.assertEqual(pendientes, [
            {"bandeja": "Atendidos", "folio": "148326"},
            {"bandeja": "Fuera de tiempo", "folio": "148326"},
        ])

    def test_extractores_reconocen_modo_solo_internos(self):
        with patch.object(sys, "argv", ["extraer_registros_documentos.py", "--solo-internos"]):
            args_extractor = extractor.parse_args()
        with patch.object(sys, "argv", [
            "automatizar_registros_diario.py", "--solo-internos",
            "--max-folios-internos", "4", "--internos-workers", "3",
        ]):
            args_monitor = monitor.construir_parser().parse_args()

        self.assertTrue(args_extractor.solo_internos)
        self.assertFalse(args_extractor.sin_internos)
        self.assertTrue(args_monitor.solo_internos)
        self.assertEqual(args_monitor.max_folios_internos, 4)
        self.assertEqual(args_monitor.internos_workers, 3)

    def test_monitor_acepta_workers_internos_sin_tope_artificial(self):
        with patch.object(sys, "argv", [
            "automatizar_registros_diario.py", "--solo-internos",
            "--internos-workers", "48",
        ]):
            args_monitor = monitor.construir_parser().parse_args()

        self.assertEqual(args_monitor.internos_workers, 48)

    def test_limita_lote_internos_sin_alterar_orden(self):
        objetivos = [
            {"bandeja": "En proceso", "folio": str(folio)}
            for folio in range(10, 15)
        ]

        self.assertEqual(monitor.limitar_objetivos_internos(objetivos, 0), objetivos)
        self.assertEqual(monitor.limitar_objetivos_internos(objetivos, 2), objetivos[:2])
        self.assertEqual(objetivos[-1]["folio"], "14")
        with self.assertRaises(ValueError):
            monitor.limitar_objetivos_internos(objetivos, -1)

    def test_lote_internos_se_reparte_entre_bandejas(self):
        objetivos = [
            {"bandeja": "En proceso", "folio": "10"},
            {"bandeja": "En proceso", "folio": "11"},
            {"bandeja": "Atendidos", "folio": "20"},
            {"bandeja": "Atendidos", "folio": "21"},
            {"bandeja": "Fuera de tiempo", "folio": "30"},
        ]

        muestra = monitor.limitar_objetivos_internos(objetivos, 4)

        self.assertEqual(
            [(item["bandeja"], item["folio"]) for item in muestra],
            [
                ("En proceso", "10"),
                ("Atendidos", "20"),
                ("Fuera de tiempo", "30"),
                ("En proceso", "11"),
            ],
        )

    def test_certificado_solo_internos_no_exige_oficialia(self):
        output = self.tmp / "registros.txt"
        output.write_text("", encoding="utf-8")
        bandejas = []
        for nombre in Parte1_descarga.BANDEJAS_INTERNOS_DEFAULT:
            folios = ["148326"] if nombre == "En proceso" else []
            bandejas.append({
                "bandeja": nombre,
                "estado": "ENCONTRADOS_COMPLETOS" if folios else "VACIO_CONFIRMADO",
                "total_reportado_satys": len(folios),
                "filas_leidas": len(folios),
                "folios": folios,
                "filas_invalidas": 0,
            })

        resumen = extractor.resumen_oficialia_omitida()
        resumen["internos"] = {
            "estado": "COMPLETO",
            "integridad": "VALIDADA",
            "vacio_confirmado": False,
            "total_folios": 1,
            "total_filas_satys": 1,
            "folios": ["148326"],
            "por_bandeja": bandejas,
        }
        output.with_suffix(".txt.json").write_text(json.dumps(resumen), encoding="utf-8")

        validacion = monitor.validar_resumen_extractor(output, [])

        self.assertTrue(validacion["ok"], validacion)
        self.assertTrue(validacion["vacio_confirmado"])
        self.assertEqual(validacion["folios_internos"], ["148326"])

    def test_lanzadores_oci_exponen_comando_internos_filtrado(self):
        root = Path(__file__).resolve().parents[1]
        podman = (root / "scripts" / "podman_satys.sh").read_text(encoding="utf-8")
        docker = (root / "scripts" / "docker_satys.sh").read_text(encoding="utf-8")

        self.assertIn("internos)", podman)
        self.assertIn("automatizar_registros_diario.py --solo-internos --headless", podman)
        self.assertIn("internos-check)", podman)
        self.assertIn("--solo-internos --no-procesar --sin-email --headless", podman)
        self.assertIn("internos)", docker)
        self.assertIn("automatizar_registros_diario.py --solo-internos --headless", docker)
        self.assertIn("internos-check)", docker)
        self.assertIn("--solo-internos --no-procesar --sin-email --headless", docker)

        powershell = (root / "scripts" / "run_satys_internos_nuevos.ps1").read_text(encoding="utf-8")
        self.assertIn("MaxFoliosInternos", powershell)
        self.assertIn("InternosWorkers", powershell)
        self.assertIn("[int]$InternosWorkers = 12", powershell)
        self.assertNotIn("ValidateRange(1, 6)", powershell)

        for contenido in (podman, docker):
            self.assertIn("folio)", contenido)
            self.assertIn('--folio-internos "$folio"', contenido)
            self.assertIn("--sin-email", contenido)

        folio_ps1 = (root / "scripts" / "procesar_folio_internos.ps1").read_text(encoding="utf-8")
        folio_sh = (root / "scripts" / "procesar_folio_internos.sh").read_text(encoding="utf-8")
        for contenido in (folio_ps1, folio_sh):
            self.assertIn("--folio-internos", contenido)
            self.assertIn("--sin-email", contenido)
        self.assertIn("Get-Command python", folio_ps1)

        monitor_source = (root / "automatizar_registros_diario.py").read_text(encoding="utf-8")
        self.assertIn('cmd_main_internos.append("--sin-sincronizar")', monitor_source)

    def test_extractor_solo_internos_no_navega_a_oficialia(self):
        output = self.tmp / "solo_internos.txt"

        class Page:
            def set_default_timeout(self, _timeout):
                return None

        class Context:
            def new_page(self):
                return Page()

        class Browser:
            def new_context(self, **_kwargs):
                return Context()

            def close(self):
                return None

        class Chromium:
            def launch(self, **_kwargs):
                return Browser()

        class Playwright:
            chromium = Chromium()

        class Manager:
            def __enter__(self):
                return Playwright()

            def __exit__(self, *_args):
                return False

        internos = {
            "estado": "COMPLETO",
            "integridad": "VALIDADA",
            "vacio_confirmado": True,
            "total_folios": 0,
            "total_filas_satys": 0,
            "folios": [],
            "por_bandeja": [],
        }
        with patch.object(
            sys,
            "argv",
            ["extraer_registros_documentos.py", "--solo-internos", "--output", str(output)],
        ), patch.object(extractor, "sync_playwright", return_value=Manager()), \
             patch.object(extractor, "cargar_credenciales_satys", return_value=("usuario", "password")), \
             patch.object(extractor, "sesion_activa", return_value=True), \
             patch.object(extractor, "extraer_folios_internos", return_value=internos) as extraer_int, \
             patch.object(extractor, "navegar_a_enlace_oficialia") as navegar_oficialia:
            rc = extractor.main()

        self.assertEqual(rc, 0)
        extraer_int.assert_called_once()
        navegar_oficialia.assert_not_called()
        resumen = json.loads(output.with_suffix(".txt.json").read_text(encoding="utf-8"))
        self.assertTrue(resumen["oficialia_omitida"])

    def test_navegacion_internos_no_usa_wait_for_function(self):
        class Page:
            def __init__(self):
                self.evaluaciones = 0

            def evaluate(self, _script, *args):
                self.evaluaciones += 1
                return True

            def wait_for_timeout(self, _timeout):
                return None

            def wait_for_function(self, *_args, **_kwargs):
                raise AssertionError("Internos no debe depender de wait_for_function")

        page = Page()
        with patch.object(extractor, "esperar_sin_spinner", return_value=True), \
             patch.object(extractor, "esperar_datatables"):
            ok = extractor.navegar_a_internos_ift(page)

        self.assertTrue(ok)
        self.assertEqual(page.evaluaciones, 3)

    def test_seleccion_bandeja_extractor_no_usa_wait_for_function(self):
        class Page:
            def __init__(self):
                self.evaluaciones = 0
                self.argumentos = []

            def evaluate(self, _script, *args):
                self.evaluaciones += 1
                self.argumentos.extend(args)
                if self.evaluaciones == 1:
                    return "CLICK"
                return True

            def wait_for_timeout(self, _timeout):
                return None

            def wait_for_function(self, *_args, **_kwargs):
                raise AssertionError("Internos no debe depender de wait_for_function")

        page = Page()
        with patch.object(extractor, "esperar_sin_spinner", return_value=True), \
             patch.object(extractor, "esperar_datatables"):
            extractor.seleccionar_bandeja_internos(page, "En proceso")

        self.assertEqual(page.evaluaciones, 3)
        self.assertEqual(page.argumentos, ["2", "2", "2"])

    def test_seleccion_bandeja_verifica_estado_tras_respuesta_vacia(self):
        class Page:
            def __init__(self):
                self.evaluaciones = 0

            def evaluate(self, _script, *_args):
                self.evaluaciones += 1
                return None if self.evaluaciones == 1 else True

            def wait_for_timeout(self, _timeout):
                return None

        page = Page()
        with patch.object(extractor, "esperar_sin_spinner", return_value=True), \
             patch.object(extractor, "esperar_datatables"):
            extractor.seleccionar_bandeja_internos(page, "Recibidos")

        self.assertEqual(page.evaluaciones, 3)

    def test_todos_internos_activa_solo_el_recorrido_completo(self):
        args = SimpleNamespace(
            todos_internos=True,
            internos=False,
            internos_bandejas=None,
            internos_objetivos="",
            solo_procesar=False,
            folios=[],
            archivo_folios="",
            archivo_registro="",
            buscar=0,
        )

        main_procesar.aplicar_modo_todos_internos(args)

        self.assertTrue(args.internos)
        self.assertIsNone(args.internos_bandejas)
        self.assertEqual(args.internos_objetivos, "")

    def test_todos_internos_rechaza_objetivos_parciales(self):
        args = SimpleNamespace(
            todos_internos=True,
            internos=False,
            internos_bandejas=None,
            internos_objetivos="pendientes.json",
            solo_procesar=False,
            folios=[],
            archivo_folios="",
            archivo_registro="",
            buscar=0,
        )

        with self.assertRaisesRegex(ValueError, "--internos-objetivos"):
            main_procesar.aplicar_modo_todos_internos(args)

    def test_selecciona_bandeja_internos_por_id_estable(self):
        class Page:
            def __init__(self):
                self.tab_id = None

            def evaluate(self, _script, tab_id):
                self.tab_id = tab_id
                return "CLICK"

        page = Page()
        with patch.object(
            Parte1_descarga,
            "_esperar_estado_bandeja_internos",
            return_value=True,
        ) as esperar:
            ok = Parte1_descarga.seleccionar_bandeja_internos(
                page,
                "Últimos Movimientos",
            )

        self.assertTrue(ok)
        self.assertEqual(page.tab_id, "5")
        esperar.assert_called_once_with(page, tab_id="5", timeout_ms=90_000)

    def test_extrae_fila_internos_desde_json_textual(self):
        class Fila:
            def evaluate(self, _script):
                return json.dumps({
                    "headers": ["Folio", "Acciones", "Tipo Trámite"],
                    "cells": ["148326", "Revisar", "CGPE-01-008"],
                    "columnas": {
                        "Folio": "148326",
                        "Acciones": "Revisar",
                        "Tipo Trámite": "CGPE-01-008",
                    },
                })

            def inner_text(self):
                return "148326 Revisar CGPE-01-008"

        meta = Parte1_descarga._extraer_datos_fila_internos(Fila())

        self.assertEqual(meta["folio"], "148326")
        self.assertEqual(meta["tipo_tramite"], "CGPE-01-008")

    def test_paraleliza_bandejas_y_conserva_el_orden_del_resumen(self):
        activas = 0
        max_activas = 0
        lock = threading.Lock()

        def worker(bandeja, _objetivos):
            nonlocal activas, max_activas
            with lock:
                activas += 1
                max_activas = max(max_activas, activas)
            time.sleep(0.03)
            with lock:
                activas -= 1
            return bandeja, [{"folio": bandeja, "ok": True}]

        bandejas = ["Recibidos", "En proceso", "Atendidos"]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos_con_reintentos", side_effect=worker), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            resultados = Parte1_descarga.descargar_internos_ift(
                bandejas=bandejas,
                workers=2,
            )

        self.assertEqual(max_activas, 2)
        self.assertEqual([item["folio"] for item in resultados], bandejas)

    def test_reparte_workers_dentro_de_una_misma_bandeja(self):
        llamadas = []
        lock = threading.Lock()

        def worker(bandeja, objetivos):
            with lock:
                llamadas.append((bandeja, list(objetivos)))
            return bandeja, [
                {"folio": folio, "ok": True}
                for folio in objetivos
            ]

        objetivos = [
            {"bandeja": "Atendidos", "folio": str(190000 + indice)}
            for indice in range(12)
        ]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos_con_reintentos", side_effect=worker), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            resultados = Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=4,
            )

        self.assertEqual(len(llamadas), 4)
        self.assertEqual(sorted(len(folios) for _bandeja, folios in llamadas), [3, 3, 3, 3])
        folios_asignados = [folio for _bandeja, folios in llamadas for folio in folios]
        self.assertCountEqual(folios_asignados, [item["folio"] for item in objetivos])
        self.assertEqual(len({item["folio"] for item in resultados}), 12)

    def test_reintenta_segmento_vacio_que_tenia_objetivos(self):
        llamadas = 0
        descarga_ok = False

        def worker(bandeja, objetivos):
            nonlocal llamadas, descarga_ok
            llamadas += 1
            if llamadas == 1:
                return bandeja, []
            descarga_ok = True
            return bandeja, [{"folio": folio, "ok": True} for folio in objetivos]

        objetivos = [{"bandeja": "Atendidos", "folio": "190823"}]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_ejecutar_worker_bandeja_internos_watchdog", side_effect=worker), \
             patch.object(Parte1_descarga, "objetivo_internos_descarga_completa", side_effect=lambda *_: descarga_ok), \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_REINTENTOS", 1), \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_ESPERA", 0), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            resultados = Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=1,
            )

        self.assertEqual(llamadas, 2)
        self.assertEqual(resultados, [{"folio": "190823", "ok": True}])

    def test_segmento_vacio_agotado_se_convierte_en_error(self):
        objetivos = [{"bandeja": "Atendidos", "folio": "190823"}]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_ejecutar_worker_bandeja_internos_watchdog", return_value=("Atendidos", [])) as worker, \
             patch.object(Parte1_descarga, "objetivo_internos_descarga_completa", return_value=False), \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_REINTENTOS", 1), \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_ESPERA", 0), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            resultados = Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=1,
            )

        self.assertEqual(worker.call_count, 2)
        self.assertEqual(resultados[0]["archivo"], "ERROR_INTERNO_INCOMPLETO_190823")
        self.assertFalse(resultados[0]["ok"])

    def test_reintento_internos_reduce_el_segmento_a_folios_aun_incompletos(self):
        llamadas = []
        completos = set()

        def watchdog(bandeja, objetivos):
            llamadas.append(list(objetivos))
            if len(llamadas) == 1:
                completos.add("190001")
            else:
                completos.add("190002")
            return bandeja, [{"folio": folio, "ok": folio in completos} for folio in objetivos]

        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_ejecutar_worker_bandeja_internos_watchdog", side_effect=watchdog), \
             patch.object(Parte1_descarga, "objetivo_internos_descarga_completa", side_effect=lambda _b, _ban, f: f in completos), \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_REINTENTOS", 1), \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_ESPERA", 0), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            Parte1_descarga.descargar_internos_ift(
                objetivos=[
                    {"bandeja": "Atendidos", "folio": "190001"},
                    {"bandeja": "Atendidos", "folio": "190002"},
                ],
                workers=1,
            )

        self.assertEqual(llamadas[0], ["190001", "190002"])
        self.assertEqual(llamadas[1], ["190002"])

    def test_reserva_dos_segmentos_y_asigna_el_resto_segun_carga(self):
        llamadas = []

        def worker(bandeja, objetivos):
            llamadas.append((bandeja, list(objetivos)))
            return bandeja, [{"folio": folio, "ok": True} for folio in objetivos]

        objetivos = [
            {"bandeja": "En proceso", "folio": str(140000 + indice)}
            for indice in range(5)
        ] + [
            {"bandeja": "Atendidos", "folio": str(190000 + indice)}
            for indice in range(19)
        ]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos_con_reintentos", side_effect=worker), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=12,
            )

        self.assertEqual(sum(1 for bandeja, _ in llamadas if bandeja == "En proceso"), 3)
        self.assertEqual(sum(1 for bandeja, _ in llamadas if bandeja == "Atendidos"), 9)
        atendidos = sorted(len(folios) for bandeja, folios in llamadas if bandeja == "Atendidos")
        self.assertEqual(atendidos, [2, 2, 2, 2, 2, 2, 2, 2, 3])

    def test_doce_workers_dan_dos_segmentos_a_cada_una_de_seis_bandejas(self):
        llamadas = []

        def worker(bandeja, objetivos):
            llamadas.append((bandeja, list(objetivos)))
            return bandeja, [{"folio": folio, "ok": True} for folio in objetivos]

        objetivos = []
        for indice_bandeja, bandeja in enumerate(Parte1_descarga.BANDEJAS_INTERNOS_DEFAULT):
            objetivos.extend([
                {"bandeja": bandeja, "folio": str(200000 + indice_bandeja * 10 + indice)}
                for indice in range(4)
            ])

        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos_con_reintentos", side_effect=worker), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=12,
            )

        for bandeja in Parte1_descarga.BANDEJAS_INTERNOS_DEFAULT:
            segmentos = [folios for nombre, folios in llamadas if nombre == bandeja]
            self.assertEqual(len(segmentos), 2, bandeja)
            self.assertEqual(sorted(len(folios) for folios in segmentos), [2, 2])

    def test_main_procesar_propaga_workers_de_internos(self):
        argv_capturado = []

        def main_falso():
            argv_capturado.extend(sys.argv)
            return 0

        with patch.object(Parte1_descarga, "main", side_effect=main_falso):
            rc = main_procesar.ejecutar_descarga_internos(
                bandejas=["Atendidos"],
                headless=True,
                workers=4,
            )

        self.assertEqual(rc, 0)
        indice = argv_capturado.index("--internos-workers")
        self.assertEqual(argv_capturado[indice + 1], "4")

    def test_reporta_fallo_al_preparar_sesion_sin_crear_workers(self):
        with patch.object(
            Parte1_descarga,
            "_validar_sesion_internos",
            side_effect=RuntimeError("Chromium no disponible"),
        ), patch.object(Parte1_descarga, "_worker_bandeja_internos_con_reintentos") as worker:
            resultados = Parte1_descarga.descargar_internos_ift(
                bandejas=["Atendidos"],
                workers=1,
            )

        worker.assert_not_called()
        self.assertEqual(resultados[0]["archivo"], "ERROR_PREPARACION_SESION_INTERNOS")
        self.assertIn("Chromium no disponible", resultados[0]["error"])

    def test_cierra_popups_sin_cerrar_la_pagina_principal(self):
        class Pagina:
            def __init__(self, cerrada=False):
                self.cerrada = cerrada
                self.close_calls = 0

            def is_closed(self):
                return self.cerrada

            def close(self, run_before_unload=False):
                self.close_calls += 1
                self.cerrada = True

        principal = Pagina()
        popup_abierto = Pagina()
        popup_ya_cerrado = Pagina(cerrada=True)
        context = SimpleNamespace(pages=[principal, popup_abierto, popup_ya_cerrado])

        cerradas = Parte1_descarga._cerrar_paginas_emergentes(
            context,
            principal,
            motivo="prueba",
        )

        self.assertEqual(cerradas, 1)
        self.assertEqual(principal.close_calls, 0)
        self.assertEqual(popup_abierto.close_calls, 1)
        self.assertEqual(popup_ya_cerrado.close_calls, 0)

    def test_descarga_directa_no_pulsa_segundo_boton(self):
        class Boton:
            def __init__(self, al_click=None):
                self.click_calls = 0
                self.al_click = al_click

            def click(self, **kwargs):
                self.click_calls += 1
                if self.al_click:
                    self.al_click()

        class Page:
            def __init__(self):
                self.handlers = {}

            def on(self, evento, callback):
                self.handlers[evento] = callback

            def remove_listener(self, evento, callback):
                self.handlers.pop(evento, None)

            def wait_for_timeout(self, milisegundos):
                pass

        class Context:
            def __init__(self):
                self.handlers = {}

            def on(self, evento, callback):
                self.handlers[evento] = callback

            def remove_listener(self, evento, callback):
                self.handlers.pop(evento, None)

        page = Page()
        context = Context()
        descarga = SimpleNamespace(suggested_filename="directo.pdf")
        boton = Boton(lambda: page.handlers["download"](descarga))

        with patch.object(
            Parte1_descarga,
            "_encontrar_boton_documento_modal",
            return_value=None,
        ) as buscar_modal:
            dl_obj, popup_obj = Parte1_descarga._click_y_esperar_descarga(page, context, boton)

        self.assertIs(dl_obj, descarga)
        self.assertIsNone(popup_obj)
        self.assertEqual(boton.click_calls, 1)
        buscar_modal.assert_called_once_with(page)

    def test_ventana_intermedia_pulsa_segundo_ver_documento_y_captura_popup(self):
        class Boton:
            def __init__(self, al_click=None):
                self.click_calls = 0
                self.al_click = al_click

            def click(self, **kwargs):
                self.click_calls += 1
                if self.al_click:
                    self.al_click()

        class Page:
            def __init__(self):
                self.handlers = {}

            def on(self, evento, callback):
                self.handlers[evento] = callback

            def remove_listener(self, evento, callback):
                self.handlers.pop(evento, None)

            def wait_for_timeout(self, milisegundos):
                pass

        class Context:
            def __init__(self):
                self.handlers = {}

            def on(self, evento, callback):
                self.handlers[evento] = callback

            def remove_listener(self, evento, callback):
                self.handlers.pop(evento, None)

        page = Page()
        context = Context()
        popup = SimpleNamespace(url="https://satys.ift.org.mx/upload/anexo.pdf")
        boton_gris = Boton()
        boton_morado = Boton(lambda: context.handlers["page"](popup))
        busquedas = iter([None, boton_morado])

        with patch.object(
            Parte1_descarga,
            "_encontrar_boton_documento_modal",
            side_effect=lambda pagina: next(busquedas),
        ):
            dl_obj, popup_obj = Parte1_descarga._click_y_esperar_descarga(
                page,
                context,
                boton_gris,
            )

        self.assertIsNone(dl_obj)
        self.assertIs(popup_obj, popup)
        self.assertEqual(boton_gris.click_calls, 1)
        self.assertEqual(boton_morado.click_calls, 1)

    def test_cierra_modal_pdf_con_boton_cerrar(self):
        class BotonCerrar:
            def __init__(self):
                self.click_calls = 0

            def is_visible(self):
                return True

            def click(self, **kwargs):
                self.click_calls += 1

        boton = BotonCerrar()
        locator = SimpleNamespace(count=lambda: 1, last=boton)
        page = SimpleNamespace(
            locator=lambda selector: locator,
            wait_for_timeout=lambda milisegundos: None,
        )

        self.assertTrue(Parte1_descarga._cerrar_modal_documento(page))
        self.assertEqual(boton.click_calls, 1)

    def test_sin_sincronizar_omite_merge_completo_depi(self):
        with patch.object(main_procesar, "sincronizar_carpeta_compartida") as sincronizar:
            main_procesar._sincronizar_si_corresponde(
                SimpleNamespace(sin_sincronizar=True)
            )

        sincronizar.assert_not_called()

    def test_certificado_valida_las_seis_bandejas_y_sus_totales(self):
        output = self.tmp / "registros.txt"
        output.write_text("CRT26-000001\n", encoding="utf-8")
        bandejas = []
        for nombre in (
            "Recibidos",
            "En proceso",
            "Copias Marcadas",
            "Atendidos",
            "Ultimos Movimientos",
            "Fuera de tiempo",
        ):
            folios = ["148326", "147390"] if nombre == "En proceso" else []
            bandejas.append({
                "bandeja": nombre,
                "estado": "ENCONTRADOS_COMPLETOS" if folios else "VACIO_CONFIRMADO",
                "total_reportado_satys": len(folios),
                "filas_leidas": len(folios),
                "folios": folios,
                "filas_invalidas": 0,
            })
        resumen = {
            "estado": "COMPLETO",
            "integridad": "VALIDADA",
            "vacio_confirmado": False,
            "total_filas_satys": 1,
            "total_registros": 1,
            "por_anio": [{
                "anio": 2026,
                "estado": "ENCONTRADOS_COMPLETOS",
                "total_reportado_satys": 1,
                "filas_leidas": 1,
                "total_guardados_anio": 1,
                "filas_invalidas": 0,
                "contador_tab": 1,
                "tamanio_pagina": 100,
            }],
            "internos": {
                "estado": "COMPLETO",
                "integridad": "VALIDADA",
                "vacio_confirmado": False,
                "total_folios": 2,
                "total_filas_satys": 2,
                "folios": ["148326", "147390"],
                "por_bandeja": bandejas,
            },
        }
        output.with_suffix(".txt.json").write_text(
            json.dumps(resumen),
            encoding="utf-8",
        )

        validacion = monitor.validar_resumen_extractor(output, ["CRT26-000001"])
        self.assertTrue(validacion["ok"], validacion)
        self.assertEqual(validacion["folios_internos"], ["148326", "147390"])

    def test_metadata_internos_conserva_el_folio_de_la_tabla(self):
        carpeta = self.tmp / "148326"
        metadata = Parte1_descarga._actualizar_metadata_internos(
            carpeta=carpeta,
            bandeja="En proceso",
            firma="firma-prueba",
            row_meta={"folio": "148326", "tipo_tramite": "CGPE-01-008"},
            meta={"registro": "CRT26-034284", "concesionario": "Operador de Prueba"},
            resultados=[],
        )

        self.assertEqual(metadata["folio_tabla_internos"], "148326")
        self.assertEqual(metadata["registro"], "CRT26-034284")
        persisted = json.loads((carpeta / "metadata_satys.json").read_text(encoding="utf-8"))
        self.assertEqual(persisted["folio_tabla_internos"], "148326")

    def test_excel_conserva_dos_folios_internos_con_el_mismo_registro(self):
        import openpyxl

        excel = self.tmp / "TramitesCRT.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Internos"
        headers = [
            "Control", "Estado", "Origen", "1711", "Memo/Volante",
            "Solicitante Promovente", "Representante Legal",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        wb.save(excel)
        wb.close()

        for folio_interno in ("148326", "190823"):
            self.assertTrue(Parte4_excel.actualizar_excel(
                folio=folio_interno,
                folio_internos=folio_interno,
                registro="CRT26-034284",
                excel_path=excel,
                sheet_name="Internos",
            ))

        wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
        ws = wb["Internos"]
        headers = {str(cell.value): cell.column for cell in ws[1] if cell.value}
        valores = [
            str(ws.cell(row=row, column=headers["Folio Internos"]).value)
            for row in range(2, ws.max_row + 1)
        ]
        registros = [
            ws.cell(row=row, column=headers["1711"]).value
            for row in range(2, ws.max_row + 1)
        ]
        wb.close()

        self.assertEqual(valores, ["148326", "190823"])
        self.assertEqual(registros, ["CRT26-034284", "CRT26-034284"])


if __name__ == "__main__":
    unittest.main(verbosity=2)
