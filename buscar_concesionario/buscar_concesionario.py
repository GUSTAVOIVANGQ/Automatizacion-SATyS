"""
=============================================================
  BUSCADOR FUZZY DE CONCESIONARIOS - RPC IFT
=============================================================
FUENTE PRIMARIA (recomendada):
  Excel de Datos Abiertos del RPC IFT
  Archivo: 03_concesiones_permisos_autorizaciones_250326.xlsx
  Descargado de: https://rpc.ift.org.mx/vrpc/visor/downloads
  Columnas usadas:
    - NOMBRE OPERADOR  → nombre/denominación social
    - ID OPERADOR      → equivalente al idBp
    - ESTATUS          → para filtrar solo vigentes (opcional)
    - FOLIO ELECTRONICO→ folio del título habilitante

  VENTAJA: ~9,000 nombres únicos sin peticiones al servidor.
           Más rápido, sin riesgo de bloqueo HTTP 403.
  DESVENTAJA: Snapshot fijo (fecha del archivo). Si el IFT
              registra concesionarios nuevos después, no
              estarán hasta descargar un Excel actualizado.

FUENTE DE RESPALDO (API en línea) — ver sección comentada:
  Se puede reactivar si el Excel está desactualizado.
  Ver funciones descargar_catalogo_modo_a() y modo_b()
  al final del archivo.

CÓMO ENCONTRAR LA URL DEL ENDPOINT (método DevTools):
  1. Abre https://rpc.ift.org.mx/vrpc/ en Chrome/Edge
  2. Presiona F12 → pestaña "Network" → filtra por "Fetch/XHR"
  3. Escribe "TEL" en el campo "Nombre o denominación social"
  4. Haz clic en la petición → copia la "Request URL" completa
     Ejemplo: https://rpc.ift.org.mx/vrpc/api/concesionarios?q=TEL
  Si no aparece ninguna petición al escribir, usa el MODO B.
=============================================================
"""

import csv
import json
import os
import unicodedata
import re
import string
import sys

# Asegurar que la consola maneje UTF-8 para evitar errores con emojis
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        import codecs
        sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

from difflib import SequenceMatcher

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN — edita aquí
# ─────────────────────────────────────────────────────────────

# ── FUENTE PRIMARIA: Excel de Datos Abiertos ──────────────────
# Ruta al archivo Excel descargado del RPC IFT.
# Puedes usar ruta absoluta o relativa al directorio del script.
EXCEL_PATH = r"Area _de_descargas\03_concesiones_permisos_autorizaciones_250326.xlsx"

# Nombre de la hoja dentro del Excel (normalmente "copeau")
EXCEL_SHEET = "copeau"

# Columnas a leer (índice 0-based, según el Excel del IFT):
#   0  = ID CONCESION
#   1  = FOLIO ELECTRONICO
#   5  = ESTATUS
#   6  = ID OPERADOR  (equivalente a idBp)
#   7  = NOMBRE OPERADOR
COL_ID_OPERADOR   = 6
COL_NOMBRE        = 7
COL_ESTATUS       = 5
COL_FOLIO         = 1

# Filtrar solo concesiones vigentes al cargar el Excel.
# True  → solo VIGENTE y EN PROCESO DE PRÓRROGA (~7,900 nombres)
# False → todos los estatus, incluye terminadas/canceladas (~9,000 nombres)
SOLO_VIGENTES = False

# ── NOMBRES A BUSCAR ─────────────────────────────────────────
# Lista de nombres (posiblemente mal escritos) que quieres
# corregir buscando su forma oficial en el catálogo del IFT.
NOMBRES_BUSCAR = [
    "TELFCOMUNICACION Y MERCADOTECNIA DE MONTERREY, S.A DE C.V.",  # → TELECOMUNICACIÓN Y MERCADOTECNIA DE MONTERREY S.A. DE C.V.
    "ROSARIOSÁNCHEZ",                                               # → MARÍA DEL ROSARIO SÁNCHEZ MEZA
    "DANIELLOERASANTILLAN_",                                        # → DANIEL LOERA SANTILLÁN
    # Agrega más aquí...
]

# Cuántos resultados mostrar por nombre
TOP_N = 5

# ── CACHE ────────────────────────────────────────────────────
# Guarda el catálogo procesado en JSON/CSV para inspección.
# No se usa como caché de descarga (el Excel ya es la fuente).
GUARDAR_CATALOGO = True
CACHE_JSON = "catalogo_rpc.json"
CACHE_CSV  = "catalogo_rpc.csv"

# ─────────────────────────────────────────────────────────────
# SUFIJOS LEGALES (para normalización)
# ─────────────────────────────────────────────────────────────

LEGAL_SUFFIXES = [
    " s a p i de c v",
    " s a s de c v",
    " s a b de c v",
    " s de r l de c v",
    " s a de c v",
    " s a p i",
    " s a s",
    " s a b",
    " s de r l",
    " s a",
    " s c",
    " a c",
]


def quitar_sufijos_legales(texto: str) -> str:
    """Elimina sufijos legales comunes al final de un nombre normalizado."""
    texto = texto.strip()
    cambio = True
    while cambio:
        cambio = False
        for sufijo in LEGAL_SUFFIXES:
            if texto.endswith(sufijo):
                texto = texto[: -len(sufijo)].strip()
                cambio = True
                break
    return texto


def normalizar(texto: str) -> str:
    """Quita acentos, pasa a minúsculas, colapsa espacios y quita sufijos legales."""
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return quitar_sufijos_legales(texto)


def compactar(texto_normalizado: str) -> str:
    """Elimina espacios del texto ya normalizado para comparaciones compactas."""
    return texto_normalizado.replace(" ", "")


# ─────────────────────────────────────────────────────────────
# FUENTE PRIMARIA — Lectura desde Excel de Datos Abiertos
# ─────────────────────────────────────────────────────────────

ESTATUS_VIGENTES = {"VIGENTE", "EN PROCESO DE PRÓRROGA", "EN PROCESO DE PRORROGA"}


def cargar_catalogo_desde_excel(ruta_excel: str, nombre_hoja: str, solo_vigentes: bool = False) -> list:
    """
    Lee el Excel de Datos Abiertos del RPC IFT y devuelve una lista
    de dicts con claves: idBp, concesionario, estatus, folio.

    Deduplica por ID OPERADOR (idBp). Si un mismo operador tiene
    múltiples concesiones, se toma la primera aparición.
    """
    try:
        import openpyxl
    except ImportError:
        print("❌ Falta la librería 'openpyxl'. Instálala con:")
        print("   pip install openpyxl")
        return []

    # Resolver ruta relativa al directorio del propio script
    if not os.path.isabs(ruta_excel):
        base = os.path.dirname(os.path.abspath(__file__))
        ruta_excel = os.path.join(base, ruta_excel)

    if not os.path.exists(ruta_excel):
        print(f"❌ No se encontró el Excel en: {ruta_excel}")
        print("   Descárgalo de: https://rpc.ift.org.mx/vrpc/visor/downloads")
        print("   → Sección 'Concesiones' → botón verde 'ÁREA DE DESCARGAS'")
        return []

    print(f"📂 Leyendo Excel: {os.path.basename(ruta_excel)}")
    print(f"   Hoja: {nombre_hoja}")
    if solo_vigentes:
        print("   Filtro: solo VIGENTES y EN PROCESO DE PRÓRROGA")
    else:
        print("   Filtro: todos los estatus (vigentes + históricos)")
    print()

    wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)

    if nombre_hoja not in wb.sheetnames:
        print(f"❌ No existe la hoja '{nombre_hoja}'. Hojas disponibles: {wb.sheetnames}")
        return []

    ws = wb[nombre_hoja]

    visto_ids  = set()   # para deduplicar por ID OPERADOR
    visto_nom  = set()   # para deduplicar por nombre (operadores sin id)
    catalogo   = []
    omitidos   = 0
    fila_num   = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        fila_num += 1

        nombre  = str(row[COL_NOMBRE]).strip()  if row[COL_NOMBRE]  else ""
        idbp    = str(row[COL_ID_OPERADOR]).strip() if row[COL_ID_OPERADOR] else ""
        estatus = str(row[COL_ESTATUS]).strip() if row[COL_ESTATUS] else ""
        folio   = str(row[COL_FOLIO]).strip()   if row[COL_FOLIO]   else ""

        # Ignorar filas sin nombre
        if not nombre or nombre.upper() == "NONE":
            continue

        # Filtro de vigentes (opcional)
        if solo_vigentes:
            estatus_norm = unicodedata.normalize("NFD", estatus.upper())
            estatus_norm = "".join(c for c in estatus_norm if unicodedata.category(c) != "Mn")
            if not any(v in estatus_norm for v in {"VIGENTE", "PRORROGA"}):
                omitidos += 1
                continue

        # Deduplicar: primero por idBp, luego por nombre exacto
        if idbp and idbp != "None":
            if idbp in visto_ids:
                continue
            visto_ids.add(idbp)
        else:
            idbp = ""
            clave_nom = nombre.upper()
            if clave_nom in visto_nom:
                continue
            visto_nom.add(clave_nom)

        catalogo.append({
            "idBp":        idbp,
            "concesionario": nombre,
            "estatus":     estatus,
            "folio":       folio,
        })

    print(f"✅ Catálogo cargado desde Excel:")
    print(f"   Filas procesadas : {fila_num:,}")
    print(f"   Nombres únicos   : {len(catalogo):,}")
    if solo_vigentes:
        print(f"   Omitidos (no vigentes): {omitidos:,}")
    print()
    return catalogo


def guardar_catalogo(catalogo: list) -> None:
    """Guarda el catálogo procesado en JSON y CSV para inspección."""
    with open(CACHE_JSON, "w", encoding="utf-8") as f:
        json.dump(catalogo, f, ensure_ascii=False, indent=2)
    with open(CACHE_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["idBp", "concesionario", "estatus", "folio"])
        writer.writeheader()
        writer.writerows(catalogo)


# ─────────────────────────────────────────────────────────────
# FUZZY MATCHING
# ─────────────────────────────────────────────────────────────

def score_similitud_precomp(a_norm: str, a_compact: str, b_norm: str, b_compact: str) -> float:
    """Combina similitud de secuencia, tokens y comparación compacta."""
    seq = SequenceMatcher(None, a_norm, b_norm).ratio()

    tokens_a = {t for t in a_norm.split() if len(t) > 2}
    tokens_b = {t for t in b_norm.split() if len(t) > 2}
    if tokens_a or tokens_b:
        token_score = len(tokens_a & tokens_b) / max(len(tokens_a | tokens_b), 1)
    else:
        token_score = 0.0

    diff_len = abs(len(a_norm) - len(b_norm)) / max(len(a_norm), len(b_norm), 1)
    penalizacion = diff_len * 0.1

    base_score = (seq * 0.5) + (token_score * 0.5) - penalizacion
    score = base_score

    if a_compact and b_compact:
        compact_score = SequenceMatcher(None, a_compact, b_compact).ratio()
        if compact_score > score:
            score = compact_score
        if a_compact in b_compact or b_compact in a_compact:
            score = max(score, min(1.0, compact_score + 0.08))

    return round(max(score, 0.0), 4)


def preparar_catalogo_para_matching(catalogo: list) -> list:
    """Pre-computa formas normalizadas y compactas para cada entrada."""
    preparado = []
    for item in catalogo:
        nombre = item.get("concesionario", "").strip()
        if not nombre:
            continue
        norm = normalizar(nombre)
        preparado.append({
            "idBp":          item.get("idBp", "").strip(),
            "concesionario": nombre,
            "estatus":       item.get("estatus", ""),
            "folio":         item.get("folio", ""),
            "norm":          norm,
            "compact":       compactar(norm),
        })
    return preparado


def buscar_coincidencias(nombre_malo: str, catalogo: list, top_n: int = 5) -> list:
    """Busca y rankea los top_n nombres más similares en el catálogo."""
    a_norm    = normalizar(nombre_malo)
    a_compact = compactar(a_norm)
    resultados = []
    for item in catalogo:
        score = score_similitud_precomp(a_norm, a_compact, item["norm"], item["compact"])
        resultados.append((score, item))
    resultados.sort(key=lambda x: x[0], reverse=True)
    return resultados[:top_n]


# ─────────────────────────────────────────────────────────────
# BÚSQUEDA EXACTA POR ID OPERADOR (id_solicitante)
# ─────────────────────────────────────────────────────────────

def buscar_por_id_solicitante(id_solicitante: str, catalogo: list) -> dict | None:
    """
    Busca el concesionario comparando el id_solicitante del JSON con la
    columna 'ID OPERADOR' (idBp) del catálogo Excel del RPC-IFT.

    Esta función reemplaza la similitud fuzzy por nombre: es una comparación
    exacta de IDs, por lo que el score resultante es siempre 1.0 (100%).

    Parámetros
    ----------
    id_solicitante : str
        Valor del campo 'id_solicitante' leído del metadata_satys.json.
    catalogo : list
        Lista de dicts preparada por preparar_catalogo_para_matching().
        Cada elemento tiene al menos: idBp, concesionario, estatus, folio.

    Retorna
    -------
    dict | None
        Si se encontró: dict con claves compatibles con rpc_resultado
        (nombre_completo, numero_rpc, idBp, score, ok).
        Si no se encontró: None.
    """
    if not id_solicitante:
        return None

    id_buscar = str(id_solicitante).strip()

    for item in catalogo:
        id_catalogo = str(item.get("idBp", "")).strip()
        if id_catalogo and id_catalogo == id_buscar:
            return {
                "nombre_completo": item.get("concesionario", ""),
                "numero_rpc":      item.get("idBp", ""),
                "idBp":            item.get("idBp", ""),
                "estatus":         item.get("estatus", ""),
                "folio_rpc":       item.get("folio", ""),
                "score":           1.0,   # coincidencia exacta = 100%
                "ok":              True,
                "empate":          False,
                "metodo":          "id_exacto",
            }

    return None


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  BUSCADOR FUZZY DE CONCESIONARIOS — RPC IFT")
    print("=" * 65 + "\n")

    # ── Cargar catálogo desde Excel ───────────────────────────
    catalogo = cargar_catalogo_desde_excel(EXCEL_PATH, EXCEL_SHEET, solo_vigentes=SOLO_VIGENTES)

    if not catalogo:
        print("❌ No se pudo cargar el catálogo desde el Excel.")
        print("   Verifica la ruta en EXCEL_PATH y que el archivo exista.")
        print()
        print("   Como alternativa puedes reactivar la descarga por API:")
        print("   busca la sección '# RESPALDO: DESCARGA POR API' al final")
        print("   de este script y sigue las instrucciones.")
        return

    # Guardar copia del catálogo para inspección (opcional)
    if GUARDAR_CATALOGO:
        guardar_catalogo(catalogo)
        print(f"💾 Catálogo guardado en '{CACHE_JSON}' y '{CACHE_CSV}'\n")

    print("=" * 65)

    # ── Fuzzy matching ────────────────────────────────────────
    catalogo_preparado = preparar_catalogo_para_matching(catalogo)
    resultados_json = []

    for nombre in NOMBRES_BUSCAR:
        print(f"\n🔍 Buscando: '{nombre}'")
        print("-" * 65)
        matches = buscar_coincidencias(nombre, catalogo_preparado, top_n=TOP_N)
        salida_matches = []
        for i, (score, match) in enumerate(matches, 1):
            barra = "█" * int(score * 20)
            if score >= 0.70:
                icono = "✅"
            elif score >= 0.50:
                icono = "⚠️ "
            else:
                icono = "❌"
            folio_str = f"  [{match['folio']}]" if match.get("folio") else ""
            print(f"  {i}. {icono} [{barra:<20}] {score:.0%}  →  {match['concesionario']}{folio_str}")
            salida_matches.append({
                "score":         score,
                "idBp":          match.get("idBp", ""),
                "concesionario": match.get("concesionario", ""),
                "estatus":       match.get("estatus", ""),
                "folio":         match.get("folio", ""),
            })

        resultados_json.append({
            "input":   nombre,
            "matches": salida_matches,
        })

    with open("matching_resultados.json", "w", encoding="utf-8") as f:
        json.dump(resultados_json, f, ensure_ascii=False, indent=2)
    print("\n💾 Resultados guardados en 'matching_resultados.json'")

    print("\n" + "=" * 65)
    print("✅ Listo. Guía de interpretación:")
    print("   ✅ > 70%  → Probablemente correcto")
    print("   ⚠️  50–70% → Revisar manualmente")
    print("   ❌ < 50%  → El nombre puede no estar en el catálogo")


# ═════════════════════════════════════════════════════════════
# RESPALDO: DESCARGA POR API (Modo A — Autocompletado)
# ═════════════════════════════════════════════════════════════
# Descomenta este bloque y ajusta main() si el Excel está
# desactualizado o no está disponible.
# ─────────────────────────────────────────────────────────────
#
# import requests, time
# from collections import deque
# from urllib.parse import quote
#
# AUTOCOMPLETE_URL   = "https://rpc.ift.org.mx/vrpc//RpcServicesController/searchBP"
# AUTOCOMPLETE_PARAM = "query"
# SEARCH_URL         = "https://rpc.ift.org.mx/vrpc/"
# ALPHABET           = string.ascii_uppercase + string.digits
# BASE_PREFIXES      = list(ALPHABET)
# MAX_PREFIX_LEN     = 5
# SLEEP_SECONDS      = 0.2
#
# HEADERS = {
#     "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0",
#     "Accept":          "application/json, text/html, */*",
#     "Accept-Language": "es-MX,es;q=0.9",
#     "Referer":         "https://rpc.ift.org.mx/vrpc/",
# }
#
# def extraer_concesionarios(data) -> list:
#     if isinstance(data, list):
#         if not data:
#             return []
#         if isinstance(data[0], str):
#             return [{"idBp": "", "concesionario": item, "estatus": "", "folio": ""} for item in data]
#         if isinstance(data[0], dict):
#             resultado = []
#             for item in data:
#                 nombre = (
#                     item.get("concesionario") or item.get("nombre")
#                     or item.get("denominacion") or item.get("label")
#                     or item.get("value") or item.get("text")
#                     or item.get("razon_social")
#                 )
#                 idbp = item.get("idBp") or item.get("id") or item.get("id_bp")
#                 if nombre:
#                     resultado.append({
#                         "idBp":          str(idbp or "").strip(),
#                         "concesionario": str(nombre).strip(),
#                         "estatus":       "",
#                         "folio":         "",
#                     })
#             return resultado
#     if isinstance(data, dict):
#         for clave in ("data", "results", "items", "concesionarios", "lista", "nombres"):
#             if clave in data:
#                 return extraer_concesionarios(data[clave])
#     return []
#
# def fetch_autocomplete(session, query: str) -> list:
#     url = AUTOCOMPLETE_URL
#     params = {AUTOCOMPLETE_PARAM: query}
#     try:
#         resp = session.get(url, params=params, timeout=10)
#     except Exception as e:
#         print(f"  {query}: error — {e}")
#         return []
#     if resp.status_code == 403:
#         print(f"  HTTP 403 en '{query}' — servidor rechaza la petición.")
#         return []
#     if resp.status_code != 200:
#         return []
#     try:
#         data = resp.json()
#     except Exception:
#         return []
#     return extraer_concesionarios(data)
#
# def descargar_catalogo_modo_a() -> list:
#     """Descarga por prefijos adaptativos con deduplicación por idBp."""
#     catalogo = {}
#     session = requests.Session()
#     session.headers.update(HEADERS)
#     try:
#         session.get(SEARCH_URL, timeout=10)
#     except Exception:
#         pass
#     # Detectar límite de resultados por query
#     muestras = ["A", "E", "T", "TE", "TEL"]
#     conteos = [len(fetch_autocomplete(session, p)) for p in muestras]
#     limite = max(conteos) if conteos and max(conteos) >= 20 else 0
#     if limite:
#         print(f"Limite detectado por consulta: {limite}\n")
#     cola = deque(BASE_PREFIXES)
#     vistos = set(BASE_PREFIXES)
#     print(f"[MODO A] Descargando catálogo ({len(BASE_PREFIXES)} prefijos base)...\n")
#     while cola:
#         prefijo = cola.popleft()
#         items = fetch_autocomplete(session, prefijo)
#         if not items:
#             continue
#         for item in items:
#             nombre = item.get("concesionario", "").strip()
#             idbp   = item.get("idBp", "").strip()
#             if not nombre:
#                 continue
#             key = idbp if idbp else f"NOID:{nombre}"
#             if key not in catalogo:
#                 catalogo[key] = item
#         if limite and len(items) >= limite and len(prefijo) < MAX_PREFIX_LEN:
#             for ch in ALPHABET:
#                 np = prefijo + ch
#                 if np not in vistos:
#                     vistos.add(np)
#                     cola.append(np)
#         time.sleep(SLEEP_SECONDS)
#     lista = sorted(catalogo.values(), key=lambda x: x["concesionario"])
#     print(f"\nCatalogo Modo A: {len(lista)} nombres unicos\n")
#     return lista
#
# ═════════════════════════════════════════════════════════════
# RESPALDO: DESCARGA POR SCRAPING (Modo B — Paginación)
# ═════════════════════════════════════════════════════════════
# Solo si el Modo A tampoco funciona. Más lento.
# ─────────────────────────────────────────────────────────────
#
# SEARCH_PARAMS = {
#     "denominacion": "{query}",
#     "tipo_servicio": "",
#     "pagina": "{page}",
# }
#
# def extraer_nombres_de_html(html: str) -> list:
#     nombres = []
#     for patron in [
#         r'class="[^"]*denominacion[^"]*"[^>]*>([^<]+)<',
#         r'class="[^"]*nombre[^"]*"[^>]*>([^<]+)<',
#         r'<td[^>]*>([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s,\.]+(?:S\.A\.|S\.C\.|DE C\.V\.|A\.C\.)?[^<]*)<\/td>',
#     ]:
#         matches = re.findall(patron, html, re.IGNORECASE)
#         nombres.extend([m.strip() for m in matches if len(m.strip()) > 5])
#     json_matches = re.findall(r'"denominacion"\s*:\s*"([^"]+)"', html)
#     nombres.extend(json_matches)
#     return list(set(nombres))
#
# def descargar_catalogo_modo_b() -> list:
#     import requests as req
#     catalogo = set()
#     letras = list(string.ascii_uppercase)
#     session = req.Session()
#     session.headers.update(HEADERS)
#     try:
#         session.get(SEARCH_URL, timeout=10)
#     except Exception as e:
#         print(f"No se pudo conectar: {e}")
#         return []
#     for letra in letras:
#         pagina = 1
#         while True:
#             params = {k: v.replace("{query}", letra).replace("{page}", str(pagina))
#                       for k, v in SEARCH_PARAMS.items()}
#             try:
#                 resp = session.post(SEARCH_URL, data=params, timeout=15)
#                 if resp.status_code != 200:
#                     break
#                 nombres_pagina = extraer_nombres_de_html(resp.text)
#                 if not nombres_pagina:
#                     break
#                 nuevos = [n for n in nombres_pagina if n not in catalogo]
#                 catalogo.update(nuevos)
#                 print(f"  {letra} p{pagina}: +{len(nuevos)} (total: {len(catalogo)})")
#                 pagina += 1
#                 time.sleep(0.3)
#             except Exception as e:
#                 print(f"  {letra} p{pagina}: error — {e}")
#                 break
#     lista = [{"idBp": "", "concesionario": n, "estatus": "", "folio": ""} for n in sorted(catalogo)]
#     print(f"\nCatalogo Modo B: {len(lista)} nombres unicos\n")
#     return lista


if __name__ == "__main__":
    main()
