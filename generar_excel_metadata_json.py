#!/usr/bin/env python3
r"""
generar_excel_metadata_json.py
──────────────────────────────
Genera output/Folios_Datos_Completos.xlsx con todos los campos de:
  - metadata_satys.json
  - metadata_tramite_nuevo.json
por cada carpeta/número de registro procesado.

También agrega las columnas:
  - output: ruta relativa donde quedaron organizados los archivos
  - descargas: ruta relativa de la carpeta fuente en descargas/
"""

from __future__ import annotations

import json
import logging
import uuid
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise RuntimeError("Falta openpyxl. Instala con: python -m pip install openpyxl") from exc

from estado_descargas import auditar_carpeta_descarga, slug_bandeja_internos
from guardado_seguro import reemplazar_desde_temporal
from rutas_salida import destino_sin_operador, folio_opc_desde_metadata

log = logging.getLogger("SATyS-ExcelMetadata")

JSON_NAMES = ("metadata_completo.json", "metadata_satys.json", "metadata_tramite_nuevo.json")


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        return data if isinstance(data, dict) else {}
    except Exception as exc:
        log.warning("No se pudo leer %s: %s", path, exc)
        return {}


def _rel_backslash(path_value: str | Path | None, project_root: Path | None = None) -> str:
    if not path_value:
        return ""
    p = Path(path_value)
    root = project_root or Path.cwd()
    try:
        rel = p.resolve().relative_to(root.resolve())
    except Exception:
        try:
            rel = p.relative_to(root)
        except Exception:
            rel = p
    text = str(rel).replace("/", "\\")
    if not text.startswith("\\"):
        text = "\\" + text
    return text


def _registro_from(meta_satys: dict[str, Any], meta_tn: dict[str, Any], fallback: str = "") -> str:
    for key in ("registro", "numero_registro", "1711"):
        val = meta_satys.get(key) or meta_tn.get(key)
        if val:
            return str(val).strip().upper()
    return str(fallback or "").strip().upper()


def _folio_from(meta_satys: dict[str, Any], meta_tn: dict[str, Any], fallback: str = "") -> str:
    # En Internos el Folio mostrado por la tabla es la identidad estable del
    # objetivo. ``folio`` dentro del detalle puede contener otro valor legacy.
    for key in ("folio_tabla_internos", "folio", "folio_opc", "memo_folio_opc"):
        val = meta_satys.get(key) or meta_tn.get(key)
        if val:
            return str(val).strip()
    return str(fallback or "").strip()


def _descargas_from_result(resultado: dict[str, Any], descargas_base: Path) -> Path:
    for key in ("descargas_dir", "carpeta", "carpeta_descarga"):
        val = resultado.get(key)
        if val:
            return Path(val)
    folio_id = resultado.get("folio_id") or resultado.get("registro") or resultado.get("folio")
    return descargas_base / str(folio_id or "")


def _output_from_result(
    resultado: dict[str, Any],
    output_base: Path,
    meta_satys: dict[str, Any] | None = None,
    meta_tn: dict[str, Any] | None = None,
) -> Path:
    for key in ("output_dir", "sin_operador_dir"):
        val = resultado.get(key)
        if val:
            return Path(val)
    rpc = resultado.get("rpc_resultado") or {}
    if isinstance(rpc, dict) and rpc.get("ok") and rpc.get("ruta"):
        return output_base / str(rpc["ruta"]).replace("\\", "/")
    folio_id = resultado.get("folio_id") or resultado.get("registro") or resultado.get("folio")
    folio_opc = resultado.get("folio_opc") or folio_opc_desde_metadata(meta_satys, meta_tn)
    return destino_sin_operador(output_base, folio_id, folio_opc)


def _scan_descargas(descargas_base: Path) -> list[dict[str, Any]]:
    resultados: list[dict[str, Any]] = []
    if not descargas_base.exists():
        return resultados
    for carpeta in sorted([p for p in descargas_base.rglob("*") if p.is_dir()], key=lambda p: str(p).upper()):
        if any((carpeta / name).exists() for name in JSON_NAMES):
            resultados.append({
                "folio": carpeta.name,
                "folio_id": carpeta.name,
                "descargas_dir": str(carpeta),
            })
    return resultados


def validar_excel_metadata_json(
    excel_path: str | Path,
    filas_esperadas: list[dict[str, Any]] | None = None,
    objetivos_esperados: list[tuple[str, str]] | None = None,
) -> dict[str, int]:
    """Reabre el XLSX y concilia filas/objetivos antes de publicarlo."""
    excel_path = Path(excel_path)
    if not excel_path.is_file() or excel_path.stat().st_size <= 0:
        raise ValueError(f"El Excel generado no existe o está vacío: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=False)
    try:
        if "Datos_Completos" not in wb.sheetnames or "Resumen" not in wb.sheetnames:
            raise ValueError(
                "El Excel no contiene las hojas obligatorias Datos_Completos y Resumen."
            )
        ws = wb["Datos_Completos"]
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        requeridos = {
            "registro",
            "folio",
            "bandeja_internos",
            "folio_tabla_internos",
            "estado_descarga",
            "descarga_completa",
            "rpc_ok",
            "output",
            "descargas",
        }
        faltan_headers = sorted(requeridos - set(headers))
        if faltan_headers:
            raise ValueError(f"Faltan encabezados obligatorios: {', '.join(faltan_headers)}")
        if len(headers) != len(set(headers)):
            raise ValueError("El Excel contiene encabezados duplicados.")

        indices = {header: idx for idx, header in enumerate(headers)}
        filas = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        esperadas = list(filas_esperadas or [])
        if filas_esperadas is not None and len(filas) != len(esperadas):
            raise ValueError(
                f"Filas generadas={len(filas)}; filas esperadas={len(esperadas)}."
            )

        rutas = [
            str(row[indices["descargas"]] or "").strip().casefold()
            for row in filas
            if str(row[indices["descargas"]] or "").strip()
        ]
        if len(rutas) != len(set(rutas)):
            raise ValueError("El Excel contiene rutas de descargas duplicadas.")

        pares_reporte = {
            (
                str(row[indices["bandeja_internos"]] or "").strip().casefold(),
                str(row[indices["folio_tabla_internos"]] or "").strip(),
            )
            for row in filas
            if str(row[indices["bandeja_internos"]] or "").strip()
            and str(row[indices["folio_tabla_internos"]] or "").strip()
        }
        pares_esperados = {
            (str(bandeja or "").strip().casefold(), str(folio or "").strip())
            for bandeja, folio in (objetivos_esperados or [])
            if str(bandeja or "").strip() and str(folio or "").strip()
        }
        faltan_objetivos = sorted(pares_esperados - pares_reporte)
        if faltan_objetivos:
            muestra = ", ".join(f"{b}/{f}" for b, f in faltan_objetivos[:20])
            raise ValueError(f"El Excel omitió objetivos esperados: {muestra}")

        ws_resumen = wb["Resumen"]
        if int(ws_resumen["B2"].value or -1) != len(filas):
            raise ValueError("La métrica Total del Resumen no coincide con Datos_Completos.")
        return {
            "total_filas": len(filas),
            "objetivos_cubiertos": len(pares_esperados & pares_reporte),
        }
    finally:
        wb.close()


def generar_excel_metadata_json(
    resultados: list[dict[str, Any]] | None = None,
    descargas_base: str | Path = "descargas",
    output_base: str | Path = "output",
    excel_salida: str | Path | None = None,
    project_root: str | Path | None = None,
    objetivos_esperados: list[dict[str, Any]] | None = None,
) -> Path:
    """
    Genera el Excel consolidado de metadatos JSON.

    Usa 'resultados' de main_procesar.py para ubicar output/_sin_operador con
    precisión. Si no se pasan resultados, escanea descargas/.
    """
    descargas_base = Path(descargas_base)
    output_base = Path(output_base)
    project_root = Path(project_root) if project_root else Path.cwd()
    excel_salida = Path(excel_salida) if excel_salida else output_base / "Folios_Datos_Completos.xlsx"
    excel_salida.parent.mkdir(parents=True, exist_ok=True)

    escanear_descargas = resultados is None
    resultados = list(resultados or [])
    objetivos_esperados = list(objetivos_esperados or [])
    if escanear_descargas:
        resultados = _scan_descargas(descargas_base)

    objetivos_normalizados: list[tuple[str, str]] = []
    objetivos_vistos_entrada: set[tuple[str, str]] = set()
    for item in objetivos_esperados:
        if not isinstance(item, dict):
            continue
        bandeja = str(item.get("bandeja") or "").strip()
        folio_objetivo = str(item.get("folio") or "").strip()
        clave = (bandeja.casefold(), folio_objetivo)
        if not bandeja or not folio_objetivo or clave in objetivos_vistos_entrada:
            continue
        objetivos_vistos_entrada.add(clave)
        objetivos_normalizados.append((bandeja, folio_objetivo))

    filas: list[dict[str, Any]] = []
    satys_keys: set[str] = set()
    tramite_keys: set[str] = set()
    vistos_descargas: set[str] = set()
    objetivos_con_fila: set[tuple[str, str]] = set()

    for resultado in resultados:
        carpeta = _descargas_from_result(resultado, descargas_base)
        try:
            key_carpeta = str(carpeta.resolve())
        except Exception:
            key_carpeta = str(carpeta)
        if key_carpeta in vistos_descargas:
            continue
        vistos_descargas.add(key_carpeta)

        meta_completo = _read_json(carpeta / "metadata_completo.json")
        meta_satys = _read_json(carpeta / "metadata_satys.json")
        meta_tn = _read_json(carpeta / "metadata_tramite_nuevo.json")
        if not meta_satys and isinstance(meta_completo.get("metadatos_satys"), dict):
            meta_satys = dict(meta_completo["metadatos_satys"])
        if not meta_tn and isinstance(meta_completo.get("metadatos_tramite"), dict):
            meta_tn = dict(meta_completo["metadatos_tramite"])
        if not meta_satys and not meta_tn and not meta_completo and not resultado:
            continue

        satys_keys.update(str(k) for k in meta_satys.keys())
        tramite_keys.update(str(k) for k in meta_tn.keys())

        registro = _registro_from(meta_satys, meta_tn, resultado.get("registro") or resultado.get("folio_id") or carpeta.name)
        folio = _folio_from(meta_satys, meta_tn, resultado.get("folio") or carpeta.name)
        rpc = resultado.get("rpc_resultado") if isinstance(resultado.get("rpc_resultado"), dict) else {}
        output_dir = _output_from_result(resultado, output_base, meta_satys, meta_tn)
        bandeja_internos = str(
            meta_satys.get("bandeja_internos")
            or meta_tn.get("bandeja_internos")
            or resultado.get("bandeja_internos")
            or ""
        ).strip()
        folio_tabla_internos = str(
            meta_satys.get("folio_tabla_internos")
            or meta_tn.get("folio_tabla_internos")
            or resultado.get("folio_tabla_internos")
            or folio
            or ""
        ).strip()
        if bandeja_internos and folio_tabla_internos:
            objetivos_con_fila.add((bandeja_internos.casefold(), folio_tabla_internos))
        try:
            total_archivos = int(
                meta_completo.get("total_archivos_encontrados")
                or meta_completo.get("total_archivos")
                or 0
            )
            total_ok = int(meta_completo.get("total_archivos_ok") or 0)
            total_error = int(meta_completo.get("total_archivos_error") or 0)
        except (TypeError, ValueError):
            total_archivos = total_ok = total_error = 0
        auditoria_descarga = auditar_carpeta_descarga(carpeta)
        descarga_completa = bool(auditoria_descarga["completo"])

        filas.append({
            "registro": registro,
            "folio": folio,
            "bandeja_internos": bandeja_internos,
            "folio_tabla_internos": folio_tabla_internos,
            "estado_descarga": meta_completo.get("estado") or "SIN_AUDITORIA",
            "descarga_completa": descarga_completa,
            "motivos_descarga_incompleta": " | ".join(auditoria_descarga["motivos"]),
            "total_archivos": total_archivos,
            "total_archivos_ok": total_ok,
            "total_archivos_error": total_error,
            "id_solicitante": meta_satys.get("id_solicitante") or meta_tn.get("id_solicitante") or resultado.get("id_solicitante") or "",
            "nombre_operador": meta_satys.get("nombre_operador") or meta_tn.get("nombre_operador") or resultado.get("nombre_operador") or "",
            "representante_legal": meta_satys.get("representante_legal") or meta_tn.get("representante_legal") or resultado.get("representante_legal") or "",
            "rpc_ok": bool(resultado.get("rpc_ok")),
            "rpc_exactitud": 100 if resultado.get("rpc_ok") else 0,
            "rpc_metodo": rpc.get("metodo") or "",
            "rpc_motivo": rpc.get("motivo") or "",
            "rpc_id_operador": rpc.get("idBp") or rpc.get("numero_rpc") or "",
            "output": _rel_backslash(output_dir, project_root),
            "descargas": _rel_backslash(carpeta, project_root),
            "_meta_satys": meta_satys,
            "_meta_tramite_nuevo": meta_tn,
        })

    # Un reporte de una corrida dirigida debe representar también los objetivos
    # que no lograron producir metadata. Así el Excel nunca aparenta que el lote
    # quedó completo sólo porque las filas fallidas desaparecieron.
    for bandeja, folio_objetivo in objetivos_normalizados:
        clave = (bandeja.casefold(), folio_objetivo)
        if clave in objetivos_con_fila:
            continue
        carpeta_esperada = (
            descargas_base
            / "internos"
            / slug_bandeja_internos(bandeja)
            / folio_objetivo
        )
        filas.append({
            "registro": folio_objetivo,
            "folio": folio_objetivo,
            "bandeja_internos": bandeja,
            "folio_tabla_internos": folio_objetivo,
            "estado_descarga": "FALTANTE",
            "descarga_completa": False,
            "motivos_descarga_incompleta": "carpeta_inexistente_o_sin_metadata_completa",
            "total_archivos": 0,
            "total_archivos_ok": 0,
            "total_archivos_error": 0,
            "id_solicitante": "",
            "nombre_operador": "",
            "representante_legal": "",
            "rpc_ok": False,
            "rpc_exactitud": 0,
            "rpc_metodo": "",
            "rpc_motivo": "sin_descarga_o_metadata_completa",
            "rpc_id_operador": "",
            "output": "",
            "descargas": _rel_backslash(carpeta_esperada, project_root),
            "_meta_satys": {},
            "_meta_tramite_nuevo": {},
        })
        objetivos_con_fila.add(clave)

    satys_cols = [f"metadata_satys.{k}" for k in sorted(satys_keys, key=str.lower)]
    tramite_cols = [f"metadata_tramite_nuevo.{k}" for k in sorted(tramite_keys, key=str.lower)]
    base_cols = [
        "registro", "folio", "bandeja_internos", "folio_tabla_internos",
        "estado_descarga", "descarga_completa", "motivos_descarga_incompleta", "total_archivos",
        "total_archivos_ok", "total_archivos_error", "id_solicitante",
        "nombre_operador", "representante_legal", "rpc_ok", "rpc_exactitud",
        "rpc_metodo", "rpc_motivo", "rpc_id_operador", "output", "descargas",
    ]
    headers = base_cols + satys_cols + tramite_cols

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos_Completos"
    ws.append(headers)

    for fila in filas:
        row = []
        for h in headers:
            if h.startswith("metadata_satys."):
                v = fila["_meta_satys"].get(h.split(".", 1)[1], "")
            elif h.startswith("metadata_tramite_nuevo."):
                v = fila["_meta_tramite_nuevo"].get(h.split(".", 1)[1], "")
            else:
                v = fila.get(h, "")
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            row.append(v)
        ws.append(row)

    # Estilo simple y legible
    header_fill = PatternFill("solid", fgColor="156E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2E3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if ws.max_row >= 2 and ws.max_column >= 1:
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName="TablaDatosCompletos", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

    # Anchos acotados
    for col_idx, h in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        if h in {"output", "descargas"} or h.startswith("metadata_"):
            ws.column_dimensions[letter].width = 34
        elif h in {"nombre_operador", "representante_legal"}:
            ws.column_dimensions[letter].width = 32
        else:
            ws.column_dimensions[letter].width = min(max(len(h) + 2, 12), 24)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    total = len(filas)
    ok = sum(1 for f in filas if f.get("rpc_ok"))
    sin = sum(
        1 for f in filas
        if not f.get("rpc_ok") and f.get("estado_descarga") != "FALTANTE"
    )
    rpc_pendiente_descarga = sum(
        1 for f in filas if f.get("estado_descarga") == "FALTANTE"
    )
    descargas_completas = sum(1 for f in filas if f.get("descarga_completa"))
    ws2.append(["Métrica", "Valor"])
    ws2.append(["Total filas del reporte", total])
    ws2.append(["Descargas completas y auditadas", descargas_completas])
    ws2.append(["Descargas faltantes o incompletas", total - descargas_completas])
    ws2.append(["RPC resuelto con evidencia segura", ok])
    ws2.append(["RPC 0% / sin operador", sin])
    ws2.append(["RPC pendiente por falta de descarga", rpc_pendiente_descarga])
    ws2.append(["Archivo generado", str(excel_salida)])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 60

    temporal = excel_salida.with_name(
        f".{excel_salida.stem}.{uuid.uuid4().hex}.tmp.xlsx"
    )
    wb.save(temporal)
    wb.close()
    try:
        validacion = validar_excel_metadata_json(
            temporal,
            filas_esperadas=filas,
            objetivos_esperados=objetivos_normalizados,
        )
        reemplazar_desde_temporal(temporal, excel_salida)
    finally:
        temporal.unlink(missing_ok=True)
    log.info(
        "Excel de metadatos JSON generado y validado: %s (%d registros, %d objetivos cubiertos)",
        excel_salida,
        validacion["total_filas"],
        validacion["objetivos_cubiertos"],
    )
    return excel_salida


if __name__ == "__main__":
    generar_excel_metadata_json()
