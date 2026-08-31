#!/usr/bin/env python3
"""Repara al final de la corrida las Rutas ``_sin_operador`` usando sólo RPC público.

Regla de esta etapa final:
- ``TrámitesCRT.xlsx`` determina qué filas siguen en ``_sin_operador``.
- La columna ``1711`` es la llave de conciliación con ``output/_sin_operador`` y
  con los expedientes originales bajo ``descargas``.
- ``metadata_satys.json`` aporta exclusivamente ``nombre_operador`` para la
  consulta al buscador público del RPC.
- NO se carga ni consulta el Excel oficial RPC en esta etapa.
- Si el RPC público devuelve una resolución segura, los documentos originales
  se fusionan en ``output/<ID>_<nombre>/01 EN/VE``. No se crean sufijos de
  colisión; la copia actual de ``descargas`` prevalece sobre ``output``.
- Tras verificar la organización local se actualiza ``Ruta`` en el Excel y se
  replica únicamente el cambio de ``output`` a la carpeta compartida DEPI.
- Después del intento RPC, cualquier fila que siga en ``_sin_operador`` y cuya
  fuente original en ``descargas`` contenga ``MEMORANDO/MEMORANDUM.pdf`` se clasifica en
  ``output/_sin_operador/(correos)/<carpeta>``. ``descargas`` nunca se mueve.

La etapa es idempotente: una fila ya reparada deja de cumplir ``Ruta=_sin_operador``
y no vuelve a procesarse en la siguiente corrida.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import re
import shutil
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl

import buscar_concesionario as bc
from configuracion_local import carpeta_compartida, ruta_configurada
from estado_descargas import es_archivo_publicable_output, iter_archivos_publicables_output
from guardado_seguro import reemplazar_desde_temporal
from Parte3_rpc import construir_ruta
from Parte4_excel import copiar_archivo_robusto, eliminar_arbol_robusto
from proceso_lock import ProcesoLock, LockOcupadoError
from sincronizacion_depi import (
    ResultadoSincronizacion,
    copiar_directorio_merge,
    validar_destino_compartido,
)

log = logging.getLogger("SATyS-SinOperador-RPCPublico")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)

PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL = ruta_configurada("excel", "TrámitesCRT.xlsx")
DEFAULT_DESCARGAS = ruta_configurada("descargas", "descargas")
DEFAULT_OUTPUT = ruta_configurada("output", "output")
DEFAULT_SHARED = carpeta_compartida()
DEFAULT_LOGS = ruta_configurada("logs", "logs")
SHEETS_DEFAULT = ("Turnados recibidos", "Internos")
METADATA_NAME = "metadata_satys.json"
CORREOS_DIR_NAME = "(correos)"


def _texto(value: Any) -> str:
    return str(value or "").strip()


def _clave(value: Any) -> str:
    texto = unicodedata.normalize("NFD", _texto(value))
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"\s+", "", texto.upper())
    return texto


def _normalizar_header(value: Any) -> str:
    texto = unicodedata.normalize("NFD", _texto(value))
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"[^A-Z0-9]+", " ", texto.upper())
    return re.sub(r"\s+", " ", texto).strip()


def _ruta_a_path(value: Any) -> Path:
    return Path(*[p for p in re.split(r"[\\/]+", _texto(value)) if p])


def _ruta_sin_operador(value: Any) -> bool:
    partes = [_normalizar_header(p).replace(" ", "_").strip("_") for p in _ruta_a_path(value).parts]
    return any(p in {"SIN_OPERADOR", "SIN_OPERADOR_CORREO"} for p in partes)


def _es_descendiente(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except (OSError, ValueError):
        return False


def _filas_reales(ws, max_col: int) -> list[int]:
    cells = getattr(ws, "_cells", None)
    if isinstance(cells, dict):
        return sorted({
            row
            for (row, col), cell in cells.items()
            if row >= 2 and col <= max_col and cell.value not in (None, "")
        })
    return list(range(2, ws.max_row + 1))


@dataclass
class PendienteExcel:
    sheet: str
    row: int
    identificador: str
    ruta_actual: str
    col_ruta: int


@dataclass
class MetadataEntrada:
    path: Path
    carpeta: Path
    data: dict[str, Any]
    claves: set[str]

    @property
    def nombre_operador(self) -> str:
        # Requisito explícito de esta etapa: usar el nombre de metadata_satys.json.
        return _texto(self.data.get("nombre_operador"))


@dataclass
class Reparacion:
    sheet: str
    row: int
    identificador: str
    ruta_anterior: str
    estado: str = "pendiente"
    motivo: str = ""
    nombre_metadata: str = ""
    nombre_rpc: str = ""
    id_operador: str = ""
    ruta_nueva: str = ""
    fuentes_descargas: list[str] = field(default_factory=list)
    carpetas_sin_operador: list[str] = field(default_factory=list)
    conflictos_legacy: int = 0
    archivos_fuente: int = 0
    archivos_compartidos: int = 0
    fuente_rpc: str = ""
    metodo_rpc: str = ""
    score_rpc: float = 0.0
    memorandums_fuente: list[str] = field(default_factory=list)
    ruta_correo: str = ""

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row": self.row,
            "identificador": self.identificador,
            "ruta_anterior": self.ruta_anterior,
            "estado": self.estado,
            "motivo": self.motivo,
            "nombre_metadata": self.nombre_metadata,
            "nombre_rpc": self.nombre_rpc,
            "id_operador": self.id_operador,
            "ruta_nueva": self.ruta_nueva,
            "fuentes_descargas": self.fuentes_descargas,
            "carpetas_sin_operador": self.carpetas_sin_operador,
            "conflictos_legacy": self.conflictos_legacy,
            "archivos_fuente": self.archivos_fuente,
            "archivos_compartidos": self.archivos_compartidos,
            "fuente_rpc": self.fuente_rpc,
            "metodo_rpc": self.metodo_rpc,
            "score_rpc": self.score_rpc,
            "memorandums_fuente": self.memorandums_fuente,
            "ruta_correo": self.ruta_correo,
        }


def leer_pendientes_excel(wb) -> list[PendienteExcel]:
    pendientes: list[PendienteExcel] = []
    for sheet_name in SHEETS_DEFAULT:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {
            _normalizar_header(cell.value): cell.column
            for cell in ws[1]
            if _texto(cell.value)
        }
        col_id = headers.get("1711")
        col_ruta = headers.get("RUTA")
        if not col_id or not col_ruta:
            log.warning("Hoja %s sin columnas 1711/Ruta; se omite.", sheet_name)
            continue
        max_col = max(col_id, col_ruta)
        for row in _filas_reales(ws, max_col):
            identificador = _texto(ws.cell(row=row, column=col_id).value)
            ruta = _texto(ws.cell(row=row, column=col_ruta).value)
            if identificador and ruta and _ruta_sin_operador(ruta):
                pendientes.append(PendienteExcel(
                    sheet=sheet_name,
                    row=row,
                    identificador=identificador,
                    ruta_actual=ruta,
                    col_ruta=col_ruta,
                ))
    return pendientes


def _leer_metadata(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8-sig", errors="replace"))
        return value if isinstance(value, dict) else {}
    except Exception as exc:
        log.warning("No se pudo leer %s: %s", path, exc)
        return {}


def _claves_metadata(path: Path, descargas_base: Path, data: dict[str, Any]) -> set[str]:
    claves: set[str] = set()
    try:
        relativo = path.parent.resolve().relative_to(descargas_base.resolve())
        partes = relativo.parts
    except (OSError, ValueError):
        partes = path.parent.parts
    for parte in partes:
        if _clave(parte):
            claves.add(_clave(parte))
    for campo in (
        "1711", "registro", "numero_registro", "folio", "folio_opc",
        "folio_tabla_internos", "folio_internos", "id_folio",
    ):
        valor = _clave(data.get(campo))
        if valor:
            claves.add(valor)
    return claves


def construir_indice_metadata(descargas_base: Path) -> tuple[list[MetadataEntrada], dict[str, list[MetadataEntrada]]]:
    entradas: list[MetadataEntrada] = []
    indice: dict[str, list[MetadataEntrada]] = defaultdict(list)
    for path in descargas_base.rglob(METADATA_NAME):
        if not path.is_file():
            continue
        data = _leer_metadata(path)
        claves = _claves_metadata(path, descargas_base, data)
        entrada = MetadataEntrada(path=path, carpeta=path.parent, data=data, claves=claves)
        entradas.append(entrada)
        for clave in claves:
            indice[clave].append(entrada)
    return entradas, indice


def _roots_sin_operador(output_base: Path) -> list[Path]:
    roots: list[Path] = []
    if not output_base.is_dir():
        return roots
    for item in output_base.iterdir():
        if not item.is_dir():
            continue
        nombre = _normalizar_header(item.name).replace(" ", "_").strip("_")
        if nombre in {"SIN_OPERADOR", "SIN_OPERADOR_CORREO"}:
            roots.append(item)
    return roots


def construir_indice_sin_operador(output_base: Path) -> list[Path]:
    carpetas: list[Path] = []
    for root in _roots_sin_operador(output_base):
        for path in root.rglob("*"):
            if path.is_dir():
                carpetas.append(path)
        # Una ruta puede apuntar directamente a la raíz, aunque normalmente no.
        carpetas.append(root)
    return carpetas


def buscar_carpetas_sin_operador(
    pendiente: PendienteExcel,
    output_base: Path,
    indice_carpetas: list[Path],
) -> list[Path]:
    salida: list[Path] = []
    preferida = output_base / _ruta_a_path(pendiente.ruta_actual)
    if preferida.is_dir() and _es_descendiente(preferida, output_base):
        salida.append(preferida)

    key = _clave(pendiente.identificador)
    if key:
        for path in indice_carpetas:
            if path in salida:
                continue
            if key in _clave(path.name):
                salida.append(path)
    # Mantener sólo hojas finales: si una coincidencia es ancestro de otra,
    # migrar la más específica para no duplicar el mismo árbol.
    especificas: list[Path] = []
    for path in sorted(set(salida), key=lambda p: len(p.parts), reverse=True):
        if any(_es_descendiente(otro, path) and otro != path for otro in especificas):
            continue
        especificas.append(path)
    return sorted(especificas, key=lambda p: str(p).casefold())


def _tokens_carpeta_sin_operador(path: Path) -> set[str]:
    """Extrae sólo identificadores útiles; nunca indexa palabras de bandeja."""
    tokens: set[str] = set()
    for token in re.split(r"__+|[\\/]+", path.name):
        key = _clave(token)
        if not key:
            continue
        if re.fullmatch(r"\d{1,15}", key) or re.fullmatch(r"[A-Z]{2,8}\d{2}-\d{3,}", key):
            tokens.add(key)
    key_nombre = _clave(path.name)
    if re.fullmatch(r"\d{1,15}", key_nombre) or re.fullmatch(r"[A-Z]{2,8}\d{2}-\d{3,}", key_nombre):
        tokens.add(key_nombre)
    return tokens


def buscar_fuentes_descargas(
    pendiente: PendienteExcel,
    carpetas_sin_operador: Iterable[Path],
    indice_metadata: dict[str, list[MetadataEntrada]],
) -> list[MetadataEntrada]:
    claves = {_clave(pendiente.identificador)}
    for carpeta in carpetas_sin_operador:
        claves.update(_tokens_carpeta_sin_operador(carpeta))

    candidatos: dict[str, MetadataEntrada] = {}
    for clave in claves:
        if not clave:
            continue
        for entrada in indice_metadata.get(clave, []):
            candidatos[str(entrada.path.resolve())] = entrada

    valores = list(candidatos.values())

    # La misma cifra puede existir como folio OPC de Oficialía y como folio de
    # Internos. La hoja del Excel aporta contexto suficiente para evitar cruces
    # entre ambos universos.
    def es_internos(entrada: MetadataEntrada) -> bool:
        partes = {_clave(p) for p in entrada.path.parts}
        data = entrada.data
        return (
            "INTERNOS" in partes
            or _clave(data.get("satys_flujo")) == "INTERNOS"
            or bool(data.get("bandeja_internos"))
            or bool(data.get("folio_tabla_internos"))
        )

    if pendiente.sheet == "Internos":
        internos = [e for e in valores if es_internos(e)]
        if internos:
            valores = internos
    else:
        oficialia = [e for e in valores if not es_internos(e)]
        if oficialia:
            valores = oficialia

    return sorted(
        valores,
        key=lambda e: (
            e.path.stat().st_mtime if e.path.exists() else 0.0,
            str(e.path).casefold(),
        ),
    )


def resolver_rpc_publico(nombres: Iterable[str], timeout_rpc: float) -> dict[str, Any]:
    """Resuelve únicamente con el buscador público; nunca consulta el Excel RPC."""
    nombres_unicos: list[str] = []
    vistos: set[str] = set()
    for nombre in nombres:
        nombre = _texto(nombre)
        k = bc.normalizar_nombre(nombre).replace(" ", "")
        if nombre and k and k not in vistos:
            vistos.add(k)
            nombres_unicos.append(nombre)
    if not nombres_unicos:
        return {"ok": False, "motivo": "metadata_satys_sin_nombre_operador", "resultados": []}

    resultados: list[dict[str, Any]] = []
    for nombre in nombres_unicos:
        diag = bc.buscar_nombre_operador_rpc_online_exacto(nombre, timeout=timeout_rpc)
        resultados.append({"nombre_consultado": nombre, **diag})

    aprobados = [r for r in resultados if r.get("ok") and bc.normalizar_id(r.get("idBp"))]
    ids = {bc.normalizar_id(r.get("idBp")) for r in aprobados}
    if not aprobados:
        return {
            "ok": False,
            "motivo": "rpc_publico_sin_coincidencia_segura",
            "resultados": resultados,
        }
    if len(ids) != 1:
        return {
            "ok": False,
            "motivo": "metadata_resuelve_a_multiples_ids_rpc_publico",
            "resultados": resultados,
        }

    # Si varias variantes del metadata apuntan al mismo ID, preferimos la de
    # mayor score; el nombre canónico siempre viene del propio RPC público.
    mejor = sorted(
        aprobados,
        key=lambda r: float(r.get("score") or 0.0),
        reverse=True,
    )[0]
    return {
        "ok": True,
        "idBp": bc.normalizar_id(mejor.get("idBp")),
        "nombre_completo": _texto(mejor.get("nombre_completo")),
        "score": float(mejor.get("score") or 0.0),
        "fuente": _texto(mejor.get("fuente")),
        "metodo": _texto(mejor.get("metodo")),
        "consulta_rpc": _texto(mejor.get("consulta_rpc")),
        "resultados": resultados,
    }


def _fusionar_legacy_sin_sobrescribir(origen: Path, destino: Path) -> tuple[int, int]:
    """Fusiona únicos históricos; nunca desplaza la copia canónica existente."""
    copiados = 0
    conflictos = 0
    if not origen.is_dir():
        return copiados, conflictos
    origen_abs = origen.resolve()
    for archivo in iter_archivos_publicables_output(origen):
        relativo = archivo.relative_to(origen_abs)
        target = destino / relativo
        if target.exists():
            conflictos += 1
            continue
        copiar_archivo_robusto(archivo, target)
        copiados += 1
    return copiados, conflictos


def _mapa_fuentes(fuentes: Iterable[MetadataEntrada]) -> dict[Path, Path]:
    """Mapa relativo->fuente. La fuente más nueva de la lista prevalece."""
    mapa: dict[Path, Path] = {}
    for entrada in fuentes:
        base_abs = entrada.carpeta.resolve()
        for archivo in iter_archivos_publicables_output(entrada.carpeta):
            mapa[archivo.relative_to(base_abs)] = archivo
    return mapa


def _copiar_fuentes_actuales(fuentes: Iterable[MetadataEntrada], destino: Path) -> dict[Path, Path]:
    """Copia fuentes en orden; en colisiones gana la más reciente (última)."""
    mapa = _mapa_fuentes(fuentes)
    for relativo, archivo in mapa.items():
        copiar_archivo_robusto(archivo, destino / relativo)
    return mapa


def _archivos_iguales(a: Path, b: Path) -> bool:
    try:
        if not a.is_file() or not b.is_file() or a.stat().st_size != b.stat().st_size:
            return False
        with a.open("rb") as fa, b.open("rb") as fb:
            while True:
                ba = fa.read(1024 * 1024)
                bb = fb.read(1024 * 1024)
                if ba != bb:
                    return False
                if not ba:
                    return True
    except OSError:
        return False


def _verificar_mapa(mapa: dict[Path, Path], destino: Path) -> bool:
    return all(_archivos_iguales(origen, destino / relativo) for relativo, origen in mapa.items())


def _retirar_sin_operador_local(carpetas: Iterable[Path], output_base: Path) -> None:
    roots = _roots_sin_operador(output_base)
    for path in sorted(set(carpetas), key=lambda p: len(p.parts), reverse=True):
        if not path.exists() or not any(_es_descendiente(path, root) and path != root for root in roots):
            continue
        eliminar_arbol_robusto(path)
        # Limpia padres vacíos hasta la raíz _sin_operador, sin retirar la raíz.
        parent = path.parent
        while parent not in roots and parent.is_dir():
            try:
                parent.rmdir()
            except OSError:
                break
            parent = parent.parent


def _sync_destino_compartido(
    *,
    ruta: str,
    destino_local: Path,
    carpetas_legacy_local: Iterable[Path],
    output_base: Path,
    shared_root: Path,
) -> tuple[int, list[str]]:
    errores: list[str] = []
    error_mount = validar_destino_compartido(shared_root)
    if error_mount:
        return 0, [error_mount]

    shared_output = shared_root / "output"
    shared_destino = shared_output / _ruta_a_path(ruta)
    shared_destino.mkdir(parents=True, exist_ok=True)

    # Primero preserva cualquier archivo único que ya hubiera sido publicado en
    # la antigua carpeta _sin_operador del recurso compartido.
    shared_legacy: list[Path] = []
    for local_old in carpetas_legacy_local:
        try:
            relativo = local_old.resolve().relative_to(output_base.resolve())
        except (OSError, ValueError):
            continue
        remote_old = shared_output / relativo
        if remote_old.is_dir():
            _fusionar_legacy_sin_sobrescribir(remote_old, shared_destino)
            shared_legacy.append(remote_old)

    resultado = ResultadoSincronizacion()
    copiar_directorio_merge(
        destino_local,
        shared_destino,
        resultado,
        excluir_json=True,
    )
    errores.extend(resultado.errores)
    if errores:
        return resultado.archivos_copiados, errores

    # La ruta nueva ya contiene la unión y la copia local actual prevaleció.
    # Retira únicamente las carpetas de revisión correspondientes a reparaciones
    # exitosas; nunca toca otras carpetas _sin_operador.
    for remote_old in sorted(shared_legacy, key=lambda p: len(p.parts), reverse=True):
        try:
            if remote_old.is_dir() and _es_descendiente(remote_old, shared_output):
                shutil.rmtree(remote_old)
        except Exception as exc:
            errores.append(f"No se pudo retirar {remote_old}: {exc}")
    return resultado.archivos_copiados, errores


def _es_memorando_memorandum_pdf(path: Path) -> bool:
    """Detecta PDF cuyo nombre sea MEMORANDO/MEMORANDUM o una variante cercana.

    La clasificación se basa sólo en el NOMBRE del archivo. Se toleran acentos,
    mayúsculas, espacios/puntuación, sufijos como ``_firmado`` y errores OCR/tecleo
    leves alrededor de la palabra memorando/memorandum. No acepta el abreviado
    genérico ``memo.pdf`` para evitar falsos positivos.
    """
    try:
        if not path.is_file() or path.suffix.strip().casefold() != ".pdf":
            return False
    except OSError:
        return False

    nombre = unicodedata.normalize("NFD", path.stem.strip())
    nombre = "".join(c for c in nombre if unicodedata.category(c) != "Mn")
    tokens = [t for t in re.split(r"[^a-z0-9]+", nombre.casefold()) if t]
    if not tokens:
        return False

    # Cubre MEMORANDO, MEMORANDUM, MEMORÁNDUM, memorando_firmado, etc.
    if any(t.startswith("memorand") for t in tokens):
        return True

    # Variantes de una letra: memorandun, memornado, etc.
    from difflib import SequenceMatcher
    objetivos = ("memorando", "memorandum")
    return any(
        len(token) >= 7
        and max(SequenceMatcher(None, token, objetivo).ratio() for objetivo in objetivos) >= 0.82
        for token in tokens
    )


def _memorandums_en_fuentes(fuentes: Iterable[MetadataEntrada]) -> list[Path]:
    encontrados: dict[str, Path] = {}
    for entrada in fuentes:
        try:
            for path in entrada.carpeta.rglob("*"):
                if _es_memorando_memorandum_pdf(path):
                    encontrados[str(path.resolve())] = path
        except OSError:
            continue
    return sorted(encontrados.values(), key=lambda p: str(p).casefold())


def _nombre_carpeta_correo(pendiente: PendienteExcel, carpetas_old: Iterable[Path]) -> str:
    """Conserva el nombre de la carpeta de revisión siempre que sea específico."""
    candidatos: list[str] = []
    partes = list(_ruta_a_path(pendiente.ruta_actual).parts)
    if partes:
        candidatos.append(partes[-1])
    candidatos.extend(path.name for path in carpetas_old)
    for candidato in candidatos:
        limpio = _texto(candidato)
        normal = _normalizar_header(limpio).replace(" ", "_").strip("_")
        if limpio and normal not in {"SIN_OPERADOR", "CORREOS"}:
            return re.sub(r"[\\/]+", "_", limpio)
    fallback = re.sub(r"[\\/]+", "_", _texto(pendiente.identificador)).strip()
    return fallback or f"fila_{pendiente.row}"


def clasificar_memorandums_restantes(
    *,
    wb,
    pendientes_iniciales: list[PendienteExcel],
    reparaciones: list[Reparacion],
    descargas_base: Path,
    output_base: Path,
    shared_root: Path,
    indice_metadata: dict[str, list[MetadataEntrada]],
    indice_sin_operador: list[Path],
    sincronizar_depi: bool,
    dry_run: bool,
) -> dict[str, Any]:
    """Clasifica pendientes con MEMORANDO/MEMORANDUM.pdf en _sin_operador/(correos).

    Se ejecuta DESPUÉS de intentar RPC público. Por eso sólo ve las filas que
    aún conservan una Ruta bajo _sin_operador. La fuente de verdad sigue siendo
    descargas y la misma política de fusión evita nombres inventados.
    """
    por_fila = {(item.sheet, item.row): item for item in reparaciones}
    restantes = leer_pendientes_excel(wb)
    detectados = 0
    clasificados = 0
    ya_clasificados = 0
    cambios_excel = 0
    errores_sync: list[str] = []

    for pendiente in restantes:
        item = por_fila.get((pendiente.sheet, pendiente.row))
        if item is None:
            item = Reparacion(
                sheet=pendiente.sheet,
                row=pendiente.row,
                identificador=pendiente.identificador,
                ruta_anterior=pendiente.ruta_actual,
            )
            reparaciones.append(item)
            por_fila[(pendiente.sheet, pendiente.row)] = item

        carpetas_old = buscar_carpetas_sin_operador(pendiente, output_base, indice_sin_operador)
        fuentes = buscar_fuentes_descargas(pendiente, carpetas_old, indice_metadata)
        if not fuentes:
            continue
        memorandums = _memorandums_en_fuentes(fuentes)
        if not memorandums:
            continue

        detectados += 1
        item.memorandums_fuente = [str(path) for path in memorandums]
        item.fuentes_descargas = [str(e.carpeta) for e in fuentes]
        item.carpetas_sin_operador = [str(p) for p in carpetas_old]

        nombre_carpeta = _nombre_carpeta_correo(pendiente, carpetas_old)
        ruta_correo = str(Path("_sin_operador") / CORREOS_DIR_NAME / nombre_carpeta).replace("/", "\\")
        item.ruta_correo = ruta_correo
        destino = output_base / _ruta_a_path(ruta_correo)
        legacy = []
        destino_resuelto = destino.resolve()
        for old in carpetas_old:
            try:
                if old.resolve() != destino_resuelto:
                    legacy.append(old)
            except OSError:
                legacy.append(old)

        if dry_run:
            item.estado = "dry_run_correo_memorandum"
            item.motivo = "permanece_sin_operador_y_contiene_memorando_memorandum_pdf"
            continue

        try:
            destino.mkdir(parents=True, exist_ok=True)
            conflictos = 0
            for old in legacy:
                _copiados_legacy, conflictos_old = _fusionar_legacy_sin_sobrescribir(old, destino)
                conflictos += conflictos_old

            # Igual que en la reparación RPC: descargas gana cualquier colisión
            # exacta, JSON no se publica y otros archivos históricos se conservan.
            mapa = _copiar_fuentes_actuales(fuentes, destino)
            item.archivos_fuente = len(mapa)
            item.conflictos_legacy += conflictos
            if not mapa:
                item.estado = "correo_sin_archivos_publicables_descargas"
                item.motivo = "fuente_descargas_sin_archivos_publicables"
                continue
            if not _verificar_mapa(mapa, destino):
                item.estado = "correo_verificacion_local_fallida"
                item.motivo = "output_correos_no_coincide_con_descargas"
                continue

            if sincronizar_depi:
                copiados_shared, sync_err = _sync_destino_compartido(
                    ruta=ruta_correo,
                    destino_local=destino,
                    carpetas_legacy_local=legacy,
                    output_base=output_base,
                    shared_root=shared_root,
                )
                item.archivos_compartidos += copiados_shared
                if sync_err:
                    errores_sync.extend(sync_err)
                    item.estado = "correo_sync_depi_fallido"
                    item.motivo = " | ".join(sync_err[:3])
                    # Igual que la reparación RPC: no cambia Ruta ni retira la
                    # carpeta anterior si DEPI no quedó publicado.
                    continue

            ws = wb[pendiente.sheet]
            ruta_actual = _texto(ws.cell(row=pendiente.row, column=pendiente.col_ruta).value)
            if ruta_actual.replace("/", "\\") != ruta_correo:
                ws.cell(row=pendiente.row, column=pendiente.col_ruta, value=ruta_correo)
                cambios_excel += 1
                item.estado = "clasificado_correo_memorandum"
            else:
                ya_clasificados += 1
                item.estado = "correo_memorandum_ya_clasificado"
            item.motivo = "permanece_sin_operador_y_contiene_memorando_memorandum_pdf"
            item.ruta_nueva = ruta_correo
            clasificados += 1
            _retirar_sin_operador_local(legacy, output_base)
        except Exception as exc:
            log.exception("Error clasificando correo %s", pendiente.identificador)
            item.estado = "error_clasificacion_correo"
            item.motivo = f"{type(exc).__name__}: {exc}"

    log.info(
        "✉️  Clasificación MEMORANDO/MEMORANDUM.pdf: %d detectado(s), %d confirmado(s) en (correos), %d cambio(s) de Ruta.",
        detectados, clasificados, cambios_excel,
    )
    return {
        "total_memorandum_detectados": detectados,
        "total_correos_confirmados": clasificados,
        "total_correos_ya_clasificados": ya_clasificados,
        "cambios_excel_correos": cambios_excel,
        "errores_sync_depi_correos": errores_sync,
    }


def _guardar_excel(wb, excel_path: Path, *, backup: bool) -> Path | None:
    backup_path: Path | None = None
    if backup and excel_path.exists():
        sello = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = excel_path.with_name(
            f"{excel_path.stem}_backup_pre_rpc_publico_sin_operador_{sello}{excel_path.suffix}"
        )
        shutil.copy2(excel_path, backup_path)
    temporal = excel_path.with_name(f".{excel_path.name}.tmp_rpc_publico")
    wb.save(temporal)
    reemplazar_desde_temporal(temporal, excel_path)
    return backup_path


def _guardar_reportes(logs_dir: Path, payload: dict[str, Any]) -> tuple[Path, Path]:
    logs_dir.mkdir(parents=True, exist_ok=True)
    sello = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = logs_dir / f"reparacion_sin_operador_rpc_publico_{sello}.json"
    csv_path = logs_dir / f"reparacion_sin_operador_rpc_publico_{sello}.csv"
    latest = logs_dir / "reparacion_sin_operador_rpc_publico_ultimo.json"
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    json_path.write_text(text, encoding="utf-8")
    latest.write_text(text, encoding="utf-8")

    fieldnames = [
        "sheet", "row", "identificador", "estado", "motivo", "nombre_metadata",
        "nombre_rpc", "id_operador", "ruta_anterior", "ruta_nueva", "fuente_rpc",
        "metodo_rpc", "score_rpc", "archivos_fuente", "archivos_compartidos",
        "conflictos_legacy", "fuentes_descargas", "carpetas_sin_operador",
        "memorandums_fuente", "ruta_correo",
    ]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for item in payload.get("resultados", []):
            row = {key: item.get(key, "") for key in fieldnames}
            for key in ("fuentes_descargas", "carpetas_sin_operador", "memorandums_fuente"):
                if isinstance(row[key], list):
                    row[key] = " | ".join(row[key])
            writer.writerow(row)
    return json_path, csv_path


def reparar(
    *,
    excel_path: Path,
    descargas_base: Path,
    output_base: Path,
    shared_root: Path,
    logs_dir: Path,
    timeout_rpc: float = 15.0,
    sincronizar_depi: bool = True,
    dry_run: bool = False,
    crear_backup: bool = True,
) -> dict[str, Any]:
    excel_path = Path(excel_path)
    descargas_base = Path(descargas_base)
    output_base = Path(output_base)
    shared_root = Path(shared_root)
    logs_dir = Path(logs_dir)

    if not excel_path.is_file():
        raise FileNotFoundError(f"No existe TrámitesCRT.xlsx: {excel_path}")
    if not descargas_base.is_dir():
        raise FileNotFoundError(f"No existe descargas/: {descargas_base}")
    if not output_base.is_dir():
        raise FileNotFoundError(f"No existe output/: {output_base}")

    wb = openpyxl.load_workbook(excel_path)
    pendientes = leer_pendientes_excel(wb)
    log.info("🔎 Rutas _sin_operador encontradas en Excel: %d", len(pendientes))

    _entradas, indice_metadata = construir_indice_metadata(descargas_base)
    indice_sin_operador = construir_indice_sin_operador(output_base)
    log.info(
        "📚 Índices: %d metadata_satys.json | %d carpeta(s) bajo revisión manual",
        len(_entradas), len(indice_sin_operador),
    )

    reparaciones: list[Reparacion] = []
    reparados = 0
    cambios_excel = 0
    errores_sync: list[str] = []

    for pos, pendiente in enumerate(pendientes, start=1):
        item = Reparacion(
            sheet=pendiente.sheet,
            row=pendiente.row,
            identificador=pendiente.identificador,
            ruta_anterior=pendiente.ruta_actual,
        )
        reparaciones.append(item)
        log.info(
            "[RPC-PUBLICO %d/%d] %s!fila %d 1711=%s",
            pos, len(pendientes), pendiente.sheet, pendiente.row, pendiente.identificador,
        )

        carpetas_old = buscar_carpetas_sin_operador(pendiente, output_base, indice_sin_operador)
        item.carpetas_sin_operador = [str(p) for p in carpetas_old]
        fuentes = buscar_fuentes_descargas(pendiente, carpetas_old, indice_metadata)
        item.fuentes_descargas = [str(e.carpeta) for e in fuentes]
        if not fuentes:
            item.estado = "sin_fuente_descargas"
            item.motivo = "no_se_encontro_metadata_satys_para_1711"
            continue

        nombres = [e.nombre_operador for e in fuentes if e.nombre_operador]
        nombres_unicos_legibles = []
        vistos_nombre: set[str] = set()
        for nombre in nombres:
            k = bc.normalizar_nombre(nombre).replace(" ", "")
            if k and k not in vistos_nombre:
                vistos_nombre.add(k)
                nombres_unicos_legibles.append(nombre)
        item.nombre_metadata = " | ".join(nombres_unicos_legibles)
        if not nombres_unicos_legibles:
            item.estado = "sin_nombre_operador_metadata"
            item.motivo = "metadata_satys_json_sin_nombre_operador"
            continue

        rpc = resolver_rpc_publico(nombres_unicos_legibles, timeout_rpc)
        if not rpc.get("ok"):
            item.estado = "sin_coincidencia_rpc_publico"
            item.motivo = _texto(rpc.get("motivo"))
            continue

        item.id_operador = _texto(rpc.get("idBp"))
        item.nombre_rpc = _texto(rpc.get("nombre_completo"))
        item.fuente_rpc = _texto(rpc.get("fuente"))
        item.metodo_rpc = _texto(rpc.get("metodo"))
        item.score_rpc = float(rpc.get("score") or 0.0)
        item.ruta_nueva = construir_ruta(item.nombre_rpc, item.id_operador)
        destino = output_base / _ruta_a_path(item.ruta_nueva)

        if dry_run:
            item.estado = "dry_run_resoluble"
            item.motivo = "sin_cambios_por_dry_run"
            continue

        destino.mkdir(parents=True, exist_ok=True)
        conflictos = 0
        try:
            # Histórico _sin_operador aporta únicamente archivos que aún no
            # existan en la carpeta canónica. Nunca reemplaza una copia canónica.
            for old in carpetas_old:
                _copiados_legacy, conflictos_old = _fusionar_legacy_sin_sobrescribir(old, destino)
                conflictos += conflictos_old

            # La fuente de verdad es descargas. Se copia al final y sobrescribe
            # exactamente el mismo nombre/ruta; nunca se inventa archivo_1.
            mapa = _copiar_fuentes_actuales(fuentes, destino)
            item.archivos_fuente = len(mapa)
            item.conflictos_legacy = conflictos
            if not mapa:
                item.estado = "sin_archivos_publicables_descargas"
                item.motivo = "descargas_solo_contiene_metadata_o_esta_vacio"
                continue
            if not _verificar_mapa(mapa, destino):
                item.estado = "verificacion_local_fallida"
                item.motivo = "output_no_coincide_con_fuente_descargas"
                continue

            if sincronizar_depi:
                copiados_shared, sync_err = _sync_destino_compartido(
                    ruta=item.ruta_nueva,
                    destino_local=destino,
                    carpetas_legacy_local=carpetas_old,
                    output_base=output_base,
                    shared_root=shared_root,
                )
                item.archivos_compartidos = copiados_shared
                if sync_err:
                    errores_sync.extend(sync_err)
                    item.estado = "sync_depi_fallido"
                    item.motivo = " | ".join(sync_err[:3])
                    # Transacción incompleta: conserva Ruta y _sin_operador para
                    # reintentar en la próxima corrida. El destino canónico local
                    # puede quedar prellenado, pero nunca se declara reparado.
                    continue
                item.estado = "reparado"
            else:
                item.estado = "reparado_sin_sync_depi"

            ws = wb[pendiente.sheet]
            ws.cell(row=pendiente.row, column=pendiente.col_ruta, value=item.ruta_nueva.replace("/", "\\"))
            cambios_excel += 1
            reparados += 1
            _retirar_sin_operador_local(carpetas_old, output_base)
        except Exception as exc:
            log.exception("Error reparando %s", pendiente.identificador)
            item.estado = "error_organizacion"
            item.motivo = f"{type(exc).__name__}: {exc}"

    # Segunda y última regla de esta misma etapa: sólo para las filas que
    # SIGUEN bajo _sin_operador después del RPC público. Si la fuente original
    # contiene MEMORANDO/MEMORANDUM.pdf, se clasifica en _sin_operador/(correos).
    resumen_correos = clasificar_memorandums_restantes(
        wb=wb,
        pendientes_iniciales=pendientes,
        reparaciones=reparaciones,
        descargas_base=descargas_base,
        output_base=output_base,
        shared_root=shared_root,
        indice_metadata=indice_metadata,
        indice_sin_operador=indice_sin_operador,
        sincronizar_depi=sincronizar_depi,
        dry_run=dry_run,
    )
    cambios_excel += int(resumen_correos.get("cambios_excel_correos") or 0)
    errores_sync.extend(resumen_correos.get("errores_sync_depi_correos") or [])

    backup_path: Path | None = None
    if cambios_excel and not dry_run:
        backup_path = _guardar_excel(wb, excel_path, backup=crear_backup)
        log.info("💾 TrámitesCRT.xlsx actualizado: %d Ruta(s)", cambios_excel)
    wb.close()

    payload = {
        "fecha": datetime.now().isoformat(),
        "excel": str(excel_path),
        "descargas": str(descargas_base),
        "output": str(output_base),
        "shared_output": str(shared_root / "output"),
        "fuente_rpc": "buscador_publico_rpc_exclusivamente",
        "usa_excel_oficial_rpc": False,
        "total_sin_operador_excel": len(pendientes),
        "total_reparados": reparados,
        "total_pendientes": len(pendientes) - reparados,
        "total_memorandum_detectados": resumen_correos.get("total_memorandum_detectados", 0),
        "total_correos_confirmados": resumen_correos.get("total_correos_confirmados", 0),
        "total_correos_ya_clasificados": resumen_correos.get("total_correos_ya_clasificados", 0),
        "cambios_excel_correos": resumen_correos.get("cambios_excel_correos", 0),
        "cambios_excel": cambios_excel,
        "backup_excel": str(backup_path) if backup_path else "",
        "errores_sync_depi": errores_sync,
        "regla_correo_memorando_memorandum": "pendiente_sin_operador + MEMORANDO/MEMORANDUM.pdf => _sin_operador/(correos)",
        "dry_run": dry_run,
        "resultados": [item.as_dict() for item in reparaciones],
    }
    json_path, csv_path = _guardar_reportes(logs_dir, payload)
    payload["reporte_json"] = str(json_path)
    payload["reporte_csv"] = str(csv_path)
    log.info(
        "✅ Reparación final RPC público: %d reparado(s), %d pendiente(s). Reporte: %s",
        reparados, len(pendientes) - reparados, json_path,
    )
    return payload


def construir_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Repara Rutas _sin_operador de TrámitesCRT.xlsx usando únicamente el buscador público RPC."
    )
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--descargas", type=Path, default=DEFAULT_DESCARGAS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--shared", type=Path, default=DEFAULT_SHARED)
    parser.add_argument("--logs-dir", type=Path, default=DEFAULT_LOGS)
    parser.add_argument("--timeout-rpc", type=float, default=15.0)
    parser.add_argument("--sin-sincronizar-depi", action="store_true")
    parser.add_argument("--sin-backup", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--sin-lock",
        action="store_true",
        help="Omite el lock global SATyS. Sólo para pruebas controladas; no usar en producción.",
    )
    return parser


def main() -> int:
    args = construir_parser().parse_args()
    lock = None
    try:
        # La ejecución independiente debe proteger el mismo Excel/output que la
        # corrida diaria. Si esta función es hija del monitor diario, ProcesoLock
        # detecta SATYS_LOCK_TOKEN y hereda el lock sin bloquearse a sí misma.
        if not args.sin_lock:
            lock = ProcesoLock(proceso="resolver_sin_operador_rpc_publico.py")
            lock.adquirir()
            log.info("🔒 Lock SATyS adquirido/heredado para reparación _sin_operador.")

        reparar(
            excel_path=args.excel,
            descargas_base=args.descargas,
            output_base=args.output,
            shared_root=args.shared,
            logs_dir=args.logs_dir,
            timeout_rpc=max(1.0, args.timeout_rpc),
            sincronizar_depi=not args.sin_sincronizar_depi,
            dry_run=args.dry_run,
            crear_backup=not args.sin_backup,
        )
        return 0
    except LockOcupadoError as exc:
        log.error("No se inicia reparación _sin_operador: %s", exc)
        return 3
    except Exception as exc:
        log.exception("Fallo fatal en reparación final _sin_operador: %s", exc)
        return 1
    finally:
        if lock is not None:
            lock.liberar()


if __name__ == "__main__":
    raise SystemExit(main())
