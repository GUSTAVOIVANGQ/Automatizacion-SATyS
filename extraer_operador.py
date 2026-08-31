#!/usr/bin/env python3
"""
extraer_operador.py

Extrae de eFormatos del IFT/CRT (R001, R011, R024, R025, R026, etc.), sin
importar en que pagina del PDF aparezcan ni si el PDF no trae la seccion,
los dos campos de la SECCION 2 "DATOS GENERALES DEL OPERADOR":

  - "Nombre o razon social del Operador"  -> columna Excel "Solicitante Promovente"
  - "Representante legal"                 -> columna Excel "Representante Legal"

Genera un unico archivo Excel (.xlsx) con una fila por PDF.

--------------------------------------------------------------------------
CONSIDERACIONES / SUPUESTOS (leer antes de confiar en los resultados)
--------------------------------------------------------------------------
1. Muchos de estos PDFs no son PDFs "nativos" (generados digitalmente),
   sino escaneos con una capa de texto de OCR de calidad variable. Por
   eso el script NO asume que "Nombre o razon social del Operador" o
   "Representante legal" van a aparecer letra por letra exactas: tolera
   errores tipicos de OCR (p. ej. "del" leido como "dei" o "de!", bordes
   de celda leidos como una "i" o "|" sueltas, etc.).
2. Sin embargo, esta tolerancia tiene un limite. Si el OCR del PDF esta
   MUY degradado (nombres con letras cambiadas, palabras clave irreco-
   nocibles), el script preferira devolver "no encontrado" en vez de
   arriesgarse a inventar un valor incorrecto. Todo resultado que no sea
   "completo" queda marcado para revision humana (columna "Estado").
3. "Representante legal" en estos formularios se imprime en 3 columnas
   separadas (Nombre(s) / Primer apellido / Segundo apellido), con los
   VALORES en un renglon y las ETIQUETAS de esas 3 columnas en el
   renglon de abajo. El script arma el nombre completo tomando el primer
   renglon de datos que encuentra despues del encabezado "Representante
   legal" y que no sea, el mismo, un renglon de puras etiquetas.
4. Si el mismo campo aparece mas de una vez en el PDF (por ejemplo, un
   PDF que trae varios eFormatos concatenados) y los valores NO
   coinciden entre si, el archivo se marca como "revisar_inconsistencia"
   en vez de quedarse arbitrariamente con el primero.
5. Si un PDF no tiene NINGUNA capa de texto extraible en ninguna pagina
   (escaneo puro, sin OCR), se marca como "requiere_ocr" en vez de
   intentar adivinar nada.
6. Se reporta, para cada campo, en que pagina y con que metodo se
   encontro (tabla / texto con layout / texto plano) como rastro de
   auditoria, por si se necesita ir a verificar el PDF original.
7. Este script NO valida que el "Solicitante Promovente" y el
   "Representante Legal" reportados correspondan realmente a la misma
   persona/empresa que el nombre del archivo o el folio del tramite;
   solo extrae lo que esta escrito en la Seccion 2 del PDF.

USO:
    python extraer_operador.py archivo1.pdf archivo2.pdf ...
    python extraer_operador.py --carpeta ./mis_pdfs
    python extraer_operador.py --carpeta ./mis_pdfs --salida resultado.xlsx

REQUISITOS:
    pip install pdfplumber openpyxl --break-system-packages
"""

import argparse
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    sys.exit(
        "Falta la libreria pdfplumber. Instalala con:\n"
        "    pip install pdfplumber --break-system-packages"
    )

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    TIENE_OPENPYXL = True
except ImportError:
    TIENE_OPENPYXL = False


# ==========================================================================
# Utilidades de texto (compartidas por ambos campos)
# ==========================================================================

def normalizar(texto):
    """Minusculas, sin acentos, espacios colapsados. Solo para COMPARAR,
    nunca para guardar (el valor final se guarda tal cual aparece)."""
    if texto is None:
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.lower()
    return re.sub(r"\s+", " ", texto).strip()


def _quitar_acentos_conservando_mayus(texto):
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


MAX_LONGITUD_VALOR = 200

# Etiquetas que indican que ya empezo el SIGUIENTE campo/seccion del
# formulario. Sirven para saber donde CORTAR un valor capturado en texto
# corrido y para saber cuando DEJAR de buscar el renglon de datos de un
# campo (si aparecen antes de encontrar datos, es que el campo esta vacio).
SIGUIENTES_ETIQUETAS = [
    "representante legal",
    "datos generales del operador",
    "domicilio para oir y recibir notificaciones",
    "autorizados",
    "seccion 3",
]

VALORES_INVALIDOS = {
    "no aplica",
    "na",
    "indique el nombre completo de la persona fisica o moral del operador",
}


def limpiar_valor_generico(valor):
    """Limpieza ligera comun a ambos campos: quita separadores de celda
    sobrantes, espacios repetidos, corta si arrastro la siguiente
    etiqueta pegada, y descarta boilerplate/instructivo conocido."""
    if not valor:
        return ""
    valor = valor.strip()
    valor = valor.lstrip(":|; ").strip()
    valor = re.sub(r"[ \t]+", " ", valor)
    valor = valor.rstrip("_ ").strip()

    valor_norm = normalizar(valor)
    for etiqueta in SIGUIENTES_ETIQUETAS:
        pos = valor_norm.find(etiqueta)
        if pos > 0:
            valor = valor[:pos].strip()
            break

    valor = valor[:MAX_LONGITUD_VALOR].strip()

    if normalizar(valor) in VALORES_INVALIDOS or len(normalizar(valor)) < 2:
        return ""
    return valor


def clave_comparacion(texto):
    """Clave 'a prueba de ruido' para decidir si dos valores encontrados en
    el mismo PDF son en realidad el mismo dato (a pesar de pequenas
    diferencias de extraccion/OCR). Solo para comparar, nunca para mostrar."""
    return re.sub(r"[^a-z0-9]", "", normalizar(texto))


def resolver_candidatos(candidatos):
    """candidatos: lista de (pagina, metodo, valor). Aplica dedupe
    tolerante a ruido y decide si el campo quedo 'encontrado',
    'inconsistente' o 'no_encontrado'."""
    resultado = {"estado": "no_encontrado", "valor": "", "pagina": "",
                 "metodo": "", "nota": ""}
    if not candidatos:
        return resultado

    vistos = {}
    for pagina_n, metodo, valor in candidatos:
        clave = clave_comparacion(valor)
        if not clave:
            continue
        if clave not in vistos or len(valor) > len(vistos[clave][2]):
            vistos[clave] = (pagina_n, metodo, valor)

    if not vistos:
        return resultado

    if len(vistos) == 1:
        pagina_n, metodo, valor = next(iter(vistos.values()))
        resultado.update(estado="encontrado", valor=valor, pagina=pagina_n,
                          metodo=metodo)
    else:
        primero = list(vistos.values())[0]
        resultado.update(
            estado="inconsistente", valor=primero[2], pagina=primero[0],
            metodo=primero[1],
            nota=(
                f"{len(vistos)} valores distintos encontrados: "
                + " | ".join(f"'{v[2]}' (pag. {v[0]})" for v in vistos.values())
            ),
        )
    return resultado


# ==========================================================================
# Campo 1: "Nombre o razon social del Operador" -> Solicitante Promovente
# ==========================================================================
#
# Los distintos eFormatos del IFT/CRT imprimen la etiqueta de dos formas:
#
#   Patron A (R001, R011, R024...):
#       "Nombre o razón social del Operador: VALOR"        (todo en 1 linea)
#
#   Patron B (R025, R026...):
#       "Nombre o razón social del VALOR"                  (linea 1)
#       "Operador:"                                        (linea 2)
#
# La palabra "del" se escribe con \S{1,4} (comodin de 1 a 4 caracteres sin
# espacio) en vez de literal, porque en PDFs escaneados el OCR a veces la
# lee como "dei", "de!", etc.

PATRON_OPERADOR_A = re.compile(
    r"nombre\s+o\s+razon\s+social\s+\S{1,4}\s+operador\s*[:;]?\s*\|?\s*",
    re.IGNORECASE,
)
PATRON_OPERADOR_B_INICIO = re.compile(
    r"nombre\s+o\s+razon\s+social\s+\S{1,4}\s+", re.IGNORECASE
)
PATRON_OPERADOR_B_SIGUIENTE = re.compile(r"^\s*operador\b", re.IGNORECASE)


def buscar_operador_en_tablas(tablas):
    for tabla in tablas:
        for fila in tabla:
            celdas = [c if c is not None else "" for c in fila]
            for i, celda in enumerate(celdas):
                if "nombre o razon social" in normalizar(celda) and \
                        "operador" in normalizar(celda):
                    if ":" in celda:
                        posible = limpiar_valor_generico(celda.split(":", 1)[1])
                        if posible:
                            return posible
                    for siguiente in celdas[i + 1:]:
                        posible = limpiar_valor_generico(siguiente)
                        if posible:
                            return posible
    return None


def buscar_operador_en_texto(texto):
    if not texto:
        return None
    texto_sa = _quitar_acentos_conservando_mayus(texto)

    m = PATRON_OPERADOR_A.search(texto_sa)
    if m:
        inicio = m.end()
        fin = texto.find("\n", inicio)
        fin = len(texto) if fin == -1 else fin
        valor = limpiar_valor_generico(texto[inicio:fin])
        if valor:
            return valor

    m2 = PATRON_OPERADOR_B_INICIO.search(texto_sa)
    if m2:
        inicio = m2.end()
        fin = texto.find("\n", inicio)
        fin = len(texto) if fin == -1 else fin
        posible_valor = texto[inicio:fin]
        resto = texto[fin:].lstrip("\n")
        resto_sa = _quitar_acentos_conservando_mayus(resto)
        if PATRON_OPERADOR_B_SIGUIENTE.match(resto_sa):
            valor = limpiar_valor_generico(posible_valor)
            if valor:
                return valor
    return None


# ==========================================================================
# Campo 2: "Representante legal" -> Representante Legal
# ==========================================================================
#
# Este campo se imprime como encabezado propio, seguido del renglon con
# los 3 componentes del nombre (Nombre(s) / Primer apellido / Segundo
# apellido) y, despues, el renglon con esas 3 etiquetas:
#
#       Representante legal
#       JANETH            GIORDANO            SERNA
#       Nombre (s)        Primer apellido     Segundo apellido
#
# En texto con layout, los espacios entre columnas quedan como multiples
# espacios; al colapsarlos a uno solo se reconstruye "JANETH GIORDANO
# SERNA". En escaneos con OCR de mala calidad, a veces una linea divisoria
# de celda se lee como una letra suelta (ej. "ADRIAN REYNA i MARTINEZ");
# esos tokens de una sola letra minuscula se descartan por ser ruido, ya
# que ningun nombre/apellido valido en estos formularios es de 1 letra.

PATRON_REP_HEADER = re.compile(r"representante\s+legal", re.IGNORECASE)


def _es_linea_header_representante(linea):
    """True si la linea es (casi) exclusivamente el encabezado
    'Representante legal' (tolera guiones bajos/':' alrededor, pero no
    texto adicional como 'Domicilio del representante legal')."""
    norm = normalizar(linea)
    m = PATRON_REP_HEADER.search(norm)
    if not m:
        return False
    resto = norm[: m.start()] + norm[m.end():]
    resto = re.sub(r"[^a-z0-9]", "", resto)
    return resto == ""


def _es_linea_de_etiquetas_representante(linea):
    """True si la linea es el renglon de etiquetas de columna (contiene
    'primer apellido' o 'segundo apellido'), no un renglon de datos."""
    norm = normalizar(linea)
    return "primer apellido" in norm or "segundo apellido" in norm


def _limpiar_valor_representante(valor):
    valor = limpiar_valor_generico(valor)
    if not valor:
        return ""
    # Descarta tokens de una sola letra minuscula (ruido de OCR de bordes
    # de tabla), preservando iniciales reales tipo "J." (con punto).
    tokens = [t for t in valor.split(" ") if not re.fullmatch(r"[a-z]", t)]
    valor = " ".join(tokens).strip()
    return valor if len(normalizar(valor)) >= 2 else ""


def buscar_representante_en_tablas(tablas):
    for tabla in tablas:
        for idx, fila in enumerate(tabla):
            celdas = [c if c is not None else "" for c in fila]
            if not any(_es_linea_header_representante(c) for c in celdas):
                continue
            # Busca, en las filas siguientes de la misma tabla, la primera
            # que traiga datos reales (no solo las etiquetas de columna).
            for fila_siguiente in tabla[idx + 1: idx + 4]:
                celdas_sig = [c if c is not None else "" for c in fila_siguiente]
                if any(_es_linea_de_etiquetas_representante(c) for c in celdas_sig):
                    continue
                partes = [limpiar_valor_generico(c) for c in celdas_sig]
                partes = [p for p in partes if p]
                if partes:
                    return _limpiar_valor_representante(" ".join(partes))
    return None


def buscar_representante_en_texto(texto):
    if not texto:
        return None
    lineas = texto.splitlines()
    for i, linea in enumerate(lineas):
        if not _es_linea_header_representante(linea):
            continue
        vistas = 0
        for j in range(i + 1, len(lineas)):
            candidata = lineas[j]
            if not candidata.strip():
                continue
            vistas += 1
            if vistas > 4:
                break
            norm = normalizar(candidata)
            if _es_linea_de_etiquetas_representante(candidata):
                continue
            if any(norm.startswith(et) for et in SIGUIENTES_ETIQUETAS
                   if et != "representante legal"):
                break
            valor = _limpiar_valor_representante(candidata)
            if valor:
                return valor
            break
    return None


# ==========================================================================
# Extraccion combinada por PDF (una sola pasada de paginas para ambos campos)
# ==========================================================================

def _extraer_campo(cache_paginas, fn_tablas, fn_texto):
    candidatos = []
    for i, (tablas, texto_layout, texto_plano) in enumerate(cache_paginas, start=1):
        valor = fn_tablas(tablas)
        if valor:
            candidatos.append((i, "tabla", valor))
            continue
        valor = fn_texto(texto_layout)
        if valor:
            candidatos.append((i, "texto_layout", valor))
            continue
        valor = fn_texto(texto_plano)
        if valor:
            candidatos.append((i, "texto_plano", valor))
    return resolver_candidatos(candidatos)


FRASE_INSTRUCTIVO = "instructivo de llenado"


def extraer_de_pdf(ruta: Path):
    """Devuelve un dict con los resultados de ambos campos y un estado
    general consolidado para la fila del Excel."""
    resultado = {
        "archivo": ruta.name,
        "estado_general": None,
        "nota_general": "",
        "solicitante": {"estado": "no_encontrado", "valor": "", "pagina": "", "metodo": "", "nota": ""},
        "representante": {"estado": "no_encontrado", "valor": "", "pagina": "", "metodo": "", "nota": ""},
    }

    try:
        with pdfplumber.open(ruta) as pdf:
            cache_paginas = []
            tiene_texto = False
            for pagina in pdf.pages:
                try:
                    tablas = pagina.extract_tables()
                except Exception:
                    tablas = []
                texto_layout = pagina.extract_text(layout=True) or ""
                texto_plano = pagina.extract_text() or ""
                if texto_layout.strip() or texto_plano.strip():
                    tiene_texto = True
                cache_paginas.append((tablas, texto_layout, texto_plano))

            # Todos estos eFormatos terminan con una seccion "INSTRUCTIVO DE
            # LLENADO" que EXPLICA que debe llevar cada campo (a modo de
            # manual), no datos reales del solicitante. Esa seccion suele
            # repetir literalmente las etiquetas "Nombre o razon social del
            # Operador" y "Representante legal", así que si se incluyera en
            # el escaneo se generarian falsos positivos / inconsistencias.
            # Se detecta la primera pagina donde aparece ese titulo y NO se
            # buscan datos en esa pagina ni en las siguientes.
            limite = None
            for i, (_, texto_layout, texto_plano) in enumerate(cache_paginas, start=1):
                if FRASE_INSTRUCTIVO in normalizar(texto_layout) or \
                        FRASE_INSTRUCTIVO in normalizar(texto_plano):
                    limite = i
                    break
            if limite is not None:
                cache_paginas = cache_paginas[: limite - 1]

            if not tiene_texto:
                resultado["estado_general"] = "requiere_ocr"
                resultado["nota_general"] = (
                    "El PDF no tiene capa de texto extraible en ninguna "
                    "pagina (probablemente escaneado sin OCR). Requiere OCR."
                )
                return resultado

            resultado["solicitante"] = _extraer_campo(
                cache_paginas, buscar_operador_en_tablas, buscar_operador_en_texto
            )
            resultado["representante"] = _extraer_campo(
                cache_paginas, buscar_representante_en_tablas, buscar_representante_en_texto
            )

    except Exception as e:
        resultado["estado_general"] = "error_lectura"
        resultado["nota_general"] = f"{type(e).__name__}: {e}"
        return resultado

    # --- Estado general consolidado a partir de los 2 campos ---
    e_sol = resultado["solicitante"]["estado"]
    e_rep = resultado["representante"]["estado"]

    if "inconsistente" in (e_sol, e_rep):
        resultado["estado_general"] = "revisar_inconsistencia"
    elif e_sol == "encontrado" and e_rep == "encontrado":
        resultado["estado_general"] = "completo"
    elif e_sol == "no_encontrado" and e_rep == "no_encontrado":
        resultado["estado_general"] = "sin_seccion"
        resultado["nota_general"] = (
            "No se encontro la seccion 'Datos generales del Operador' "
            "(ni Solicitante Promovente ni Representante Legal) en "
            "ninguna pagina del PDF."
        )
    else:
        resultado["estado_general"] = "parcial"
        faltante = "Solicitante Promovente" if e_sol == "no_encontrado" else "Representante Legal"
        resultado["nota_general"] = f"No se encontro el campo '{faltante}' en ninguna pagina."

    return resultado


# ==========================================================================
# Diagnostico manual (para PDFs sin resultado)
# ==========================================================================

def diagnosticar_pdf(ruta: Path, max_paginas_reportadas=15):
    print(f"\n--- DIAGNOSTICO: {ruta.name} ---")
    encontro_algo = False
    try:
        with pdfplumber.open(ruta) as pdf:
            paginas_reportadas = 0
            for i, pagina in enumerate(pdf.pages, start=1):
                texto = pagina.extract_text() or ""
                if not texto.strip():
                    print(f"  pag. {i}: (sin texto extraible; puede ser una imagen/escaneo)")
                    continue
                relevantes = [
                    ln for ln in texto.splitlines()
                    if any(p in normalizar(ln) for p in
                           ("operador", "razon social", "representante"))
                ]
                if relevantes:
                    encontro_algo = True
                    paginas_reportadas += 1
                    print(f"  pag. {i}:")
                    for ln in relevantes[:8]:
                        print(f"      {ln!r}")
                    if paginas_reportadas >= max_paginas_reportadas:
                        print("  ... (se alcanzo el limite de paginas a mostrar)")
                        break
        if not encontro_algo:
            print("  No aparecen las palabras 'operador', 'razón social' ni "
                  "'representante' en NINGUNA pagina con texto extraible. "
                  "Es probable que este PDF simplemente no traiga esa "
                  "seccion (p. ej. es solo un oficio/memo de acompañamiento).")
    except Exception as e:
        print(f"  [error al diagnosticar: {type(e).__name__}: {e}]")


# ==========================================================================
# Salida: Excel (unificado) con fallback a CSV si falta openpyxl
# ==========================================================================

ESTADOS_A_REVISAR = {"revisar_inconsistencia", "parcial", "sin_seccion", "requiere_ocr", "error_lectura"}

ENCABEZADOS = [
    "Archivo",
    "Solicitante Promovente",
    "Ubicacion Solicitante Promovente",
    "Representante Legal",
    "Ubicacion Representante Legal",
    "Estado",
    "Notas",
]


def _ubicacion(campo):
    if not campo["pagina"]:
        return ""
    return f"pag. {campo['pagina']} ({campo['metodo']})"


def _fila_para(r):
    notas = []
    if r["nota_general"]:
        notas.append(r["nota_general"])
    if r["solicitante"]["nota"]:
        notas.append("Solicitante Promovente: " + r["solicitante"]["nota"])
    if r["representante"]["nota"]:
        notas.append("Representante Legal: " + r["representante"]["nota"])
    return [
        r["archivo"],
        r["solicitante"]["valor"],
        _ubicacion(r["solicitante"]),
        r["representante"]["valor"],
        _ubicacion(r["representante"]),
        r["estado_general"],
        " | ".join(notas),
    ]


def escribir_excel(filas, ruta_salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Operadores"

    ws.append(ENCABEZADOS)
    fuente_encabezado = Font(bold=True, color="FFFFFF")
    relleno_encabezado = PatternFill("solid", fgColor="2F5496")
    for celda in ws[1]:
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    relleno_revisar = PatternFill("solid", fgColor="FFF2CC")

    for r in filas:
        fila = _fila_para(r)
        ws.append(fila)
        if r["estado_general"] in ESTADOS_A_REVISAR:
            for celda in ws[ws.max_row]:
                celda.fill = relleno_revisar

    anchos = [28, 34, 26, 28, 26, 22, 60]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    for fila_celdas in ws.iter_rows(min_row=2):
        for celda in fila_celdas:
            celda.alignment = Alignment(vertical="top", wrap_text=(celda.column_letter == "G"))

    wb.save(ruta_salida)


def escribir_csv(filas, ruta_salida):
    import csv

    def sanear_csv(v):
        v = "" if v is None else str(v)
        return "'" + v if v and v[0] in ("=", "+", "-", "@") else v

    with open(ruta_salida, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(ENCABEZADOS)
        for r in filas:
            writer.writerow([sanear_csv(v) for v in _fila_para(r)])


# ==========================================================================
# CLI / procesamiento por lote
# ==========================================================================

def recolectar_pdfs(argumentos):
    rutas = []
    if argumentos.carpeta:
        carpeta = Path(argumentos.carpeta)
        if not carpeta.is_dir():
            sys.exit(f"La carpeta no existe: {carpeta}")
        rutas.extend(sorted(carpeta.glob("*.pdf")))
        rutas.extend(sorted(carpeta.glob("*.PDF")))
    for archivo in argumentos.archivos:
        p = Path(archivo)
        if not p.exists():
            print(f"[AVISO] No existe, se omite: {p}", file=sys.stderr)
            continue
        rutas.append(p)

    vistos, unicos = set(), []
    for r in rutas:
        rr = r.resolve()
        if rr not in vistos:
            vistos.add(rr)
            unicos.append(r)
    return unicos


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Extrae 'Solicitante Promovente' (Nombre o razon social del "
            "Operador) y 'Representante Legal' de eFormatos IFT/CRT."
        )
    )
    parser.add_argument("archivos", nargs="*", help="Rutas de PDFs a procesar")
    parser.add_argument("--carpeta", help="Carpeta con PDFs a procesar (busca *.pdf)")
    parser.add_argument(
        "--salida", default="resultado_operadores.xlsx",
        help="Ruta del archivo de salida. Si termina en .csv se genera CSV, "
             "cualquier otra extension (o ninguna) genera Excel (default: "
             "resultado_operadores.xlsx)",
    )
    parser.add_argument(
        "--debug", action="store_true",
        help=(
            "Para cada PDF marcado como 'sin_seccion' o 'requiere_ocr', "
            "imprime en pantalla las lineas que mencionan 'operador', "
            "'razon social' o 'representante' (con su pagina)."
        ),
    )
    args = parser.parse_args()

    pdfs = recolectar_pdfs(args)
    if not pdfs:
        sys.exit("No se especificaron PDFs. Usa argumentos posicionales o --carpeta.")

    filas = []
    for ruta in pdfs:
        print(f"Procesando: {ruta.name} ...", file=sys.stderr)
        filas.append(extraer_de_pdf(ruta))

    print("\n" + "=" * 100)
    for r in filas:
        print(f"{r['archivo']:30s} | {r['estado_general']:22s} | "
              f"Solicitante: {r['solicitante']['valor']!r:45s} | "
              f"Representante: {r['representante']['valor']!r}")
        if r["nota_general"]:
            print(f"   -> {r['nota_general']}")
    print("=" * 100 + "\n")

    ruta_salida = Path(args.salida)
    usar_csv = ruta_salida.suffix.lower() == ".csv"

    if usar_csv:
        escribir_csv(filas, ruta_salida)
    elif TIENE_OPENPYXL:
        if ruta_salida.suffix.lower() != ".xlsx":
            ruta_salida = ruta_salida.with_suffix(".xlsx")
        escribir_excel(filas, ruta_salida)
    else:
        print(
            "[AVISO] No esta instalado openpyxl (pip install openpyxl "
            "--break-system-packages); se genera un CSV en su lugar.",
            file=sys.stderr,
        )
        ruta_salida = ruta_salida.with_suffix(".csv")
        escribir_csv(filas, ruta_salida)

    print(f"Resultado guardado en: {ruta_salida}")

    if args.debug:
        for ruta, r in zip(pdfs, filas):
            if r["estado_general"] in ("sin_seccion", "requiere_ocr", "parcial"):
                diagnosticar_pdf(ruta)

    print("Resumen:", dict(Counter(r["estado_general"] for r in filas)))


if __name__ == "__main__":
    main()