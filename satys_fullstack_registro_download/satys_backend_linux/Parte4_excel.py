#!/usr/bin/env python3
r"""
=============================================================
  PARTE 4 — ACTUALIZACIÓN DE EXCEL Y ORGANIZACIÓN DE ARCHIVOS
=============================================================
Actualiza el archivo TrámitesCRT.xlsx con los datos extraídos
y organiza los archivos descargados en carpetas estandarizadas.

Uso como módulo:
  from Parte4_excel import actualizar_excel, organizar_archivos

Uso independiente:
  .\python_portable\python.exe Parte4_excel.py  (no se usa solo normalmente)
=============================================================
"""

import sys
import io
import os
import re
import shutil
import logging
import traceback
from pathlib import Path
from datetime import datetime

# Forzar UTF-8 (solo si no está ya configurado)
if hasattr(sys.stdout, "buffer") and getattr(sys.stdout, 'encoding', '') != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "buffer") and getattr(sys.stderr, 'encoding', '') != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("ERROR: Instala openpyxl con:")
    print("  .\\python_portable\\python.exe -m pip install openpyxl")
    sys.exit(1)

# ════════════════════════════════════════════════════════
#  CONFIGURACIÓN
# ════════════════════════════════════════════════════════

DESCARGA_BASE = Path("descargas")
OUTPUT_BASE = Path("output")
EXCEL_PATH = Path("TrámitesCRT.xlsx")
SHEET_NAME = "Turnados recibidos"

ORGANIZAR_DESCARGAS = True
BORRAR_CARPETA_FOLIO_VACIA = False

# Carpeta compartida/salida consolidada donde se sincroniza el Excel.
# En Linux queda deshabilitada si SATYS_CARPETA_COMPARTIDA no está definida.
CARPETA_COMPARTIDA_EXCEL: Path | None = (
    Path(os.getenv("SATYS_CARPETA_COMPARTIDA")).expanduser()
    if os.getenv("SATYS_CARPETA_COMPARTIDA", "").strip()
    else None
)

EXCEL_EXTS = {".xls", ".xlsx", ".xlsm", ".xlsb", ".csv"}
WORD_EXTS = {".doc", ".docx"}

# ════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("SATyS-Excel")


# ────────────────────────────────────────────────────────
#  NOTAS VÍCTOR (tipos de archivo)
# ────────────────────────────────────────────────────────

def obtener_nota_victor(carpeta: Path) -> str:
    """
    Devuelve nota sobre formatos entregados, solo si hay archivos
    distintos de .pdf, .csv y .png (imágenes de sello).

    Ejemplo de salida:
      "Formato entregado en .docx"
      "Formato entregado en .xlsx, .docx"
      ""  (si solo hay PDF/CSV o la carpeta está vacía)
    """
    # Extensiones de archivo que son "solo infraestructura" del pipeline
    _EXCLUIR = {".pdf", ".csv", ".png", ".jpg", ".jpeg", ".tmp", ".json"}

    exts_relevantes = set()
    for archivo in carpeta.glob("*"):
        if archivo.is_file():
            ext = archivo.suffix.lower()
            if ext and ext not in _EXCLUIR:
                exts_relevantes.add(ext)

    if not exts_relevantes:
        return ""

    exts_ordenadas = ", ".join(sorted(exts_relevantes))
    return f"Formato entregado en {exts_ordenadas}"


# ────────────────────────────────────────────────────────
#  ORGANIZACIÓN DE ARCHIVOS
# ────────────────────────────────────────────────────────

def _ruta_a_path(ruta: str) -> Path:
    """Convierte una ruta con separadores mixtos a Path."""
    partes = [p for p in re.split(r"[\\/]+", ruta) if p]
    return Path(*partes)


def _destino_sin_colision(destino: Path, item: Path) -> Path:
    """Retorna la ruta destino. En esta versión, siempre sobrescribe el archivo original."""
    return destino / item.name


def organizar_archivos(carpeta_folio: Path, ruta: str) -> Path | None:
    """
    Mueve los archivos de la carpeta del folio a la ruta estandarizada.
    Retorna la ruta destino o None.
    """
    if not ruta:
        return None

    destino = OUTPUT_BASE / _ruta_a_path(ruta)
    destino.mkdir(parents=True, exist_ok=True)

    copiados = 0
    for item in carpeta_folio.iterdir():
        if item.resolve() == destino.resolve():
            continue
            
        # Excluir archivos JSON de la organización
        if item.suffix.lower() == ".json":
            continue

        target = _destino_sin_colision(destino, item)
        try:
            # Si el destino ya existe, lo eliminamos para que shutil.copy sobrescriba limpiamente
            if target.exists():
                if target.is_file():
                    target.unlink()
                else:
                    shutil.rmtree(target)
            
            if item.is_file():
                shutil.copy2(str(item), str(target))
            else:
                shutil.copytree(str(item), str(target), dirs_exist_ok=True)
            copiados += 1
        except Exception as e:
            log.error("❌ No se pudo copiar %s → %s: %s", item.name, target, e)

    if copiados:
        log.info("📁 %d archivos copiados a: %s", copiados, destino)

    return destino


# ────────────────────────────────────────────────────────
#  BÚSQUEDA DE FILA EN EXCEL
# ────────────────────────────────────────────────────────

def _buscar_fila(ws, folio: str, registro: str = None, col_registro: int = None) -> int | None:
    """
    Busca la fila que corresponde a un (folio, registro).

    Un mismo folio puede tener MAS DE UN tramite/registro asociado en SATyS
    (ej. folio 1660 con registros CRT26-020606 y CRT26-002483, cada uno con
    su propio operador/representante). Por eso, cuando se pasa 'registro':

      1. Si ya existe una fila para este folio CON ESE MISMO registro
         -> se reutiliza esa fila (idempotente: volver a correr el script
         no duplica filas).
      2. Si existe una fila para este folio pero SIN registro asignado aun
         -> se reutiliza esa fila (primera vez que se completa el dato).
      3. Si todas las filas existentes de este folio ya tienen un registro
         DISTINTO -> se devuelve None para que el caller cree una fila nueva
         (es un tramite distinto del mismo folio, necesita su propia fila).

    Si no se pasa 'registro', se conserva el comportamiento original:
    primera fila cuyo texto en columna D o E contenga el folio.
    """
    candidatos = []  # (fila, valor_registro_en_esa_fila)
    for row in range(2, ws.max_row + 1):
        # Columna D (1711)
        celda_d = ws.cell(row=row, column=4).value
        coincide = bool(celda_d and str(folio) in str(celda_d).upper())
        # Columna E (Memo/Volante)
        celda_e = ws.cell(row=row, column=5).value
        coincide = coincide or bool(celda_e and str(folio) in str(celda_e))

        if coincide:
            valor_registro = None
            if col_registro:
                valor_registro = ws.cell(row=row, column=col_registro).value
            candidatos.append((row, valor_registro))

    if not candidatos:
        return None

    if not registro:
        # Comportamiento original: primera coincidencia por folio.
        return candidatos[0][0]

    registro_norm = str(registro).strip().upper()

    # 1) Coincidencia exacta de registro -> misma fila (idempotente)
    for row, valor_registro in candidatos:
        if valor_registro and str(valor_registro).strip().upper() == registro_norm:
            return row

    # 2) Fila de este folio que aun no tiene registro asignado -> reusar
    for row, valor_registro in candidatos:
        if not valor_registro:
            return row

    # 3) Todas las filas de este folio ya tienen un registro DISTINTO
    #    -> es un tramite distinto del mismo folio, necesita fila nueva.
    return None


# ────────────────────────────────────────────────────────
#  ACTUALIZACIÓN DEL EXCEL
# ────────────────────────────────────────────────────────

def actualizar_excel(
    folio: str,
    registro: str = "",
    nombre_operador: str = "",
    representante_legal: str = "",
    formatos: dict = None,
    rpc_resultado: dict = None,
    nota_victor: str = "",
    imagen_sello: Path = None,
    fecha_sello: str = "",
    excel_path: Path = None,
    sheet_name: str = None,
    asunto: str = "",
    tipo_tramite: str = "",
    fecha_limite: str = "",
) -> bool:
    """
    Actualiza la fila correspondiente al folio en el Excel.
    Retorna True si fue exitoso.
    """
    excel = excel_path or EXCEL_PATH
    sheet = sheet_name or SHEET_NAME
    formatos = formatos or {}

    try:
        wb = openpyxl.load_workbook(excel)
        ws = wb[sheet]

        # Leer encabezados dinámicamente
        encabezados = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                encabezados[str(header).strip()] = col

        # Mapeo de columnas
        col_1711        = encabezados.get("1711", 4)
        col_memo        = encabezados.get("Memo/Volante", 5)
        col_solicitante = encabezados.get("Solicitante Promovente", 6)
        col_rep         = encabezados.get("Representante Legal", 7)
        col_fecha       = encabezados.get("Fecha de creación", 10)
        col_ruta        = encabezados.get("Ruta", 13)
        col_notas       = encabezados.get("NOTAS_VICTOR", 42)
        col_asunto      = encabezados.get("Asunto", None)
        col_tipo        = encabezados.get("Tipo Trámite", None)
        col_fecha_limite = encabezados.get("FECHA LÍMITE", None)

        # Idempotencia: si el registro/folio ya existe, actualizar esa fila
        # en lugar de duplicar. Esto permite reprocesar TODO descargas/ sin
        # generar filas repetidas en TrámitesCRT.xlsx.
        def _norm_excel(v):
            return str(v or "").strip().upper()

        fila = None
        registro_norm = _norm_excel(registro)
        folio_norm = _norm_excel(folio)

        # Prioridad 1: misma columna 1711 (registro), si hay registro.
        if registro_norm:
            for r in range(2, ws.max_row + 1):
                if _norm_excel(ws.cell(row=r, column=col_1711).value) == registro_norm:
                    fila = r
                    break

        # Prioridad 2: mismo Memo/Volante, solo si no se encontró por registro.
        if fila is None and folio_norm:
            for r in range(2, ws.max_row + 1):
                if _norm_excel(ws.cell(row=r, column=col_memo).value) == folio_norm:
                    fila = r
                    break

        if fila is None:
            fila = ws.max_row + 1
            log.info("➕ Agregando nueva fila %d para folio %s (registro %s)", fila, folio, registro or "N/A")
        else:
            log.info("♻️  Actualizando fila existente %d para folio %s (registro %s)", fila, folio, registro or "N/A")

        ws.cell(row=fila, column=col_memo, value=folio)

        # Escribir datos
        if registro:
            ws.cell(row=fila, column=col_1711, value=registro)
        if nombre_operador:
            ws.cell(row=fila, column=col_solicitante, value=nombre_operador)
        if representante_legal:
            ws.cell(row=fila, column=col_rep, value=representante_legal)
        if fecha_sello:
            fecha_sello = fecha_sello.replace("-", "/")
            ws.cell(row=fila, column=col_fecha, value=fecha_sello)
            log.info("   📅 Fecha sello → col %s: %s", get_column_letter(col_fecha), fecha_sello)
        if rpc_resultado and rpc_resultado.get("ok"):
            ws.cell(row=fila, column=col_ruta, value=rpc_resultado["ruta"])

        # Asunto
        if asunto and col_asunto:
            ws.cell(row=fila, column=col_asunto, value=asunto)
            log.info("   📝 Asunto → col %s", get_column_letter(col_asunto))

        # Tipo Trámite
        if tipo_tramite and col_tipo:
            ws.cell(row=fila, column=col_tipo, value=tipo_tramite)
            log.info("   📋 Tipo Trámite → col %s", get_column_letter(col_tipo))

        # FECHA LÍMITE: solo si viene plazo_atencion de metadata_tramite_nuevo.json
        if fecha_limite and col_fecha_limite:
            ws.cell(row=fila, column=col_fecha_limite, value=fecha_limite)
            log.info("   🗓️ FECHA LÍMITE → col %s: %s", get_column_letter(col_fecha_limite), fecha_limite)
        elif fecha_limite and not col_fecha_limite:
            log.warning("⚠️  Se tiene plazo_atencion (%s) pero no se encontró columna 'FECHA LÍMITE' en el Excel.", fecha_limite)

        # Formatos R001–R027
        for fmt, presente in formatos.items():
            if presente and fmt in encabezados:
                col = encabezados[fmt]
                ws.cell(row=fila, column=col, value=1)
                log.info("   ✅ Marcado %s en columna %s", fmt, get_column_letter(col))

        # NOTAS_VICTOR: solo formato entregado (nunca score RPC, nunca .json)
        if nota_victor:
            valor_actual = ws.cell(row=fila, column=col_notas).value
            if valor_actual and nota_victor not in str(valor_actual):
                nota_victor = f"{valor_actual}; {nota_victor}"
            elif valor_actual and nota_victor in str(valor_actual):
                nota_victor = str(valor_actual)  # ya está, no duplicar
            ws.cell(row=fila, column=col_notas, value=nota_victor)

        wb.save(excel)
        wb.close()
        log.info("💾 Excel guardado: %s", excel)

        # Sincronizar inmediatamente a la carpeta de red (si está configurada)
        sincronizar_excel_a_red(excel)

        return True

    except PermissionError:
        log.error("❌ El archivo Excel está abierto. Ciérralo e inténtalo de nuevo.")
        return False
    except Exception as e:
        log.error("❌ Error actualizando Excel: %s", e)
        traceback.print_exc()
        return False


# ────────────────────────────────────────────────────────
#  SINCRONIZACIÓN INMEDIATA DEL EXCEL A LA RED
# ────────────────────────────────────────────────────────

def sincronizar_excel_a_red(excel_local: Path = None) -> None:
    """
    Copia el Excel local a SATYS_CARPETA_COMPARTIDA, si está definido
    inmediatamente después de cada guardado.

    - Si el archivo de red no existe, lo crea.
    - Si ya existe, lo sobreescribe (la versión local es siempre la fuente
      de verdad: las filas solo se agregan al final, nunca se eliminan).
    - Los errores se loguean pero NO detienen el flujo principal.
    """
    if CARPETA_COMPARTIDA_EXCEL is None:
        return

    local = excel_local or EXCEL_PATH
    if not local.exists():
        log.warning("⚠️  [Red] Excel local no encontrado, no se puede sincronizar: %s", local)
        return

    destino_dir = Path(CARPETA_COMPARTIDA_EXCEL)
    destino = destino_dir / local.name

    try:
        destino_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(local), str(destino))
        log.info("🌐 Excel sincronizado a red: %s", destino)
    except PermissionError:
        log.warning(
            "⚠️  [Red] El Excel en la red está abierto por otro usuario. "
            "Se sincronizará en la próxima ejecución: %s", destino
        )
    except Exception as e:
        log.warning("⚠️  [Red] No se pudo sincronizar el Excel a la red (%s): %s", destino, e)


# ────────────────────────────────────────────────────────
#  CLI
# ────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Parte4_excel.py — Módulo de actualización de Excel.")
    print("Normalmente se ejecuta desde main_procesar.py")
    print()
    print("Funciones disponibles:")
    print("  actualizar_excel(folio, pdf_nombre, nombre_operador, ...)")
    print("  organizar_archivos(carpeta_folio, ruta)")
    print("  obtener_nota_victor(carpeta)")
    print("  sincronizar_excel_a_red(excel_local)")