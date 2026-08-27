#!/usr/bin/env python3
r"""
buscar_concesionario.py
───────────────────────
Compatibilidad Linux para el cruce original SATyS/RPC.

Regla de negocio:
  - metadata_satys.json aporta id_solicitante.
  - El Excel oficial del RPC aporta la columna ID OPERADOR.
  - Si id_solicitante == ID OPERADOR: coincidencia 100%.
  - Si no existe/no coincide, se permite nombre exacto normalizado.
  - Como respaldo se consultan los resultados y el autocompletado públicos del
    RPC, que devuelven ID OPERADOR y nombres actuales.
  - Una variante no exacta sólo se acepta con umbral alto, cobertura de palabras,
    margen suficiente y un único ID; las ambigüedades quedan para revisión.

Este módulo conserva los nombres de funciones que main_procesar.py esperaba en
la versión Windows: cargar_catalogo_desde_excel, preparar_catalogo_para_matching,
buscar_por_id_solicitante y buscar_coincidencias.
"""

from __future__ import annotations

import logging
import os
import re
import time
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:
    raise RuntimeError("Falta openpyxl. Instala con: python -m pip install openpyxl") from exc

try:
    import requests
except ImportError:  # La lectura del Excel sigue disponible en modo offline.
    requests = None

try:
    from bs4 import BeautifulSoup
except ImportError:  # El autocompletado sigue disponible sin parsear resultados HTML.
    BeautifulSoup = None

try:
    from rapidfuzz import fuzz
except ImportError:  # Fallback determinista con difflib.
    fuzz = None

log = logging.getLogger("SATyS-BuscarConcesionario")

RPC_AUTOCOMPLETE_URL = "https://rpc.ift.org.mx/vrpc/RpcServicesController/searchBP"
RPC_RESULTADOS_URL = "https://rpc.ift.org.mx/vrpc/RpcSearchController/searchConcesiones"
RPC_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0",
    "Accept": "application/json, text/html, */*",
    "Accept-Language": "es-MX,es;q=0.9",
    "Referer": "https://rpc.ift.org.mx/vrpc/",
}

# Evita consultar repetidamente el RPC cuando varios registros pertenecen al
# mismo operador durante una corrida. No se persiste para no volver obsoleta la
# fuente en línea entre ejecuciones.
_CACHE_RPC_NOMBRE: dict[str, dict[str, Any]] = {}
_SESION_RPC = None


def _sin_acentos(valor: Any) -> str:
    texto = "" if valor is None else str(valor)
    texto = unicodedata.normalize("NFD", texto)
    return "".join(c for c in texto if unicodedata.category(c) != "Mn")


def normalizar_header(valor: Any) -> str:
    """Normaliza encabezados para comparar sin acentos ni signos."""
    texto = _sin_acentos(valor).upper().strip()
    texto = re.sub(r"[^A-Z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def normalizar_id(valor: Any) -> str:
    """Normaliza IDs de SATyS/Excel: 518858, 518858.0, ' 518858 '."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip()
    if not texto:
        return ""
    if re.fullmatch(r"\d+\.0+", texto):
        texto = texto.split(".", 1)[0]
    return re.sub(r"\s+", "", texto).upper()


def normalizar_nombre(valor: Any) -> str:
    texto = _sin_acentos(valor).lower()
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _buscar_hoja(wb, nombre_preferido: str | None):
    if nombre_preferido and nombre_preferido in wb.sheetnames:
        return wb[nombre_preferido]
    if nombre_preferido:
        objetivo = normalizar_header(nombre_preferido)
        for nombre in wb.sheetnames:
            if normalizar_header(nombre) == objetivo:
                return wb[nombre]
    return wb[wb.sheetnames[0]]


def _detectar_encabezados(ws, max_filas: int = 30) -> tuple[int, dict[str, int]]:
    """Detecta fila de encabezados sin acceso aleatorio ``ws.cell()``.

    En hojas abiertas con ``read_only=True``, ``ws.cell(fila, columna)`` puede
    recorrer repetidamente el XML interno. Iterar secuencialmente evita que la
    detección se vuelva innecesariamente costosa en archivos grandes.
    """
    claves_id = {
        "ID OPERADOR", "ID DE OPERADOR", "IDOPERADOR", "ID BP", "IDBP", "ID_BP",
        "ID CONCESIONARIO", "IDCONCESIONARIO",
    }
    claves_id_compactas = {valor.replace(" ", "") for valor in claves_id}
    mejor: tuple[int, dict[str, int], int] | None = None

    limite = min(ws.max_row or max_filas, max_filas)
    filas = ws.iter_rows(min_row=1, max_row=limite, values_only=True)

    for row_idx, valores in enumerate(filas, start=1):
        headers: dict[str, int] = {}
        puntaje = 0
        for col_idx, valor in enumerate(valores, start=1):
            h = normalizar_header(valor)
            if not h:
                continue
            headers[h] = col_idx
            h_sin = h.replace(" ", "")
            if h in claves_id or h_sin in claves_id_compactas:
                puntaje += 100
            if any(k in h for k in ("OPERADOR", "CONCESIONARIO", "RAZON SOCIAL", "NOMBRE", "TITULAR", "DENOMINACION")):
                puntaje += 2
            if h in {"ESTATUS", "ESTADO", "VIGENCIA", "SITUACION"}:
                puntaje += 1
        if puntaje and (mejor is None or puntaje > mejor[2]):
            mejor = (row_idx, headers, puntaje)

    if mejor is None or mejor[2] < 100:
        raise ValueError("No se encontró encabezado con columna 'ID OPERADOR' en el Excel RPC")
    return mejor[0], mejor[1]


def _columna_por_alias(headers: dict[str, int], aliases: list[str]) -> int | None:
    aliases_norm = [normalizar_header(a) for a in aliases]
    aliases_sin = {a.replace(" ", "") for a in aliases_norm}
    for alias in aliases_norm:
        if alias in headers:
            return headers[alias]
    for h, col in headers.items():
        if h.replace(" ", "") in aliases_sin:
            return col
    return None


def cargar_catalogo_desde_excel(
    excel_path: str | Path,
    hoja: str = "copeau",
    solo_vigentes: bool = False,
) -> list[dict[str, Any]]:
    """
    Lee el Excel oficial del RPC y devuelve operadores normalizados.

    Requiere columna ID OPERADOR. Para el nombre busca, en orden flexible,
    columnas como OPERADOR, CONCESIONARIO, RAZON SOCIAL, DENOMINACION, TITULAR.

    La lectura es secuencial mediante ``iter_rows(values_only=True)``. Esto es
    crítico con ``read_only=True``: el acceso repetido con ``ws.cell()`` puede
    convertir una lectura lineal en una operación aproximadamente cuadrática.
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el Excel RPC: {excel_path}")

    inicio = time.perf_counter()
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = _buscar_hoja(wb, hoja)
        header_row, headers = _detectar_encabezados(ws)

        col_id = _columna_por_alias(headers, [
            "ID OPERADOR", "ID DE OPERADOR", "IDOPERADOR", "ID BP", "IDBP", "ID_BP",
            "ID CONCESIONARIO", "IDCONCESIONARIO",
        ])
        if not col_id:
            raise ValueError("No se encontró la columna 'ID OPERADOR' en el Excel RPC")

        col_nombre = _columna_por_alias(headers, [
            "OPERADOR", "NOMBRE OPERADOR", "NOMBRE DEL OPERADOR", "CONCESIONARIO",
            "NOMBRE CONCESIONARIO", "NOMBRE DEL CONCESIONARIO", "RAZON SOCIAL", "RAZÓN SOCIAL",
            "DENOMINACION", "DENOMINACIÓN", "TITULAR", "NOMBRE", "NOMBRE COMPLETO",
        ])
        if not col_nombre:
            for h, col in headers.items():
                if col != col_id and any(k in h for k in ("OPERADOR", "CONCESIONARIO", "RAZON", "DENOMINACION", "TITULAR", "NOMBRE")):
                    col_nombre = col
                    break
        if not col_nombre:
            raise ValueError("No se encontró columna de nombre de operador/concesionario en el Excel RPC")

        col_estatus = _columna_por_alias(headers, ["ESTATUS", "ESTADO", "SITUACION", "SITUACIÓN", "VIGENCIA"])
        columnas_requeridas = [col_id, col_nombre]
        if col_estatus:
            columnas_requeridas.append(col_estatus)
        max_col_requerida = max(columnas_requeridas)

        log.info(
            "Leyendo catálogo RPC secuencialmente desde %s/%s "
            "(encabezado fila %d, columnas hasta %d)...",
            excel_path.name,
            ws.title,
            header_row,
            max_col_requerida,
        )

        catalogo: list[dict[str, Any]] = []
        ids_vistos: set[str] = set()
        duplicados = 0
        filas_vacias = 0

        filas = ws.iter_rows(
            min_row=header_row + 1,
            max_col=max_col_requerida,
            values_only=True,
        )
        for row, valores in enumerate(filas, start=header_row + 1):
            id_bp = normalizar_id(valores[col_id - 1])
            nombre = str(valores[col_nombre - 1] or "").strip()
            if not id_bp or not nombre:
                filas_vacias += 1
                continue

            if solo_vigentes and col_estatus:
                estatus = normalizar_header(valores[col_estatus - 1])
                if estatus and "VIGENTE" not in estatus:
                    continue

            if id_bp in ids_vistos:
                duplicados += 1
                continue
            ids_vistos.add(id_bp)

            catalogo.append({
                "idBp": id_bp,
                "ID OPERADOR": id_bp,
                "concesionario": nombre,
                "nombre_completo": nombre,
                "_fila_excel_rpc": row,
                "_hoja_excel_rpc": ws.title,
            })

        duracion = time.perf_counter() - inicio
        log.info(
            "Catálogo RPC cargado desde %s/%s: %d operadores en %.2f segundos "
            "(%d duplicados omitidos, %d filas vacías/incompletas)",
            excel_path.name,
            ws.title,
            len(catalogo),
            duracion,
            duplicados,
            filas_vacias,
        )
        return catalogo
    finally:
        wb.close()


def preparar_catalogo_para_matching(catalogo: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Compatibilidad con main_procesar.py: agrega norm/compact para identificar
    que el catálogo viene de Excel RPC, aunque el cruce operativo sea por ID.
    """
    preparado: list[dict[str, Any]] = []
    for item in catalogo or []:
        id_bp = normalizar_id(item.get("idBp") or item.get("ID OPERADOR"))
        nombre = str(item.get("nombre_completo") or item.get("concesionario") or "").strip()
        if not id_bp or not nombre:
            continue
        norm = normalizar_nombre(nombre)
        nuevo = dict(item)
        nuevo.update({
            "idBp": id_bp,
            "ID OPERADOR": id_bp,
            "concesionario": nombre,
            "nombre_completo": nombre,
            "norm": norm,
            "compact": norm.replace(" ", ""),
        })
        preparado.append(nuevo)
    return preparado


def buscar_por_id_solicitante(id_solicitante: Any, catalogo: list[dict[str, Any]]) -> dict[str, Any] | None:
    """Cruce exacto: id_solicitante de SATyS == ID OPERADOR del Excel RPC."""
    objetivo = normalizar_id(id_solicitante)
    if not objetivo:
        return None
    for item in catalogo or []:
        id_bp = normalizar_id(item.get("idBp") or item.get("ID OPERADOR"))
        if id_bp == objetivo:
            nombre = str(item.get("nombre_completo") or item.get("concesionario") or "").strip()
            return {
                "idBp": id_bp,
                "ID OPERADOR": id_bp,
                "concesionario": nombre,
                "nombre_completo": nombre,
                "score": 1.0,
            }
    return None


def buscar_por_nombre_operador_exacto(nombre_operador: Any, catalogo: list[dict[str, Any]]) -> dict[str, Any] | None:
    """Cruce exacto normalizado y único por nombre contra el Excel RPC."""
    diagnostico = diagnosticar_nombre_operador_exacto(nombre_operador, catalogo)
    if not diagnostico.get("ok"):
        return None
    return {
        "idBp": diagnostico["idBp"],
        "ID OPERADOR": diagnostico["idBp"],
        "concesionario": diagnostico["nombre_completo"],
        "nombre_completo": diagnostico["nombre_completo"],
        "score": 1.0,
    }


def diagnosticar_nombre_operador_exacto(
    nombre_operador: Any,
    catalogo: list[dict[str, Any]],
) -> dict[str, Any]:
    """Devuelve coincidencia exacta sólo cuando conduce a un único ID.

    Se eliminan acentos, puntuación y diferencias de espacios/mayúsculas. Por
    ello ``Wal-Mart.S.A.`` y ``WalMart S.A.`` comparten la clave
    ``walmartsa``. Si esa clave aparece con dos IDs diferentes, la función no
    elige arbitrariamente y devuelve ``estado=ambiguo``.
    """
    objetivo = normalizar_nombre(nombre_operador)
    if not objetivo:
        return {"ok": False, "estado": "nombre_vacio", "candidatos": []}
    objetivo_compact = objetivo.replace(" ", "")

    coincidencias: dict[tuple[str, str], dict[str, str]] = {}
    for item in catalogo or []:
        nombre = str(item.get("nombre_completo") or item.get("concesionario") or "").strip()
        norm = item.get("norm") or normalizar_nombre(nombre)
        compact = item.get("compact") or norm.replace(" ", "")
        if norm == objetivo or (objetivo_compact and compact == objetivo_compact):
            id_bp = normalizar_id(item.get("idBp") or item.get("ID OPERADOR"))
            if id_bp:
                coincidencias[(id_bp, nombre)] = {"idBp": id_bp, "nombre_completo": nombre}

    candidatos = list(coincidencias.values())
    ids = {item["idBp"] for item in candidatos}
    if not candidatos:
        return {
            "ok": False,
            "estado": "sin_coincidencia",
            "nombre_normalizado": objetivo_compact,
            "candidatos": [],
        }
    if len(ids) != 1:
        return {
            "ok": False,
            "estado": "ambiguo",
            "nombre_normalizado": objetivo_compact,
            "candidatos": candidatos,
        }

    elegido = candidatos[0]
    return {
        "ok": True,
        "estado": "coincidencia_unica",
        "idBp": elegido["idBp"],
        "nombre_completo": elegido["nombre_completo"],
        "nombre_normalizado": objetivo_compact,
        "candidatos": candidatos,
    }


def _extraer_respuesta_rpc(data: Any) -> list[dict[str, str]]:
    """Convierte la respuesta de ``searchBP`` a una lista uniforme."""
    if data is None:
        return []
    if isinstance(data, dict):
        for clave in ("data", "results", "items", "concesionarios", "lista"):
            if clave in data:
                return _extraer_respuesta_rpc(data[clave])
        data = [data]
    if not isinstance(data, list):
        return []

    items: list[dict[str, str]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        nombre = str(
            item.get("concesionario")
            or item.get("nombre")
            or item.get("denominacion")
            or item.get("label")
            or ""
        ).strip()
        id_bp = normalizar_id(item.get("idBp") or item.get("id") or item.get("id_bp"))
        if nombre and id_bp:
            items.append({"idBp": id_bp, "nombre_completo": nombre})
    return items


_SUFIJOS_LEGALES_TOKENS = (
    ("s", "a", "p", "i", "de", "c", "v"),
    ("s", "a", "s", "de", "c", "v"),
    ("s", "a", "b", "de", "c", "v"),
    ("s", "de", "r", "l", "de", "c", "v"),
    ("s", "a", "de", "c", "v"),
    ("s", "de", "r", "l"),
    ("s", "a", "p", "i"),
    ("s", "a", "s"),
    ("s", "a", "b"),
    ("s", "a"),
    ("s", "c"),
    ("a", "c"),
)


def _nombre_base_sin_sufijo_legal(nombre: Any) -> str:
    tokens = normalizar_nombre(nombre).split()
    cambio = True
    while tokens and cambio:
        cambio = False
        for sufijo in _SUFIJOS_LEGALES_TOKENS:
            if len(tokens) >= len(sufijo) and tuple(tokens[-len(sufijo):]) == sufijo:
                del tokens[-len(sufijo):]
                cambio = True
                break
    return " ".join(tokens)


def _obtener_sesion_rpc(session: Any = None):
    """Reutiliza conexiones HTTP durante una corrida con muchos registros."""
    if session is not None:
        session.headers.update(RPC_HEADERS)
        return session
    if requests is None:
        return None

    global _SESION_RPC
    if _SESION_RPC is None:
        _SESION_RPC = requests.Session()
        _SESION_RPC.headers.update(RPC_HEADERS)
    return _SESION_RPC


def dividir_nombres_operador(valor: Any) -> list[str]:
    """Separa una celda que enumera varias razones sociales completas."""
    texto = re.sub(r"\s+", " ", str(valor or "")).strip()
    if not texto:
        return []
    partes = re.split(
        r"(?i)(?<=C\.V\.),\s*(?=[A-ZÁÉÍÓÚÜÑ&])|\s*;\s*",
        texto,
    )
    return [parte.strip(" ,;") for parte in partes if parte.strip(" ,;")]


def _porcentaje_similitud_rpc(entrada: Any, candidato: Any) -> float:
    """Replica el balance de similitud usado por ``rpc_operador.py``."""
    a = normalizar_nombre(entrada)
    b = normalizar_nombre(candidato)
    if not a or not b:
        return 0.0
    if a == b:
        return 100.0
    if fuzz is not None:
        return round(0.70 * fuzz.WRatio(a, b) + 0.30 * fuzz.ratio(a, b), 2)
    return round(100.0 * SequenceMatcher(None, a, b).ratio(), 2)


def _candidatos_unicos(candidatos: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unicos: dict[tuple[str, str], dict[str, Any]] = {}
    for candidato in candidatos:
        id_bp = normalizar_id(candidato.get("idBp") or candidato.get("numero_rpc"))
        nombre = str(
            candidato.get("nombre_completo")
            or candidato.get("concesionario")
            or ""
        ).strip()
        if not id_bp or not nombre:
            continue
        clave = (id_bp, normalizar_nombre(nombre).replace(" ", ""))
        combinado = dict(candidato)
        combinado["idBp"] = id_bp
        combinado["nombre_completo"] = nombre
        unicos.setdefault(clave, combinado)
    return list(unicos.values())


def seleccionar_candidato_rpc_seguro(
    nombre_operador: Any,
    candidatos: list[dict[str, Any]],
    *,
    similitud_minima: float | None = None,
    margen_minimo: float | None = None,
) -> dict[str, Any]:
    """Elige sólo una coincidencia exacta, de base legal o inequívocamente alta.

    La puntuación por sí sola no basta: también se exige cobertura de tokens y
    una separación clara frente al segundo ID. Así se aprovechan variantes como
    ``VALDEZ``/``VALDES`` sin convertir un parecido débil en asignación automática.
    """
    objetivo = normalizar_nombre(nombre_operador)
    objetivo_compacto = objetivo.replace(" ", "")
    objetivo_base = _nombre_base_sin_sufijo_legal(nombre_operador)
    objetivo_base_compacto = objetivo_base.replace(" ", "")
    candidatos = _candidatos_unicos(candidatos)
    if not objetivo_compacto or not candidatos:
        return {"ok": False, "estado": "sin_coincidencia", "candidatos": []}

    def resolver_unicos(coincidencias: list[dict[str, Any]], metodo: str):
        ids = {item["idBp"] for item in coincidencias}
        if len(ids) == 1:
            elegido = coincidencias[0]
            return {
                "ok": True,
                "estado": "coincidencia_unica",
                "metodo": metodo,
                "idBp": elegido["idBp"],
                "nombre_completo": elegido["nombre_completo"],
                "score": 1.0,
                "candidatos": coincidencias,
            }
        return {
            "ok": False,
            "estado": "ambiguo",
            "motivo": "nombre_corresponde_a_varios_id_rpc",
            "candidatos": coincidencias,
        }

    exactos = [
        item for item in candidatos
        if normalizar_nombre(item["nombre_completo"]).replace(" ", "") == objetivo_compacto
    ]
    if exactos:
        return resolver_unicos(exactos, "nombre_exacto_rpc")

    if len(objetivo_base_compacto) >= 5:
        exactos_base = [
            item for item in candidatos
            if _nombre_base_sin_sufijo_legal(item["nombre_completo"]).replace(" ", "")
            == objetivo_base_compacto
        ]
        if exactos_base:
            return resolver_unicos(exactos_base, "nombre_base_legal_rpc")

    minimo = float(
        similitud_minima
        if similitud_minima is not None
        else os.getenv("SATYS_RPC_SIMILITUD_MINIMA", "96")
    )
    margen_requerido = float(
        margen_minimo
        if margen_minimo is not None
        else os.getenv("SATYS_RPC_MARGEN_MINIMO", "5")
    )
    tokens_objetivo = set(objetivo_base.split())
    ranking: list[dict[str, Any]] = []
    for item in candidatos:
        nombre = item["nombre_completo"]
        nombre_base = _nombre_base_sin_sufijo_legal(nombre)
        tokens_candidato = set(nombre_base.split())
        cobertura = (
            len(tokens_objetivo & tokens_candidato) / max(len(tokens_objetivo), len(tokens_candidato))
            if tokens_objetivo and tokens_candidato
            else 0.0
        )
        fila = dict(item)
        fila["similitud"] = max(
            _porcentaje_similitud_rpc(nombre_operador, nombre),
            _porcentaje_similitud_rpc(objetivo_base, nombre_base),
        )
        fila["cobertura_tokens"] = round(cobertura, 4)
        ranking.append(fila)
    ranking.sort(key=lambda item: (item["similitud"], item["cobertura_tokens"]), reverse=True)

    mejor = ranking[0]
    segundo_score = next(
        (item["similitud"] for item in ranking[1:] if item["idBp"] != mejor["idBp"]),
        0.0,
    )
    margen = round(mejor["similitud"] - segundo_score, 2)
    if (
        len(objetivo_base_compacto) >= 8
        and mejor["similitud"] >= minimo
        and mejor["cobertura_tokens"] >= 0.75
        and margen >= margen_requerido
    ):
        return {
            "ok": True,
            "estado": "coincidencia_alta_confianza",
            "metodo": "nombre_alta_confianza_rpc",
            "idBp": mejor["idBp"],
            "nombre_completo": mejor["nombre_completo"],
            "score": mejor["similitud"] / 100.0,
            "margen": margen,
            "candidatos": ranking[:10],
        }

    return {
        "ok": False,
        "estado": "sin_coincidencia_segura",
        "motivo": "candidatos_rpc_sin_confianza_suficiente",
        "score": mejor["similitud"] / 100.0,
        "margen": margen,
        "candidatos": ranking[:10],
    }


def _extraer_resultados_concesiones_html(html: str) -> list[dict[str, str]]:
    """Extrae ``FOLIO-ID - NOMBRE`` de las tarjetas de resultados del RPC."""
    if BeautifulSoup is None:
        return []
    soup = BeautifulSoup(html or "", "html.parser")
    resultados: list[dict[str, str]] = []
    for encabezado in soup.select(".strip_all_tour_list h3"):
        texto = re.sub(r"\s+", " ", encabezado.get_text(" ", strip=True)).strip()
        match = re.match(r"^([A-Z0-9]+)-(\d+)\s+-\s+(.+)$", texto, re.IGNORECASE)
        if not match:
            continue
        resultados.append({
            "folio_electronico": match.group(1).upper(),
            "idBp": normalizar_id(match.group(2)),
            "nombre_completo": match.group(3).strip(),
        })
    return _candidatos_unicos(resultados)


def buscar_nombre_operador_rpc_resultados(
    nombre_operador: Any,
    *,
    timeout: float = 15.0,
    session: Any = None,
) -> dict[str, Any]:
    """Consulta la sección RESULTADOS, que puede encontrar nombres omitidos por autocomplete."""
    if requests is None or BeautifulSoup is None:
        return {
            "ok": False,
            "estado": "error_rpc",
            "motivo": "dependencias_rpc_html_no_disponibles",
            "candidatos": [],
        }

    nombres = dividir_nombres_operador(nombre_operador)
    if not nombres:
        return {"ok": False, "estado": "nombre_vacio", "candidatos": []}

    cliente = _obtener_sesion_rpc(session)
    candidatos: list[dict[str, Any]] = []
    errores: list[str] = []
    hubo_respuesta = False
    payload_base = {
        "txtBPConcesionario": "",
        "strServicios": "",
        "strFET": "",
        "strCobertura": "",
        "strExpediente": "",
        "strEstatus": "",
        "strCanal": "",
        "strTipo": "",
        "strSatelite": "",
        "strPosicion": "",
        "strRangoSegmentosFrom": "",
        "strRangoSegmentosTo": "",
        "strTipoUso": "",
        "strFrecuencia": "",
        "cbComercializadoraChecked": "0",
        "cbOMVChecked": "0",
    }
    for nombre in nombres:
        try:
            respuesta = cliente.post(
                RPC_RESULTADOS_URL,
                data={**payload_base, "strConcesionario": nombre},
                timeout=timeout,
            )
            respuesta.raise_for_status()
            hubo_respuesta = True
            for item in _extraer_resultados_concesiones_html(respuesta.text):
                item["consulta_rpc"] = nombre
                candidatos.append(item)
        except Exception as exc:
            errores.append(f"{type(exc).__name__}: {exc}")

    candidatos = _candidatos_unicos(candidatos)
    if len(nombres) > 1:
        ids = {item["idBp"] for item in candidatos}
        if len(ids) > 1:
            return {
                "ok": False,
                "estado": "ambiguo",
                "motivo": "metadata_contiene_varios_operadores_rpc",
                "candidatos": candidatos,
            }
        if not candidatos:
            return {
                "ok": False,
                "estado": "sin_coincidencia",
                "motivo": "nombres_compuestos_no_encontrados_en_resultados_rpc",
                "candidatos": [],
            }
        diagnostico = seleccionar_candidato_rpc_seguro(nombres[0], candidatos)
    else:
        diagnostico = seleccionar_candidato_rpc_seguro(nombres[0], candidatos)

    if diagnostico.get("ok"):
        metodo_base = diagnostico.get("metodo", "nombre_alta_confianza_rpc")
        diagnostico["metodo"] = {
            "nombre_exacto_rpc": "nombre_exacto_rpc_resultados",
            "nombre_base_legal_rpc": "nombre_base_legal_rpc_resultados",
            "nombre_alta_confianza_rpc": "nombre_alta_confianza_rpc_resultados",
        }.get(metodo_base, f"{metodo_base}_resultados")
        diagnostico["fuente"] = "rpc_online_resultados"
        diagnostico["consulta_rpc"] = nombres[0]
        return diagnostico
    if candidatos:
        diagnostico["fuente"] = "rpc_online_resultados"
        return diagnostico
    if hubo_respuesta:
        return {
            "ok": False,
            "estado": "sin_coincidencia",
            "motivo": "nombre_no_encontrado_en_resultados_rpc",
            "candidatos": [],
        }
    return {
        "ok": False,
        "estado": "error_rpc",
        "motivo": "rpc_resultados_no_disponible",
        "detalle_error": " | ".join(errores[-2:]),
        "candidatos": [],
    }


def _consultas_rpc_para_nombre(nombre: Any) -> list[str]:
    """Genera pocas consultas amplias; el resultado aún se valida exactamente."""
    original = re.sub(r"\s+", " ", str(nombre or "")).strip()
    base = _nombre_base_sin_sufijo_legal(original)
    tokens_base = base.split()

    # Conserva un posible guion del primer token (WAL-MART), que el buscador
    # público puede distinguir de WAL MART.
    primer_token_original = re.split(r"\s+", original, maxsplit=1)[0] if original else ""
    primer_token_original = re.sub(
        r"(?i)[,.;]*(?:s\.?a\.?|s\.?c\.?|a\.?c\.?).*$",
        "",
        primer_token_original,
    ).strip(" ,.;")

    propuestas = [
        original,
        base,
        " ".join(tokens_base[:5]),
        " ".join(tokens_base[:3]),
        " ".join(tokens_base[:2]),
        primer_token_original,
    ]
    consultas: list[str] = []
    vistos: set[str] = set()
    for propuesta in propuestas:
        propuesta = propuesta.strip()
        clave = propuesta.casefold()
        if len(propuesta) >= 2 and clave not in vistos:
            vistos.add(clave)
            consultas.append(propuesta)
    return consultas


def buscar_nombre_operador_rpc_online_exacto(
    nombre_operador: Any,
    *,
    timeout: float = 10.0,
    session: Any = None,
) -> dict[str, Any]:
    """Consulta resultados y autocomplete del RPC con selección conservadora."""
    objetivo = normalizar_nombre(nombre_operador)
    objetivo_compact = objetivo.replace(" ", "")
    if not objetivo_compact:
        return {"ok": False, "estado": "nombre_vacio", "candidatos": []}

    cacheado = _CACHE_RPC_NOMBRE.get(objetivo_compact)
    if cacheado is not None:
        return dict(cacheado)

    if requests is None:
        resultado = {
            "ok": False,
            "estado": "error_rpc",
            "motivo": "requests_no_disponible",
            "candidatos": [],
        }
        _CACHE_RPC_NOMBRE[objetivo_compact] = resultado
        return dict(resultado)

    cliente = _obtener_sesion_rpc(session)
    hubo_respuesta = False
    errores: list[str] = []
    candidatos_acumulados: list[dict[str, Any]] = []

    # El formulario de resultados encuentra nombres que searchBP omite. Es la
    # misma sección que muestra encabezados como FET...-ID - RAZÓN SOCIAL.
    diagnostico_resultados = buscar_nombre_operador_rpc_resultados(
        nombre_operador,
        timeout=max(timeout, 15.0),
        session=cliente,
    )
    if diagnostico_resultados.get("ok") or diagnostico_resultados.get("estado") == "ambiguo":
        diagnostico_resultados["nombre_normalizado"] = objetivo_compact
        _CACHE_RPC_NOMBRE[objetivo_compact] = diagnostico_resultados
        return dict(diagnostico_resultados)
    candidatos_acumulados.extend(diagnostico_resultados.get("candidatos") or [])
    if diagnostico_resultados.get("estado") == "error_rpc":
        detalle = diagnostico_resultados.get("detalle_error")
        if detalle:
            errores.append(str(detalle))

    for consulta in _consultas_rpc_para_nombre(nombre_operador):
        try:
            respuesta = cliente.get(
                RPC_AUTOCOMPLETE_URL,
                params={"query": consulta},
                timeout=timeout,
            )
            respuesta.raise_for_status()
            data = respuesta.json()
            hubo_respuesta = True
        except Exception as exc:
            errores.append(f"{type(exc).__name__}: {exc}")
            continue

        for item in _extraer_respuesta_rpc(data):
            item["consulta_rpc"] = consulta
            candidatos_acumulados.append(item)

    diagnostico_sugerencias = seleccionar_candidato_rpc_seguro(
        nombre_operador,
        candidatos_acumulados,
    )
    if diagnostico_sugerencias.get("ok"):
        metodo_base = diagnostico_sugerencias.get("metodo", "nombre_alta_confianza_rpc")
        diagnostico_sugerencias.update({
            "metodo": f"{metodo_base}_autocomplete",
            "fuente": "rpc_online_searchBP",
            "nombre_normalizado": objetivo_compact,
            "consulta_rpc": (diagnostico_sugerencias.get("candidatos") or [{}])[0].get(
                "consulta_rpc", ""
            ),
        })
        _CACHE_RPC_NOMBRE[objetivo_compact] = diagnostico_sugerencias
        return dict(diagnostico_sugerencias)
    if diagnostico_sugerencias.get("estado") == "ambiguo":
        diagnostico_sugerencias["nombre_normalizado"] = objetivo_compact
        _CACHE_RPC_NOMBRE[objetivo_compact] = diagnostico_sugerencias
        return dict(diagnostico_sugerencias)

    if hubo_respuesta:
        resultado = {
            "ok": False,
            "estado": diagnostico_sugerencias.get("estado", "sin_coincidencia"),
            "motivo": diagnostico_sugerencias.get("motivo", "nombre_no_encontrado_exacto_en_rpc_online"),
            "nombre_normalizado": objetivo_compact,
            "score": diagnostico_sugerencias.get("score", 0.0),
            "margen": diagnostico_sugerencias.get("margen", 0.0),
            "candidatos": diagnostico_sugerencias.get("candidatos", []),
        }
    else:
        resultado = {
            "ok": False,
            "estado": "error_rpc",
            "motivo": "rpc_online_no_disponible",
            "detalle_error": " | ".join(errores[-2:]),
            "nombre_normalizado": objetivo_compact,
            "candidatos": [],
        }
    _CACHE_RPC_NOMBRE[objetivo_compact] = resultado
    return dict(resultado)


def resolver_operador_seguro(
    id_solicitante: Any,
    nombre_operador: Any,
    catalogo: list[dict[str, Any]],
    *,
    permitir_rpc_online: bool = True,
    timeout_rpc: float = 10.0,
) -> dict[str, Any]:
    """Resuelve operador por evidencia exacta y deja trazabilidad del método.

    Prioridad:
      1. ID SATyS == ID OPERADOR del Excel.
      2. Si falta ID, nombre canónico único en el Excel.
      3. Nombre canónico único en el RPC en línea.
      4. Si el RPC no está disponible, nombre canónico único del Excel.

    Nunca basta con ser el "más parecido" ni se toma el primer resultado:
    las variantes deben cumplir umbral alto, cobertura, margen y unicidad de ID.
    """
    id_normalizado = normalizar_id(id_solicitante)
    nombre_original = str(nombre_operador or "").strip()
    nombre_normalizado = normalizar_nombre(nombre_original).replace(" ", "")

    # Algunos expedientes de Internos contienen varias razones sociales en la
    # misma celda. Resolver el texto completo como si fuera una sola empresa
    # produce una ambigüedad artificial. Se consulta cada razón por separado,
    # se conserva el orden de SATyS y nunca se inventa un ID para las que no
    # tengan una coincidencia verificable.
    razones_sociales = dividir_nombres_operador(nombre_original)
    if len(razones_sociales) > 1:
        indice_id = 0
        if id_normalizado:
            match_id_multiple = buscar_por_id_solicitante(id_normalizado, catalogo)
            if match_id_multiple:
                nombre_id = normalizar_nombre(
                    match_id_multiple.get("nombre_completo")
                    or match_id_multiple.get("concesionario")
                ).replace(" ", "")
                for indice, razon in enumerate(razones_sociales):
                    if normalizar_nombre(razon).replace(" ", "") == nombre_id:
                        indice_id = indice
                        break

        operadores: list[dict[str, Any]] = []
        candidatos: list[dict[str, Any]] = []
        consultas_rpc: list[str] = []
        for indice, razon in enumerate(razones_sociales):
            resolucion = resolver_operador_seguro(
                id_normalizado if id_normalizado and indice == indice_id else "",
                razon,
                catalogo,
                permitir_rpc_online=permitir_rpc_online,
                timeout_rpc=timeout_rpc,
            )
            id_operador = normalizar_id(
                resolucion.get("idBp") or resolucion.get("numero_rpc")
            )
            nombre_oficial = str(
                resolucion.get("nombre_completo") or razon
            ).strip()
            operadores.append({
                "idBp": id_operador,
                "nombre_completo": nombre_oficial,
                "nombre_original": razon,
                "ok": bool(resolucion.get("ok") and id_operador),
                "metodo": resolucion.get("metodo", ""),
                "fuente": resolucion.get("fuente", ""),
                "score": resolucion.get("score", 0.0) or 0.0,
                "motivo": resolucion.get("motivo", ""),
            })
            if resolucion.get("consulta_rpc"):
                consultas_rpc.append(str(resolucion["consulta_rpc"]))
            for candidato in resolucion.get("candidatos") or []:
                candidatos.append({"razon_social": razon, **candidato})

        operadores_resueltos = [item for item in operadores if item["ok"]]
        razones_sin_id = [
            item["nombre_original"] for item in operadores if not item["ok"]
        ]
        todas_resueltas = len(operadores_resueltos) == len(operadores)
        alguna_resuelta = bool(operadores_resueltos)
        scores = [float(item.get("score") or 0.0) for item in operadores]
        motivo = ""
        if razones_sin_id:
            motivo = "razones_sociales_sin_id: " + " | ".join(razones_sin_id)
        if not alguna_resuelta:
            motivo = "ninguna_razon_social_resuelta"

        return {
            "ok": alguna_resuelta,
            "score": min(scores) if scores else 0.0,
            "margen": "",
            "empate": False,
            "metodo": (
                "razones_sociales_multiples_todas_resueltas"
                if todas_resueltas
                else "razones_sociales_multiples_parcial"
            ),
            "fuente": "rpc_compuesto",
            "id_solicitante": id_normalizado,
            "idBp": operadores_resueltos[0]["idBp"] if alguna_resuelta else "",
            "numero_rpc": operadores_resueltos[0]["idBp"] if alguna_resuelta else "",
            "ids_operador": [item["idBp"] for item in operadores_resueltos],
            "nombre_completo": " | ".join(
                item["nombre_completo"] for item in operadores
            ),
            "nombre_operador_satys": nombre_original,
            "nombre_normalizado": nombre_normalizado,
            "operadores": operadores,
            "razones_sin_id": razones_sin_id,
            "consulta_rpc": " | ".join(consultas_rpc),
            "candidatos": candidatos,
            "motivo": motivo,
        }

    if id_normalizado:
        match_id = buscar_por_id_solicitante(id_normalizado, catalogo)
        if match_id:
            return {
                "ok": True,
                "score": 1.0,
                "empate": False,
                "metodo": "id_exacto_excel",
                "fuente": "excel_rpc",
                "id_solicitante": id_normalizado,
                "idBp": match_id["idBp"],
                "numero_rpc": match_id["idBp"],
                "nombre_completo": match_id["nombre_completo"],
                "nombre_operador_satys": nombre_original,
                "nombre_normalizado": nombre_normalizado,
            }

    diagnostico_excel = diagnosticar_nombre_operador_exacto(nombre_original, catalogo)
    if not id_normalizado and diagnostico_excel.get("ok"):
        return {
            "ok": True,
            "score": 1.0,
            "empate": False,
            "metodo": "nombre_exacto_excel",
            "fuente": "excel_rpc",
            "id_solicitante": "",
            "idBp": diagnostico_excel["idBp"],
            "numero_rpc": diagnostico_excel["idBp"],
            "nombre_completo": diagnostico_excel["nombre_completo"],
            "nombre_operador_satys": nombre_original,
            "nombre_normalizado": nombre_normalizado,
        }

    diagnostico_online: dict[str, Any] = {
        "ok": False,
        "estado": "consulta_online_desactivada",
        "motivo": "consulta_online_desactivada",
    }
    if permitir_rpc_online and nombre_original:
        diagnostico_online = buscar_nombre_operador_rpc_online_exacto(
            nombre_original,
            timeout=timeout_rpc,
        )
        if diagnostico_online.get("ok"):
            return {
                "ok": True,
                "score": diagnostico_online.get("score", 1.0),
                "empate": False,
                "metodo": diagnostico_online.get("metodo", "nombre_exacto_rpc_online"),
                "fuente": diagnostico_online.get("fuente", "rpc_online_searchBP"),
                "id_solicitante": id_normalizado,
                "idBp": diagnostico_online["idBp"],
                "numero_rpc": diagnostico_online["idBp"],
                "nombre_completo": diagnostico_online["nombre_completo"],
                "nombre_operador_satys": nombre_original,
                "nombre_normalizado": nombre_normalizado,
                "consulta_rpc": diagnostico_online.get("consulta_rpc", ""),
                "margen": diagnostico_online.get("margen", ""),
                "candidatos": diagnostico_online.get("candidatos", []),
            }

    # Un ID proporcionado pero ausente del Excel puede coexistir con una fila
    # vieja identificada por nombre. Sólo se usa esa fila como respaldo cuando
    # el RPC en línea falló técnicamente o fue desactivado; un "no encontrado"
    # actual se conserva para revisión manual.
    if (
        diagnostico_excel.get("ok")
        and diagnostico_online.get("estado") in {"error_rpc", "consulta_online_desactivada"}
    ):
        return {
            "ok": True,
            "score": 1.0,
            "empate": False,
            "metodo": "nombre_exacto_excel_respaldo",
            "fuente": "excel_rpc",
            "id_solicitante": id_normalizado,
            "idBp": diagnostico_excel["idBp"],
            "numero_rpc": diagnostico_excel["idBp"],
            "nombre_completo": diagnostico_excel["nombre_completo"],
            "nombre_operador_satys": nombre_original,
            "nombre_normalizado": nombre_normalizado,
            "advertencia": "rpc_online_no_disponible",
        }

    estados = {diagnostico_excel.get("estado"), diagnostico_online.get("estado")}
    if "ambiguo" in estados:
        motivo = "nombre_ambiguo_requiere_revision"
    elif not id_normalizado and not nombre_original:
        motivo = "metadata_sin_id_ni_nombre_operador"
    elif id_normalizado and not nombre_original:
        motivo = "id_solicitante_no_encontrado_y_sin_nombre"
    else:
        motivo = diagnostico_online.get("motivo") or "sin_coincidencia_exacta"

    candidatos = diagnostico_online.get("candidatos") or diagnostico_excel.get("candidatos") or []
    return {
        "ok": False,
        # Una coincidencia no aprobada sigue siendo 0% de confianza operativa,
        # pero se conserva su similitud diagnóstica para el CSV de revisión.
        "score": diagnostico_online.get("score", 0.0) or 0.0,
        "margen": diagnostico_online.get("margen", 0.0) or 0.0,
        "empate": motivo == "nombre_ambiguo_requiere_revision",
        "metodo": "resolucion_exacta_segura",
        "fuente": "rpc_online_y_excel",
        "id_solicitante": id_normalizado,
        "idBp": "",
        "numero_rpc": "",
        "nombre_completo": "",
        "nombre_operador_satys": nombre_original,
        "nombre_normalizado": nombre_normalizado,
        "motivo": motivo,
        "estado_excel": diagnostico_excel.get("estado", ""),
        "estado_rpc_online": diagnostico_online.get("estado", ""),
        "detalle_error": diagnostico_online.get("detalle_error", ""),
        "candidatos": candidatos,
    }


def buscar_coincidencias(nombre: str, catalogo: list[dict[str, Any]], top_n: int = 5):
    """
    Compatibilidad con código viejo. El flujo de producción usa el selector
    conservador con umbral, cobertura, margen y unicidad de ID; esta función
    sólo se conserva para no romper utilerías.
    """
    q = normalizar_nombre(nombre)
    if not q:
        return []
    resultados = []
    for item in catalogo or []:
        norm = item.get("norm") or normalizar_nombre(item.get("nombre_completo") or item.get("concesionario"))
        score = SequenceMatcher(None, q, norm).ratio()
        resultados.append((score, item))
    resultados.sort(key=lambda x: x[0], reverse=True)
    return resultados[:top_n]
