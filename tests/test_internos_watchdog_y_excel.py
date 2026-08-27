import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

import Parte1_descarga
from generar_excel_metadata_json import (
    generar_excel_metadata_json,
    validar_excel_metadata_json,
)


class _ProcesoColgado:
    siguiente_pid = 41000

    def __init__(self, *_args, **_kwargs):
        type(self).siguiente_pid += 1
        self.pid = type(self).siguiente_pid
        self.returncode = None

    def poll(self):
        return self.returncode

    def wait(self, timeout=None):
        return self.returncode


class InternosWatchdogYExcelTest(unittest.TestCase):
    def test_worker_hijo_publica_resultado_atomico_del_segmento(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            progreso = root / "progreso.json"
            payload = root / "payload.json"
            resultado = root / "resultado.json"
            payload.write_text(json.dumps({
                "bandeja": "Atendidos",
                "folios": ["190823"],
                "progreso_path": str(progreso),
            }), encoding="utf-8")
            filas = [{
                "folio": "190823",
                "folio_tabla_internos": "190823",
                "bandeja_internos": "Atendidos",
                "archivo": "190823.pdf",
                "ok": True,
            }]
            with patch.object(
                Parte1_descarga,
                "_worker_bandeja_internos",
                return_value=("Atendidos", filas),
            ) as worker, patch.object(
                Parte1_descarga,
                "objetivo_internos_esta_completo",
                return_value=True,
            ):
                rc = Parte1_descarga._ejecutar_worker_segmento_internos(
                    payload, resultado
                )

            publicado = json.loads(resultado.read_text(encoding="utf-8"))
            self.assertEqual(rc, 0)
            self.assertTrue(publicado["ok"])
            self.assertEqual(publicado["folios_pendientes"], [])
            self.assertEqual(publicado["resultados"], filas)
            worker.assert_called_once_with(
                "Atendidos", ["190823"], progreso_path=progreso
            )

    def test_timeout_mata_arbol_y_reintenta_el_segmento(self):
        procesos = []
        reloj = {"valor": 0.0}

        def popen(*args, **kwargs):
            proceso = _ProcesoColgado(*args, **kwargs)
            procesos.append(proceso)
            return proceso

        def monotonic():
            reloj["valor"] += 61.0
            return reloj["valor"]

        def matar(proceso, _identificador, motivo="timeout"):
            proceso.returncode = -9

        with tempfile.TemporaryDirectory() as tmp, \
             patch.object(Parte1_descarga, "log_dir", Path(tmp)), \
             patch.object(Parte1_descarga.subprocess, "Popen", side_effect=popen), \
             patch.object(Parte1_descarga.time, "monotonic", side_effect=monotonic), \
             patch.object(Parte1_descarga.time, "sleep"), \
             patch.object(Parte1_descarga, "_relay_worker_log"), \
             patch.object(Parte1_descarga, "_kill_process_tree", side_effect=matar) as kill, \
             patch.object(Parte1_descarga, "objetivo_internos_esta_completo", return_value=False):
            resultados = Parte1_descarga._procesar_tareas_internos_subprocesos(
                [("atendidos#1/1", "Atendidos", ["190823"])],
                workers_activos=1,
                timeout_registro=60,
                reintentos=1,
            )

        self.assertEqual(len(procesos), 2)
        self.assertEqual(kill.call_count, 2)
        self.assertEqual(resultados[0]["archivo"], "TIMEOUT_INTERNO")
        self.assertEqual(resultados[0]["folio_tabla_internos"], "190823")

    def test_reintento_conserva_completos_y_reencola_solo_incompletos(self):
        payloads = []
        completos = set()

        class ProcesoInstantaneo:
            siguiente_pid = 42000

            def __init__(self, cmd, **_kwargs):
                type(self).siguiente_pid += 1
                self.pid = type(self).siguiente_pid
                payload_path = Path(cmd[cmd.index("--_internos-worker-payload") + 1])
                resultado_path = Path(cmd[cmd.index("--_resultado-json") + 1])
                payload = json.loads(payload_path.read_text(encoding="utf-8"))
                folios = list(payload["folios"])
                payloads.append(folios)
                if len(payloads) == 1:
                    completos.add("190001")
                else:
                    completos.update(folios)
                resultado_path.write_text(json.dumps({
                    "resultados": [
                        {
                            "folio": folio,
                            "folio_tabla_internos": folio,
                            "bandeja_internos": "Atendidos",
                            "archivo": f"{folio}.pdf",
                            "ok": folio in completos,
                        }
                        for folio in folios
                    ]
                }), encoding="utf-8")
                self.returncode = 0 if all(folio in completos for folio in folios) else 2

            def poll(self):
                return self.returncode

        def esta_completo(_base, _bandeja, folio):
            return folio in completos

        with tempfile.TemporaryDirectory() as tmp, \
             patch.object(Parte1_descarga, "log_dir", Path(tmp)), \
             patch.object(Parte1_descarga.subprocess, "Popen", ProcesoInstantaneo), \
             patch.object(Parte1_descarga.time, "sleep"), \
             patch.object(Parte1_descarga, "_relay_worker_log"), \
             patch.object(Parte1_descarga, "objetivo_internos_esta_completo", side_effect=esta_completo):
            resultados = Parte1_descarga._procesar_tareas_internos_subprocesos(
                [("atendidos#1/1", "Atendidos", ["190001", "190002"])],
                workers_activos=1,
                timeout_registro=60,
                reintentos=1,
            )

        self.assertEqual(payloads, [["190001", "190002"], ["190002"]])
        self.assertEqual(
            {item["folio_tabla_internos"] for item in resultados},
            {"190001", "190002"},
        )
        self.assertTrue(all(item["ok"] for item in resultados))

    def test_excel_internos_incluye_faltantes_y_se_valida_antes_de_publicar(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            descargas = root / "descargas"
            output = root / "output"
            carpeta = descargas / "internos" / "atendidos" / "190001"
            carpeta.mkdir(parents=True)
            meta = {
                "folio": "190001",
                "folio_tabla_internos": "190001",
                "registro": "190001",
                "bandeja_internos": "Atendidos",
                "nombre_operador": "OPERADOR DE PRUEBA",
            }
            (carpeta / "metadata_satys.json").write_text(
                json.dumps(meta), encoding="utf-8"
            )
            (carpeta / "metadata_tramite_nuevo.json").write_text(
                json.dumps(meta), encoding="utf-8"
            )
            (carpeta / "metadata_completo.json").write_text(json.dumps({
                "estado": "OK",
                "coincide": True,
                "total_archivos_encontrados": 1,
                "total_archivos_ok": 1,
                "total_archivos_error": 0,
                "metadatos_satys": meta,
                "metadatos_tramite": meta,
                "archivos": [{"archivo": "190001.pdf", "ok": True}],
            }), encoding="utf-8")
            (carpeta / "190001.pdf").write_bytes(b"PDF")

            excel = output / "Folios_Datos_Completos_Internos.xlsx"
            objetivos = [
                {"bandeja": "Atendidos", "folio": "190001"},
                {"bandeja": "En proceso", "folio": "190002"},
            ]
            generado = generar_excel_metadata_json(
                resultados=[{
                    "folio": "190001",
                    "folio_id": "internos__atendidos__190001",
                    "descargas_dir": str(carpeta),
                    "rpc_ok": True,
                    "rpc_resultado": {
                        "ok": True,
                        "metodo": "id_exacto",
                        "idBp": "100028",
                    },
                }],
                descargas_base=descargas,
                output_base=output,
                excel_salida=excel,
                project_root=root,
                objetivos_esperados=objetivos,
            )

            validacion = validar_excel_metadata_json(
                generado,
                objetivos_esperados=[("Atendidos", "190001"), ("En proceso", "190002")],
            )
            wb = openpyxl.load_workbook(generado, read_only=True, data_only=True)
            try:
                ws = wb["Datos_Completos"]
                headers = [cell.value for cell in ws[1]]
                idx = {header: pos for pos, header in enumerate(headers)}
                rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
                estados = {
                    row[idx["folio_tabla_internos"]]: row[idx["estado_descarga"]]
                    for row in rows
                }
                self.assertEqual(estados, {"190001": "OK", "190002": "FALTANTE"})
                self.assertEqual(wb["Resumen"]["B2"].value, 2)
            finally:
                wb.close()

        self.assertEqual(validacion["total_filas"], 2)
        self.assertEqual(validacion["objetivos_cubiertos"], 2)

    def test_excel_con_lista_vacia_publica_todos_los_objetivos_como_faltantes(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            excel = root / "output" / "Folios_Datos_Completos_Internos.xlsx"
            objetivos = [
                {"bandeja": "Atendidos", "folio": "190001"},
                {"bandeja": "En proceso", "folio": "190002"},
            ]
            generar_excel_metadata_json(
                resultados=[],
                descargas_base=root / "descargas",
                output_base=root / "output",
                excel_salida=excel,
                project_root=root,
                objetivos_esperados=objetivos,
            )
            wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
            try:
                ws = wb["Datos_Completos"]
                headers = [cell.value for cell in ws[1]]
                idx_estado = headers.index("estado_descarga")
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                self.assertEqual(len(rows), 2)
                self.assertTrue(all(row[idx_estado] == "FALTANTE" for row in rows))
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
