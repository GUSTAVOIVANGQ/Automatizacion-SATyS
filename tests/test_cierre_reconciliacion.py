from __future__ import annotations

import json
import sys
import tempfile
import time
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from openpyxl.styles import Font

import automatizar_registros_diario as diario
import reconciliar_metadata_global as global_recon
from reconciliar_tramites_desde_folios import reconciliar


class CierreReconciliacionTests(unittest.TestCase):
    def test_reconciliacion_ignora_dimension_fantasma_del_excel(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            maestro = root / "TrámitesCRT.xlsx"
            consolidado = root / "Folios_Datos_Completos.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Turnados recibidos"
            headers = [
                "1711", "Memo/Volante", "Solicitante Promovente",
                "Representante Legal", "Asunto", "Tipo Trámite",
                "Fecha de creación", "FECHA LÍMITE", "Ruta", "NOTAS_VICTOR",
            ] + [f"R{i:03d}" for i in range(1, 28)]
            ws.append(headers)
            ws.append(["CRT26-000001", "1", "OPERADOR", "", "", "", "", "", "vieja"])
            # Simula formato residual de Excel/SharePoint en la última fila.
            ws.cell(1_048_576, 1).font = Font(bold=True)
            wb.save(maestro)
            wb.close()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Datos_Completos"
            ws.append([
                "registro", "folio", "nombre_operador", "output",
                "metadata_satys.asunto", "metadata_satys.tipo_tramite",
                "metadata_satys.fecha_registro",
                "metadata_tramite_nuevo.plazo_atencion",
            ])
            ws.append([
                "CRT26-000001", "1", "OPERADOR",
                r"output\100_operador\01 EN\VE",
                "FORMATO R001", "TIPO", "01/01/2026", "02/01/2026",
            ])
            wb.save(consolidado)
            wb.close()

            inicio = time.monotonic()
            resultado = reconciliar(maestro, consolidado, crear_backup=False)
            duracion = time.monotonic() - inicio

            self.assertLess(duracion, 5)
            self.assertEqual(resultado["target_rows_scanned"], 1)
            self.assertEqual(resultado["routes_blank"], 0)

    def test_global_sin_reorganizar_no_recopia_historico(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            descargas = root / "descargas"
            output = root / "output"
            carpeta = descargas / "CRT26-000001"
            carpeta.mkdir(parents=True)
            output.mkdir()
            (carpeta / "metadata_satys.json").write_text(
                json.dumps({
                    "registro": "CRT26-000001",
                    "folio": "1",
                    "id_solicitante": "123",
                    "nombre_operador": "OPERADOR DEMO",
                }),
                encoding="utf-8",
            )
            (carpeta / "documento.pdf").write_bytes(b"contenido")
            resolucion = {
                "ok": True,
                "idBp": "123",
                "numero_rpc": "123",
                "nombre_completo": "OPERADOR DEMO",
                "metodo": "id_exacto_excel",
            }
            with patch.object(
                global_recon.bc,
                "resolver_operador_seguro",
                return_value=resolucion,
            ):
                resultados, _ = global_recon.construir_resultados(
                    descargas,
                    output,
                    {"123": {"idBp": "123", "nombre_completo": "OPERADOR DEMO"}},
                    reorganizar_output=False,
                )

            self.assertEqual(len(resultados), 1)
            self.assertEqual(resultados[0]["archivos_copiados"], 0)
            self.assertFalse(any(output.rglob("documento.pdf")))
            self.assertTrue(resultados[0]["output_dir"].endswith("01 EN/VE"))

    def test_subproceso_silencioso_tiene_timeout_y_monitor_continua(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            log_path = root / "timeout.log"
            inicio = time.monotonic()
            rc = diario.ejecutar_comando(
                [sys.executable, "-c", "import time; time.sleep(10)"],
                root,
                log_path,
                "TEST TIMEOUT",
                timeout_segundos=1,
            )
            duracion = time.monotonic() - inicio
            self.assertEqual(rc, 124)
            self.assertLess(duracion, 5)
            self.assertIn("[TIMEOUT]", log_path.read_text(encoding="utf-8"))


if __name__ == "__main__":
    unittest.main()
