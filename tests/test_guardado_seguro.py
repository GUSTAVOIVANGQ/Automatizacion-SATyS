from __future__ import annotations

import errno
import tempfile
import unittest
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from unittest.mock import patch

import openpyxl

import guardado_seguro
from Parte4_excel import actualizar_excel


class GuardadoSeguroTests(unittest.TestCase):
    def test_ebusy_de_bind_mount_conserva_inode_y_copia_contenido(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            destino = root / "destino.xlsx"
            temporal = root / ".destino.xlsx.tmp"
            destino.write_bytes(b"anterior")
            temporal.write_bytes(b"nuevo-contenido")

            error = OSError(errno.EBUSY, "Device or resource busy")
            with patch.object(guardado_seguro.os, "replace", side_effect=error):
                fallback = guardado_seguro.reemplazar_desde_temporal(temporal, destino)

            self.assertTrue(fallback)
            self.assertEqual(destino.read_bytes(), b"nuevo-contenido")
            self.assertFalse(temporal.exists())

    def test_error_distinto_de_ebusy_no_se_oculta(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            destino = root / "destino.xlsx"
            temporal = root / ".destino.xlsx.tmp"
            destino.write_bytes(b"anterior")
            temporal.write_bytes(b"nuevo")

            error = OSError(errno.EACCES, "Permission denied")
            with patch.object(guardado_seguro.os, "replace", side_effect=error):
                with self.assertRaises(OSError):
                    guardado_seguro.reemplazar_desde_temporal(temporal, destino)

            self.assertEqual(destino.read_bytes(), b"anterior")
            self.assertTrue(temporal.exists())

    def test_workers_serializan_actualizaciones_del_mismo_excel(self):
        with tempfile.TemporaryDirectory() as td:
            excel = Path(td) / "TramitesCRT.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Internos"
            for col, header in enumerate([
                "Control", "Estado", "Origen", "1711", "Memo/Volante",
                "Solicitante Promovente", "Representante Legal",
            ], 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(excel)
            wb.close()

            folios = [str(190000 + indice) for indice in range(8)]
            with ThreadPoolExecutor(max_workers=8) as pool:
                resultados = list(pool.map(
                    lambda folio: actualizar_excel(
                        folio=folio,
                        folio_internos=folio,
                        registro=f"CRT26-{folio}",
                        excel_path=excel,
                        sheet_name="Internos",
                    ),
                    folios,
                ))

            self.assertTrue(all(resultados))
            wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
            ws = wb["Internos"]
            headers = {str(cell.value): cell.column for cell in ws[1] if cell.value}
            guardados = {
                str(ws.cell(row=row, column=headers["Folio Internos"]).value)
                for row in range(2, ws.max_row + 1)
            }
            wb.close()
            self.assertEqual(guardados, set(folios))


if __name__ == "__main__":
    unittest.main(verbosity=2)
