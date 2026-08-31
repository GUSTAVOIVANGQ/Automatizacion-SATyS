from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

import completar_remitentes_desde_pdfs as mod
import reconciliar_tramites_desde_folios as recon
import resolver_sin_operador_rpc_publico as sinop


class CompletarRemitentesPdfTests(unittest.TestCase):
    def _excel(self, path: Path, rows, *, sheet="Turnados recibidos"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet
        ws.append(["1711", "Solicitante Promovente", "Representante   Legal", "Ruta"])
        for row in rows:
            ws.append(row)
        if sheet != "Internos":
            wi = wb.create_sheet("Internos")
            wi.append(["1711", "Solicitante Promovente", "Representante Legal", "Ruta"])
        wb.save(path)
        wb.close()

    def _resultado(self, *, sol="", rep="", general="parcial"):
        return {
            "estado_general": general,
            "solicitante": {
                "estado": "encontrado" if sol else "no_encontrado",
                "valor": sol,
                "pagina": 1 if sol else "",
                "metodo": "texto_layout" if sol else "",
                "nota": "",
            },
            "representante": {
                "estado": "encontrado" if rep else "no_encontrado",
                "valor": rep,
                "pagina": 1 if rep else "",
                "metodo": "tabla" if rep else "",
                "nota": "",
            },
        }

    @patch("completar_remitentes_desde_pdfs.extractor.extraer_de_pdf")
    def test_busca_todos_los_pdf_y_combina_campos_en_archivos_distintos(self, extraer):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "descargas" / "CRT26-022937"
            (source / "sub").mkdir(parents=True)
            (source / "metadata_satys.json").write_text(
                json.dumps({"registro": "CRT26-022937", "nombre_operador": "X"}),
                encoding="utf-8",
            )
            (source / "uno.pdf").write_bytes(b"1")
            (source / "sub" / "dos.PDF").write_bytes(b"2")
            (source / "tres.pdf").write_bytes(b"3")
            excel = root / "TrámitesCRT.xlsx"
            self._excel(excel, [["CRT26-022937", "SIN REMITENTE", "", "_sin_operador\\CRT26-022937"]])

            def fake(path):
                if path.name.lower() == "uno.pdf":
                    return self._resultado(sol="TECNOLOGÍA Y REDES DE DATOS, S.A. DE C.V.")
                if path.name.lower() == "dos.pdf":
                    return self._resultado(rep="JUAN PÉREZ LÓPEZ")
                return self._resultado(general="requiere_ocr")

            extraer.side_effect = fake
            payload = mod.completar(
                excel_path=excel,
                descargas_base=root / "descargas",
                logs_dir=root / "logs",
            )
            self.assertEqual(payload["cambios_solicitante"], 1)
            self.assertEqual(payload["cambios_representante"], 1)
            self.assertEqual(payload["resultados"][0]["pdfs_revisados"], 3)
            wb = openpyxl.load_workbook(excel)
            ws = wb["Turnados recibidos"]
            self.assertEqual(ws.cell(2, 2).value, "TECNOLOGÍA Y REDES DE DATOS, S.A. DE C.V.")
            self.assertEqual(ws.cell(2, 3).value, "JUAN PÉREZ LÓPEZ")
            wb.close()

    @patch("completar_remitentes_desde_pdfs.extractor.extraer_de_pdf")
    def test_no_sobrescribe_valor_valido_y_si_completa_el_otro(self, extraer):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "descargas" / "CRT26-000001"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({"registro": "CRT26-000001"}), encoding="utf-8")
            (source / "a.pdf").write_bytes(b"a")
            excel = root / "TrámitesCRT.xlsx"
            self._excel(excel, [["CRT26-000001", "VALOR YA CORRECTO", "SIN REMITENTE", "x"]])
            extraer.return_value = self._resultado(sol="OTRA EMPRESA", rep="REP CORRECTO")
            payload = mod.completar(excel_path=excel, descargas_base=root / "descargas", logs_dir=root / "logs")
            self.assertEqual(payload["cambios_solicitante"], 0)
            self.assertEqual(payload["cambios_representante"], 1)
            wb = openpyxl.load_workbook(excel)
            ws = wb["Turnados recibidos"]
            self.assertEqual(ws.cell(2, 2).value, "VALOR YA CORRECTO")
            self.assertEqual(ws.cell(2, 3).value, "REP CORRECTO")
            wb.close()

    @patch("completar_remitentes_desde_pdfs.extractor.extraer_de_pdf")
    def test_conflicto_entre_pdfs_no_inventa_valor(self, extraer):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "descargas" / "CRT26-000002"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({"registro": "CRT26-000002"}), encoding="utf-8")
            (source / "a.pdf").write_bytes(b"a")
            (source / "b.pdf").write_bytes(b"b")
            excel = root / "TrámitesCRT.xlsx"
            self._excel(excel, [["CRT26-000002", "", "REP EXISTENTE", "x"]])
            extraer.side_effect = [self._resultado(sol="EMPRESA UNO"), self._resultado(sol="EMPRESA DOS")]
            payload = mod.completar(excel_path=excel, descargas_base=root / "descargas", logs_dir=root / "logs")
            self.assertEqual(payload["cambios_totales"], 0)
            self.assertEqual(payload["filas_conflicto"], 1)
            wb = openpyxl.load_workbook(excel)
            self.assertIsNone(wb["Turnados recibidos"].cell(2, 2).value)
            wb.close()

    @patch("completar_remitentes_desde_pdfs.extractor.extraer_de_pdf")
    def test_internos_concilia_folio_numerico(self, extraer):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "descargas" / "internos" / "atendidos" / "103369"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(
                json.dumps({"folio": "103369", "satys_flujo": "internos"}), encoding="utf-8"
            )
            (source / "formato.pdf").write_bytes(b"pdf")
            excel = root / "TrámitesCRT.xlsx"
            self._excel(excel, [["103369", "", "", "_sin_operador\\internos__Atendidos__103369"]], sheet="Internos")
            extraer.return_value = self._resultado(sol="OPERADOR INTERNO", rep="REP INTERNO")
            payload = mod.completar(excel_path=excel, descargas_base=root / "descargas", logs_dir=root / "logs")
            self.assertEqual(payload["cambios_totales"], 2)
            wb = openpyxl.load_workbook(excel)
            ws = wb["Internos"]
            self.assertEqual(ws.cell(2, 2).value, "OPERADOR INTERNO")
            self.assertEqual(ws.cell(2, 3).value, "REP INTERNO")
            wb.close()

    def test_reconciliacion_global_preserva_valores_pdf_validos(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            maestro = root / "TrámitesCRT.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Turnados recibidos"
            ws.append([
                "1711", "Memo/Volante", "Solicitante Promovente", "Representante Legal",
                "Asunto", "Tipo Trámite", "Fecha de creación", "FECHA LÍMITE", "Ruta",
            ])
            ws.append(["CRT26-000003", "1", "EMPRESA DESDE PDF", "REP DESDE PDF", "", "", "", "", "_sin_operador\\CRT26-000003"])
            wb.save(maestro)
            wb.close()

            fuente = root / "Folios_Datos_Completos.xlsx"
            wf = openpyxl.Workbook()
            sf = wf.active
            sf.title = "Datos_Completos"
            sf.append(["registro", "folio", "metadata_satys.solicitante", "metadata_satys.representante_legal", "output"])
            sf.append(["CRT26-000003", "1", "SIN REMITENTE", "", "output/_sin_operador/CRT26-000003"])
            wf.save(fuente)
            wf.close()

            recon.reconciliar(maestro, fuente, crear_backup=False)
            wb = openpyxl.load_workbook(maestro)
            ws = wb["Turnados recibidos"]
            self.assertEqual(ws.cell(2, 3).value, "EMPRESA DESDE PDF")
            self.assertEqual(ws.cell(2, 4).value, "REP DESDE PDF")
            wb.close()


class MemorandoNombreTests(unittest.TestCase):
    def test_variantes_memorando_memorandum(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            buenos = [
                "MEMORANDO.pdf",
                "memorandum.pdf",
                "MEMORÁNDUM.pdf",
                "Memorando firmado.pdf",
                "memorandun.pdf",
                "MEMORANDUM_2026.pdf",
            ]
            malos = ["memo.pdf", "memoria.pdf", "oficio.pdf"]
            for name in buenos + malos:
                (root / name).write_bytes(b"x")
            for name in buenos:
                self.assertTrue(sinop._es_memorando_memorandum_pdf(root / name), name)
            for name in malos:
                self.assertFalse(sinop._es_memorando_memorandum_pdf(root / name), name)


if __name__ == "__main__":
    unittest.main()
