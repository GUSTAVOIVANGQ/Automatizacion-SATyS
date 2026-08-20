#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import sys
import tempfile
import threading
import time
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import automatizar_registros_diario as monitor
import extraer_registros_documentos as extractor
import main_procesar
import Parte1_descarga
import Parte4_excel
import estado_descargas


class InternosDiarioTest(unittest.TestCase):
    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def test_objetivos_conservan_cada_pareja_bandeja_folio_entre_etapas(self):
        resumen = {
            "por_bandeja": [
                {"bandeja": "En proceso", "folios": ["148326", "147390"]},
                {"bandeja": "Atendidos", "folios": ["148326", "190823"]},
            ]
        }
        objetivos = monitor.construir_objetivos_internos(
            resumen,
            ["148326", "190823"],
        )
        self.assertEqual(objetivos, [
            {"bandeja": "En proceso", "folio": "148326"},
            {"bandeja": "Atendidos", "folio": "148326"},
            {"bandeja": "Atendidos", "folio": "190823"},
        ])

        path = self.tmp / "objetivos.json"
        monitor.guardar_objetivos_internos(path, objetivos)
        self.assertEqual(main_procesar.cargar_objetivos_internos(path), objetivos)
        self.assertEqual(Parte1_descarga._cargar_objetivos_internos(path), objetivos)

    def test_auditoria_internos_exige_metadata_y_todos_los_archivos(self):
        carpeta = self.tmp / "descargas" / "internos" / "atendidos" / "190823"
        carpeta.mkdir(parents=True)
        (carpeta / "documento_1.pdf").write_bytes(b"pdf-1")
        (carpeta / "documento_2.csv").write_bytes(b"csv-2")
        metadata = {
            "folio": "CRT26-000001",
            "estado": "OK",
            "coincide": True,
            "total_archivos_encontrados": 2,
            "total_archivos_ok": 2,
            "total_archivos_error": 0,
            "metadatos_satys": {
                "bandeja_internos": "Atendidos",
                "folio_tabla_internos": "190823",
            },
            "archivos": [
                {"archivo": "documento_1.pdf", "ok": True},
                {"archivo": "documento_2.csv", "ok": True},
            ],
        }
        (carpeta / "metadata_completo.json").write_text(
            json.dumps(metadata),
            encoding="utf-8",
        )

        self.assertTrue(estado_descargas.objetivo_internos_esta_completo(
            self.tmp / "descargas", "Atendidos", "190823"
        ))
        (carpeta / "documento_2.csv").unlink()
        self.assertFalse(estado_descargas.objetivo_internos_esta_completo(
            self.tmp / "descargas", "Atendidos", "190823"
        ))

    def test_auditoria_internos_reintenta_zip_residual(self):
        carpeta = self.tmp / "descargas" / "internos" / "fuera_de_tiempo" / "121195"
        carpeta.mkdir(parents=True)
        (carpeta / "pendiente.zip").write_bytes(b"zip-incompleto")
        metadata = {
            "folio": "121195",
            "estado": "OK",
            "coincide": True,
            "total_archivos_encontrados": 1,
            "total_archivos_ok": 1,
            "total_archivos_error": 0,
            "metadatos_satys": {
                "bandeja_internos": "Fuera de tiempo",
                "folio_tabla_internos": "121195",
            },
            "archivos": [{"archivo": "pendiente.zip", "ok": True}],
        }
        (carpeta / "metadata_completo.json").write_text(
            json.dumps(metadata),
            encoding="utf-8",
        )

        self.assertFalse(estado_descargas.objetivo_internos_esta_completo(
            self.tmp / "descargas", "Fuera de tiempo", "121195"
        ))

    def test_metadata_distingue_documentos_portal_de_archivos_extraidos_del_zip(self):
        carpeta = self.tmp / "zip_expandido"
        resultados = [
            {
                "archivo": f"miembro_{indice}.csv",
                "ok": True,
                "indice_documento_portal": 1,
                "total_documentos_portal": 1,
            }
            for indice in range(3)
        ]

        metadata = Parte1_descarga.guardar_metadata_completo(
            "190823", "190823", carpeta,
            {"bandeja_internos": "Atendidos", "folio_tabla_internos": "190823"},
            {}, resultados, "INTERNOS_DOCUMENTOS_ANEXOS",
        )

        self.assertEqual(metadata["estado"], "OK")
        self.assertEqual(metadata["total_archivos_ok"], 3)
        self.assertEqual(metadata["total_documentos_portal"], 1)
        self.assertEqual(metadata["total_documentos_portal_ok"], 1)
        self.assertTrue(metadata["documentos_portal_completos"])

    def test_metadata_marca_parcial_si_no_recorrio_todos_los_documentos_portal(self):
        carpeta = self.tmp / "conteo_incompleto"
        resultados = [{
            "archivo": "documento_1.pdf",
            "ok": True,
            "indice_documento_portal": 1,
            "total_documentos_portal": 2,
        }]

        metadata = Parte1_descarga.guardar_metadata_completo(
            "190823", "190823", carpeta,
            {"bandeja_internos": "Atendidos", "folio_tabla_internos": "190823"},
            {}, resultados, "INTERNOS_DOCUMENTOS_ANEXOS",
        )

        self.assertEqual(metadata["estado"], "PARCIAL")
        self.assertFalse(metadata["coincide"])
        self.assertFalse(metadata["documentos_portal_completos"])

    def test_clasificacion_internos_no_usa_excel_y_mantiene_bandejas(self):
        resumen = {
            "por_bandeja": [
                {"bandeja": "En proceso", "folios": ["148326"]},
                {"bandeja": "Atendidos", "folios": ["148326"]},
                {"bandeja": "Fuera de tiempo", "folios": ["148326"]},
            ]
        }
        with patch.object(
            monitor,
            "objetivo_internos_esta_completo",
            side_effect=lambda _base, bandeja, _folio: bandeja == "En proceso",
        ):
            inventario, completos, pendientes = monitor.clasificar_objetivos_internos(
                resumen,
                self.tmp / "descargas",
            )

        self.assertEqual(len(inventario), 3)
        self.assertEqual(completos, [{"bandeja": "En proceso", "folio": "148326"}])
        self.assertEqual(pendientes, [
            {"bandeja": "Atendidos", "folio": "148326"},
            {"bandeja": "Fuera de tiempo", "folio": "148326"},
        ])

    def test_extractores_reconocen_modo_solo_internos(self):
        with patch.object(sys, "argv", ["extraer_registros_documentos.py", "--solo-internos"]):
            args_extractor = extractor.parse_args()
        with patch.object(sys, "argv", [
            "automatizar_registros_diario.py", "--solo-internos",
            "--max-folios-internos", "4", "--internos-workers", "3",
        ]):
            args_monitor = monitor.construir_parser().parse_args()

        self.assertTrue(args_extractor.solo_internos)
        self.assertFalse(args_extractor.sin_internos)
        self.assertTrue(args_monitor.solo_internos)
        self.assertEqual(args_monitor.max_folios_internos, 4)
        self.assertEqual(args_monitor.internos_workers, 3)

    def test_monitor_acepta_workers_internos_sin_tope_artificial(self):
        with patch.object(sys, "argv", [
            "automatizar_registros_diario.py", "--solo-internos",
            "--internos-workers", "48",
        ]):
            args_monitor = monitor.construir_parser().parse_args()

        self.assertEqual(args_monitor.internos_workers, 48)

    def test_limita_lote_internos_sin_alterar_orden(self):
        objetivos = [
            {"bandeja": "En proceso", "folio": str(folio)}
            for folio in range(10, 15)
        ]

        self.assertEqual(monitor.limitar_objetivos_internos(objetivos, 0), objetivos)
        self.assertEqual(monitor.limitar_objetivos_internos(objetivos, 2), objetivos[:2])
        self.assertEqual(objetivos[-1]["folio"], "14")
        with self.assertRaises(ValueError):
            monitor.limitar_objetivos_internos(objetivos, -1)

    def test_lote_internos_se_reparte_entre_bandejas(self):
        objetivos = [
            {"bandeja": "En proceso", "folio": "10"},
            {"bandeja": "En proceso", "folio": "11"},
            {"bandeja": "Atendidos", "folio": "20"},
            {"bandeja": "Atendidos", "folio": "21"},
            {"bandeja": "Fuera de tiempo", "folio": "30"},
        ]

        muestra = monitor.limitar_objetivos_internos(objetivos, 4)

        self.assertEqual(
            [(item["bandeja"], item["folio"]) for item in muestra],
            [
                ("En proceso", "10"),
                ("Atendidos", "20"),
                ("Fuera de tiempo", "30"),
                ("En proceso", "11"),
            ],
        )

    def test_certificado_solo_internos_no_exige_oficialia(self):
        output = self.tmp / "registros.txt"
        output.write_text("", encoding="utf-8")
        bandejas = []
        for nombre in Parte1_descarga.BANDEJAS_INTERNOS_DEFAULT:
            folios = ["148326"] if nombre == "En proceso" else []
            bandejas.append({
                "bandeja": nombre,
                "estado": "ENCONTRADOS_COMPLETOS" if folios else "VACIO_CONFIRMADO",
                "total_reportado_satys": len(folios),
                "filas_leidas": len(folios),
                "folios": folios,
                "filas_invalidas": 0,
            })

        resumen = extractor.resumen_oficialia_omitida()
        resumen["internos"] = {
            "estado": "COMPLETO",
            "integridad": "VALIDADA",
            "vacio_confirmado": False,
            "total_folios": 1,
            "total_filas_satys": 1,
            "folios": ["148326"],
            "por_bandeja": bandejas,
        }
        output.with_suffix(".txt.json").write_text(json.dumps(resumen), encoding="utf-8")

        validacion = monitor.validar_resumen_extractor(output, [])

        self.assertTrue(validacion["ok"], validacion)
        self.assertTrue(validacion["vacio_confirmado"])
        self.assertEqual(validacion["folios_internos"], ["148326"])

    def test_lanzadores_oci_exponen_comando_internos_filtrado(self):
        root = Path(__file__).resolve().parents[1]
        podman = (root / "scripts" / "podman_satys.sh").read_text(encoding="utf-8")
        docker = (root / "scripts" / "docker_satys.sh").read_text(encoding="utf-8")

        self.assertIn("internos)", podman)
        self.assertIn("automatizar_registros_diario.py --solo-internos --headless", podman)
        self.assertIn("internos-check)", podman)
        self.assertIn("--solo-internos --no-procesar --sin-email --headless", podman)
        self.assertIn("internos)", docker)
        self.assertIn("automatizar_registros_diario.py --solo-internos --headless", docker)
        self.assertIn("internos-check)", docker)
        self.assertIn("--solo-internos --no-procesar --sin-email --headless", docker)

        powershell = (root / "scripts" / "run_satys_internos_nuevos.ps1").read_text(encoding="utf-8")
        self.assertIn("MaxFoliosInternos", powershell)
        self.assertIn("InternosWorkers", powershell)
        self.assertIn("[int]$InternosWorkers = 12", powershell)
        self.assertNotIn("ValidateRange(1, 6)", powershell)

    def test_extractor_solo_internos_no_navega_a_oficialia(self):
        output = self.tmp / "solo_internos.txt"

        class Page:
            def set_default_timeout(self, _timeout):
                return None

        class Context:
            def new_page(self):
                return Page()

        class Browser:
            def new_context(self, **_kwargs):
                return Context()

            def close(self):
                return None

        class Chromium:
            def launch(self, **_kwargs):
                return Browser()

        class Playwright:
            chromium = Chromium()

        class Manager:
            def __enter__(self):
                return Playwright()

            def __exit__(self, *_args):
                return False

        internos = {
            "estado": "COMPLETO",
            "integridad": "VALIDADA",
            "vacio_confirmado": True,
            "total_folios": 0,
            "total_filas_satys": 0,
            "folios": [],
            "por_bandeja": [],
        }
        with patch.object(
            sys,
            "argv",
            ["extraer_registros_documentos.py", "--solo-internos", "--output", str(output)],
        ), patch.object(extractor, "sync_playwright", return_value=Manager()), \
             patch.object(extractor, "cargar_credenciales_satys", return_value=("usuario", "password")), \
             patch.object(extractor, "sesion_activa", return_value=True), \
             patch.object(extractor, "extraer_folios_internos", return_value=internos) as extraer_int, \
             patch.object(extractor, "navegar_a_enlace_oficialia") as navegar_oficialia:
            rc = extractor.main()

        self.assertEqual(rc, 0)
        extraer_int.assert_called_once()
        navegar_oficialia.assert_not_called()
        resumen = json.loads(output.with_suffix(".txt.json").read_text(encoding="utf-8"))
        self.assertTrue(resumen["oficialia_omitida"])

    def test_navegacion_internos_no_usa_wait_for_function(self):
        class Page:
            def __init__(self):
                self.evaluaciones = 0

            def evaluate(self, _script, *args):
                self.evaluaciones += 1
                return True

            def wait_for_timeout(self, _timeout):
                return None

            def wait_for_function(self, *_args, **_kwargs):
                raise AssertionError("Internos no debe depender de wait_for_function")

        page = Page()
        with patch.object(extractor, "esperar_sin_spinner", return_value=True), \
             patch.object(extractor, "esperar_datatables"):
            ok = extractor.navegar_a_internos_ift(page)

        self.assertTrue(ok)
        self.assertEqual(page.evaluaciones, 3)

    def test_seleccion_bandeja_extractor_no_usa_wait_for_function(self):
        class Page:
            def __init__(self):
                self.evaluaciones = 0
                self.argumentos = []

            def evaluate(self, _script, *args):
                self.evaluaciones += 1
                self.argumentos.extend(args)
                if self.evaluaciones == 1:
                    return "CLICK"
                return True

            def wait_for_timeout(self, _timeout):
                return None

            def wait_for_function(self, *_args, **_kwargs):
                raise AssertionError("Internos no debe depender de wait_for_function")

        page = Page()
        with patch.object(extractor, "esperar_sin_spinner", return_value=True), \
             patch.object(extractor, "esperar_datatables"):
            extractor.seleccionar_bandeja_internos(page, "En proceso")

        self.assertEqual(page.evaluaciones, 3)
        self.assertEqual(page.argumentos, ["2", "2", "2"])

    def test_seleccion_bandeja_verifica_estado_tras_respuesta_vacia(self):
        class Page:
            def __init__(self):
                self.evaluaciones = 0

            def evaluate(self, _script, *_args):
                self.evaluaciones += 1
                return None if self.evaluaciones == 1 else True

            def wait_for_timeout(self, _timeout):
                return None

        page = Page()
        with patch.object(extractor, "esperar_sin_spinner", return_value=True), \
             patch.object(extractor, "esperar_datatables"):
            extractor.seleccionar_bandeja_internos(page, "Recibidos")

        self.assertEqual(page.evaluaciones, 3)

    def test_todos_internos_activa_solo_el_recorrido_completo(self):
        args = SimpleNamespace(
            todos_internos=True,
            internos=False,
            internos_bandejas=None,
            internos_objetivos="",
            solo_procesar=False,
            folios=[],
            archivo_folios="",
            archivo_registro="",
            buscar=0,
        )

        main_procesar.aplicar_modo_todos_internos(args)

        self.assertTrue(args.internos)
        self.assertIsNone(args.internos_bandejas)
        self.assertEqual(args.internos_objetivos, "")

    def test_todos_internos_rechaza_objetivos_parciales(self):
        args = SimpleNamespace(
            todos_internos=True,
            internos=False,
            internos_bandejas=None,
            internos_objetivos="pendientes.json",
            solo_procesar=False,
            folios=[],
            archivo_folios="",
            archivo_registro="",
            buscar=0,
        )

        with self.assertRaisesRegex(ValueError, "--internos-objetivos"):
            main_procesar.aplicar_modo_todos_internos(args)

    def test_selecciona_bandeja_internos_por_id_estable(self):
        class Page:
            def __init__(self):
                self.tab_id = None

            def evaluate(self, _script, tab_id):
                self.tab_id = tab_id
                return "CLICK"

        page = Page()
        with patch.object(
            Parte1_descarga,
            "_esperar_estado_bandeja_internos",
            return_value=True,
        ) as esperar:
            ok = Parte1_descarga.seleccionar_bandeja_internos(
                page,
                "Últimos Movimientos",
            )

        self.assertTrue(ok)
        self.assertEqual(page.tab_id, "5")
        esperar.assert_called_once_with(page, tab_id="5", timeout_ms=90_000)

    def test_extrae_fila_internos_desde_json_textual(self):
        class Fila:
            def evaluate(self, _script):
                return json.dumps({
                    "headers": ["Folio", "Acciones", "Tipo Trámite"],
                    "cells": ["148326", "Revisar", "CGPE-01-008"],
                    "columnas": {
                        "Folio": "148326",
                        "Acciones": "Revisar",
                        "Tipo Trámite": "CGPE-01-008",
                    },
                })

            def inner_text(self):
                return "148326 Revisar CGPE-01-008"

        meta = Parte1_descarga._extraer_datos_fila_internos(Fila())

        self.assertEqual(meta["folio"], "148326")
        self.assertEqual(meta["tipo_tramite"], "CGPE-01-008")

    def test_paraleliza_bandejas_y_conserva_el_orden_del_resumen(self):
        activas = 0
        max_activas = 0
        lock = threading.Lock()

        def worker(bandeja, _objetivos):
            nonlocal activas, max_activas
            with lock:
                activas += 1
                max_activas = max(max_activas, activas)
            time.sleep(0.03)
            with lock:
                activas -= 1
            return bandeja, [{"folio": bandeja, "ok": True}]

        bandejas = ["Recibidos", "En proceso", "Atendidos"]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos", side_effect=worker), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            resultados = Parte1_descarga.descargar_internos_ift(
                bandejas=bandejas,
                workers=2,
            )

        self.assertEqual(max_activas, 2)
        self.assertEqual([item["folio"] for item in resultados], bandejas)

    def test_reparte_workers_dentro_de_una_misma_bandeja(self):
        llamadas = []
        lock = threading.Lock()

        def worker(bandeja, objetivos):
            with lock:
                llamadas.append((bandeja, list(objetivos)))
            return bandeja, [
                {"folio": folio, "ok": True}
                for folio in objetivos
            ]

        objetivos = [
            {"bandeja": "Atendidos", "folio": str(190000 + indice)}
            for indice in range(12)
        ]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos", side_effect=worker), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            resultados = Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=4,
            )

        self.assertEqual(len(llamadas), 4)
        self.assertEqual(sorted(len(folios) for _bandeja, folios in llamadas), [3, 3, 3, 3])
        folios_asignados = [folio for _bandeja, folios in llamadas for folio in folios]
        self.assertCountEqual(folios_asignados, [item["folio"] for item in objetivos])
        self.assertEqual(len({item["folio"] for item in resultados}), 12)

    def test_reintenta_segmento_vacio_que_tenia_objetivos(self):
        llamadas = 0

        def worker(bandeja, objetivos):
            nonlocal llamadas
            llamadas += 1
            if llamadas == 1:
                return bandeja, []
            return bandeja, [{"folio": folio, "ok": True} for folio in objetivos]

        objetivos = [{"bandeja": "Atendidos", "folio": "190823"}]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos", side_effect=worker), \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_REINTENTOS", 1), \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_ESPERA", 0), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            resultados = Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=1,
            )

        self.assertEqual(llamadas, 2)
        self.assertEqual(resultados, [{"folio": "190823", "ok": True}])

    def test_segmento_vacio_agotado_se_convierte_en_error(self):
        objetivos = [{"bandeja": "Atendidos", "folio": "190823"}]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos", return_value=("Atendidos", [])) as worker, \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_REINTENTOS", 1), \
             patch.object(Parte1_descarga, "INTERNOS_WORKER_ESPERA", 0), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            resultados = Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=1,
            )

        self.assertEqual(worker.call_count, 2)
        self.assertEqual(resultados[0]["archivo"], "ERROR_SEGMENTO_INTERNO_VACIO")
        self.assertFalse(resultados[0]["ok"])

    def test_reserva_dos_segmentos_y_asigna_el_resto_segun_carga(self):
        llamadas = []

        def worker(bandeja, objetivos):
            llamadas.append((bandeja, list(objetivos)))
            return bandeja, [{"folio": folio, "ok": True} for folio in objetivos]

        objetivos = [
            {"bandeja": "En proceso", "folio": str(140000 + indice)}
            for indice in range(5)
        ] + [
            {"bandeja": "Atendidos", "folio": str(190000 + indice)}
            for indice in range(19)
        ]
        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos", side_effect=worker), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=12,
            )

        self.assertEqual(sum(1 for bandeja, _ in llamadas if bandeja == "En proceso"), 3)
        self.assertEqual(sum(1 for bandeja, _ in llamadas if bandeja == "Atendidos"), 9)
        atendidos = sorted(len(folios) for bandeja, folios in llamadas if bandeja == "Atendidos")
        self.assertEqual(atendidos, [2, 2, 2, 2, 2, 2, 2, 2, 3])

    def test_doce_workers_dan_dos_segmentos_a_cada_una_de_seis_bandejas(self):
        llamadas = []

        def worker(bandeja, objetivos):
            llamadas.append((bandeja, list(objetivos)))
            return bandeja, [{"folio": folio, "ok": True} for folio in objetivos]

        objetivos = []
        for indice_bandeja, bandeja in enumerate(Parte1_descarga.BANDEJAS_INTERNOS_DEFAULT):
            objetivos.extend([
                {"bandeja": bandeja, "folio": str(200000 + indice_bandeja * 10 + indice)}
                for indice in range(4)
            ])

        with patch.object(Parte1_descarga, "_validar_sesion_internos", return_value=True), \
             patch.object(Parte1_descarga, "_worker_bandeja_internos", side_effect=worker), \
             patch.object(Parte1_descarga, "guardar_resumen_global"):
            Parte1_descarga.descargar_internos_ift(
                objetivos=objetivos,
                workers=12,
            )

        for bandeja in Parte1_descarga.BANDEJAS_INTERNOS_DEFAULT:
            segmentos = [folios for nombre, folios in llamadas if nombre == bandeja]
            self.assertEqual(len(segmentos), 2, bandeja)
            self.assertEqual(sorted(len(folios) for folios in segmentos), [2, 2])

    def test_main_procesar_propaga_workers_de_internos(self):
        argv_capturado = []

        def main_falso():
            argv_capturado.extend(sys.argv)
            return 0

        with patch.object(Parte1_descarga, "main", side_effect=main_falso):
            rc = main_procesar.ejecutar_descarga_internos(
                bandejas=["Atendidos"],
                headless=True,
                workers=4,
            )

        self.assertEqual(rc, 0)
        indice = argv_capturado.index("--internos-workers")
        self.assertEqual(argv_capturado[indice + 1], "4")

    def test_reporta_fallo_al_preparar_sesion_sin_crear_workers(self):
        with patch.object(
            Parte1_descarga,
            "_validar_sesion_internos",
            side_effect=RuntimeError("Chromium no disponible"),
        ), patch.object(Parte1_descarga, "_worker_bandeja_internos") as worker:
            resultados = Parte1_descarga.descargar_internos_ift(
                bandejas=["Atendidos"],
                workers=1,
            )

        worker.assert_not_called()
        self.assertEqual(resultados[0]["archivo"], "ERROR_PREPARACION_SESION_INTERNOS")
        self.assertIn("Chromium no disponible", resultados[0]["error"])

    def test_cierra_popups_sin_cerrar_la_pagina_principal(self):
        class Pagina:
            def __init__(self, cerrada=False):
                self.cerrada = cerrada
                self.close_calls = 0

            def is_closed(self):
                return self.cerrada

            def close(self, run_before_unload=False):
                self.close_calls += 1
                self.cerrada = True

        principal = Pagina()
        popup_abierto = Pagina()
        popup_ya_cerrado = Pagina(cerrada=True)
        context = SimpleNamespace(pages=[principal, popup_abierto, popup_ya_cerrado])

        cerradas = Parte1_descarga._cerrar_paginas_emergentes(
            context,
            principal,
            motivo="prueba",
        )

        self.assertEqual(cerradas, 1)
        self.assertEqual(principal.close_calls, 0)
        self.assertEqual(popup_abierto.close_calls, 1)
        self.assertEqual(popup_ya_cerrado.close_calls, 0)

    def test_certificado_valida_las_seis_bandejas_y_sus_totales(self):
        output = self.tmp / "registros.txt"
        output.write_text("CRT26-000001\n", encoding="utf-8")
        bandejas = []
        for nombre in (
            "Recibidos",
            "En proceso",
            "Copias Marcadas",
            "Atendidos",
            "Ultimos Movimientos",
            "Fuera de tiempo",
        ):
            folios = ["148326", "147390"] if nombre == "En proceso" else []
            bandejas.append({
                "bandeja": nombre,
                "estado": "ENCONTRADOS_COMPLETOS" if folios else "VACIO_CONFIRMADO",
                "total_reportado_satys": len(folios),
                "filas_leidas": len(folios),
                "folios": folios,
                "filas_invalidas": 0,
            })
        resumen = {
            "estado": "COMPLETO",
            "integridad": "VALIDADA",
            "vacio_confirmado": False,
            "total_filas_satys": 1,
            "total_registros": 1,
            "por_anio": [{
                "anio": 2026,
                "estado": "ENCONTRADOS_COMPLETOS",
                "total_reportado_satys": 1,
                "filas_leidas": 1,
                "total_guardados_anio": 1,
                "filas_invalidas": 0,
                "contador_tab": 1,
                "tamanio_pagina": 100,
            }],
            "internos": {
                "estado": "COMPLETO",
                "integridad": "VALIDADA",
                "vacio_confirmado": False,
                "total_folios": 2,
                "total_filas_satys": 2,
                "folios": ["148326", "147390"],
                "por_bandeja": bandejas,
            },
        }
        output.with_suffix(".txt.json").write_text(
            json.dumps(resumen),
            encoding="utf-8",
        )

        validacion = monitor.validar_resumen_extractor(output, ["CRT26-000001"])
        self.assertTrue(validacion["ok"], validacion)
        self.assertEqual(validacion["folios_internos"], ["148326", "147390"])

    def test_metadata_internos_conserva_el_folio_de_la_tabla(self):
        carpeta = self.tmp / "148326"
        metadata = Parte1_descarga._actualizar_metadata_internos(
            carpeta=carpeta,
            bandeja="En proceso",
            firma="firma-prueba",
            row_meta={"folio": "148326", "tipo_tramite": "CGPE-01-008"},
            meta={"registro": "CRT26-034284", "concesionario": "Operador de Prueba"},
            resultados=[],
        )

        self.assertEqual(metadata["folio_tabla_internos"], "148326")
        self.assertEqual(metadata["registro"], "CRT26-034284")
        persisted = json.loads((carpeta / "metadata_satys.json").read_text(encoding="utf-8"))
        self.assertEqual(persisted["folio_tabla_internos"], "148326")

    def test_excel_conserva_dos_folios_internos_con_el_mismo_registro(self):
        import openpyxl

        excel = self.tmp / "TramitesCRT.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Internos"
        headers = [
            "Control", "Estado", "Origen", "1711", "Memo/Volante",
            "Solicitante Promovente", "Representante Legal",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        wb.save(excel)
        wb.close()

        for folio_interno in ("148326", "190823"):
            self.assertTrue(Parte4_excel.actualizar_excel(
                folio=folio_interno,
                folio_internos=folio_interno,
                registro="CRT26-034284",
                excel_path=excel,
                sheet_name="Internos",
            ))

        wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
        ws = wb["Internos"]
        headers = {str(cell.value): cell.column for cell in ws[1] if cell.value}
        valores = [
            str(ws.cell(row=row, column=headers["Folio Internos"]).value)
            for row in range(2, ws.max_row + 1)
        ]
        registros = [
            ws.cell(row=row, column=headers["1711"]).value
            for row in range(2, ws.max_row + 1)
        ]
        wb.close()

        self.assertEqual(valores, ["148326", "190823"])
        self.assertEqual(registros, ["CRT26-034284", "CRT26-034284"])


if __name__ == "__main__":
    unittest.main(verbosity=2)
