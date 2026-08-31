import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

import resolver_sin_operador_rpc_publico as reparador


class ReparacionSinOperadorRpcPublicoTest(unittest.TestCase):
    def _crear_excel(self, path: Path, identificador: str, ruta: str, *, sheet="Turnados recibidos"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet
        ws.append(["1711", "Ruta", "Solicitante Promovente"])
        ws.append([identificador, ruta, "NOMBRE SATYS"])
        if sheet != "Internos":
            wi = wb.create_sheet("Internos")
            wi.append(["1711", "Ruta", "Solicitante Promovente"])
        wb.save(path)
        wb.close()

    @patch("resolver_sin_operador_rpc_publico.bc.buscar_nombre_operador_rpc_online_exacto")
    def test_repara_con_rpc_publico_fusiona_y_sincroniza_sin_duplicados(self, rpc):
        rpc.return_value = {
            "ok": True,
            "idBp": "520537",
            "nombre_completo": "TECNOLOGÍA Y REDES DE DATOS, S.A. DE C.V.",
            "score": 1.0,
            "fuente": "rpc_online_resultados",
            "metodo": "nombre_exacto_rpc_resultados",
        }
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            descargas = root / "descargas"
            output = root / "output"
            shared = root / "shared"
            logs = root / "logs"
            excel = root / "TrámitesCRT.xlsx"
            source = descargas / "CRT26-022937"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({
                "registro": "CRT26-022937",
                "nombre_operador": "Tecnologia y Redes de Datos SA de CV",
            }), encoding="utf-8")
            (source / "solicitud.pdf").write_bytes(b"FUENTE-ACTUAL")
            (source / "anexos").mkdir()
            (source / "anexos" / "formato.xlsx").write_bytes(b"XLSX")

            old = output / "_sin_operador" / "CRT26-022937"
            old.mkdir(parents=True)
            (old / "historico.txt").write_bytes(b"HISTORICO")
            (old / "solicitud.pdf").write_bytes(b"VIEJO")
            (old / "metadata_satys.json").write_text("{}", encoding="utf-8")

            remote_old = shared / "output" / "_sin_operador" / "CRT26-022937"
            remote_old.mkdir(parents=True)
            (remote_old / "solo_red.txt").write_bytes(b"RED-HISTORICA")

            self._crear_excel(excel, "CRT26-022937", "_sin_operador\\CRT26-022937")
            old_require = os.environ.get("SATYS_REQUIRE_SHARED_MOUNT")
            os.environ["SATYS_REQUIRE_SHARED_MOUNT"] = "0"
            try:
                payload = reparador.reparar(
                    excel_path=excel,
                    descargas_base=descargas,
                    output_base=output,
                    shared_root=shared,
                    logs_dir=logs,
                    timeout_rpc=1,
                )
            finally:
                if old_require is None:
                    os.environ.pop("SATYS_REQUIRE_SHARED_MOUNT", None)
                else:
                    os.environ["SATYS_REQUIRE_SHARED_MOUNT"] = old_require

            self.assertEqual(payload["total_reparados"], 1)
            destino = output / "520537_tecnologia_y_redes_de_datos_s_a_de_c_v" / "01 EN" / "VE"
            self.assertEqual((destino / "solicitud.pdf").read_bytes(), b"FUENTE-ACTUAL")
            self.assertEqual((destino / "historico.txt").read_bytes(), b"HISTORICO")
            self.assertEqual((destino / "anexos" / "formato.xlsx").read_bytes(), b"XLSX")
            self.assertFalse(any(destino.rglob("*.json")))
            self.assertFalse(old.exists())
            self.assertFalse((destino / "solicitud_1.pdf").exists())

            remote = shared / "output" / "520537_tecnologia_y_redes_de_datos_s_a_de_c_v" / "01 EN" / "VE"
            self.assertEqual((remote / "solicitud.pdf").read_bytes(), b"FUENTE-ACTUAL")
            self.assertEqual((remote / "historico.txt").read_bytes(), b"HISTORICO")
            self.assertEqual((remote / "solo_red.txt").read_bytes(), b"RED-HISTORICA")
            self.assertFalse(remote_old.exists())

            wb = openpyxl.load_workbook(excel, data_only=False)
            self.assertEqual(
                wb["Turnados recibidos"]["B2"].value,
                "520537_tecnologia_y_redes_de_datos_s_a_de_c_v\\01 EN\\VE",
            )
            wb.close()

    @patch("resolver_sin_operador_rpc_publico.bc.buscar_nombre_operador_rpc_online_exacto")
    def test_internos_concilia_1711_con_nombre_de_carpeta(self, rpc):
        rpc.return_value = {
            "ok": True,
            "idBp": "555001",
            "nombre_completo": "SIERRA NORTE TELEVISIÓN POR CABLE, S.A. DE C.V.",
            "score": 1.0,
            "fuente": "rpc_online_searchBP",
            "metodo": "nombre_exacto_rpc_autocomplete",
        }
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            descargas = root / "descargas"
            output = root / "output"
            shared = root / "shared"
            excel = root / "TrámitesCRT.xlsx"
            source = descargas / "internos" / "fuera_de_tiempo" / "135531"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({
                "folio": "135531",
                "nombre_operador": "SIERRA NORTE TELEVISIÓN POR CABLE, S.A. DE C.V.",
            }), encoding="utf-8")
            (source / "documento.pdf").write_bytes(b"PDF")
            old = output / "_sin_operador" / "internos__Fuera_de_tiempo__135531"
            old.mkdir(parents=True)
            (old / "nota.txt").write_bytes(b"NOTA")
            self._crear_excel(
                excel,
                "135531",
                "_sin_operador\\internos__Fuera_de_tiempo__135531",
                sheet="Internos",
            )
            old_require = os.environ.get("SATYS_REQUIRE_SHARED_MOUNT")
            os.environ["SATYS_REQUIRE_SHARED_MOUNT"] = "0"
            try:
                payload = reparador.reparar(
                    excel_path=excel,
                    descargas_base=descargas,
                    output_base=output,
                    shared_root=shared,
                    logs_dir=root / "logs",
                    timeout_rpc=1,
                )
            finally:
                if old_require is None:
                    os.environ.pop("SATYS_REQUIRE_SHARED_MOUNT", None)
                else:
                    os.environ["SATYS_REQUIRE_SHARED_MOUNT"] = old_require
            self.assertEqual(payload["total_reparados"], 1)
            destino = output / "555001_sierra_norte_television_por_cable_s_a_de_c_v" / "01 EN" / "VE"
            self.assertTrue((destino / "documento.pdf").exists())
            self.assertTrue((destino / "nota.txt").exists())
            self.assertFalse(old.exists())

    @patch("resolver_sin_operador_rpc_publico.bc.buscar_nombre_operador_rpc_online_exacto")
    def test_sin_coincidencia_publica_deja_sin_operador_intacto(self, rpc):
        rpc.return_value = {
            "ok": False,
            "estado": "sin_coincidencia",
            "motivo": "nombre_no_encontrado_en_resultados_rpc",
            "candidatos": [],
        }
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "descargas" / "CRT26-099999"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({
                "registro": "CRT26-099999",
                "nombre_operador": "OPERADOR INEXISTENTE",
            }), encoding="utf-8")
            (source / "a.pdf").write_bytes(b"A")
            old = root / "output" / "_sin_operador" / "CRT26-099999"
            old.mkdir(parents=True)
            (old / "a.pdf").write_bytes(b"A")
            excel = root / "TrámitesCRT.xlsx"
            self._crear_excel(excel, "CRT26-099999", "_sin_operador\\CRT26-099999")
            payload = reparador.reparar(
                excel_path=excel,
                descargas_base=root / "descargas",
                output_base=root / "output",
                shared_root=root / "shared",
                logs_dir=root / "logs",
                timeout_rpc=1,
                sincronizar_depi=False,
            )
            self.assertEqual(payload["total_reparados"], 0)
            self.assertTrue(old.exists())
            wb = openpyxl.load_workbook(excel)
            self.assertEqual(wb["Turnados recibidos"]["B2"].value, "_sin_operador\\CRT26-099999")
            wb.close()

    def test_no_llama_resolver_con_excel_oficial(self):
        with patch.object(reparador.bc, "resolver_operador_seguro", side_effect=AssertionError("no debe usarse")):
            with patch.object(reparador.bc, "buscar_nombre_operador_rpc_online_exacto") as publico:
                publico.return_value = {
                    "ok": True,
                    "idBp": "123",
                    "nombre_completo": "EMPRESA, S.A. DE C.V.",
                    "score": 1.0,
                }
                res = reparador.resolver_rpc_publico(["EMPRESA SA DE CV"], 1)
                self.assertTrue(res["ok"])
                publico.assert_called_once()

    @patch("resolver_sin_operador_rpc_publico.bc.buscar_nombre_operador_rpc_online_exacto")
    def test_internos_no_confunde_folio_numerico_con_oficialia(self, rpc):
        rpc.return_value = {
            "ok": True, "idBp": "555001",
            "nombre_completo": "OPERADOR INTERNO, S.A. DE C.V.",
            "score": 1.0,
        }
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            interno = root / "descargas" / "internos" / "atendidos" / "113035"
            interno.mkdir(parents=True)
            (interno / "metadata_satys.json").write_text(json.dumps({
                "folio": "113035", "satys_flujo": "internos",
                "nombre_operador": "OPERADOR INTERNO",
            }), encoding="utf-8")
            (interno / "i.pdf").write_bytes(b"I")
            oficial = root / "descargas" / "CRT26-000001"
            oficial.mkdir(parents=True)
            (oficial / "metadata_satys.json").write_text(json.dumps({
                "folio": "113035", "registro": "CRT26-000001",
                "nombre_operador": "OTRO OPERADOR",
            }), encoding="utf-8")
            old = root / "output" / "_sin_operador" / "internos__Atendidos__113035"
            old.mkdir(parents=True)
            (old / "i.pdf").write_bytes(b"I")
            excel = root / "TrámitesCRT.xlsx"
            self._crear_excel(excel, "113035", "_sin_operador\\internos__Atendidos__113035", sheet="Internos")
            old_require = os.environ.get("SATYS_REQUIRE_SHARED_MOUNT")
            os.environ["SATYS_REQUIRE_SHARED_MOUNT"] = "0"
            try:
                payload = reparador.reparar(
                    excel_path=excel, descargas_base=root / "descargas",
                    output_base=root / "output", shared_root=root / "shared",
                    logs_dir=root / "logs", timeout_rpc=1,
                )
            finally:
                if old_require is None:
                    os.environ.pop("SATYS_REQUIRE_SHARED_MOUNT", None)
                else:
                    os.environ["SATYS_REQUIRE_SHARED_MOUNT"] = old_require
            self.assertEqual(payload["total_reparados"], 1)
            rpc.assert_called_once_with("OPERADOR INTERNO", timeout=1)

    @patch("resolver_sin_operador_rpc_publico.bc.buscar_nombre_operador_rpc_online_exacto")
    def test_sync_fallido_no_actualiza_ruta_ni_retira_revision(self, rpc):
        rpc.return_value = {
            "ok": True, "idBp": "520537",
            "nombre_completo": "TECNOLOGÍA Y REDES DE DATOS, S.A. DE C.V.",
            "score": 1.0,
        }
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "descargas" / "CRT26-022937"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({
                "registro": "CRT26-022937", "nombre_operador": "TECNOLOGIA Y REDES DE DATOS",
            }), encoding="utf-8")
            (source / "a.pdf").write_bytes(b"A")
            old = root / "output" / "_sin_operador" / "CRT26-022937"
            old.mkdir(parents=True)
            (old / "a.pdf").write_bytes(b"A")
            excel = root / "TrámitesCRT.xlsx"
            self._crear_excel(excel, "CRT26-022937", "_sin_operador\\CRT26-022937")
            with patch("resolver_sin_operador_rpc_publico._sync_destino_compartido", return_value=(0, ["DEPI no disponible"])):
                payload = reparador.reparar(
                    excel_path=excel, descargas_base=root / "descargas",
                    output_base=root / "output", shared_root=root / "shared",
                    logs_dir=root / "logs", timeout_rpc=1,
                )
            self.assertEqual(payload["total_reparados"], 0)
            self.assertTrue(old.exists())
            wb = openpyxl.load_workbook(excel)
            self.assertEqual(wb["Turnados recibidos"]["B2"].value, "_sin_operador\\CRT26-022937")
            wb.close()


    def test_correo_refleja_reparacion_final(self):
        import automatizar_registros_diario as diario
        resultados = [{
            "registro": "CRT26-022937",
            "rpc_ok": False,
            "organizado_ok": False,
            "nombre_operador": "Tecnologia y Redes",
        }]
        resumen = {
            "resultados": [{
                "identificador": "CRT26-022937",
                "estado": "reparado",
                "ruta_nueva": "520537_tecnologia_y_redes_de_datos_s_a_de_c_v\\01 EN\\VE",
                "nombre_rpc": "TECNOLOGÍA Y REDES DE DATOS, S.A. DE C.V.",
                "id_operador": "520537",
                "fuente_rpc": "rpc_online_resultados",
                "metodo_rpc": "nombre_exacto_rpc_resultados",
            }]
        }
        salida = diario.aplicar_reparaciones_a_resultados_email(resultados, resumen)
        self.assertTrue(salida[0]["rpc_ok"])
        self.assertTrue(salida[0]["organizado_ok"])
        self.assertEqual(salida[0]["rpc_resultado"]["idBp"], "520537")
        self.assertTrue(salida[0]["_reparado_rpc_publico_final"])


    @patch("resolver_sin_operador_rpc_publico.reparar")
    @patch("resolver_sin_operador_rpc_publico.ProcesoLock")
    def test_main_independiente_adquiere_y_libera_lock(self, lock_cls, reparar_mock):
        lock = lock_cls.return_value
        with patch("sys.argv", ["resolver_sin_operador_rpc_publico.py"]):
            rc = reparador.main()
        self.assertEqual(rc, 0)
        lock_cls.assert_called_once_with(proceso="resolver_sin_operador_rpc_publico.py")
        lock.adquirir.assert_called_once_with()
        lock.liberar.assert_called_once_with()
        reparar_mock.assert_called_once()

    @patch("resolver_sin_operador_rpc_publico.reparar")
    @patch("resolver_sin_operador_rpc_publico.ProcesoLock")
    def test_main_lock_ocupado_no_repara(self, lock_cls, reparar_mock):
        lock_cls.return_value.adquirir.side_effect = reparador.LockOcupadoError("ocupado")
        with patch("sys.argv", ["resolver_sin_operador_rpc_publico.py"]):
            rc = reparador.main()
        self.assertEqual(rc, 3)
        reparar_mock.assert_not_called()
        lock_cls.return_value.liberar.assert_called_once_with()


if __name__ == "__main__":
    unittest.main()


class ClasificacionCorreosMemorandumTest(unittest.TestCase):
    def _crear_excel(self, path: Path, identificador: str, ruta: str, *, sheet="Turnados recibidos"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet
        ws.append(["1711", "Ruta"])
        ws.append([identificador, ruta])
        if sheet != "Internos":
            wi = wb.create_sheet("Internos")
            wi.append(["1711", "Ruta"])
        wb.save(path)
        wb.close()

    @patch("resolver_sin_operador_rpc_publico.bc.buscar_nombre_operador_rpc_online_exacto")
    def test_rpc_no_resuelve_y_memorandum_mueve_a_correos_y_depi(self, rpc):
        rpc.return_value = {"ok": False, "motivo": "sin_coincidencia"}
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            descargas = root / "descargas"
            output = root / "output"
            shared = root / "shared"
            logs = root / "logs"
            excel = root / "TrámitesCRT.xlsx"
            source = descargas / "CRT26-020606"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({
                "registro": "CRT26-020606",
                "nombre_operador": "PEGASO PCS",
            }), encoding="utf-8")
            # También valida el espacio accidental al final del nombre.
            (source / "MEMORANDUM.pdf ").write_bytes(b"MEMO-ACTUAL")
            (source / "otro.pdf").write_bytes(b"FUENTE-ACTUAL")

            old = output / "_sin_operador" / "CRT26-020606"
            old.mkdir(parents=True)
            (old / "historico.txt").write_bytes(b"HISTORICO")
            (old / "otro.pdf").write_bytes(b"VIEJO")

            remote_old = shared / "output" / "_sin_operador" / "CRT26-020606"
            remote_old.mkdir(parents=True)
            (remote_old / "solo_red.txt").write_bytes(b"RED")
            self._crear_excel(excel, "CRT26-020606", "_sin_operador\\CRT26-020606")

            old_require = os.environ.get("SATYS_REQUIRE_SHARED_MOUNT")
            os.environ["SATYS_REQUIRE_SHARED_MOUNT"] = "0"
            try:
                payload = reparador.reparar(
                    excel_path=excel,
                    descargas_base=descargas,
                    output_base=output,
                    shared_root=shared,
                    logs_dir=logs,
                    timeout_rpc=1,
                )
            finally:
                if old_require is None:
                    os.environ.pop("SATYS_REQUIRE_SHARED_MOUNT", None)
                else:
                    os.environ["SATYS_REQUIRE_SHARED_MOUNT"] = old_require

            self.assertEqual(payload["total_reparados"], 0)
            self.assertEqual(payload["total_memorandum_detectados"], 1)
            self.assertEqual(payload["total_correos_confirmados"], 1)
            self.assertEqual(payload["cambios_excel_correos"], 1)
            destino = output / "_sin_operador" / "(correos)" / "CRT26-020606"
            self.assertEqual((destino / "MEMORANDUM.pdf ").read_bytes(), b"MEMO-ACTUAL")
            self.assertEqual((destino / "otro.pdf").read_bytes(), b"FUENTE-ACTUAL")
            self.assertEqual((destino / "historico.txt").read_bytes(), b"HISTORICO")
            self.assertFalse((destino / "metadata_satys.json").exists())
            self.assertFalse(old.exists())
            self.assertTrue((source / "MEMORANDUM.pdf ").exists(), "descargas nunca debe moverse")

            remote = shared / "output" / "_sin_operador" / "(correos)" / "CRT26-020606"
            self.assertEqual((remote / "MEMORANDUM.pdf ").read_bytes(), b"MEMO-ACTUAL")
            self.assertEqual((remote / "solo_red.txt").read_bytes(), b"RED")
            self.assertFalse(remote_old.exists())

            wb = openpyxl.load_workbook(excel)
            self.assertEqual(
                wb["Turnados recibidos"]["B2"].value,
                "_sin_operador\\(correos)\\CRT26-020606",
            )
            wb.close()

    @patch("resolver_sin_operador_rpc_publico.bc.buscar_nombre_operador_rpc_online_exacto")
    def test_sin_memorandum_permanece_en_sin_operador(self, rpc):
        rpc.return_value = {"ok": False, "motivo": "sin_coincidencia"}
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "descargas" / "CRT26-099998"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({
                "registro": "CRT26-099998", "nombre_operador": "SIN RPC",
            }), encoding="utf-8")
            (source / "documento.pdf").write_bytes(b"DOC")
            old = root / "output" / "_sin_operador" / "CRT26-099998"
            old.mkdir(parents=True)
            (old / "documento.pdf").write_bytes(b"DOC")
            excel = root / "TrámitesCRT.xlsx"
            self._crear_excel(excel, "CRT26-099998", "_sin_operador\\CRT26-099998")
            payload = reparador.reparar(
                excel_path=excel,
                descargas_base=root / "descargas",
                output_base=root / "output",
                shared_root=root / "shared",
                logs_dir=root / "logs",
                timeout_rpc=1,
                sincronizar_depi=False,
            )
            self.assertEqual(payload["total_memorandum_detectados"], 0)
            self.assertFalse((root / "output" / "_sin_operador" / "(correos)").exists())
            self.assertTrue(old.exists())
            wb = openpyxl.load_workbook(excel)
            self.assertEqual(wb["Turnados recibidos"]["B2"].value, "_sin_operador\\CRT26-099998")
            wb.close()

    @patch("resolver_sin_operador_rpc_publico.bc.buscar_nombre_operador_rpc_online_exacto")
    def test_memorandum_sync_depi_fallido_no_cambia_ruta_ni_retira_origen_output(self, rpc):
        rpc.return_value = {"ok": False, "motivo": "sin_coincidencia"}
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "descargas" / "CRT26-020607"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({
                "registro": "CRT26-020607", "nombre_operador": "SIN RPC",
            }), encoding="utf-8")
            (source / "MEMORANDUM.pdf").write_bytes(b"MEMO")
            old = root / "output" / "_sin_operador" / "CRT26-020607"
            old.mkdir(parents=True)
            (old / "MEMORANDUM.pdf").write_bytes(b"MEMO")
            excel = root / "TrámitesCRT.xlsx"
            self._crear_excel(excel, "CRT26-020607", "_sin_operador\\CRT26-020607")
            with patch("resolver_sin_operador_rpc_publico._sync_destino_compartido", return_value=(0, ["DEPI no disponible"])):
                payload = reparador.reparar(
                    excel_path=excel,
                    descargas_base=root / "descargas",
                    output_base=root / "output",
                    shared_root=root / "shared",
                    logs_dir=root / "logs",
                    timeout_rpc=1,
                )
            self.assertEqual(payload["total_correos_confirmados"], 0)
            self.assertTrue(old.exists())
            wb = openpyxl.load_workbook(excel)
            self.assertEqual(wb["Turnados recibidos"]["B2"].value, "_sin_operador\\CRT26-020607")
            wb.close()

    @patch("resolver_sin_operador_rpc_publico.bc.buscar_nombre_operador_rpc_online_exacto")
    def test_correos_es_idempotente(self, rpc):
        rpc.return_value = {"ok": False, "motivo": "sin_coincidencia"}
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            source = root / "descargas" / "CRT26-020608"
            source.mkdir(parents=True)
            (source / "metadata_satys.json").write_text(json.dumps({
                "registro": "CRT26-020608", "nombre_operador": "SIN RPC",
            }), encoding="utf-8")
            (source / "MEMORANDUM.pdf").write_bytes(b"MEMO")
            destino = root / "output" / "_sin_operador" / "(correos)" / "CRT26-020608"
            destino.mkdir(parents=True)
            (destino / "MEMORANDUM.pdf").write_bytes(b"MEMO")
            excel = root / "TrámitesCRT.xlsx"
            self._crear_excel(excel, "CRT26-020608", "_sin_operador\\(correos)\\CRT26-020608")
            payload = reparador.reparar(
                excel_path=excel,
                descargas_base=root / "descargas",
                output_base=root / "output",
                shared_root=root / "shared",
                logs_dir=root / "logs",
                timeout_rpc=1,
                sincronizar_depi=False,
            )
            self.assertEqual(payload["total_correos_confirmados"], 1)
            self.assertEqual(payload["total_correos_ya_clasificados"], 1)
            self.assertEqual(payload["cambios_excel_correos"], 0)
            self.assertTrue(destino.exists())
