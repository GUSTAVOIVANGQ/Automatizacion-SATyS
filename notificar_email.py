#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
notificar_email.py
==================
Notificación final por correo electrónico para el sistema SATyS.

Las credenciales y destinatarios se leen de config/configuracion_local.json.
No se conservan contraseñas dentro de este módulo.

Se integra con main_procesar.py y se ejecuta al terminar cualquier corrida:
  - TXT de números de registro.
  - Folios SATyS.
  - Corrida diaria invocada por automatizar_registros_diario.py.

El correo incluye los datos esenciales de salida:
  1) output/Folios_Datos_Completos.xlsx
  2) output/
  3) descargas/
  4) TrámitesCRT.xlsx

Uso independiente:
  python notificar_email.py --test
  python notificar_email.py --log descargas/procesamiento_log_registros.json
"""

from __future__ import annotations

import html
import json
import mimetypes
import re
import smtplib
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import openpyxl

from configuracion_local import carpeta_compartida, configuracion_email

# ══════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN LOCAL DE CORREO
# ══════════════════════════════════════════════════════════════════════

_EMAIL_CFG = configuracion_email()
EMAIL_ENABLED = bool(_EMAIL_CFG.get("enabled", True))
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 465
SMTP_SSL = True
SMTP_TIMEOUT = 30

GMAIL_REMITENTE = str(_EMAIL_CFG.get("remitente", "")).strip()
GMAIL_APP_PASSWORD = str(_EMAIL_CFG.get("app_password", "")).strip()
EMAIL_FROM_NAME = str(_EMAIL_CFG.get("from_name", "Sistema SATyS DEI")).strip()
DESTINATARIOS = [str(x).strip() for x in _EMAIL_CFG.get("destinatarios", []) if str(x).strip()]
CC = list(_EMAIL_CFG.get("cc", []))
BCC = list(_EMAIL_CFG.get("bcc", []))

AUTOR = "David Palestina Ramirez y equipo"
ORGANIZACION = "Direccion Ejecutiva de Indicadores (DEI)"
CONSEJO = ""
CARPETA_SALIDA = str(carpeta_compartida())
ADJUNTAR_TRAMITES_CRT = True
MAX_ADJUNTO_MB = 25

# ══════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════


def _split_list(value: Any) -> list[str]:
    if not value:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(x).strip() for x in value if str(x).strip()]
    return [x.strip() for x in str(value).replace(";", ",").split(",") if x.strip()]


def _existe(path: Path | str | None) -> str:
    if not path:
        return "N/D"
    try:
        p = Path(path)
        if p.exists():
            if p.is_dir():
                return "Disponible"
            return f"Disponible ({p.stat().st_size / 1024:,.1f} KB)"
        return "No encontrado"
    except Exception:
        return "N/D"


def _abs(project_root: Path, p: Path | str | None) -> str:
    if p is None:
        return ""
    path = Path(p)
    if not path.is_absolute():
        path = project_root / path
    try:
        return str(path.resolve())
    except Exception:
        return str(path)


def _norm_rel(path: Path | str | None) -> str:
    r"""Convierte una ruta a forma legible tipo \output\... cuando es relativa."""
    if not path:
        return ""
    s = str(path).replace("/", "\\")
    if len(s) >= 2 and s[1] == ":":
        return s
    if not s.startswith("\\"):
        s = "\\" + s
    return s


def _candidatos_excel_tramites(
    *,
    project_root: Path,
    excel_path: Path | str = "TrámitesCRT.xlsx",
    outputs: dict[str, str] | None = None,
    carpeta_compartida: Path | str | None = None,
) -> list[Path]:
    """Devuelve rutas candidatas para adjuntar TrámitesCRT.xlsx.

    Orden de prioridad:
      1. Ruta explícita en outputs["TrámitesCRT.xlsx"].
      2. excel_path recibido por main_procesar.py.
      3. Carpeta compartida configurada.
      4. Excel local del proyecto.
    """
    candidatos: list[Path] = []

    def add(value: Path | str | None) -> None:
        if not value:
            return
        p = Path(value)
        if not p.is_absolute():
            p = project_root / p
        if p.name != "TrámitesCRT.xlsx":
            # Evita adjuntar por error otro Excel, por ejemplo Folios_Datos_Completos.xlsx.
            return
        if p not in candidatos:
            candidatos.append(p)

    if outputs:
        add(outputs.get("TrámitesCRT.xlsx"))

    add(excel_path)

    if carpeta_compartida:
        add(Path(carpeta_compartida) / "TrámitesCRT.xlsx")

    add(Path(CARPETA_SALIDA) / "TrámitesCRT.xlsx")
    add(project_root / "TrámitesCRT.xlsx")

    return candidatos


def _adjuntar_archivo(msg: EmailMessage, path: Path) -> bool:
    """Adjunta un archivo al mensaje. Retorna True si quedó adjunto."""
    try:
        if not path.exists() or not path.is_file():
            return False

        size_mb = path.stat().st_size / (1024 * 1024)
        if size_mb > MAX_ADJUNTO_MB:
            print(f"  ⚠️  No se adjunta {path.name}: {size_mb:.1f} MB supera límite de {MAX_ADJUNTO_MB} MB.")
            return False

        ctype, encoding = mimetypes.guess_type(str(path))
        if ctype is None or encoding is not None:
            ctype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        maintype, subtype = ctype.split("/", 1)

        with path.open("rb") as f:
            msg.add_attachment(
                f.read(),
                maintype=maintype,
                subtype=subtype,
                filename=path.name,
            )
        print(f"  Adjunto      : {path} ({path.stat().st_size / 1024:,.1f} KB)")
        return True
    except Exception as exc:
        print(f"  ⚠️  No se pudo adjuntar {path}: {exc}")
        return False


def _ruta_resultado(r: dict[str, Any]) -> str:
    """Devuelve la mejor Ruta conocida para clasificar el estado final."""
    rpc = r.get("rpc_resultado") or {}
    for value in (
        r.get("ruta"),
        rpc.get("ruta"),
        r.get("output_dir"),
        r.get("sin_operador_dir"),
    ):
        text = str(value or "").strip()
        if text:
            return text.replace("/", "\\")
    return ""


def _ruta_es_correos(value: Any) -> bool:
    """True sólo para _sin_operador\\(correos)\\... (también ruta absoluta)."""
    s = str(value or "").strip().replace("/", "\\").casefold()
    s = re.sub(r"\\+", r"\\", s)
    return bool(re.search(r"(?:^|\\)_sin_operador\\\(correos\)(?:\\|$)", s))


def _ruta_es_revision(value: Any) -> bool:
    """True para _sin_operador pendiente, excluyendo expresamente (correos)."""
    s = str(value or "").strip().replace("/", "\\").casefold()
    s = re.sub(r"\\+", r"\\", s)
    if not re.search(r"(?:^|\\)_sin_operador(?:\\|$)", s):
        return False
    return not _ruta_es_correos(s)


def conteos_revision_desde_excel(excel_path: str | Path) -> dict[str, int]:
    """Cuenta el estado FINAL directamente desde TrámitesCRT.xlsx.

    ``en_revision`` incluye únicamente filas cuya Ruta permanece bajo
    ``_sin_operador`` y excluye ``_sin_operador\\(correos)``. Así el correo no
    reutiliza contadores intermedios de workers ni duplica folios que aparecieron
    en varias bandejas de Internos.
    """
    path = Path(excel_path)
    salida = {
        "total_excel": 0,
        "en_revision": 0,
        "correos_clasificados": 0,
        "organizados_excel": 0,
    }
    if not path.is_file():
        return salida

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return salida
    try:
        for sheet in ("Turnados recibidos", "Internos"):
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            rows = ws.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration:
                continue
            header_map = {
                re.sub(r"\\s+", " ", str(v or "").strip()).casefold(): i
                for i, v in enumerate(headers)
                if str(v or "").strip()
            }
            idx_ruta = header_map.get("ruta")
            idx_id = header_map.get("1711")
            if idx_ruta is None:
                continue
            for row in rows:
                identificador = ""
                if idx_id is not None and idx_id < len(row):
                    identificador = str(row[idx_id] or "").strip()
                ruta = row[idx_ruta] if idx_ruta < len(row) else ""
                # Una fila real debe tener al menos 1711 o Ruta.
                if not identificador and not str(ruta or "").strip():
                    continue
                salida["total_excel"] += 1
                if _ruta_es_correos(ruta):
                    salida["correos_clasificados"] += 1
                elif _ruta_es_revision(ruta):
                    salida["en_revision"] += 1
                else:
                    salida["organizados_excel"] += 1
    finally:
        wb.close()
    return salida


def conteos_desde_resultados(resultados: list[dict[str, Any]]) -> dict[str, int]:
    """Calcula los indicadores esenciales de una corrida consolidada."""
    exitosos = 0
    sin_operador = 0
    errores = 0
    correos = 0
    internos = 0
    rpc_excel = 0
    rpc_online = 0
    for r in resultados or []:
        es_correo = bool(r.get("es_correo"))
        ruta_final = _ruta_resultado(r)
        correo_clasificado = _ruta_es_correos(ruta_final)
        revision_final = _ruta_es_revision(ruta_final)
        if es_correo:
            correos += 1
        if r.get("bandeja_internos") or r.get("_origen_proceso") == "internos":
            internos += 1

        rpc = r.get("rpc_resultado") or {}
        fuente = str(rpc.get("fuente") or "").lower()
        metodo = str(rpc.get("metodo") or "").lower()
        if r.get("rpc_ok") and (fuente == "excel_rpc" or "_excel" in metodo):
            rpc_excel += 1
        elif r.get("rpc_ok") and (
            fuente.startswith("rpc_online")
            or "_rpc_" in metodo
            or metodo.endswith("_autocomplete")
        ):
            rpc_online += 1

        ok = bool(
            correo_clasificado
            or (
                not revision_final
                and r.get("rpc_ok")
                and r.get("organizado_ok")
                and r.get("excel_ok")
            )
        )
        if ok:
            exitosos += 1
        elif revision_final or (
            r.get("_fallido_controlado")
            and bool(r.get("output_dir") or r.get("sin_operador_dir"))
        ) or (
            not r.get("rpc_ok")
            and _ruta_es_revision(r.get("output_dir") or r.get("sin_operador_dir"))
        ):
            sin_operador += 1
        else:
            errores += 1
    return {
        "total": len(resultados or []),
        "exitosos": exitosos,
        "sin_operador": sin_operador,
        "revision_manual": sin_operador,
        "errores": errores,
        "correos": correos,
        "internos": internos,
        "oficialia_otros": max(0, len(resultados or []) - internos),
        "rpc_excel": rpc_excel,
        "rpc_online": rpc_online,
    }


def _estado_texto(r: dict[str, Any]) -> tuple[str, str, str]:
    ruta_final = _ruta_resultado(r)
    if _ruta_es_correos(ruta_final):
        return "Correo clasificado", "#075985", "#e0f2fe"
    if (
        not _ruta_es_revision(ruta_final)
        and r.get("rpc_ok")
        and r.get("organizado_ok")
        and r.get("excel_ok")
    ):
        return "Éxito", "#166534", "#dcfce7"
    if _ruta_es_revision(ruta_final) or (
        r.get("_fallido_controlado")
        and bool(r.get("output_dir") or r.get("sin_operador_dir"))
    ):
        return "En revisión", "#92400e", "#fef3c7"
    return "Error", "#991b1b", "#fee2e2"


def _registro_label(r: dict[str, Any]) -> str:
    return str(r.get("registro") or r.get("folio") or r.get("folio_id") or "N/D")


def _rpc_exactitud(r: dict[str, Any]) -> str:
    rpc = r.get("rpc_resultado") or {}
    score = rpc.get("score")
    if score is None:
        if r.get("rpc_ok"):
            return "100%"
        if r.get("nombre_operador") or r.get("id_solicitante"):
            return "0%"
        return ""
    try:
        return f"{float(score) * 100:.0f}%"
    except Exception:
        return ""


def _motivo_revision(r: dict[str, Any]) -> str:
    if r.get("_fallido_controlado"):
        return "Descarga agotó sus reintentos; requiere revisión manual"
    if r.get("es_correo"):
        return f"Folio OPC {r.get('folio_opc') or 'CORREO'}"
    rpc = r.get("rpc_resultado") or {}
    if not r.get("rpc_ok"):
        return str(rpc.get("motivo") or "Operador no resuelto por Excel ni RPC web")
    if not r.get("organizado_ok"):
        return "No se verificó la organización de archivos"
    if not r.get("excel_ok"):
        return "No se confirmó la actualización del Excel"
    return "Revisión requerida"


def _tabla_resultados_html(resultados: list[dict[str, Any]], max_mostrar: int = 25) -> str:
    pendientes = [r for r in resultados or [] if _estado_texto(r)[0] in {"En revisión", "Error"}]
    if not pendientes:
        return "<tr><td colspan='4' style='padding:12px;color:#64748b;text-align:center;'>Sin pendientes de revisión.</td></tr>"

    filas: list[str] = []
    for r in pendientes[:max_mostrar]:
        estado, color, bg = _estado_texto(r)
        output_dir = r.get("output_dir") or r.get("sin_operador_dir") or ""
        tipo = "Internos" if r.get("bandeja_internos") or r.get("_origen_proceso") == "internos" else "Oficialía/otro"
        filas.append(
            "<tr>"
            f"<td style='padding:8px 10px;border-bottom:1px solid #e5e7eb;font-family:Consolas,monospace;'>{html.escape(_registro_label(r))}</td>"
            f"<td style='padding:8px 10px;border-bottom:1px solid #e5e7eb;'>{html.escape(tipo)}<br><span style='background:{bg};color:{color};border-radius:999px;padding:3px 8px;font-size:11px;font-weight:700;'>{estado}</span></td>"
            f"<td style='padding:8px 10px;border-bottom:1px solid #e5e7eb;'>{html.escape(_motivo_revision(r))}</td>"
            f"<td style='padding:8px 10px;border-bottom:1px solid #e5e7eb;font-family:Consolas,monospace;font-size:11px;color:#475569;'>{html.escape(_norm_rel(output_dir) or '-')}</td>"
            "</tr>"
        )

    if len(pendientes) > max_mostrar:
        filas.append(
            f"<tr><td colspan='4' style='padding:10px;text-align:center;color:#64748b;'>"
            f"... y {len(pendientes) - max_mostrar:,} pendientes más. Ver los CSV de auditoría y el JSON de log.</td></tr>"
        )
    return "\n".join(filas)


def _outputs_default(project_root: Path,
                     descargas_base: Path | str = "descargas",
                     output_base: Path | str = "output",
                     excel_path: Path | str = "TrámitesCRT.xlsx",
                     excel_metadata_path: Path | str | None = None,
                     carpeta_compartida: Path | str | None = None) -> dict[str, str]:
    excel_metadata_path = excel_metadata_path or (Path(output_base) / "Folios_Datos_Completos.xlsx")
    outputs = {
        "Folios_Datos_Completos.xlsx": _abs(project_root, excel_metadata_path),
        "Carpeta /output": _abs(project_root, output_base),
        "Carpeta /descargas": _abs(project_root, descargas_base),
        "TrámitesCRT.xlsx": _abs(project_root, excel_path),
    }
    if carpeta_compartida:
        outputs["Carpeta compartida sincronizada"] = str(Path(carpeta_compartida))
    return outputs


def _tabla_outputs_html(outputs: dict[str, str]) -> str:
    filas = []
    for nombre, ruta in outputs.items():
        filas.append(
            "<tr>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #e5e7eb;font-weight:700;color:#0f172a;'>{html.escape(nombre)}</td>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #e5e7eb;font-family:Consolas,monospace;color:#1d4ed8;font-size:12px;'>{html.escape(str(ruta))}</td>"
            f"<td style='padding:10px 12px;border-bottom:1px solid #e5e7eb;color:#475569;'>{html.escape(_existe(ruta))}</td>"
            "</tr>"
        )
    return "\n".join(filas)


# ══════════════════════════════════════════════════════════════════════
# HTML / TEXTO
# ══════════════════════════════════════════════════════════════════════


def construir_html(fecha_ejecucion: str,
                   modo: str,
                   conteos: dict[str, int],
                   resultados: list[dict[str, Any]],
                   outputs: dict[str, str],
                   log_path: str | None = None) -> str:
    total = int(conteos.get("total", 0) or 0)
    exitosos = int(conteos.get("exitosos", 0) or 0)
    sin_operador = int(conteos.get("sin_operador", 0) or 0)
    errores = int(conteos.get("errores", 0) or 0)
    internos = int(conteos.get("internos", 0) or 0)
    oficialia_otros = int(conteos.get("oficialia_otros", 0) or 0)
    correos = int(conteos.get("correos", 0) or 0)
    rpc_excel = int(conteos.get("rpc_excel", 0) or 0)
    rpc_online = int(conteos.get("rpc_online", 0) or 0)
    correos_clasificados = int(conteos.get("correos_clasificados_excel", 0) or 0)
    pct = lambda n: round((n / total) * 100) if total else 0

    fecha_fmt = fecha_ejecucion
    try:
        fecha_fmt = datetime.fromisoformat(str(fecha_ejecucion)).strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        pass

    tabla_resultados = _tabla_resultados_html(resultados)
    tabla_outputs = _tabla_outputs_html(outputs)
    log_html = html.escape(str(log_path)) if log_path else "N/D"

    return f"""<!doctype html>
<html lang="es">
<head><meta charset="utf-8"><title>Reporte SATyS</title></head>
<body style="margin:0;background:#eef2f7;font-family:Segoe UI,Arial,sans-serif;color:#0f172a;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#eef2f7;padding:28px 0;">
    <tr><td align="center">
      <table width="960" cellpadding="0" cellspacing="0" style="max-width:960px;background:#ffffff;border-radius:18px;overflow:hidden;box-shadow:0 10px 30px rgba(15,23,42,.14);">
        <tr>
          <td style="background:linear-gradient(135deg,#0f5860,#156e78 55%,#0f172a);padding:34px 40px;color:white;">
            <div style="font-size:12px;letter-spacing:2px;text-transform:uppercase;color:#bff3f7;font-weight:700;">Automatización SATyS</div>
            <h1 style="margin:8px 0 6px;font-size:28px;line-height:1.2;">Resultado del proceso</h1>
            <div style="font-size:14px;color:#d7fbff;">Modo: <b>{html.escape(modo)}</b> &nbsp;|&nbsp; Fecha: <b>{html.escape(fecha_fmt)}</b></div>
            <div style="font-size:12px;color:#bff3f7;margin-top:8px;">Resolución RPC: <b>Excel oficial primero</b> y, si no resuelve, <b>buscador público del RPC</b>.</div>
          </td>
        </tr>
        <tr><td style="background:#0f172a;color:#cbd5e1;text-align:center;padding:12px 40px;font-size:12px;">
          {html.escape(AUTOR)} &middot; {html.escape(ORGANIZACION)} &middot; {html.escape(CONSEJO)}
        </td></tr>
        <tr><td style="padding:28px 40px 10px;">
          <table width="100%" cellpadding="0" cellspacing="10">
            <tr>
              <td style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:14px;padding:18px;text-align:center;">
                <div style="font-size:30px;font-weight:800;color:#1d4ed8;">{total:,}</div><div style="font-size:12px;color:#2563eb;font-weight:700;">TOTAL</div>
              </td>
              <td style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:14px;padding:18px;text-align:center;">
                <div style="font-size:30px;font-weight:800;color:#16a34a;">{exitosos:,}</div><div style="font-size:12px;color:#15803d;font-weight:700;">EXITOSOS ({pct(exitosos)}%)</div>
              </td>
              <td style="background:#fffbeb;border:1px solid #fde68a;border-radius:14px;padding:18px;text-align:center;">
                <div style="font-size:30px;font-weight:800;color:#d97706;">{sin_operador:,}</div><div style="font-size:12px;color:#92400e;font-weight:700;">EN REVISIÓN ({pct(sin_operador)}%)</div>
              </td>
              <td style="background:#fef2f2;border:1px solid #fecaca;border-radius:14px;padding:18px;text-align:center;">
                <div style="font-size:30px;font-weight:800;color:#dc2626;">{errores:,}</div><div style="font-size:12px;color:#991b1b;font-weight:700;">ERRORES ({pct(errores)}%)</div>
              </td>
            </tr>
          </table>
          <div style="margin:4px 10px 0;background:#f8fafc;border:1px solid #e5e7eb;border-radius:10px;padding:12px 14px;color:#334155;font-size:13px;">
            <b>Tipos:</b> Internos {internos:,} &nbsp;|&nbsp; Oficialía/otros {oficialia_otros:,} &nbsp;|&nbsp; Folio OPC CORREO {correos:,}<br>
            <b>Operadores resueltos:</b> Excel RPC {rpc_excel:,} &nbsp;|&nbsp; buscador web RPC {rpc_online:,}<br>
            <b>Correos clasificados fuera de revisión:</b> {correos_clasificados:,}
          </div>
        </td></tr>
        <tr><td style="padding:8px 40px 22px;">
          <h2 style="font-size:18px;margin:16px 0 10px;border-left:5px solid #156e78;padding-left:12px;">Salidas principales</h2>
          <p style="margin:0 0 12px;color:#475569;font-size:13px;">Estas son las 4 ubicaciones/documentos que deben revisarse al terminar cualquier corrida.</p>
          <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;border-collapse:separate;border-spacing:0;font-size:13px;">
            <thead><tr style="background:#f8fafc;"><th align="left" style="padding:10px 12px;color:#334155;">Salida</th><th align="left" style="padding:10px 12px;color:#334155;">Ruta</th><th align="left" style="padding:10px 12px;color:#334155;">Estado</th></tr></thead>
            <tbody>{tabla_outputs}</tbody>
          </table>
          <div style="margin-top:12px;background:#f8fafc;border:1px solid #e5e7eb;border-radius:10px;padding:12px 14px;color:#475569;font-size:12px;">
            Log del proceso: <span style="font-family:Consolas,monospace;color:#1d4ed8;">{log_html}</span>
          </div>
        </td></tr>
        <tr><td style="padding:0 40px 34px;">
          <h2 style="font-size:18px;margin:16px 0 10px;border-left:5px solid #156e78;padding-left:12px;">Pendientes que requieren atención</h2>
          <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;border-collapse:separate;border-spacing:0;font-size:13px;">
            <thead><tr style="background:#0f172a;color:#f8fafc;"><th align="left" style="padding:10px;">Registro/Folio</th><th align="left" style="padding:10px;">Tipo/estado</th><th align="left" style="padding:10px;">Motivo</th><th align="left" style="padding:10px;">Ruta</th></tr></thead>
            <tbody>{tabla_resultados}</tbody>
          </table>
        </td></tr>
        <tr><td style="background:#0f172a;color:#94a3b8;text-align:center;padding:20px 40px;font-size:12px;">
          Correo generado automáticamente por SATyS. Configuración de correo leída desde config/configuracion_local.json.
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>"""


def construir_texto(fecha_ejecucion: str,
                    modo: str,
                    conteos: dict[str, int],
                    resultados: list[dict[str, Any]],
                    outputs: dict[str, str],
                    log_path: str | None = None) -> str:
    lines = [
        "Resultado del proceso SATyS",
        "=" * 60,
        f"Modo: {modo}",
        f"Fecha: {fecha_ejecucion}",
        "RPC: Excel oficial primero; buscador público del RPC como alternativa.",
        "",
        f"Total: {conteos.get('total', 0)}",
        f"Exitosos: {conteos.get('exitosos', 0)}",
        f"En revisión: {conteos.get('sin_operador', 0)}",
        f"Errores: {conteos.get('errores', 0)}",
        f"Tipos: Internos={conteos.get('internos', 0)}, Oficialía/otros={conteos.get('oficialia_otros', 0)}, CORREO={conteos.get('correos', 0)}",
        f"Operadores resueltos: Excel RPC={conteos.get('rpc_excel', 0)}, buscador web RPC={conteos.get('rpc_online', 0)}",
        f"Correos clasificados fuera de revisión: {conteos.get('correos_clasificados_excel', 0)}",
        "",
        "Salidas principales:",
    ]
    for nombre, ruta in outputs.items():
        lines.append(f"- {nombre}: {ruta} [{_existe(ruta)}]")
    if log_path:
        lines.extend(["", f"Log: {log_path}"])
    lines.extend(["", "Pendientes que requieren atención:"])
    pendientes = [r for r in (resultados or []) if _estado_texto(r)[0] in {"En revisión", "Error"}]
    if not pendientes:
        lines.append("- Ninguno")
    for r in pendientes[:25]:
        estado, _, _ = _estado_texto(r)
        lines.append(
            f"- {_registro_label(r)} | {estado} | {_motivo_revision(r)} "
            f"| {_norm_rel(r.get('output_dir') or r.get('sin_operador_dir') or '-') }"
        )
    if len(pendientes) > 25:
        lines.append(f"- ... y {len(pendientes) - 25} pendientes más (ver log/CSV).")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════
# ENVÍO
# ══════════════════════════════════════════════════════════════════════


def enviar_notificacion(total_registros: int | None = None,
                        exitosos: int | None = None,
                        sin_operador: int | None = None,
                        errores: int | None = None,
                        registros: list[dict[str, Any]] | None = None,
                        fecha_ejecucion: str | None = None,
                        carpeta_salida: str | None = None,
                        destinatarios: list[str] | str | None = None,
                        remitente: str | None = None,
                        app_password: str | None = None,
                        modo: str = "SATyS",
                        outputs: dict[str, str] | None = None,
                        log_path: str | None = None,
                        project_root: str | Path | None = None,
                        descargas_base: str | Path = "descargas",
                        output_base: str | Path = "output",
                        excel_path: str | Path = "TrámitesCRT.xlsx",
                        excel_metadata_path: str | Path | None = None,
                        carpeta_compartida: str | Path | None = None,
                        habilitado: bool | None = None,
                        usar_conteo_revision_excel: bool = False) -> bool:
    """Envía correo de notificación final.

    Mantiene compatibilidad con llamadas antiguas y nuevas de main_procesar.py.
    Por defecto usa la configuración local de este archivo.
    """
    registros = registros or []

    if habilitado is None:
        habilitado = EMAIL_ENABLED
    if not habilitado:
        print("ℹ️  Notificación por correo deshabilitada por configuración local.")
        return False

    tos = _split_list(destinatarios) if destinatarios is not None else list(DESTINATARIOS)
    if not tos:
        print("⚠️  No hay destinatarios configurados en config/configuracion_local.json.")
        return False

    from_email = (remitente or GMAIL_REMITENTE).strip()
    password = (app_password if app_password is not None else GMAIL_APP_PASSWORD).replace(" ", "")
    if not from_email or not password:
        print("⚠️  Falta GMAIL_REMITENTE o GMAIL_APP_PASSWORD en notificar_email.py.")
        return False

    conteos_calc = conteos_desde_resultados(registros)
    conteos = dict(conteos_calc)
    conteos.update({
        "total": int(total_registros if total_registros is not None else conteos_calc["total"]),
        "exitosos": int(exitosos if exitosos is not None else conteos_calc["exitosos"]),
        "sin_operador": int(sin_operador if sin_operador is not None else conteos_calc["sin_operador"]),
        "errores": int(errores if errores is not None else conteos_calc["errores"]),
    })
    if conteos["total"] == 0 and registros:
        conteos["total"] = len(registros)

    if usar_conteo_revision_excel:
        final_excel = conteos_revision_desde_excel(excel_path)
        if final_excel.get("total_excel", 0):
            conteos["sin_operador"] = int(final_excel["en_revision"])
            conteos["revision_manual"] = int(final_excel["en_revision"])
            conteos["correos_clasificados_excel"] = int(final_excel["correos_clasificados"])
            # Si no hay resultados de worker (por ejemplo postproceso manual),
            # el Excel final es también la fuente del total. En una corrida
            # diaria se conserva el total de resultados, pero el bloque verde
            # se recalcula como complemento para que las tarjetas sumen 100%.
            if conteos["total"] == 0:
                conteos["total"] = int(final_excel["total_excel"])
            conteos["exitosos"] = max(
                0,
                int(conteos["total"]) - int(conteos["errores"]) - int(conteos["sin_operador"]),
            )
        else:
            conteos["correos_clasificados_excel"] = 0

    fecha_ejecucion = fecha_ejecucion or datetime.now().isoformat()
    project_root = Path(project_root or Path.cwd())
    outputs = outputs or _outputs_default(
        project_root=project_root,
        descargas_base=descargas_base,
        output_base=output_base,
        excel_path=excel_path,
        excel_metadata_path=excel_metadata_path,
        carpeta_compartida=carpeta_compartida or carpeta_salida or CARPETA_SALIDA,
    )

    estado_asunto = "sin errores" if conteos["errores"] == 0 and conteos["sin_operador"] == 0 else "con revisión"
    if conteos["errores"]:
        estado_asunto = f"con {conteos['errores']} error(es)"
    asunto = f"[SATyS-DEI] Proceso {estado_asunto} - {conteos['total']:,} registro(s) - {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    html_body = construir_html(fecha_ejecucion, modo, conteos, registros, outputs, log_path)
    text_body = construir_texto(fecha_ejecucion, modo, conteos, registros, outputs, log_path)

    msg = EmailMessage()
    msg["Subject"] = asunto
    msg["From"] = f"{EMAIL_FROM_NAME} <{from_email}>"
    msg["To"] = ", ".join(tos)
    if CC:
        msg["Cc"] = ", ".join(CC)
    msg.set_content(text_body, subtype="plain", charset="utf-8")
    msg.add_alternative(html_body, subtype="html", charset="utf-8")

    adjunto_excel = None
    if ADJUNTAR_TRAMITES_CRT:
        for candidato in _candidatos_excel_tramites(
            project_root=project_root,
            excel_path=excel_path,
            outputs=outputs,
            carpeta_compartida=carpeta_compartida or carpeta_salida or CARPETA_SALIDA,
        ):
            if candidato.exists() and candidato.is_file():
                if _adjuntar_archivo(msg, candidato):
                    adjunto_excel = candidato
                break
        if adjunto_excel is None:
            print("  ⚠️  No se adjuntó TrámitesCRT.xlsx: archivo no encontrado en rutas esperadas.")

    all_recipients = tos + list(CC) + list(BCC)

    try:
        print("\n" + "=" * 60)
        print("  NOTIFICACIÓN POR CORREO ELECTRÓNICO SATyS")
        print("=" * 60)
        print(f"  Remitente    : {from_email}")
        print(f"  Destinatarios: {', '.join(tos)}")
        print(f"  Asunto       : {asunto}")
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=SMTP_TIMEOUT) as smtp:
            smtp.login(from_email, password)
            smtp.send_message(msg, from_addr=from_email, to_addrs=all_recipients)
        print("  ✅ Correo enviado exitosamente.")
        print("=" * 60 + "\n")
        return True
    except smtplib.SMTPAuthenticationError as exc:
        print(f"  ❌ Error de autenticación Gmail: {exc}")
        return False
    except Exception as exc:
        print(f"  ❌ Error al enviar correo: {exc}")
        return False


def enviar_desde_log_json(log_json_path: str | Path, destinatarios: list[str] | str | None = None) -> bool:
    """Lee procesamiento_log*.json y envía el correo usando los destinatarios configurados por defecto."""
    path = Path(log_json_path)
    if not path.exists():
        print(f"⚠️  No se encontró el log JSON: {path}")
        return enviar_notificacion(0, 0, 0, 0, [], destinatarios=destinatarios, log_path=str(path))

    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as exc:
        print(f"❌ Error al leer {path}: {exc}")
        return False

    resultados = data.get("resultados") or data.get("detalle") or []
    conteos = conteos_desde_resultados(resultados)

    return enviar_notificacion(
        total_registros=data.get("total_registros", data.get("total", conteos["total"])),
        exitosos=data.get("total_exitosos", data.get("exitosos", conteos["exitosos"])),
        sin_operador=data.get("total_sin_operador", data.get("sin_operador", conteos["sin_operador"])),
        errores=data.get("total_errores", data.get("errores", conteos["errores"])),
        registros=resultados,
        fecha_ejecucion=data.get("fecha_ejecucion", datetime.now().isoformat()),
        destinatarios=destinatarios,
        modo=data.get("modo", "SATyS"),
        log_path=str(path),
    )


# ══════════════════════════════════════════════════════════════════════
# CLI DE PRUEBA
# ══════════════════════════════════════════════════════════════════════


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Notificación correo SATyS con configuración local")
    ap.add_argument("--test", action="store_true", help="Enviar correo de prueba")
    ap.add_argument("--log", default="", help="Ruta al JSON de log")
    ap.add_argument("--to", default="", help="Destinatarios opcionales separados por coma; si se omite usa DESTINATARIOS")
    args = ap.parse_args()

    if args.log:
        ok = enviar_desde_log_json(args.log, destinatarios=args.to or None)
    else:
        ok = enviar_notificacion(
            total_registros=4,
            exitosos=2,
            sin_operador=1,
            errores=1,
            registros=[
                {"folio": "CRT26-029502", "registro": "CRT26-029502", "id_solicitante": "518858", "rpc_ok": True, "organizado_ok": True, "excel_ok": True, "nombre_operador": "GRUPO MAGERCA COMUNICACIONES S.A. DE C.V.", "output_dir": r"output\518858_grupo_magerca_comunicaciones_s_a_de_c_v\01 EN\VE", "rpc_resultado": {"score": 1.0, "metodo": "id_exacto"}},
                {"folio": "CRT25-004721", "registro": "CRT25-004721", "id_solicitante": "999999", "rpc_ok": False, "organizado_ok": False, "excel_ok": True, "nombre_operador": "OPERADOR NO EN CATÁLOGO", "output_dir": r"output\_sin_operador\CRT25-004721", "rpc_resultado": {"score": 0.0, "metodo": "id_exacto"}},
                {"folio": "CRT26-000000", "registro": "CRT26-000000", "rpc_ok": False, "organizado_ok": False, "excel_ok": False, "nombre_operador": ""},
                {"folio": "CRT26-012241", "registro": "CRT26-012241", "id_solicitante": "522954", "rpc_ok": True, "organizado_ok": True, "excel_ok": True, "nombre_operador": "JAIME ROJAS RAMÍREZ", "output_dir": r"output\522954_jaime_rojas_ramirez\01 EN\VE", "rpc_resultado": {"score": 1.0, "metodo": "id_exacto"}},
            ],
            fecha_ejecucion=datetime.now().isoformat(),
            modo="Prueba de configuración local",
            destinatarios=args.to or None,
        )
    raise SystemExit(0 if ok else 1)
