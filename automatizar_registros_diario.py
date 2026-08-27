#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
=============================================================
  SATyS CRT — Monitor diario de Registros nuevos
=============================================================

Objetivo:
  1. Entra a SATyS con extraer_registros_documentos.py.
  2. Extrae todos los valores de la columna "Registro" en Documentos en Proceso.
  3. Lee TrámitesCRT.xlsx y toma como evidencia la columna cuyo encabezado es 1711.
  4. Genera registros.txt SOLO con registros nuevos no encontrados en Excel.
  5. Ejecuta main_procesar.py --archivo-registro registros.txt.
  6. Guarda logs, resumen JSON y estado vivo para monitoreo.

Uso manual recomendado en Linux/RHEL:
  /data/gustavo.garcia/satys/venv/bin/python automatizar_registros_diario.py --headless --workers 10

Primera prueba visible en una estación con navegador:
  python automatizar_registros_diario.py --visible --workers 1
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
import signal
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Iterable


def configurar_salida_utf8() -> None:
    """Permite imprimir el log Unicode también con el Python embebido de Windows."""
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if not callable(reconfigure):
            continue
        try:
            reconfigure(encoding="utf-8", errors="replace")
        except (OSError, ValueError):
            pass


configurar_salida_utf8()


from proceso_lock import ProcesoLock, LockOcupadoError
from estado_ejecucion import EstadoEjecucion
from estado_descargas import objetivo_internos_esta_completo, registro_esta_completo
from configuracion_local import carpeta_compartida, configuracion_procesamiento, ruta_configurada
from sincronizacion_depi import sincronizar_salidas

try:
    import notificar_email as _email_mod
    _EMAIL_DISPONIBLE = True
except Exception:
    _EMAIL_DISPONIBLE = False

REGISTRO_RE = re.compile(r"\b[A-Z]{2,8}\d{2}-\d{3,}\b", re.IGNORECASE)
FOLIO_INTERNO_RE = re.compile(r"^\d{1,15}$")

PROJECT_DIR = Path(__file__).resolve().parent
PYTHON_EXE_DEFAULT = Path(os.getenv("SATYS_PYTHON", sys.executable))
EXTRAER_SCRIPT_DEFAULT = PROJECT_DIR / "extraer_registros_documentos.py"
MAIN_SCRIPT_DEFAULT = PROJECT_DIR / "main_procesar.py"
RECONCILIAR_SCRIPT_DEFAULT = PROJECT_DIR / "reconciliar_metadata_global.py"
EXCEL_DEFAULT = ruta_configurada("excel", "TrámitesCRT.xlsx")
REGISTROS_LATEST_DEFAULT = PROJECT_DIR / "registros.txt"
INTERNOS_LATEST_DEFAULT = PROJECT_DIR / "folios_internos_nuevos.json"
REGISTROS_DIR_DEFAULT = PROJECT_DIR / "registros_diarios"
LOG_DIR_DEFAULT = PROJECT_DIR / "logs"
SHEET_DEFAULT = "Turnados recibidos"
HEADER_REGISTRO_DEFAULT = "1711"
SHEET_INTERNOS_DEFAULT = "Internos"
HEADER_FOLIO_INTERNOS_DEFAULT = "Folio Internos"
PROCESAMIENTO_CFG = configuracion_procesamiento()
WORKERS_DEFAULT = int(PROCESAMIENTO_CFG.get("workers", 10))
INTERNOS_WORKERS_DEFAULT = int(PROCESAMIENTO_CFG.get("internos_workers", 12))
TIMEOUT_REGISTRO_DEFAULT = int(PROCESAMIENTO_CFG.get("timeout_registro", 900))
REINTENTOS_REGISTRO_DEFAULT = int(PROCESAMIENTO_CFG.get("reintentos_registro", 2))
WORKERS_REINTENTO_DEFAULT = int(PROCESAMIENTO_CFG.get("workers_reintento", 2))
TIMEOUT_TABLA_DEFAULT = 120
INTENTOS_ANIO_EXTRACCION_DEFAULT = 3
INTENTOS_PAGINA_EXTRACCION_DEFAULT = 3
# Reintentos exclusivos de la etapa inicial de consulta a SATyS. No modifican
# los reintentos por Registro de main_procesar.py.
REINTENTOS_EXTRACCION_DEFAULT = max(0, int(PROCESAMIENTO_CFG.get("reintentos_extraccion", 2)))
ESPERA_REINTENTO_EXTRACCION_DEFAULT = 0


def normalizar_registro(valor: object) -> str:
    """Devuelve un Registro CRT normalizado o cadena vacía si no parece Registro."""
    if valor is None:
        return ""
    texto = str(valor).replace("\u00a0", " ").strip().upper()
    texto = re.sub(r"\s+", "", texto)
    m = REGISTRO_RE.search(texto)
    return m.group(0).upper() if m else ""


def normalizar_folio_interno(valor: object) -> str:
    """Return a numeric Internos Folio or an empty string."""
    texto = re.sub(r"\s+", "", str(valor or "").strip())
    return texto if FOLIO_INTERNO_RE.fullmatch(texto) else ""


def unicos_folios_internos(items: Iterable[object]) -> list[str]:
    vistos: set[str] = set()
    salida: list[str] = []
    for item in items:
        folio = normalizar_folio_interno(item)
        if folio and folio not in vistos:
            vistos.add(folio)
            salida.append(folio)
    return salida


def unicos_preservando_orden(items: Iterable[str]) -> list[str]:
    vistos: set[str] = set()
    salida: list[str] = []
    for item in items:
        item = normalizar_registro(item)
        if item and item not in vistos:
            vistos.add(item)
            salida.append(item)
    return salida


def leer_registros_txt(path: Path) -> list[str]:
    if not path.exists():
        return []
    texto = path.read_text(encoding="utf-8-sig", errors="replace")
    return unicos_preservando_orden(REGISTRO_RE.findall(texto))


def validar_txt_un_registro_por_linea(path: Path, esperados: int) -> dict:
    """
    Valida que el TXT que consumirá main_procesar.py no quede como una sola línea gigante.
    """
    if not path.exists():
        return {"ok": False, "error": f"No existe el TXT: {path}"}

    texto = path.read_text(encoding="utf-8-sig", errors="replace")
    registros = unicos_preservando_orden(REGISTRO_RE.findall(texto))
    lineas_con_registro = [
        linea for linea in texto.splitlines()
        if normalizar_registro(linea)
    ]

    lineas_multiples = []
    for idx, linea in enumerate(texto.splitlines(), start=1):
        encontrados = REGISTRO_RE.findall(linea)
        if len(encontrados) > 1:
            lineas_multiples.append(idx)

    return {
        "ok": len(registros) == esperados and len(lineas_multiples) == 0,
        "registros_detectados": len(registros),
        "lineas_con_registro": len(lineas_con_registro),
        "esperados": esperados,
        "lineas_con_multiples_registros": lineas_multiples[:20],
    }


def guardar_txt_lineas(path: Path, registros: list[str]) -> None:
    """
    Guarda un registro por línea.

    IMPORTANTE:
    main_procesar.py debe recibir el TXT con saltos de línea reales.
    Antes se guardaba con espacios (" ".join), y eso provocaba que
    main_procesar.py leyera todos los Registros como 1 solo elemento.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    registros = unicos_preservando_orden(registros)
    contenido = "\n".join(registros)
    if contenido:
        contenido += "\n"
    path.write_text(contenido, encoding="utf-8")


def cargar_registros_procesados_excel(excel_path: Path, sheet_name: str, header_registro: str) -> tuple[set[str], dict]:
    """
    Busca la columna cuyo encabezado es 1711 y regresa todos los registros ya procesados.
    El Excel enviado tiene el encabezado 1711 en la hoja 'Turnados recibidos'.
    """
    try:
        import openpyxl  # dependencia ya usada por Parte4_excel.py
    except ImportError as exc:
        raise RuntimeError(
            "No se pudo importar openpyxl. Instálalo en el Python portable o revisa que Parte4_excel.py funcione."
        ) from exc

    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el Excel de evidencia: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    header_col = None
    header_row = None
    header_target = str(header_registro).strip()

    # Normalmente está en la primera fila, pero se busca en las primeras 20 por seguridad.
    max_scan_row = min(ws.max_row or 1, 20)
    for row in ws.iter_rows(min_row=1, max_row=max_scan_row):
        for cell in row:
            value = cell.value
            if value is None:
                continue
            value_text = str(value).strip()
            if value_text == header_target or value == int(header_target) if header_target.isdigit() else value_text == header_target:
                header_col = cell.column
                header_row = cell.row
                break
        if header_col is not None:
            break

    if header_col is None:
        # Fallback controlado: en el Excel revisado la columna 1711 corresponde a D.
        header_col = 4
        header_row = 1

    procesados: set[str] = set()
    for row in ws.iter_rows(min_row=(header_row or 1) + 1, min_col=header_col, max_col=header_col):
        registro = normalizar_registro(row[0].value)
        if registro:
            procesados.add(registro)

    info = {
        "excel": str(excel_path),
        "sheet": ws.title,
        "header_registro": header_registro,
        "header_row": header_row,
        "header_col": header_col,
        "total_procesados_excel": len(procesados),
    }
    wb.close()
    return procesados, info


def cargar_folios_internos_procesados_excel(
    excel_path: Path,
    sheet_name: str = SHEET_INTERNOS_DEFAULT,
    header_folio: str = HEADER_FOLIO_INTERNOS_DEFAULT,
) -> tuple[set[str], dict]:
    """Read processed numeric Folio values from the dedicated Internos sheet."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("No se pudo importar openpyxl para leer la hoja Internos.") from exc

    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el Excel de evidencia: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        info = {
            "excel": str(excel_path),
            "sheet": sheet_name,
            "sheet_exists": False,
            "header_folio": header_folio,
            "header_row": None,
            "header_col": None,
            "total_procesados_excel": 0,
        }
        wb.close()
        return set(), info

    ws = wb[sheet_name]

    def _norm_header(value: object) -> str:
        texto = str(value or "").strip().lower()
        reemplazos = str.maketrans("áéíóúüñ", "aeiouun")
        return re.sub(r"[^a-z0-9]", "", texto.translate(reemplazos))

    aliases = {
        _norm_header(header_folio),
        _norm_header("Folio Interno"),
        _norm_header("Folio SATyS Internos"),
    }
    header_col = None
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 20)):
        for cell in row:
            if _norm_header(cell.value) in aliases:
                header_col = cell.column
                header_row = cell.row
                break
        if header_col is not None:
            break

    procesados: set[str] = set()
    if header_col is not None:
        for row in ws.iter_rows(min_row=(header_row or 1) + 1, min_col=header_col, max_col=header_col):
            folio = normalizar_folio_interno(row[0].value)
            if folio:
                procesados.add(folio)

    info = {
        "excel": str(excel_path),
        "sheet": ws.title,
        "sheet_exists": True,
        "header_folio": header_folio,
        "header_row": header_row,
        "header_col": header_col,
        "total_procesados_excel": len(procesados),
    }
    wb.close()
    return procesados, info


def construir_objetivos_internos(
    resumen_internos: dict,
    folios: Iterable[str] | None = None,
) -> list[dict]:
    """Conserva cada pareja (bandeja, folio), incluso si el folio se repite."""
    seleccionados = None if folios is None else set(unicos_folios_internos(folios))
    objetivos: list[dict] = []
    vistos: set[tuple[str, str]] = set()
    for item in resumen_internos.get("por_bandeja") or []:
        bandeja = str(item.get("bandeja") or "").strip()
        for folio in unicos_folios_internos(item.get("folios") or []):
            clave = (bandeja.casefold(), folio)
            if (seleccionados is None or folio in seleccionados) and clave not in vistos:
                vistos.add(clave)
                objetivos.append({"bandeja": bandeja, "folio": folio})
    return objetivos


def seleccionar_objetivos_folio_internos(
    resumen_internos: dict,
    folio: object,
) -> list[dict]:
    """Localiza un Folio exacto en todas las bandejas de Internos.

    El mismo Folio puede aparecer en más de una bandeja. El modo de revisión
    individual debe conservar todas esas apariciones para descargar y auditar
    sus archivos de manera independiente.
    """
    folio_normalizado = normalizar_folio_interno(folio)
    if not folio_normalizado:
        raise ValueError("--folio-internos debe contener entre 1 y 15 dígitos.")

    objetivos = construir_objetivos_internos(
        resumen_internos,
        [folio_normalizado],
    )
    if not objetivos:
        raise LookupError(
            f"El Folio Internos {folio_normalizado} no aparece en ninguna de "
            "las seis bandejas inventariadas de SATyS."
        )
    return objetivos


def validar_salidas_folio_internos(
    *,
    folio: str,
    objetivos: list[dict],
    procesamiento_log: Path,
    excel_path: Path,
    sheet: str,
    header_folio: str,
    project_dir: Path = PROJECT_DIR,
) -> dict:
    """Comprueba que el Folio terminó en Excel y en una carpeta de output."""
    errores: list[str] = []
    salidas: list[str] = []
    resultados: list[dict] = []

    try:
        payload = json.loads(
            procesamiento_log.read_text(encoding="utf-8-sig", errors="replace")
        )
        resultados = payload.get("resultados") or []
        if not isinstance(resultados, list):
            raise ValueError("el campo resultados no es una lista")
    except Exception as exc:
        return {
            "ok": False,
            "folio": folio,
            "procesamiento_log": str(procesamiento_log),
            "excel": str(excel_path),
            "output_dirs": [],
            "errores": [f"No se pudo leer el log de procesamiento: {exc}"],
        }

    por_clave: dict[tuple[str, str], list[dict]] = {}
    for resultado in resultados:
        if not isinstance(resultado, dict):
            continue
        folio_resultado = normalizar_folio_interno(
            resultado.get("folio_tabla_internos") or resultado.get("folio")
        )
        bandeja_resultado = str(resultado.get("bandeja_internos") or "").strip().casefold()
        if folio_resultado and bandeja_resultado:
            por_clave.setdefault((bandeja_resultado, folio_resultado), []).append(resultado)

    for objetivo in objetivos:
        bandeja = str(objetivo.get("bandeja") or "").strip()
        folio_objetivo = normalizar_folio_interno(objetivo.get("folio"))
        clave = (bandeja.casefold(), folio_objetivo)
        candidatos = por_clave.get(clave, [])
        if not candidatos:
            errores.append(f"Sin resultado final para {bandeja}/{folio_objetivo}.")
            continue

        resultado = candidatos[-1]
        if resultado.get("excel_ok") is not True:
            errores.append(f"Excel no fue actualizado para {bandeja}/{folio_objetivo}.")

        output_raw = str(
            resultado.get("output_dir") or resultado.get("sin_operador_dir") or ""
        ).strip()
        if not output_raw:
            errores.append(f"Sin carpeta output para {bandeja}/{folio_objetivo}.")
            continue
        output_path = Path(output_raw)
        if not output_path.is_absolute():
            output_path = project_dir / output_path
        output_path = output_path.resolve()
        if not output_path.is_dir():
            errores.append(
                f"La carpeta output no existe para {bandeja}/{folio_objetivo}: {output_path}"
            )
        else:
            salidas.append(str(output_path))

    try:
        folios_excel, _info_excel = cargar_folios_internos_procesados_excel(
            excel_path,
            sheet,
            header_folio,
        )
        if folio not in folios_excel:
            errores.append(
                f"El Folio {folio} no quedó escrito en la hoja {sheet!r} del Excel."
            )
    except Exception as exc:
        errores.append(f"No se pudo verificar el Excel final: {exc}")

    return {
        "ok": not errores,
        "folio": folio,
        "objetivos_esperados": len(objetivos),
        "resultados_en_log": len(resultados),
        "procesamiento_log": str(procesamiento_log),
        "excel": str(excel_path),
        "output_dirs": list(dict.fromkeys(salidas)),
        "errores": errores,
    }


def configurar_modo_folio_internos(args: argparse.Namespace) -> None:
    """Normaliza el Folio y aplica las garantías del comando individual."""
    if not args.folio_internos:
        return
    args.folio_internos = normalizar_folio_interno(args.folio_internos)
    if not args.folio_internos:
        raise ValueError("--folio-internos debe contener entre 1 y 15 dígitos.")
    args.solo_internos = True
    args.sin_email = True


def clasificar_objetivos_internos(
    resumen_internos: dict,
    descargas_base: Path,
) -> tuple[list[dict], list[dict], list[dict]]:
    """Devuelve inventario, objetivos completos y pendientes por bandeja+folio."""
    inventario = construir_objetivos_internos(resumen_internos)
    completos: list[dict] = []
    pendientes: list[dict] = []
    for objetivo in inventario:
        destino = completos if objetivo_internos_esta_completo(
            descargas_base,
            objetivo["bandeja"],
            objetivo["folio"],
        ) else pendientes
        destino.append(objetivo)
    return inventario, completos, pendientes


def limitar_objetivos_internos(objetivos: Iterable[dict], limite: int) -> list[dict]:
    """Limita el lote y lo reparte entre bandejas para aprovechar paralelismo."""
    if limite < 0:
        raise ValueError("El limite de Folios Internos no puede ser negativo.")
    items = list(objetivos)
    if limite == 0 or limite >= len(items):
        return items

    orden_bandejas: list[str] = []
    por_bandeja: dict[str, list[dict]] = {}
    for item in items:
        bandeja = str(item.get("bandeja") or "").strip()
        if bandeja not in por_bandeja:
            orden_bandejas.append(bandeja)
            por_bandeja[bandeja] = []
        por_bandeja[bandeja].append(item)

    seleccionados: list[dict] = []
    indice = 0
    while len(seleccionados) < limite:
        agregados = 0
        for bandeja in orden_bandejas:
            cola = por_bandeja[bandeja]
            if indice < len(cola):
                seleccionados.append(cola[indice])
                agregados += 1
                if len(seleccionados) == limite:
                    break
        if agregados == 0:
            break
        indice += 1
    return seleccionados


def guardar_objetivos_internos(path: Path, objetivos: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps({"objetivos": objetivos}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def ejecutar_comando(
    cmd: list[str],
    cwd: Path,
    log_path: Path,
    titulo: str,
    estado: EstadoEjecucion | None = None,
    etapa: str = "",
) -> int:
    """Ejecuta un comando mostrando/guardando salida y actualizando estado vivo."""
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8", errors="replace") as log_file:
        separador = "=" * 90
        log_file.write(f"\n{separador}\n{titulo}\n{datetime.now().isoformat()}\n")
        log_file.write("CMD: " + " ".join(f'\"{x}\"' if " " in str(x) else str(x) for x in cmd) + "\n")
        log_file.write(separador + "\n")
        log_file.flush()

        print(f"\n{separador}\n{titulo}\n{separador}")
        print("CMD:", " ".join(str(x) for x in cmd))

        if estado is not None:
            estado.actualizar(stage=etapa or titulo, comando=" ".join(str(x) for x in cmd))

        proc = subprocess.Popen(
            [str(x) for x in cmd],
            cwd=str(cwd),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
        )

        assert proc.stdout is not None
        ultima_actualizacion_estado = 0.0
        for line in proc.stdout:
            print(line, end="")
            log_file.write(line)
            log_file.flush()
            if estado is not None and time.time() - ultima_actualizacion_estado >= 5:
                estado.actualizar(stage=etapa or titulo, ultima_linea=line.strip()[:500])
                ultima_actualizacion_estado = time.time()

        rc = proc.wait()
        if estado is not None:
            estado.actualizar(stage=f"{etapa or titulo}: terminado", return_code=rc)
        log_file.write(f"\n[{titulo}] return_code={rc}\n")
        log_file.flush()
        print(f"\n[{titulo}] return_code={rc}")
        return rc


class ExtraccionSatysAgotada(RuntimeError):
    """Error con el historial de todos los intentos de extracción agotados."""

    def __init__(self, mensaje: str, historial: list[dict]):
        super().__init__(mensaje)
        self.historial = historial


def limpiar_salida_extraccion(output_path: Path) -> None:
    """Elimina la salida de un intento anterior para evitar aceptar datos obsoletos."""
    candidatos = (
        output_path,
        output_path.with_suffix(output_path.suffix + ".json"),
    )
    for candidato in candidatos:
        try:
            candidato.unlink()
        except FileNotFoundError:
            pass


def validar_resumen_extractor(output_path: Path, registros: list[str]) -> dict:
    """Valida el certificado JSON del extractor contra el TXT publicado."""
    resumen_path = output_path.with_suffix(output_path.suffix + ".json")
    resultado = {
        "ok": False,
        "path": str(resumen_path),
        "vacio_confirmado": False,
        "error": "",
        "resumen": None,
    }
    if not resumen_path.exists():
        resultado["error"] = "falta el resumen JSON del extractor"
        return resultado
    try:
        resumen = json.loads(resumen_path.read_text(encoding="utf-8-sig", errors="replace"))
    except Exception as exc:
        resultado["error"] = f"resumen JSON ilegible: {exc}"
        return resultado
    resultado["resumen"] = resumen
    if resumen.get("estado") != "COMPLETO" or resumen.get("integridad") != "VALIDADA":
        resultado["error"] = (
            f"estado/integridad no válidos: estado={resumen.get('estado')!r}, "
            f"integridad={resumen.get('integridad')!r}"
        )
        return resultado
    if int(resumen.get("total_registros", -1)) != len(registros):
        resultado["error"] = (
            f"TXT y resumen no coinciden: TXT={len(registros)}, "
            f"resumen={resumen.get('total_registros')!r}"
        )
        return resultado

    estados_validos = {"ENCONTRADOS_COMPLETOS", "VACIO_CONFIRMADO"}
    oficialia_omitida = bool(
        resumen.get("oficialia_omitida")
        or resumen.get("modo") == "solo_internos"
    )
    if oficialia_omitida:
        if registros or resumen.get("registros") or int(resumen.get("total_filas_satys", -1)) != 0:
            resultado["error"] = "Oficialia omitida contiene registros o filas inesperadas"
            return resultado
        if resumen.get("por_anio") not in (None, []):
            resultado["error"] = "Oficialia omitida no debe contener detalle por año"
            return resultado
        if not bool(resumen.get("vacio_confirmado")):
            resultado["error"] = "Oficialia omitida debe declarar vacio_confirmado"
            return resultado
        todos_vacios = True
    else:
        por_anio = resumen.get("por_anio")
        if not isinstance(por_anio, list) or not por_anio:
            resultado["error"] = "el resumen no contiene detalle por año"
            return resultado
        total_filas = 0
        todos_vacios = True
        for item in por_anio:
            estado = item.get("estado")
            if estado not in estados_validos:
                resultado["error"] = f"año {item.get('anio')}: estado indeterminado {estado!r}"
                return resultado
            total = int(item.get("total_reportado_satys") or 0)
            filas = int(item.get("filas_leidas") or 0)
            guardados = int(item.get("total_guardados_anio") or 0)
            invalidas = int(item.get("filas_invalidas") or 0)
            contador_tab = item.get("contador_tab")
            tamanio_pagina = item.get("tamanio_pagina")
            if contador_tab is None or int(contador_tab) != total:
                resultado["error"] = (
                    f"año {item.get('anio')}: contador de pestaña inconsistente "
                    f"contador={contador_tab!r}, total={total}"
                )
                return resultado
            if tamanio_pagina is None or int(tamanio_pagina) != 100:
                resultado["error"] = (
                    f"año {item.get('anio')}: el selector Mostrar no quedó en 100 "
                    f"(valor={tamanio_pagina!r})"
                )
                return resultado
            if invalidas != 0:
                resultado["error"] = f"año {item.get('anio')}: {invalidas} fila(s) inválida(s)"
                return resultado
            if estado == "ENCONTRADOS_COMPLETOS":
                todos_vacios = False
                if total <= 0 or filas != total or guardados <= 0:
                    resultado["error"] = (
                        f"año {item.get('anio')}: conciliación incompleta "
                        f"total={total}, filas={filas}, guardados={guardados}"
                    )
                    return resultado
            elif total != 0 or filas != 0 or guardados != 0:
                resultado["error"] = (
                    f"año {item.get('anio')}: VACIO_CONFIRMADO inconsistente "
                    f"total={total}, filas={filas}, guardados={guardados}"
                )
                return resultado
            total_filas += filas

        if int(resumen.get("total_filas_satys", -1)) != total_filas:
            resultado["error"] = (
                f"total_filas_satys inconsistente: resumen={resumen.get('total_filas_satys')!r}, "
                f"suma={total_filas}"
            )
            return resultado
        vacio_declarado = bool(resumen.get("vacio_confirmado"))
        if vacio_declarado != todos_vacios:
            resultado["error"] = (
                f"bandera vacio_confirmado inconsistente: declarada={vacio_declarado}, calculada={todos_vacios}"
            )
            return resultado
        if todos_vacios and registros:
            resultado["error"] = "el resumen confirma vacío pero el TXT contiene registros"
            return resultado
        if not todos_vacios and not registros:
            resultado["error"] = "el resumen reporta filas pero el TXT quedó vacío"
            return resultado

    internos = resumen.get("internos")
    if not isinstance(internos, dict):
        resultado["error"] = "el resumen no contiene la extraccion de Internos IFT"
        return resultado
    if internos.get("estado") != "COMPLETO" or internos.get("integridad") != "VALIDADA":
        resultado["error"] = (
            f"Internos con estado/integridad no validos: estado={internos.get('estado')!r}, "
            f"integridad={internos.get('integridad')!r}"
        )
        return resultado

    por_bandeja = internos.get("por_bandeja")
    if not isinstance(por_bandeja, list) or len(por_bandeja) != 6:
        resultado["error"] = "Internos no contiene exactamente las seis bandejas esperadas"
        return resultado
    total_filas_internos = 0
    todos_vacios_internos = True
    for item in por_bandeja:
        estado_bandeja = item.get("estado")
        if estado_bandeja not in estados_validos:
            resultado["error"] = (
                f"Internos/{item.get('bandeja')}: estado indeterminado {estado_bandeja!r}"
            )
            return resultado
        total = int(item.get("total_reportado_satys") or 0)
        filas = int(item.get("filas_leidas") or 0)
        invalidas = int(item.get("filas_invalidas") or 0)
        folios = unicos_folios_internos(item.get("folios") or [])
        if invalidas:
            resultado["error"] = f"Internos/{item.get('bandeja')}: {invalidas} fila(s) invalidas"
            return resultado
        if estado_bandeja == "ENCONTRADOS_COMPLETOS":
            todos_vacios_internos = False
            if total <= 0 or filas != total or not folios:
                resultado["error"] = (
                    f"Internos/{item.get('bandeja')}: conciliacion incompleta "
                    f"total={total}, filas={filas}, folios={len(folios)}"
                )
                return resultado
        elif total != 0 or filas != 0 or folios:
            resultado["error"] = (
                f"Internos/{item.get('bandeja')}: VACIO_CONFIRMADO inconsistente"
            )
            return resultado
        total_filas_internos += filas

    folios_internos = unicos_folios_internos(internos.get("folios") or [])
    if int(internos.get("total_folios", -1)) != len(folios_internos):
        resultado["error"] = "total_folios de Internos no coincide con la lista deduplicada"
        return resultado
    if int(internos.get("total_filas_satys", -1)) != total_filas_internos:
        resultado["error"] = "total_filas_satys de Internos no coincide con sus bandejas"
        return resultado
    if bool(internos.get("vacio_confirmado")) != todos_vacios_internos:
        resultado["error"] = "bandera vacio_confirmado de Internos inconsistente"
        return resultado

    resultado["ok"] = True
    resultado["vacio_confirmado"] = todos_vacios
    resultado["folios_internos"] = folios_internos
    return resultado


def extraer_registros_satys_con_reintentos(
    *,
    cmd_extraer: list[str],
    output_path: Path,
    cwd: Path,
    log_path: Path,
    estado: EstadoEjecucion | None,
    reintentos: int,
    espera_segundos: int,
) -> tuple[list[str], list[dict]]:
    """
    Ejecuta la extracción hasta obtener un resultado certificado y conciliado.

    Cada intento lanza un proceso nuevo de extraer_registros_documentos.py. Ese
    proceso abre su propio navegador y siempre lo cierra en su bloque ``finally``;
    por ello cada reintento empieza con un navegador limpio.

    Se reintenta cuando:
      * el extractor termina con código distinto de cero;
      * falta el resumen JSON de integridad; o
      * TXT y resumen no concilian.

    Un TXT vacío solo es válido si el resumen declara VACIO_CONFIRMADO para todos
    los años. ``reintentos=2`` significa un máximo de tres intentos totales. Una
    vez que el resultado queda certificado, el flujo regresa al procesamiento sin alterar
    Parte 1, Parte 2, Parte 3, Parte 4 ni sus reintentos por Registro.
    """
    reintentos = max(0, int(reintentos))
    espera_segundos = max(0, int(espera_segundos))
    total_intentos = reintentos + 1
    historial: list[dict] = []

    for numero_intento in range(1, total_intentos + 1):
        limpiar_salida_extraccion(output_path)
        titulo = (
            "1) EXTRAER REGISTROS DESDE SATyS "
            f"(intento {numero_intento}/{total_intentos})"
        )
        inicio = datetime.now().isoformat()
        if estado is not None:
            estado.actualizar(
                stage="extrayendo_registros_satys",
                intento_extraccion=numero_intento,
                total_intentos_extraccion=total_intentos,
            )

        rc = ejecutar_comando(
            cmd_extraer,
            cwd,
            log_path,
            titulo,
            estado=estado,
            etapa="extrayendo_registros_satys",
        )
        registros = leer_registros_txt(output_path) if rc == 0 else []
        validacion = validar_resumen_extractor(output_path, registros) if rc == 0 else {
            "ok": False, "vacio_confirmado": False, "error": "el extractor terminó con error"
        }
        resumen_validado = validacion.get("resumen") or {}
        solo_internos_validado = bool(
            resumen_validado.get("oficialia_omitida")
            or resumen_validado.get("modo") == "solo_internos"
        )
        folios_internos_validados = validacion.get("folios_internos") or []
        motivo = "ok"
        if rc != 0:
            motivo = f"extractor terminó con código {rc}"
        elif not validacion.get("ok"):
            motivo = f"certificado de integridad inválido: {validacion.get('error')}"
        elif solo_internos_validado:
            motivo = "INTERNOS_VALIDADO"
        elif validacion.get("vacio_confirmado"):
            motivo = "VACIO_CONFIRMADO"

        detalle = {
            "intento": numero_intento,
            "total_intentos": total_intentos,
            "fecha_inicio": inicio,
            "fecha_fin": datetime.now().isoformat(),
            "return_code": rc,
            "total_registros": len(registros),
            "total_folios_internos": len(folios_internos_validados),
            "vacio_confirmado": bool(validacion.get("vacio_confirmado")),
            "integridad_ok": bool(validacion.get("ok")),
            "error_integridad": validacion.get("error", ""),
            "resultado": motivo,
        }
        historial.append(detalle)

        if rc == 0 and validacion.get("ok"):
            if solo_internos_validado:
                print(
                    f"✅ Extracción Internos válida en intento {numero_intento}/{total_intentos}: "
                    f"{len(folios_internos_validados)} Folio(s) en seis bandejas, integridad conciliada."
                )
            elif registros:
                print(
                    f"✅ Extracción SATyS válida en intento {numero_intento}/{total_intentos}: "
                    f"{len(registros)} Registro(s), integridad conciliada."
                )
            else:
                print(
                    f"✅ Extracción SATyS válida en intento {numero_intento}/{total_intentos}: "
                    "cero Registros confirmado por DataTables y certificado JSON."
                )
            return registros, historial

        aviso = (
            f"⚠️  Extracción SATyS no válida en intento {numero_intento}/{total_intentos}: "
            f"{motivo}."
        )
        print(aviso)
        with log_path.open("a", encoding="utf-8", errors="replace") as log_file:
            log_file.write("\n" + aviso + "\n")

        if numero_intento < total_intentos:
            if estado is not None:
                estado.actualizar(
                    stage="esperando_reintento_extraccion_satys",
                    intento_extraccion=numero_intento,
                    siguiente_intento=numero_intento + 1,
                    espera_segundos=espera_segundos,
                    motivo_reintento=motivo,
                )
            if espera_segundos:
                print(
                    f"   El navegador del intento anterior ya terminó. "
                    f"Nuevo intento en {espera_segundos} segundo(s)..."
                )
                time.sleep(espera_segundos)
            else:
                print(
                    "   El navegador del intento anterior ya terminó. "
                    "Iniciando el siguiente intento de inmediato..."
                )

    ultimo = historial[-1] if historial else {}
    raise ExtraccionSatysAgotada(
        "No fue posible obtener Registros desde SATyS después de "
        f"{total_intentos} intento(s). Último resultado: "
        f"{ultimo.get('resultado', 'sin resultado')}; "
        f"return_code={ultimo.get('return_code', 'N/D')}; "
        f"registros={ultimo.get('total_registros', 0)}.",
        historial,
    )


def ps_quote(text: str) -> str:
    return "'" + text.replace("'", "''") + "'"


def notificar_windows(titulo: str, mensaje: str, habilitado: bool = True) -> bool:
    """Notificación tipo globo en Windows, sin dependencias externas. No bloquea el proceso."""
    if not habilitado or os.name != "nt":
        return False
    try:
        mensaje = mensaje[:250]
        script = f"""
Add-Type -AssemblyName System.Windows.Forms
Add-Type -AssemblyName System.Drawing
$n = New-Object System.Windows.Forms.NotifyIcon
$n.Icon = [System.Drawing.SystemIcons]::Information
$n.BalloonTipTitle = {ps_quote(titulo)}
$n.BalloonTipText = {ps_quote(mensaje)}
$n.Visible = $true
$n.ShowBalloonTip(10000)
Start-Sleep -Seconds 11
$n.Dispose()
""".strip()
        subprocess.Popen(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-WindowStyle", "Hidden", "-Command", script],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return True
    except Exception:
        return False




def sincronizar_estado_diario_depi() -> None:
    """Sincroniza únicamente TrámitesCRT.xlsx, output/ y descargas/ hacia DEPI."""
    resultado = sincronizar_salidas(
        PROJECT_DIR,
        carpeta_compartida(),
    )
    for error in resultado.errores:
        print(f"⚠️  Sincronización diaria DEPI: {error}")


def cargar_resultados_procesamiento(
    path: Path,
    *,
    origen: str,
    mtime_minimo: float | None = None,
) -> list[dict]:
    """Lee sólo el log producido por el subproceso de esta corrida diaria."""
    try:
        if not path.is_file():
            return []
        if mtime_minimo is not None and path.stat().st_mtime < (mtime_minimo - 2.0):
            return []
        payload = json.loads(path.read_text(encoding="utf-8-sig", errors="replace"))
        resultados = payload.get("resultados") or []
        if not isinstance(resultados, list):
            return []
        salida: list[dict] = []
        for item in resultados:
            if isinstance(item, dict):
                salida.append({**item, "_origen_proceso": origen})
        return salida
    except Exception as exc:
        print(f"⚠️  No se pudo incorporar {path.name} al correo consolidado: {exc}")
        return []


def agregar_fallidos_controlados(
    resultados: list[dict],
    fallidos: Iterable[str],
    *,
    origen: str = "oficialia",
) -> list[dict]:
    """Agrega al resumen los agotados por watchdog que aún no estén en el log."""
    salida = list(resultados)
    existentes = {
        str(
            item.get("folio_tabla_internos")
            or item.get("registro")
            or item.get("folio")
            or item.get("folio_id")
            or ""
        ).strip().upper()
        for item in salida
    }
    for valor in fallidos:
        identificador = str(valor or "").strip()
        if not identificador or identificador.upper() in existentes:
            continue
        salida.append({
            "registro": identificador,
            "rpc_ok": False,
            "organizado_ok": False,
            "excel_ok": False,
            "_fallido_controlado": True,
            "_origen_proceso": origen,
        })
        existentes.add(identificador.upper())
    return salida


def enviar_resumen_email_diario(
    *,
    resultados: list[dict],
    log_path: Path,
    excel_path: Path,
    modo: str = "CORRIDA DIARIA CONSOLIDADA",
    error_general: str = "",
) -> bool:
    """Emite la única notificación por correo autorizada de la corrida diaria."""
    if not _EMAIL_DISPONIBLE:
        print("⚠️  Módulo de correo no disponible; no se envió el resumen diario.")
        return False

    descargas_base = ruta_configurada("descargas", "descargas")
    output_base = ruta_configurada("output", "output")
    conteos = _email_mod.conteos_desde_resultados(resultados)
    if error_general:
        conteos["errores"] = max(1, conteos["errores"])
        conteos["total"] = max(1, conteos["total"])

    outputs = {
        "Carpeta output": str(output_base.resolve()),
        "Carpeta descargas": str(descargas_base.resolve()),
        "TrámitesCRT.xlsx": str(excel_path.resolve()),
        "Folios_Datos_Completos.xlsx": str((output_base / "Folios_Datos_Completos.xlsx").resolve()),
        "Folios_Datos_Completos_Internos.xlsx": str((output_base / "Folios_Datos_Completos_Internos.xlsx").resolve()),
    }
    try:
        return bool(_email_mod.enviar_notificacion(
            total_registros=conteos["total"],
            exitosos=conteos["exitosos"],
            sin_operador=conteos["sin_operador"],
            errores=conteos["errores"],
            registros=resultados,
            fecha_ejecucion=datetime.now().isoformat(),
            modo=modo if not error_general else f"ERROR — {modo}",
            outputs=outputs,
            log_path=str(log_path),
            project_root=PROJECT_DIR,
            descargas_base=descargas_base,
            output_base=output_base,
            excel_path=excel_path,
            excel_metadata_path=output_base / "Folios_Datos_Completos.xlsx",
            carpeta_compartida=carpeta_compartida(),
        ))
    except Exception as exc:
        print(f"⚠️  Error no crítico al construir/enviar el correo diario: {exc}")
        return False


def consolidar_revision_manual_final() -> dict:
    """Aplica al final de la corrida la ruta única de todos los sin_operador."""
    from Parte4_excel import consolidar_sin_operador_legacy

    resumen = consolidar_sin_operador_legacy(ruta_configurada("output", "output"))
    if resumen["errores"]:
        for error in resumen["errores"]:
            print(f"⚠️  Consolidación final sin_operador: {error}")
    else:
        print(
            "✅ Revisión manual consolidada en _sin_operador/(correos): "
            f"{resumen['carpetas_migradas']} carpeta(s) heredada(s) migrada(s)."
        )
    return resumen


def ejecutar_reconciliacion_global(
    *,
    python_exe: Path,
    script: Path,
    excel: Path,
    log_path: Path,
    estado: EstadoEjecucion,
    sin_backup: bool = False,
) -> int:
    """Reconcilia siempre el Excel maestro desde todos los metadata JSON."""
    cmd = [
        str(python_exe),
        str(script),
        "--excel", str(excel),
        "--resumen-json", str(LOG_DIR_DEFAULT / "reconciliacion_global_ultimo.json"),
    ]
    if sin_backup:
        cmd.append("--sin-backup")
    return ejecutar_comando(
        cmd,
        PROJECT_DIR,
        log_path,
        "4) RECONCILIAR TRÁMITESCRT DESDE TODOS LOS METADATA",
        estado=estado,
        etapa="reconciliando_excel_global",
    )

def construir_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Monitor diario: extrae Registros SATyS, compara contra TrámitesCRT.xlsx y procesa solo nuevos."
    )
    parser.add_argument("--python", dest="python_exe", type=Path, default=PYTHON_EXE_DEFAULT,
                        help="Ruta al intérprete Python. Default: sys.executable o SATYS_PYTHON.")
    parser.add_argument("--extraer-script", type=Path, default=EXTRAER_SCRIPT_DEFAULT,
                        help="Ruta a extraer_registros_documentos.py.")
    parser.add_argument("--main-script", type=Path, default=MAIN_SCRIPT_DEFAULT,
                        help="Ruta a main_procesar.py.")
    parser.add_argument("--reconciliar-script", type=Path, default=RECONCILIAR_SCRIPT_DEFAULT,
                        help="Ruta a reconciliar_metadata_global.py.")
    parser.add_argument("--sin-reconciliacion-global", action="store_true",
                        help="Desactiva la reconciliación completa de TrámitesCRT.xlsx; solo diagnóstico.")
    parser.add_argument("--solo-internos", action="store_true",
                        help="Omite Oficialia; inventaria las seis bandejas de Internos y procesa solo Folios nuevos.")
    parser.add_argument("--folio-internos", default="", metavar="FOLIO",
                        help=(
                            "Procesa de principio a fin un único Folio numérico de Internos. "
                            "Implica --solo-internos y --sin-email, y conserva todas sus "
                            "apariciones entre las seis bandejas."
                        ))
    parser.add_argument("--excel", type=Path, default=EXCEL_DEFAULT,
                        help="Ruta a TrámitesCRT.xlsx.")
    parser.add_argument("--sheet", default=SHEET_DEFAULT,
                        help="Hoja del Excel donde se encuentra la columna 1711.")
    parser.add_argument("--header-registro", default=HEADER_REGISTRO_DEFAULT,
                        help="Encabezado de la columna de registros ya procesados.")
    parser.add_argument("--sheet-internos", default=SHEET_INTERNOS_DEFAULT,
                        help="Hoja donde se controlan los Folios de Internos IFT.")
    parser.add_argument("--header-folio-internos", default=HEADER_FOLIO_INTERNOS_DEFAULT,
                        help="Encabezado de la columna numerica usada para Folios de Internos IFT.")
    parser.add_argument("--registros-latest", type=Path, default=REGISTROS_LATEST_DEFAULT,
                        help="TXT que consumirá main_procesar.py. Default: registros.txt")
    parser.add_argument("--internos-latest", type=Path, default=INTERNOS_LATEST_DEFAULT,
                        help="JSON de objetivos bandeja+folio pendientes que consumira main_procesar.py.")
    parser.add_argument("--registros-dir", type=Path, default=REGISTROS_DIR_DEFAULT,
                        help="Carpeta donde se guardan copias históricas de TXT.")
    parser.add_argument("--logs-dir", type=Path, default=LOG_DIR_DEFAULT,
                        help="Carpeta donde se guardan logs y resúmenes.")
    parser.add_argument("--workers", type=int, default=WORKERS_DEFAULT,
                        help="Workers iniciales de Playwright. Configuración local: 10.")
    parser.add_argument("--internos-workers", type=int, default=INTERNOS_WORKERS_DEFAULT,
                        help="Navegadores paralelos para Internos IFT. Default: 12; acepta cualquier entero positivo.")
    parser.add_argument("--max-folios-internos", type=int, default=0,
                        help="Tamaño máximo del lote de Internos. 0 procesa todos; para prueba local usa 4.")
    parser.add_argument("--timeout-registro", type=int, default=TIMEOUT_REGISTRO_DEFAULT,
                        help="Timeout duro por Registro en segundos. Si un Registro se traba, se mata su proceso hijo y sigue el lote.")
    parser.add_argument("--reintentos-registro", type=int, default=REINTENTOS_REGISTRO_DEFAULT,
                        help="Reintentos automáticos solo para registros incompletos. 2 = hasta 3 intentos totales.")
    parser.add_argument("--workers-reintento", type=int, default=WORKERS_REINTENTO_DEFAULT,
                        help="Workers usados en reintentos de registros fallidos/incompletos. Default: 2.")
    parser.add_argument("--reintentos-extraccion", type=int, default=REINTENTOS_EXTRACCION_DEFAULT,
                        help="Reintentos adicionales solo para extraer la lista inicial desde SATyS. 2 = hasta 3 intentos totales.")
    parser.add_argument("--espera-reintento-extraccion", type=int, default=ESPERA_REINTENTO_EXTRACCION_DEFAULT,
                        help="Segundos de espera entre intentos de extracción SATyS. Default: 0 (reintento inmediato).")
    parser.add_argument("--headless", action="store_true", default=True,
                        help="Ejecuta Playwright sin navegador visible. Default: activo.")
    parser.add_argument("--visible", action="store_true",
                        help="Fuerza navegador visible para depuración; desactiva --headless.")
    parser.add_argument("--max-paginas", type=int, default=100,
                        help="Máximo de páginas DataTables al extraer registros.")
    parser.add_argument("--timeout-tabla", type=int, default=TIMEOUT_TABLA_DEFAULT,
                        help="Tiempo máximo por año/página para esperar al menos un Registro. Default: 120 segundos.")
    parser.add_argument("--intentos-anio-extraccion", type=int, default=INTENTOS_ANIO_EXTRACCION_DEFAULT,
                        help="Intentos totales por año antes de fallar. Default: 3.")
    parser.add_argument("--intentos-pagina-extraccion", type=int, default=INTENTOS_PAGINA_EXTRACCION_DEFAULT,
                        help="Intentos para confirmar el avance de cada página. Default: 3.")
    parser.add_argument("--no-procesar", action="store_true",
                        help="Solo genera TXT de nuevos registros; no ejecuta main_procesar.py.")
    parser.add_argument("--sin-notificacion", action="store_true",
                        help="No intenta mostrar notificación de Windows.")
    parser.add_argument("--sin-email", action="store_true",
                        help="No envía correo al finalizar; útil para pruebas o despliegue inicial.")
    parser.add_argument("--estado-json", type=Path, default=LOG_DIR_DEFAULT / "estado_actual.json",
                        help="Archivo JSON de estado vivo para dashboard/monitoreo.")
    return parser


def main() -> int:
    args = construir_parser().parse_args()
    try:
        # Este modo existe para una revisión manual individual. Nunca debe
        # generar correo, incluso si quien lo invoca omite --sin-email.
        configurar_modo_folio_internos(args)
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc
    if args.internos_workers < 1:
        raise SystemExit("--internos-workers debe ser un entero positivo.")
    if args.max_folios_internos < 0:
        raise SystemExit("--max-folios-internos debe ser 0 o un entero positivo.")
    # La producción diaria siempre aplica el segundo nivel de resolución:
    # Excel RPC y, si no basta, el buscador público actual del propio RPC.
    os.environ["SATYS_RPC_CONSULTA_ONLINE"] = "1"
    headless = bool(args.headless and not args.visible)
    reconciliacion_habilitada = bool(
        not args.sin_reconciliacion_global and not args.solo_internos
    )

    fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
    args.registros_dir.mkdir(parents=True, exist_ok=True)
    args.logs_dir.mkdir(parents=True, exist_ok=True)

    log_path = args.logs_dir / f"monitor_registros_{fecha}.log"
    resumen_json = args.logs_dir / f"monitor_registros_{fecha}.json"
    resumen_latest = args.logs_dir / "monitor_registros_ultimo.json"
    estado = EstadoEjecucion(args.estado_json, proceso="automatizar_registros_diario.py")

    registros_satys_hist = args.registros_dir / f"registros_satys_{fecha}.txt"
    registros_nuevos_hist = args.registros_dir / f"registros_nuevos_{fecha}.txt"
    internos_nuevos_hist = args.registros_dir / f"folios_internos_nuevos_{fecha}.json"

    resumen: dict = {
        "fecha_ejecucion": datetime.now().isoformat(),
        "modo": (
            "folio_internos"
            if args.folio_internos
            else "solo_internos"
            if args.solo_internos
            else "diario_completo"
        ),
        "folio_internos_objetivo": args.folio_internos or None,
        "reconciliacion_global_habilitada": reconciliacion_habilitada,
        "rpc_online_alternativo": True,
        "headless": headless,
        "workers": args.workers,
        "internos_workers": args.internos_workers,
        "max_folios_internos": args.max_folios_internos,
        "timeout_registro_segundos": args.timeout_registro,
        "reintentos_registro": args.reintentos_registro,
        "workers_reintento": args.workers_reintento,
        "reintentos_extraccion": args.reintentos_extraccion,
        "espera_reintento_extraccion_segundos": args.espera_reintento_extraccion,
        "timeout_tabla_segundos": args.timeout_tabla,
        "intentos_anio_extraccion": args.intentos_anio_extraccion,
        "intentos_pagina_extraccion": args.intentos_pagina_extraccion,
        "paths": {
            "project_dir": str(PROJECT_DIR),
            "excel": str(args.excel),
            "registros_latest": str(args.registros_latest),
            "registros_satys_hist": str(registros_satys_hist),
            "registros_nuevos_hist": str(registros_nuevos_hist),
            "internos_latest": str(args.internos_latest),
            "internos_nuevos_hist": str(internos_nuevos_hist),
            "log": str(log_path),
        },
        "ok": False,
        "errores": [],
    }
    correo_diario_intentado = False

    estado.actualizar(
        running=True,
        stage="inicializando",
        log=str(log_path),
        resumen=str(resumen_json),
        resumen_latest=str(resumen_latest),
        workers=args.workers,
        timeout_registro_segundos=args.timeout_registro,
    )

    # ──── Bloqueo compartido: evita que 2+ laptops corran el monitor a la vez ────
    # Cubre también a extraer_registros_documentos.py y main_procesar.py, que se
    # lanzan como subprocesos y heredan este mismo bloqueo automáticamente.
    lock = ProcesoLock(proceso="automatizar_registros_diario.py")
    try:
        estado.actualizar(stage="tomando_lock")
        lock.adquirir()

        def _salir_limpiamente(signum, frame):
            try:
                lock.liberar()
            finally:
                estado.finalizar(ok=False, mensaje=f"Proceso detenido por señal {signum}")
                raise SystemExit(128 + signum)

        try:
            signal.signal(signal.SIGTERM, _salir_limpiamente)
            signal.signal(signal.SIGINT, _salir_limpiamente)
        except Exception:
            pass

        estado.actualizar(stage="lock_adquirido")
    except LockOcupadoError as exc:
        # No es un error real del monitor: simplemente otra laptop ya está
        # trabajando. Se omite esta corrida sin marcarla como fallo.
        resumen["ok"] = True
        resumen["omitido_por_bloqueo"] = True
        resumen["mensaje"] = f"Se omitió esta corrida: {exc}"
        resumen["fecha_fin"] = datetime.now().isoformat()
        print(f"🔒 {exc}")
        print("   Se omite esta corrida del monitor diario en esta laptop.")
        try:
            resumen_json.write_text(json.dumps(resumen, ensure_ascii=False, indent=2), encoding="utf-8")
            resumen_latest.write_text(json.dumps(resumen, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass
        notificar_windows(
            "SATyS CRT — corrida diaria omitida",
            str(exc),
            habilitado=not args.sin_notificacion,
        )
        estado.finalizar(ok=True, mensaje=resumen["mensaje"], omitido_por_bloqueo=True)
        return 0

    try:
        if not args.python_exe.exists():
            raise FileNotFoundError(f"No existe python.exe: {args.python_exe}")
        if not args.extraer_script.exists():
            raise FileNotFoundError(f"No existe script extractor: {args.extraer_script}")
        if not args.main_script.exists():
            raise FileNotFoundError(f"No existe main_procesar.py: {args.main_script}")

        # 1) Extraer todos los registros visibles desde SATyS.
        cmd_extraer = [
            str(args.python_exe),
            str(args.extraer_script),
            "--output", str(registros_satys_hist),
            "--separador", "linea",
            "--max-paginas", str(args.max_paginas),
            "--timeout-tabla", str(args.timeout_tabla),
            "--intentos-anio", str(args.intentos_anio_extraccion),
            "--intentos-pagina", str(args.intentos_pagina_extraccion),
            "--modo-anios", "todos",
        ]
        if args.solo_internos:
            cmd_extraer.append("--solo-internos")
        cmd_extraer.append("--headless" if headless else "--visible")
        try:
            registros_satys, intentos_extraccion = extraer_registros_satys_con_reintentos(
                cmd_extraer=cmd_extraer,
                output_path=registros_satys_hist,
                cwd=PROJECT_DIR,
                log_path=log_path,
                estado=estado,
                reintentos=args.reintentos_extraccion,
                espera_segundos=args.espera_reintento_extraccion,
            )
        except ExtraccionSatysAgotada as exc:
            resumen["intentos_extraccion_satys"] = exc.historial
            resumen["total_intentos_extraccion_satys"] = len(exc.historial)
            if exc.historial:
                resumen["return_code_extraer"] = exc.historial[-1]["return_code"]
            raise
        resumen["intentos_extraccion_satys"] = intentos_extraccion
        resumen["total_intentos_extraccion_satys"] = len(intentos_extraccion)
        resumen["return_code_extraer"] = intentos_extraccion[-1]["return_code"]
        resumen["total_registros_satys"] = len(registros_satys)
        resumen["primeros_registros_satys"] = registros_satys[:15]
        resumen_extraer_path = registros_satys_hist.with_suffix(registros_satys_hist.suffix + ".json")
        if resumen_extraer_path.exists():
            try:
                resumen["extraccion_satys"] = json.loads(
                    resumen_extraer_path.read_text(encoding="utf-8-sig", errors="replace")
                )
            except Exception as exc:
                resumen["errores"].append(f"No se pudo leer resumen del extractor: {exc}")
        vacio_confirmado_satys = bool(
            isinstance(resumen.get("extraccion_satys"), dict)
            and resumen["extraccion_satys"].get("estado") == "COMPLETO"
            and resumen["extraccion_satys"].get("integridad") == "VALIDADA"
            and resumen["extraccion_satys"].get("vacio_confirmado")
        )
        resumen["vacio_confirmado_satys"] = vacio_confirmado_satys
        if not registros_satys and not vacio_confirmado_satys:
            raise RuntimeError(
                "El TXT quedó vacío sin una confirmación auditable de SATyS. "
                "Se trata como error indeterminado, no como cero registros."
            )

        extraccion_satys = resumen.get("extraccion_satys") or {}
        resumen_internos = extraccion_satys.get("internos") or {}
        folios_internos_satys = unicos_folios_internos(resumen_internos.get("folios") or [])
        resumen["total_folios_internos_satys"] = len(folios_internos_satys)
        resumen["primeros_folios_internos_satys"] = folios_internos_satys[:15]

        # 2) Leer evidencia Excel, columna 1711.
        estado.actualizar(stage="leyendo_excel_control", excel=str(args.excel))
        if args.solo_internos:
            procesados_excel = set()
            excel_info = {
                "omitido": True,
                "motivo": "modo solo_internos",
                "sheet": args.sheet,
            }
        else:
            procesados_excel, excel_info = cargar_registros_procesados_excel(
                args.excel, args.sheet, args.header_registro
            )
        resumen["excel_info"] = excel_info
        procesados_internos_excel, excel_internos_info = cargar_folios_internos_procesados_excel(
            args.excel,
            args.sheet_internos,
            args.header_folio_internos,
        )
        resumen["excel_internos_info"] = excel_internos_info

        estado.actualizar(
            stage="comparando_registros",
            total_registros_satys=len(registros_satys),
            total_procesados_excel=len(procesados_excel),
        )
        # 3) Comparar y guardar nuevos.
        nuevos_excel = [registro for registro in registros_satys if registro not in procesados_excel]
        folios_internos_nuevos_excel = [
            folio for folio in folios_internos_satys
            if folio not in procesados_internos_excel
        ]

        # Excel sólo demuestra que un folio fue procesado alguna vez; no demuestra
        # que se descargaron sus archivos en todas las bandejas donde aparece.
        # Por eso Internos se concilia por la pareja (bandeja, folio) y contra la
        # evidencia física de metadata_completo.json + todos sus archivos.
        DESCARGA_BASE_DIARIO = ruta_configurada("descargas", "descargas")
        (
            objetivos_internos_inventario,
            objetivos_internos_completos,
            objetivos_internos_descarga_pendiente,
        ) = clasificar_objetivos_internos(
            resumen_internos,
            DESCARGA_BASE_DIARIO,
        )
        objetivos_internos_sin_excel = construir_objetivos_internos(
            resumen_internos,
            folios_internos_nuevos_excel,
        )
        if args.folio_internos:
            # Una revisión individual es intencionalmente forzada: vuelve a
            # recorrer el detalle y los anexos aunque el Folio ya exista en
            # Excel o conserve una descarga previa completa.
            objetivos_internos_detectados = seleccionar_objetivos_folio_internos(
                resumen_internos,
                args.folio_internos,
            )
            objetivos_internos = list(objetivos_internos_detectados)
            folios_internos_nuevos = [args.folio_internos]
        else:
            objetivos_internos_detectados = []
            claves_objetivos_internos: set[tuple[str, str]] = set()
            for objetivo in objetivos_internos_descarga_pendiente + objetivos_internos_sin_excel:
                clave = (objetivo["bandeja"].casefold(), objetivo["folio"])
                if clave not in claves_objetivos_internos:
                    claves_objetivos_internos.add(clave)
                    objetivos_internos_detectados.append(objetivo)
            folios_internos_nuevos = unicos_folios_internos(
                item["folio"] for item in objetivos_internos_detectados
            )
            objetivos_internos = limitar_objetivos_internos(
                objetivos_internos_detectados,
                args.max_folios_internos,
            )
        guardar_objetivos_internos(internos_nuevos_hist, objetivos_internos)
        guardar_objetivos_internos(args.internos_latest, objetivos_internos)

        # 3b) También incluir registros que ya están en el Excel pero tienen
        #     carpeta sin archivos reales (solo los JSONs generados por el programa).
        #     Esto garantiza que los re-intentos ocurran en cada corrida diaria.
        # Registros en Excel pero sin una descarga real válida. La decisión usa
        # la misma regla central que Parte 1 y main_procesar.py: los JSON de
        # metadata, archivos vacíos, temporales y auxiliares nunca cuentan.
        incompletos_en_excel = [] if args.solo_internos else [
            reg for reg in registros_satys
            if reg in procesados_excel
            and not registro_esta_completo(DESCARGA_BASE_DIARIO, reg)
        ]

        nuevos = unicos_preservando_orden(nuevos_excel + incompletos_en_excel)

        if incompletos_en_excel:
            print(f"⚠️  Registros en Excel pero con descarga incompleta (se re-intentarán): {len(incompletos_en_excel)}")
            print("   ", ", ".join(incompletos_en_excel[:30]) + ("..." if len(incompletos_en_excel) > 30 else ""))

        guardar_txt_lineas(registros_nuevos_hist, nuevos)
        guardar_txt_lineas(args.registros_latest, nuevos)

        validacion_txt = validar_txt_un_registro_por_linea(args.registros_latest, len(nuevos))
        resumen["validacion_registros_txt"] = validacion_txt
        if nuevos and not validacion_txt.get("ok"):
            raise RuntimeError(
                "El TXT de registros nuevos no quedó en formato de un Registro por línea: "
                f"{validacion_txt}"
            )

        resumen["total_procesados_excel"] = len(procesados_excel)
        resumen["total_nuevos"] = len(nuevos)
        resumen["total_nuevos_excel"] = len(nuevos_excel)
        resumen["total_incompletos_reintento"] = len(incompletos_en_excel)
        resumen["registros_nuevos"] = nuevos
        resumen["total_folios_internos_procesados_excel"] = len(procesados_internos_excel)
        resumen["total_folios_internos_nuevos_excel"] = len(folios_internos_nuevos_excel)
        resumen["total_folios_internos_nuevos"] = len(folios_internos_nuevos)
        resumen["total_objetivos_internos_satys"] = len(objetivos_internos_inventario)
        resumen["total_objetivos_internos_completos_local"] = len(objetivos_internos_completos)
        resumen["total_objetivos_internos_descarga_pendiente"] = len(
            objetivos_internos_descarga_pendiente
        )
        resumen["total_objetivos_internos_sin_excel"] = len(objetivos_internos_sin_excel)
        resumen["total_objetivos_internos_pendientes"] = len(objetivos_internos_detectados)
        resumen["total_objetivos_internos_seleccionados"] = len(objetivos_internos)
        resumen["folios_internos_nuevos"] = folios_internos_nuevos
        resumen["objetivos_internos_pendientes"] = objetivos_internos_detectados
        resumen["objetivos_internos_nuevos"] = objetivos_internos
        estado.actualizar(
            stage="comparacion_lista",
            total_registros_satys=len(registros_satys),
            total_procesados_excel=len(procesados_excel),
            total_nuevos=len(nuevos),
            total_nuevos_excel=len(nuevos_excel),
            total_incompletos_reintento=len(incompletos_en_excel),
            registros_nuevos_preview=nuevos[:30],
            total_folios_internos_satys=len(folios_internos_satys),
            total_folios_internos_nuevos=len(folios_internos_nuevos),
            total_objetivos_internos_satys=len(objetivos_internos_inventario),
            total_objetivos_internos_completos_local=len(objetivos_internos_completos),
            total_objetivos_internos_descarga_pendiente=len(
                objetivos_internos_descarga_pendiente
            ),
            total_objetivos_internos_sin_excel=len(objetivos_internos_sin_excel),
            total_objetivos_internos_pendientes=len(objetivos_internos_detectados),
        )

        print("\n" + "=" * 90)
        print("RESULTADO DE COMPARACIÓN")
        print("=" * 90)
        print(f"Registros en SATyS:           {len(registros_satys)}")
        print(f"Ya procesados en Excel:        {len(procesados_excel)}")
        print(f"Nuevos (no en Excel):          {len(nuevos_excel)}")
        print(f"Re-intentos (descarga incompleta): {len(incompletos_en_excel)}")
        print(f"Total a procesar:              {len(nuevos)}")
        print(f"TXT nuevos para main:          {args.registros_latest}")
        print(f"Copia histórica nuevos:        {registros_nuevos_hist}")
        if nuevos:
            print("Nuevos:", ", ".join(nuevos[:50]) + ("..." if len(nuevos) > 50 else ""))
        print(f"Folios Internos en SATyS:      {len(folios_internos_satys)}")
        print(f"Folios Internos en Excel:      {len(procesados_internos_excel)}")
        print(f"Folios Internos no en Excel:   {len(folios_internos_nuevos_excel)}")
        print(f"Objetivos bandeja+folio SATyS: {len(objetivos_internos_inventario)}")
        print(f"Objetivos completos locales:   {len(objetivos_internos_completos)}")
        print(f"Descargas incompletas/faltantes: {len(objetivos_internos_descarga_pendiente)}")
        print(f"Objetivos sin fila en Excel:   {len(objetivos_internos_sin_excel)}")
        print(f"Objetivos pendientes (unión):  {len(objetivos_internos_detectados)}")
        print(f"Objetivos seleccionados lote:  {len(objetivos_internos)}")
        print(f"JSON Internos para main:       {args.internos_latest}")
        if args.folio_internos:
            ubicaciones = ", ".join(
                item["bandeja"] for item in objetivos_internos
            )
            print(f"Modo Folio único:              {args.folio_internos}")
            print(f"Bandeja(s) donde aparece:      {ubicaciones}")
            print("Correo electrónico:            DESHABILITADO obligatoriamente")
        if folios_internos_nuevos:
            print(
                "Folios Internos pendientes (únicos):",
                ", ".join(folios_internos_nuevos[:50])
                + ("..." if len(folios_internos_nuevos) > 50 else ""),
            )

        if not nuevos and not objetivos_internos:
            rc_reconciliacion = 0
            if reconciliacion_habilitada:
                rc_reconciliacion = ejecutar_reconciliacion_global(
                    python_exe=args.python_exe,
                    script=args.reconciliar_script,
                    excel=args.excel,
                    log_path=log_path,
                    estado=estado,
                )
            consolidacion_revision = consolidar_revision_manual_final()
            resumen["consolidacion_sin_operador"] = consolidacion_revision
            rc_consolidacion = 4 if consolidacion_revision["errores"] else 0
            rc_sin_pendientes = rc_reconciliacion or rc_consolidacion
            resumen["return_code_reconciliacion_global"] = rc_reconciliacion
            resumen["return_code_consolidacion_sin_operador"] = rc_consolidacion
            resumen["ok"] = rc_sin_pendientes == 0
            if args.solo_internos:
                base_mensaje = "No hay descargas pendientes de Internos."
                mensaje_notificacion = (
                    f"Se auditaron {len(objetivos_internos_inventario)} objetivos "
                    "bandeja+folio de Internos; todos conservan metadata completa "
                    "y sus archivos físicos."
                )
            elif vacio_confirmado_satys and not folios_internos_satys:
                base_mensaje = "SATyS confirmó cero registros."
                mensaje_notificacion = "SATyS confirmó cero registros en todos los años consultados."
            else:
                base_mensaje = "No hay registros nuevos."
                mensaje_notificacion = (
                    f"Se revisaron {len(registros_satys)} registros de Oficialia y "
                    f"{len(folios_internos_satys)} folios de Internos; "
                    "todos existen en TrámitesCRT.xlsx."
                )
            detalle_reconciliacion = (
                f"Reconciliación global de TrámitesCRT.xlsx: código {rc_reconciliacion}."
                if reconciliacion_habilitada
                else "Reconciliación global omitida en modo solo Internos."
            )
            resumen["mensaje"] = f"{base_mensaje} {detalle_reconciliacion}"
            notificar_windows(
                "SATyS CRT — sin registros nuevos",
                mensaje_notificacion,
                habilitado=not args.sin_notificacion,
            )
            estado.finalizar(
                ok=rc_sin_pendientes == 0,
                mensaje=resumen["mensaje"],
                total_registros_satys=len(registros_satys),
                total_nuevos=0,
                return_code_reconciliacion_global=rc_reconciliacion,
                return_code_consolidacion_sin_operador=rc_consolidacion,
            )
            if args.sin_email:
                print("\nℹ️  Correo diario deshabilitado por --sin-email.")
            else:
                correo_diario_intentado = True
                enviar_resumen_email_diario(
                    resultados=[],
                    log_path=log_path,
                    excel_path=args.excel,
                    modo="CORRIDA DIARIA — SIN PENDIENTES",
                    error_general=resumen["mensaje"] if rc_sin_pendientes else "",
                )
            sincronizar_estado_diario_depi()
            return rc_sin_pendientes

        # 4) Ejecutar main_procesar.py por Registro.
        if args.no_procesar:
            rc_reconciliacion = 0
            if reconciliacion_habilitada:
                rc_reconciliacion = ejecutar_reconciliacion_global(
                    python_exe=args.python_exe,
                    script=args.reconciliar_script,
                    excel=args.excel,
                    log_path=log_path,
                    estado=estado,
                )
            resumen["return_code_reconciliacion_global"] = rc_reconciliacion
            resumen["ok"] = rc_reconciliacion == 0
            alcance_listas = "Internos" if args.solo_internos else "Oficialia e Internos"
            resumen["mensaje"] = (
                f"Se generaron las listas de nuevos de {alcance_listas}, "
                "pero no se procesaron por --no-procesar. "
                + (
                    f"Reconciliación global: código {rc_reconciliacion}."
                    if reconciliacion_habilitada
                    else "Reconciliación global omitida en modo solo Internos."
                )
            )
            detalle_nuevos = (
                f"Internos: {len(objetivos_internos)} nuevo(s)."
                if args.solo_internos
                else f"Oficialia: {len(nuevos)} nuevo(s). Internos: {len(objetivos_internos)} nuevo(s)."
            )
            notificar_windows(
                "SATyS CRT — registros nuevos detectados",
                detalle_nuevos,
                habilitado=not args.sin_notificacion,
            )
            estado.finalizar(
                ok=rc_reconciliacion == 0,
                mensaje=resumen["mensaje"],
                total_nuevos=len(nuevos),
                no_procesar=True,
                return_code_reconciliacion_global=rc_reconciliacion,
            )
            # Un inventario con --no-procesar no creó descargas ni actualizó el
            # Excel. Evitamos copiar de nuevo todo descargas/ a la red, operación
            # que puede tardar horas y contradice el objetivo de una comprobación rápida.
            resumen["sincronizacion_depi_omitida"] = True
            resumen["motivo_sincronizacion_depi_omitida"] = "--no-procesar"
            print("ℹ️  Sincronización DEPI omitida: --no-procesar sólo generó el inventario.")
            if args.sin_email:
                print("\nℹ️  Correo diario deshabilitado por --sin-email.")
            else:
                correo_diario_intentado = True
                enviar_resumen_email_diario(
                    resultados=[],
                    log_path=log_path,
                    excel_path=args.excel,
                    modo="INVENTARIO DIARIO — SIN PROCESAR",
                    error_general=resumen["mensaje"] if rc_reconciliacion else "",
                )
            return rc_reconciliacion

        # Internos se procesa antes que Oficialía. En el servidor una corrida de
        # Oficialía puede durar varias horas; no debe volver a dejar sin atender
        # el inventario de las seis bandejas de Internos.
        resultados_email: list[dict] = []
        rc_main_internos = 0
        if objetivos_internos:
            cmd_main_internos = [
                str(args.python_exe),
                str(args.main_script),
                "--internos",
                "--internos-objetivos", str(args.internos_latest),
                "--internos-workers", str(args.internos_workers),
                "--timeout-registro", str(args.timeout_registro),
                "--reintentos-registro", str(args.reintentos_registro),
                "--sin-lock",
                "--sin-email",
                "--rpc-online",
            ]
            if headless:
                cmd_main_internos.append("--headless")
            if args.folio_internos:
                # La validación de un Folio termina en Excel y output locales.
                # Evitar el merge completo de miles de archivos a DEPI mantiene
                # esta revisión puntual rápida y evita una espera engañosa al final.
                cmd_main_internos.append("--sin-sincronizar")
            estado.actualizar(
                stage="procesando_folios_internos_nuevos",
                total_folios_internos_nuevos=len(objetivos_internos),
            )
            inicio_main_internos = time.time()
            rc_main_internos = ejecutar_comando(
                cmd_main_internos,
                PROJECT_DIR,
                log_path,
                "2) PROCESAR OBJETIVOS BANDEJA+FOLIO DE INTERNOS IFT",
                estado=estado,
                etapa="procesando_folios_internos_nuevos",
            )
            resultados_email.extend(cargar_resultados_procesamiento(
                ruta_configurada("descargas", "descargas")
                / "internos"
                / "procesamiento_log_internos.json",
                origen="internos",
                mtime_minimo=inicio_main_internos,
            ))
        resumen["return_code_main_internos"] = rc_main_internos

        if args.folio_internos and rc_main_internos == 0:
            procesamiento_log_folio = (
                DESCARGA_BASE_DIARIO / "internos" / "procesamiento_log_internos.json"
            )
            validacion_folio = validar_salidas_folio_internos(
                folio=args.folio_internos,
                objetivos=objetivos_internos,
                procesamiento_log=procesamiento_log_folio,
                excel_path=args.excel,
                sheet=args.sheet_internos,
                header_folio=args.header_folio_internos,
            )
            resumen["validacion_folio_internos"] = validacion_folio
            print("\n" + "=" * 90)
            print(f"SALIDAS VERIFICADAS DEL FOLIO {args.folio_internos}")
            print("=" * 90)
            print(f"Excel actualizado:             {validacion_folio['excel']}")
            for output_dir in validacion_folio["output_dirs"]:
                print(f"Carpeta organizada:            {output_dir}")
            print(f"Log de procesamiento:          {validacion_folio['procesamiento_log']}")
            if not validacion_folio["ok"]:
                for error in validacion_folio["errores"]:
                    print(f"ERROR VALIDACIÓN:              {error}")
                rc_main_internos = 3
                resumen["return_code_main_internos"] = rc_main_internos

        rc_main = 0
        inicio_main_registros: float | None = None
        if nuevos:
            cmd_main = [
                str(args.python_exe),
                str(args.main_script),
                "--archivo-registro", str(args.registros_latest),
                "--workers", str(args.workers),
                "--timeout-registro", str(args.timeout_registro),
                "--reintentos-registro", str(args.reintentos_registro),
                "--workers-reintento", str(args.workers_reintento),
                "--sin-lock",
                "--sin-email",
                "--rpc-online",
            ]
            if headless:
                cmd_main.append("--headless")

            estado.actualizar(stage="procesando_registros_nuevos", total_nuevos=len(nuevos))
            inicio_main_registros = time.time()
            rc_main = ejecutar_comando(
                cmd_main, PROJECT_DIR, log_path, "3) PROCESAR REGISTROS NUEVOS",
                estado=estado, etapa="procesando_registros_nuevos"
            )
            resultados_email.extend(cargar_resultados_procesamiento(
                ruta_configurada("descargas", "descargas") / "procesamiento_log_registros.json",
                origen="oficialia",
                mtime_minimo=inicio_main_registros,
            ))
        resumen["return_code_main"] = rc_main

        rc_reconciliacion = 0
        if reconciliacion_habilitada:
            rc_reconciliacion = ejecutar_reconciliacion_global(
                python_exe=args.python_exe,
                script=args.reconciliar_script,
                excel=args.excel,
                log_path=log_path,
                estado=estado,
                sin_backup=(rc_main == 0 and rc_main_internos == 0),
            )
        resumen["return_code_reconciliacion_global"] = rc_reconciliacion
        consolidacion_revision = consolidar_revision_manual_final()
        resumen["consolidacion_sin_operador"] = consolidacion_revision
        rc_consolidacion = 4 if consolidacion_revision["errores"] else 0
        resumen["return_code_consolidacion_sin_operador"] = rc_consolidacion
        rc_final = rc_main or rc_main_internos or rc_reconciliacion or rc_consolidacion

        fallidos_latest = PROJECT_DIR / "registros_fallidos" / "registros_fallidos_latest.txt"
        fallidos = []
        if (
            nuevos
            and inicio_main_registros is not None
            and fallidos_latest.is_file()
            and fallidos_latest.stat().st_mtime >= inicio_main_registros - 2.0
        ):
            fallidos = leer_registros_txt(fallidos_latest)
        resumen["registros_fallidos_controlados"] = fallidos
        resumen["total_fallidos_controlados"] = len(fallidos)
        resultados_email = agregar_fallidos_controlados(resultados_email, fallidos)
        if rc_main_internos:
            resultados_email = agregar_fallidos_controlados(
                resultados_email,
                [item["folio"] for item in objetivos_internos],
                origen="internos",
            )
        conteos_email = (
            _email_mod.conteos_desde_resultados(resultados_email)
            if _EMAIL_DISPONIBLE
            else {"total": len(resultados_email)}
        )
        resumen["resumen_email_consolidado"] = conteos_email
        resumen["ok"] = rc_final == 0
        if args.solo_internos:
            resumen["mensaje"] = (
                f"Procesados {len(objetivos_internos)} objetivo(s) bandeja+folio de Internos. "
                f"Código main Internos={rc_main_internos}. "
                f"Fallidos controlados: {len(fallidos)}."
            )
        else:
            resumen["mensaje"] = (
                f"Procesados {len(nuevos)} registro(s) de Oficialia y "
                f"{len(objetivos_internos)} folio(s) de Internos. "
                f"Codigos main: Oficialia={rc_main}, Internos={rc_main_internos}. "
                f"Reconciliación global: {rc_reconciliacion}. Fallidos controlados: {len(fallidos)}."
            )

        detalle_final = (
            f"Internos: {len(objetivos_internos)} | Fallidos: {len(fallidos)} | "
            f"Código main: {rc_main_internos} | Log: {log_path.name}"
            if args.solo_internos
            else f"Oficialia: {len(nuevos)} | Internos: {len(objetivos_internos)} | "
                 f"Fallidos: {len(fallidos)} | Codigos main: {rc_main}/{rc_main_internos} | Log: {log_path.name}"
        )
        notificar_windows(
            "SATyS CRT — proceso diario finalizado",
            detalle_final,
            habilitado=not args.sin_notificacion,
        )

        # Los dos main_procesar.py siempre reciben --sin-email. Únicamente este
        # proceso padre combina Internos + Oficialía y envía un correo final.
        if args.sin_email:
            print("\nℹ️  Correo diario consolidado deshabilitado por --sin-email.")
        else:
            correo_diario_intentado = True
            enviar_resumen_email_diario(
                resultados=resultados_email,
                log_path=log_path,
                excel_path=args.excel,
                error_general=resumen["mensaje"] if rc_final else "",
            )

        estado.finalizar(
            ok=rc_final == 0,
            mensaje=resumen["mensaje"],
            total_nuevos=len(nuevos),
            total_folios_internos_nuevos=len(objetivos_internos),
            total_fallidos_controlados=len(fallidos),
            return_code_main=rc_main,
            return_code_main_internos=rc_main_internos,
            return_code_reconciliacion_global=rc_reconciliacion,
            return_code_consolidacion_sin_operador=rc_consolidacion,
        )
        # ─────────────────────────────────────────────────────────────────────
        if args.folio_internos:
            resumen["sincronizacion_depi_omitida"] = True
            resumen["motivo_sincronizacion_depi_omitida"] = "--folio-internos"
            print("ℹ️  Sincronización DEPI omitida: la revisión de un Folio conserva salidas locales.")
        else:
            sincronizar_estado_diario_depi()

        return rc_final

    except Exception as exc:
        resumen["ok"] = False
        resumen["errores"].append(str(exc))
        resumen["traceback"] = traceback.format_exc()
        print("\nERROR EN MONITOR DIARIO:", exc)
        print(resumen["traceback"])
        notificar_windows(
            "SATyS CRT — error en monitor diario",
            str(exc),
            habilitado=not args.sin_notificacion,
        )
        # Único correo de aviso de fallo; los subprocesos siempre están silenciados.
        if _EMAIL_DISPONIBLE and not args.sin_email and not correo_diario_intentado:
            correo_diario_intentado = True
            enviar_resumen_email_diario(
                resultados=[],
                log_path=log_path,
                excel_path=args.excel,
                modo="MONITOR DIARIO",
                error_general=str(exc),
            )
        estado.finalizar(ok=False, mensaje=str(exc), errores=resumen.get("errores", []), traceback=resumen.get("traceback", ""))
        return 1

    finally:
        try:
            lock.liberar()
            print("🔓 Lock compartido liberado al finalizar automatizar_registros_diario.py.")
        except Exception:
            pass
        try:
            resumen["fecha_fin"] = datetime.now().isoformat()
            resumen_json.write_text(json.dumps(resumen, ensure_ascii=False, indent=2), encoding="utf-8")
            resumen_latest.write_text(json.dumps(resumen, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"\nResumen JSON: {resumen_json}")
            print(f"Último resumen: {resumen_latest}")
            print(f"Log completo: {log_path}")
        except Exception:
            pass


if __name__ == "__main__":
    raise SystemExit(main())
